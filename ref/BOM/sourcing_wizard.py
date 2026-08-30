import tkinter as tk
from tkinter import ttk, StringVar, IntVar
from utils import SemanticMessageBox as messagebox
from tkinter import Frame, Label, Button, Entry, Checkbutton, Radiobutton, Listbox, Scrollbar, Canvas, W, E, EW, MULTIPLE, LabelFrame
from tkinter.ttk import Combobox, Treeview
from utils import show_info, show_error, MASTER_DATA_DIR
import os
from dialogs import BasePanel, MultiColumnSelectionDialog, CategoryInputDialog, SourcingCancelWarningDialog, apply_panel_theme
import json

def format_to_sig_figs(val, sig_figs=2):
    import math
    if val is None or str(val).strip() == "":
        return ""
    try:
        f_val = float(val)
        if f_val == 0.0:
            return "0"
        if f_val == int(f_val):
            return str(int(f_val))
        dec = -int(math.floor(math.log10(abs(f_val))))
        places = max(0, dec + sig_figs - 1)
        fmt = f"{{:.{places}f}}"
        res = fmt.format(f_val)
        if "." in res:
            res = res.rstrip('0').rstrip('.')
        return res
    except:
        return str(val)

def apply_uom_conversion_to_row(row_uom, row_qty):
    """Applies UOM conversion rule from uom_conversions.json if row_uom matches a known rule."""
    try:
        from utils import load_uom_conversions
        from bomformatter import round_up_to_2_sig_figs
        uom_config = load_uom_conversions()
        tolerance_pct = uom_config.get("tolerance_pct", 5.0)
        rules = uom_config.get("rules", {})
        upper_rules = {k.strip().upper(): v for k, v in rules.items()}

        uom_str = str(row_uom).strip().upper()
        if uom_str in upper_rules:
            rule = upper_rules[uom_str]
            to_uom = rule.get("to_uom", uom_str)
            factor = float(rule.get("factor", 1.0))
            apply_tolerance = rule.get("apply_tolerance", False)

            try:
                qty_val = float(row_qty)
            except (ValueError, TypeError):
                qty_val = 0.0

            new_qty = qty_val * factor
            if apply_tolerance:
                rule_tol_pct = float(rule.get("tolerance_pct", tolerance_pct))
                new_qty = new_qty * (1.0 + rule_tol_pct / 100.0)

            return to_uom, round_up_to_2_sig_figs(new_qty)
    except Exception as e:
        print(f"Error applying UOM conversion: {e}")
    return row_uom, row_qty

def _get_shared_commodities():

    default_groups = ["BoxBuild", "FIBER Optic", "Module", "PCBA", "Wire Harness"]
    try:
        import configparser, json
        possible_paths = [
            r"D:\RadysisAsia MockServer\Costing\AppData\Quotation_Data\Master Data\master_data.json",
            r"D:\RadysisAsia MockServer\BOM\AppData\Quotation_Data\Master Data\master_data.json",
        ]
        curr_dir = os.path.dirname(os.path.abspath(__file__))
        for script_dir in [curr_dir, os.path.join(os.path.dirname(curr_dir), "Costing")]:
            config_path = os.path.join(script_dir, "config.ini")
            if os.path.exists(config_path):
                try:
                    config = configparser.ConfigParser()
                    config.read(config_path, encoding='utf-8')
                    if 'PATHS' in config:
                        for k, v in config['PATHS'].items():
                            if 'path' in k.lower():
                                possible_paths.append(os.path.normpath(os.path.join(v, "Quotation_Data", "Master Data", "master_data.json")))
                                possible_paths.append(os.path.normpath(os.path.join(v, "Master Data", "master_data.json")))
                except: pass

        for fp in possible_paths:
            if os.path.exists(fp):
                try:
                    with open(fp, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    pg_config = data.get('process_groups_config', {})
                    if pg_config:
                        return sorted(list(pg_config.keys()))
                except: pass
    except: pass
    return default_groups

class CombinedMappingPanel(BasePanel):

    def __init__(self, master, actual_headers, special_columns, default_cust_name, standard_columns, mandatory_columns, special_columns_to_skip, multi_source_columns, initial_mapping=None, initial_special=None, initial_rfq="", initial_email=""):
        super().__init__(master)
        
        self.actual_headers = actual_headers
        self.special_columns = special_columns
        self.default_cust_name = default_cust_name
        self.standard_columns = standard_columns
        self.mandatory_columns = mandatory_columns
        self.special_columns_to_skip = special_columns_to_skip
        self.multi_source_columns = multi_source_columns
        self.initial_mapping = initial_mapping
        self.initial_special = initial_special or {}
        self.initial_rfq = initial_rfq
        self.initial_email = initial_email
        
        self.special_results = {}
        self.mapping_vars = {}
        self.multi_mapping_labels = {}
        self.result_mapping = {}
        
        # Instantiate inner frames
        self.special_frame = tk.Frame(self, bg="#EBF8FF")
        self.special_frame.pack(fill="x", pady=10)
        
        self.map_frame = tk.Frame(self, bg="#EBF8FF")
        self.map_frame.pack(fill="both", expand=True)

        self._init_special()
        self._init_mapper()
        
        # Bottom Buttons
        btn_f = tk.Frame(self, bg="#EBF8FF")
        btn_f.pack(fill="x", side="bottom")
        tk.Button(btn_f, text="Cancel Workflow", command=self._on_cancel, width=20, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2").pack(side="left", padx=10, pady=10)
        tk.Button(btn_f, text="Confirm Special & Mapping", command=self._on_combined_confirm, width=30, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2").pack(side="right", padx=10, pady=10)

    def _on_cancel(self):
        dialog = SourcingCancelWarningDialog(self.winfo_toplevel(), msg_type="verification")
        self.winfo_toplevel().wait_window(dialog)
        if dialog.result:
            self.result = None
            self._wait_var.set(1)

    def _init_special(self):
        self.column_method_vars = {col: StringVar(self, value="map_column") for col in self.special_columns}
        self.column_selection_vars = {col: StringVar(self) for col in self.special_columns}
        self.fixed_value_entry_vars = {col: StringVar(self) for col in self.special_columns}
        self.cust_name_var = StringVar(self, value=self.default_cust_name)
        self.commodity_var = StringVar(self, value=getattr(self, 'initial_commodity', 'Wire Harness') or 'Wire Harness')
        self.rfq_id_var = StringVar(self, value=self.initial_rfq)
        self.email_subject_var = StringVar(self, value=self.initial_email)
        self.input_widgets = {}
        

        for col in self.special_columns:
            if col in self.initial_special:
                spec = self.initial_special[col]
                if spec['method'] == 'fixed':
                    self.column_method_vars[col].set('fixed_value')
                    self.fixed_value_entry_vars[col].set(spec['value'])
                else:
                    self.column_method_vars[col].set('map_column')
                    self.column_selection_vars[col].set(spec['source_column'])

        main_frame = tk.LabelFrame(self.special_frame, text="Select Special Columns & Customer Info", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 9, "bold"), padx=10, pady=5)
        main_frame.pack(fill="x", expand=True, padx=10)

        process_groups = _get_shared_commodities()


        Label(main_frame, text="Customer Name:", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10)).grid(row=0, column=0, sticky=W, pady=2)
        cust_en = Entry(main_frame, textvariable=self.cust_name_var, width=40)
        cust_en.grid(row=0, column=1, sticky=W, padx=5, pady=2)

        Label(main_frame, text="Commodity:", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10)).grid(row=1, column=0, sticky=W, pady=2)
        comm_cb = Combobox(main_frame, textvariable=self.commodity_var, values=process_groups, state="readonly", width=38)
        comm_cb.grid(row=1, column=1, sticky=W, padx=5, pady=2)

        Label(main_frame, text="RFQ Number:", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10)).grid(row=2, column=0, sticky=W, pady=2)
        rfq_en = Entry(main_frame, textvariable=self.rfq_id_var, width=40)
        rfq_en.grid(row=2, column=1, sticky=W, padx=5, pady=2)

        Label(main_frame, text="Project Title:", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10)).grid(row=3, column=0, sticky=W, pady=2)
        email_en = Entry(main_frame, textvariable=self.email_subject_var, width=40)
        email_en.grid(row=3, column=1, sticky=W, padx=5, pady=2)
        
        cust_en.bind("<Down>", lambda e: comm_cb.focus_set())
        comm_cb.bind("<Up>", lambda e: cust_en.focus_set())
        comm_cb.bind("<Down>", lambda e: rfq_en.focus_set())
        rfq_en.bind("<Up>", lambda e: comm_cb.focus_set())
        rfq_en.bind("<Down>", lambda e: email_en.focus_set())
        email_en.bind("<Up>", lambda e: rfq_en.focus_set())

        row_num = 4
        combobox_options = [""] + sorted(list(set(self.actual_headers)))

        for col_name in self.special_columns:
            Label(main_frame, text=f"For '{col_name}':", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D").grid(row=row_num, column=0, sticky=W, pady=(5, 0))
            
            f = tk.Frame(main_frame, bg="#EBF8FF")
            f.grid(row=row_num, column=1, columnspan=3, sticky="w", pady=(5,0))
            
            rb_map = Radiobutton(f, text="Map", variable=self.column_method_vars[col_name], value="map_column", command=lambda c=col_name: self._on_method_select(c), bg="#EBF8FF", fg="#1A365D", activebackground="#EBF8FF", font=("Segoe UI", 9))
            rb_map.pack(side="left")
            
            col_combobox = Combobox(f, textvariable=self.column_selection_vars[col_name], values=combobox_options, state="readonly", width=25)
            col_combobox.pack(side="left", padx=5)
            self.input_widgets[f'{col_name}_combobox'] = col_combobox
            
            rb_fixed = Radiobutton(f, text="Enter a fixed value", variable=self.column_method_vars[col_name], value="fixed_value", command=lambda c=col_name: self._on_method_select(c), bg="#EBF8FF", fg="#1A365D", activebackground="#EBF8FF", font=("Segoe UI", 9))
            rb_fixed.pack(side="left", padx=10)

            fixed_entry = Entry(f, textvariable=self.fixed_value_entry_vars[col_name], width=20)
            fixed_entry.pack(side="left")
            self.input_widgets[f'{col_name}_entry'] = fixed_entry
            
            row_num += 1

        for col in self.special_columns:
            self._on_method_select(col)

    def _init_mapper(self):
        row_num = 0
        main_frame = tk.LabelFrame(self.map_frame, text="Map Excel Columns", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 9, "bold"), padx=10, pady=5)
        main_frame.pack(fill="both", expand=True, padx=10)

        combobox_options = [""] + sorted(list(set(self.actual_headers)))

        for col_name in self.standard_columns:
            if col_name in self.special_columns_to_skip:
                continue

            Label(main_frame, text=f"'{col_name}':", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10)).grid(row=row_num, column=0, padx=5, pady=2, sticky="w")

            if col_name in self.multi_source_columns:
                current_selection_label = Label(main_frame, text="No sources selected", wraplength=200, bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10))
                current_selection_label.grid(row=row_num, column=1, padx=5, pady=2, sticky="w")
                self.multi_mapping_labels[col_name] = current_selection_label
                
                initial_sel = []
                if self.initial_mapping and col_name in self.initial_mapping:
                    raw_val = self.initial_mapping[col_name]
                    if isinstance(raw_val, list):
                        initial_sel = [str(x) for x in raw_val]
                    elif isinstance(raw_val, str) and raw_val:
                        initial_sel = [raw_val]
                    current_selection_label.config(text=", ".join(initial_sel) if initial_sel else "No sources selected")
                self.result_mapping[col_name] = initial_sel

                select_btn = Button(main_frame, text="Select Sources", bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2", command=lambda c=col_name: self._open_multi_selection_dialog(c))
                select_btn.grid(row=row_num, column=2, padx=5, pady=2, sticky="w")
            else:
                var = StringVar(self)
                initial_selection = ""
                if self.initial_mapping:
                    for act_h, std_col in self.initial_mapping.items():
                        if std_col == col_name:
                            initial_selection = act_h
                            break

                if initial_selection == "" and not self.initial_mapping:
                    for actual_h in self.actual_headers:
                        if actual_h.lower() == col_name.lower():
                            initial_selection = actual_h
                            break
                            
                var.set(initial_selection)
                combobox = Combobox(main_frame, textvariable=var, values=combobox_options, state="readonly", width=40)
                combobox.grid(row=row_num, column=1, columnspan=2, padx=5, pady=2, sticky="w")
                self.mapping_vars[col_name] = var

            row_num += 1

        # Add a highly visible, premium warning notice at the bottom of the map frame
        warning_frame = Frame(main_frame, bg="#FFF5F5", bd=1, relief="solid")
        warning_frame.grid(row=row_num, column=0, columnspan=3, padx=5, pady=(15, 5), sticky="we")
        
        warning_label = Label(
            warning_frame,
            text="⚠️ CRITICAL MAPPING NOTICE: PLEASE ASSIGN ALL COLUMNS ACCURATELY!\n"
                 "• Ensure standard 'MPN' is mapped to your Excel column containing Manufacturer Part Numbers (e.g. GRM31CR61C475KA01L).\n"
                 "• Ensure standard 'MFR' is mapped to your Excel column containing Manufacturer Names (e.g. MURATA, AVX).\n"
                 "• Reversing or mis-assigning these columns will prevent the system from matching any sourcing quotes,\n"
                 "  or will cause serious calculation errors during BOM generation.",
            font=("Segoe UI", 9, "bold"),
            fg="#C53030",  # Premium warning crimson
            bg="#FFF5F5",
            justify="left",
            anchor="w",
            padx=12,
            pady=12
        )
        warning_label.pack(fill="both", expand=True)

    def _on_method_select(self, col_name):
        current_method = self.column_method_vars[col_name].get()
        combobox = self.input_widgets[f'{col_name}_combobox']
        entry = self.input_widgets[f'{col_name}_entry']
        if current_method == "map_column":
            combobox.config(state="readonly")
            entry.config(state="disabled")
            self.fixed_value_entry_vars[col_name].set("")
        else:
            combobox.config(state="disabled")
            self.column_selection_vars[col_name].set("")
            entry.config(state="normal")

    def _open_multi_selection_dialog(self, col_name):
        initial_selections = self.result_mapping.get(col_name, [])
        multi_dialog = MultiColumnSelectionDialog(self.winfo_toplevel(), self.actual_headers, col_name, initial_selections)
        selected_sources = multi_dialog.get_selection()
        if selected_sources is not None:
            self.result_mapping[col_name] = selected_sources
            display_text = ", ".join(selected_sources) if selected_sources else "No sources selected"
            self.multi_mapping_labels[col_name].config(text=display_text)

    def _on_combined_confirm(self):
        missing_fields = []
        
        # 1. Validate Special Columns & Customer Info
        self.cust_name = self.cust_name_var.get().strip()
        self.commodity = getattr(self, 'commodity_var', None).get().strip() if hasattr(self, 'commodity_var') else "Wire Harness"
        self.RFQ_ID = self.rfq_id_var.get().strip()
        self.email_subject = self.email_subject_var.get().strip()
        
        if not self.cust_name:
            missing_fields.append("Customer Name")
        if not self.commodity:
            missing_fields.append("Commodity")
        if not self.RFQ_ID:
            missing_fields.append("RFQ Number")
        if not self.email_subject:
            missing_fields.append("Email Subject")

        for col_name in self.special_columns:
            method = self.column_method_vars[col_name].get()
            if method == "map_column":
                selected_column = self.column_selection_vars[col_name].get().strip()
                if not selected_column:
                    missing_fields.append(f"'{col_name}'")
                else:
                    self.special_results[col_name] = {'method': 'map', 'source_column': selected_column}
            else:
                fixed_val = self.fixed_value_entry_vars[col_name].get().strip()
                if not fixed_val:
                    missing_fields.append(f"'{col_name}'")
                else:
                    self.special_results[col_name] = {'method': 'fixed', 'value': fixed_val}

        # 2. Validate Mapping
        final_mapping = {}
        all_selected_actual_headers = set()

        for standard_col, var in self.mapping_vars.items():
            actual_header = var.get().strip()
            if actual_header:
                if actual_header in all_selected_actual_headers:
                    show_error("Duplicate Selection", f"The Excel header '{actual_header}' has been mapped to more than one standard column.", parent=self)
                    return
                final_mapping[actual_header] = standard_col
                all_selected_actual_headers.add(actual_header)
            elif standard_col in self.mandatory_columns:
                missing_fields.append(f"'{standard_col}'")

        for standard_col, actual_headers_list in self.result_mapping.items():
            if actual_headers_list:
                final_mapping[standard_col] = actual_headers_list
            elif standard_col in self.mandatory_columns:
                missing_fields.append(f"'{standard_col}'")

        if missing_fields:
            missing_str = "\n".join(f"• {field}" for field in missing_fields)
            messagebox.showwarning("Missing Selection", f"Please fill in or map the following mandatory fields:\n\n{missing_str}", parent=self)
            return

        # 3. Validate MFR and MPN Count Match
        mpn_sources = self.result_mapping.get('MPN', [])
        mfr_sources = self.result_mapping.get('MFR', [])
        if len(mpn_sources) != len(mfr_sources):
            messagebox.showwarning("Column Count Mismatch", f"The number of MPN columns ({len(mpn_sources)}) must match the number of MFR columns ({len(mfr_sources)}).\n\nPlease ensure you have mapped an equal number of sources for both.", parent=self)
            return

        # 4. Check if RFQ already exists in system
        init_rfq_norm = str(getattr(self, 'initial_rfq', '') or '').strip().lower()
        if not init_rfq_norm or self.RFQ_ID.lower() != init_rfq_norm:
            from utils import check_rfq_exists
            existing_cust = check_rfq_exists(self.RFQ_ID)
            if existing_cust:
                messagebox.showwarning(
                    "RFQ Number Already Exists",
                    f"The RFQ Number '{self.RFQ_ID}' has already been used in the system under Customer '{existing_cust}'.\n\nPlease specify a new, unique RFQ number to proceed.",
                    parent=self
                )
                return

        self.result = (self.special_results, self.cust_name, self.RFQ_ID, self.commodity, self.email_subject, final_mapping)
        self._wait_var.set(1)

class BOMHistoryDialog(tk.Toplevel):
    def __init__(self, parent, raw_data, title="Display Changes - BOM Verification"):
        super().__init__(parent)
        self._skip_autofit = True
        self.dialog_title = title
        self.title(title)
        self.geometry("1200x700")
        self.resizable(True, True)
        self.minsize(900, 450)
        self.configure(bg="#EBF8FF")
        self.raw_data = raw_data

        # Center on screen in normal/restored mode (1200x700)
        try:
            self.update_idletasks()
            sw = self.winfo_screenwidth()
            sh = self.winfo_screenheight()
            x = max(0, (sw - 1200) // 2)
            y = max(0, (sh - 700) // 2)
            self.geometry(f"1200x700+{x}+{y}")
        except Exception:
            pass
        
        self._build_ui()
        
    def _build_ui(self):
        import tkinter as tk
        from tkinter import ttk
        import textwrap
        
        # Header banner
        header_frame = tk.Frame(self, bg="#dcedf5", height=50)
        header_frame.pack(fill="x")
        tk.Label(header_frame, text=f"📋  {self.dialog_title}",
                 font=("Segoe UI", 14, "bold", "italic"), bg="#dcedf5", fg="#1A365D").pack(side="left", padx=15, pady=10)
        
        # Info subheader row
        info_frame = tk.Frame(self, bg="#EBF8FF")
        info_frame.pack(fill="x", padx=15, pady=(8, 2))
        cust = self.raw_data.get("Customer", "")
        rfq = self.raw_data.get("RFQ", "")
        tk.Label(
            info_frame,
            text=f"Customer: {cust}   |   RFQ: {rfq}",
            font=("Segoe UI", 10, "bold"),
            fg="#1A365D", bg="#EBF8FF"
        ).pack(anchor="w")

        # Configure treeview style with larger rowheight to support text wrapping & unlock tag colors
        style = ttk.Style(self)
        style.configure("History.Treeview", font=("Segoe UI", 10), rowheight=38)
        style.configure("History.Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#dcedf5", foreground="#1A365D")
        style.map("History.Treeview", foreground=[("selected", "white")])

        tree_frame = tk.Frame(self, bg="#EBF8FF")
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(6, 5))
        
        cols = ("Date", "Time", "Changed By", "Field Name", "New Value", "Old Value")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", style="History.Treeview", selectmode="none")
        
        col_widths = {
            "Date": 110,
            "Time": 90,
            "Changed By": 170,
            "Field Name": 350,
            "New Value": 240,
            "Old Value": 240
        }
        for col in cols:
            self.tree.heading(col, text=col, anchor="w")
            self.tree.column(col, width=col_widths.get(col, 150), stretch=True, anchor="w")
            
        y_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        x_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        self.tree.pack(side="left", fill="both", expand=True)
        y_scroll.pack(side="right", fill="y")
        
        self.tree.tag_configure("oddrow", background="#ffffff")
        self.tree.tag_configure("evenrow", background="#f0f4f8")
        self.tree.tag_configure("separator", background="#dcdcdc")
        
        history = self.raw_data.get("history", [])
        
        # Filter history records based on dialog title context
        is_moq_dialog = "MOQ Assignation" in self.dialog_title
        is_verification_dialog = "BOM Verification" in self.dialog_title

        filtered_history = []
        for h in history:
            field_name = str(h.get("Field Name", ""))
            # Never include dispatch records in Display Changes
            if "dispatch" in field_name.lower() or field_name == "Stage Dispatch":
                continue
            is_moq_record = ("MOQ" in field_name) or (field_name == "Assigned MOQs")
            is_tp_eau_record = ("Target Price" in field_name) or ("EAU" in field_name)
            
            if is_moq_dialog:
                if is_moq_record:
                    filtered_history.append(h)
            elif is_verification_dialog:
                if not is_moq_record and not is_tp_eau_record:
                    filtered_history.append(h)
            else:
                filtered_history.append(h)

        def get_sort_key(h):
            d_str = str(h.get("Date", "")).strip()
            t_str = str(h.get("Time", "")).strip()
            dt_str = f"{d_str} {t_str}".strip()
            from datetime import datetime
            try:
                return datetime.strptime(dt_str, "%d.%m.%Y %H:%M:%S")
            except Exception:
                try:
                    return datetime.strptime(d_str, "%d.%m.%Y")
                except Exception:
                    return dt_str

        # Sort newest first
        try:
            filtered_history.sort(key=get_sort_key, reverse=True)
        except Exception:
            filtered_history = list(reversed(filtered_history))

        current_date = None
        row_idx = 0
        for h in filtered_history:
            h_date = h.get("Date", "")
            if current_date != h_date:
                if current_date is not None:
                    self.tree.insert("", "end", values=("", "", "", "", "", ""), tags=("separator",))
                current_date = h_date
                
            row_tag = "evenrow" if row_idx % 2 == 0 else "oddrow"
            user_text = str(h.get("Changed By", "Unknown"))
            field_wrapped = textwrap.fill(str(h.get("Field Name", "")), width=45)
            new_val_wrapped = textwrap.fill(str(h.get("New Value", "")), width=30)
            old_val_wrapped = textwrap.fill(str(h.get("Old Value", "")), width=30)

            self.tree.insert("", "end", values=(
                h.get("Date", ""),
                h.get("Time", ""),
                user_text,
                field_wrapped,
                new_val_wrapped,
                old_val_wrapped
            ), tags=(row_tag,))
            row_idx += 1

        # Footer
        footer = tk.Frame(self, bg="#EBF8FF", pady=10)
        footer.pack(fill="x", side="bottom")
        tk.Label(
            footer,
            text=f"Total records: {len(filtered_history)}",
            font=("Segoe UI", 9),
            fg="#4a4a4a", bg="#EBF8FF"
        ).pack(side="left", padx=15)

        tk.Button(footer, text="Close", command=self.destroy, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), width=15, pady=6, bd=0, relief="flat", cursor="hand2").pack(side="right", padx=15)

from dialogs import CategoryInputDialog
class AssemblyMOQPanel(BasePanel):

    def __init__(self, master, unique_assemblies, initial_global_moqs=None, initial_assembly_moqs=None, title="", raw_data=None, current_user="Admin", read_only=False, **kwargs):
        super().__init__(master)
        self.unique_assemblies = list(unique_assemblies)
        self.raw_data = raw_data if raw_data else {}
        self.current_user = current_user
        self.read_only = read_only

        # Load Global MOQs from initial_global_moqs or raw_data
        if initial_global_moqs:
            self.global_moqs = [int(x) for x in initial_global_moqs]
        elif self.raw_data.get("Global MOQs"):
            self.global_moqs = [int(x) for x in self.raw_data.get("Global MOQs", [])]
        else:
            self.global_moqs = []

        # Ensure all assemblies in unique_assemblies are populated in assembly_moqs
        base_moqs = initial_assembly_moqs if initial_assembly_moqs else {}
        self.assembly_moqs = {}
        for assy in self.unique_assemblies:
            if assy in base_moqs:
                self.assembly_moqs[assy] = list(base_moqs[assy])
            elif self.global_moqs:
                self.assembly_moqs[assy] = list(self.global_moqs)
            else:
                self.assembly_moqs[assy] = []
        self.title_str = title
        
        # Track whether each assembly MOQ was set as Custom vs Global
        self.assembly_is_custom = {}
        
        saved_moq_types = {}
        for a_item in self.raw_data.get("Assemblies", []):
            if a_item.get("Assy #") and a_item.get("MOQ Type"):
                saved_moq_types[a_item.get("Assy #")] = a_item.get("MOQ Type")

        for assy in self.unique_assemblies:
            moqs = self.assembly_moqs.get(assy, [])
            if assy in saved_moq_types:
                self.assembly_is_custom[assy] = (saved_moq_types[assy] == "Custom")
            elif not moqs:
                self.assembly_is_custom[assy] = False
            elif self.global_moqs and sorted(moqs) == sorted(self.global_moqs):
                self.assembly_is_custom[assy] = False
            else:
                self.assembly_is_custom[assy] = bool(moqs)
        
        self.current_page = 0
        self.page_size = 100
        
        import copy
        self.original_assembly_moqs = copy.deepcopy(self.assembly_moqs)
        self.global_moq_lbl_var = StringVar(value="Current Global MOQs: " + (self.format_moq_str(self.global_moqs) if self.global_moqs else "- None -"))
        
        import tkinter as tk
        self._wait_var = tk.IntVar()
        self._create_widgets()
        
        self.status_bar = Label(self, text="", font=("Arial", 10, "bold"), anchor="center", pady=5)
        self.status_bar.pack(side="bottom", fill="x")
        
        self._populate_tree()
        
    def _on_cancel_moq(self):
        any_assigned = any(len(moqs) > 0 for moqs in self.assembly_moqs.values())
        has_amended = False
        for assy, moqs in self.assembly_moqs.items():
            orig_moqs = self.original_assembly_moqs.get(assy, [])
            if sorted(moqs) != sorted(orig_moqs):
                has_amended = True
                break
                
        if has_amended:
            # User amended MOQ, show "Exit Without Saving"
            dialog = SourcingCancelWarningDialog(self.winfo_toplevel(), msg_type="moq")
            self.winfo_toplevel().wait_window(dialog)
            if dialog.result:
                self.result = "CANCEL"
                self._wait_var.set(1)
        elif not any_assigned:
            # No MOQs assigned at all, show "No MOQ Assigned" warning
            dialog = SourcingCancelWarningDialog(self.winfo_toplevel(), msg_type="no_moq")
            self.winfo_toplevel().wait_window(dialog)
            if dialog.result:
                self.result = "CANCEL"
                self._wait_var.set(1)
        else:
            # No changes made, but MOQs are already assigned, exit directly without warning!
            self.result = "CANCEL"
            self._wait_var.set(1)

    def _create_widgets(self):
        # Apply premium background to self panel
        self.configure(bg="#EBF8FF")

        main_frame = Frame(self, padx=15, pady=15, bg="#EBF8FF")
        main_frame.pack(fill="both", expand=True)

        # Pack bottom buttons first so they are always visible!
        btn_frame = Frame(main_frame, bg="#EBF8FF")
        btn_frame.pack(fill="x", side="bottom", pady=(10, 0))
        
        # Standard buttons following the UI Standards Template
        cancel_text = "Close" if self.read_only else "Cancel"
        Button(btn_frame, text=cancel_text, command=self._on_cancel_moq, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", width=15, pady=6).pack(side="left", padx=5)
        
        if self.read_only:
            Button(btn_frame, text="◄ Previous", command=self._on_back, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", width=15, pady=6).pack(side="left", padx=5)
        else:
            Button(btn_frame, text="Apply Global MOQs to Selected", command=self._apply_global_to_selected, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", pady=6).pack(side="left", padx=5)
            Button(btn_frame, text="Set Custom MOQs for Selected", command=self._set_custom_to_selected, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", pady=6).pack(side="left", padx=5)
        
        confirm_text = "Next ➔" if self.read_only else "💾 Confirm"
        Button(btn_frame, text=confirm_text, bg="#2ead4e", fg="white", font=("Segoe UI", 10, "bold"), command=self._on_confirm, width=15 if self.read_only else 18, bd=0, relief="flat", cursor="hand2", pady=6).pack(side="right", padx=5)
        Button(btn_frame, text="View History", command=self._on_display_changes, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), width=15, bd=0, relief="flat", cursor="hand2", pady=6).pack(side="right", padx=5)

        if self.title_str:
            title_lbl = Label(main_frame, text=self.title_str, font=("Segoe UI", 16, "bold"), fg="#1a365d", bg="#EBF8FF", pady=5)
            title_lbl.pack(fill="x", side="top", pady=(0, 15))

        cust_name = self.raw_data.get("Customer", "")
        rfq_num = self.raw_data.get("RFQ", "")
        proj_title = self.raw_data.get("description", "") or self.raw_data.get("project_title", "") or self.raw_data.get("email_subject", "")
        from utils import get_bom_creation_date
        created_at = get_bom_creation_date(self.raw_data)
        if cust_name or rfq_num:
            info_frame = Frame(main_frame, bg="#dcedf5", relief="ridge", borderwidth=1)
            info_frame.pack(fill="x", side="top", pady=(0, 10))
            if rfq_num:
                Label(info_frame, text="RFQ:", font=('Segoe UI', 10, 'bold'), bg="#dcedf5", fg="#1A365D").pack(side="left", padx=(15, 2), pady=8)
                Label(info_frame, text=f"{rfq_num}", font=('Segoe UI', 10, 'bold'), bg="#dcedf5", fg="#cc0000").pack(side="left", padx=(0, 15), pady=8)
            if cust_name:
                Label(info_frame, text="Customer:", font=('Segoe UI', 10, 'bold'), bg="#dcedf5", fg="#1A365D").pack(side="left", padx=(0, 2), pady=8)
                Label(info_frame, text=f"{cust_name}", font=('Segoe UI', 10), bg="#dcedf5", fg="#1A365D").pack(side="left", padx=(0, 15), pady=8)
            Label(info_frame, text="Project / Email Subject:", font=('Segoe UI', 10, 'bold'), bg="#dcedf5", fg="#1A365D").pack(side="left", padx=(0, 2), pady=8)
            Label(info_frame, text=f"{proj_title or '-'}", font=('Segoe UI', 10), bg="#dcedf5", fg="#1A365D").pack(side="left", padx=(0, 15), pady=8)
            Label(info_frame, text="BOM Created:", font=('Segoe UI', 10, 'bold'), bg="#dcedf5", fg="#1A365D").pack(side="left", padx=(0, 2), pady=8)
            Label(info_frame, text=f"{created_at}", font=('Segoe UI', 10), bg="#dcedf5", fg="#1A365D").pack(side="left", padx=(0, 15), pady=8)

        # Pack center_frame second to occupy the remaining space
        center_frame = Frame(main_frame, bg="#EBF8FF")
        center_frame.pack(expand=True, fill="both", side="top")

        # --- Top Section: Global MOQs ---
        top_frame = LabelFrame(center_frame, text="1. Define Global MOQs", padx=10, pady=10, bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10, "bold"))
        top_frame.pack(side="top", fill="x", pady=(0, 10))
        
        if not self.read_only:
            Button(top_frame, text="Set Global MOQs", command=self._set_global_moq, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2").pack(side="left", padx=(0, 10))
        Label(top_frame, textvariable=self.global_moq_lbl_var, font=("Segoe UI", 10), bg="#EBF8FF", fg="#1A365D").pack(side="left")

        # Pagination Controls
        pag_frame = Frame(center_frame, bg="#EBF8FF", pady=5)
        pag_frame.pack(side="bottom", fill="x")
        
        pag_center = Frame(pag_frame, bg="#EBF8FF")
        pag_center.pack(anchor="center")
        
        self.btn_first = Button(pag_center, text="|<", command=self.goto_first_page, width=4, bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2")
        self.btn_first.pack(side="left", padx=2)
        
        self.btn_prev = Button(pag_center, text="<", command=self.goto_prev_page, width=4, bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2")
        self.btn_prev.pack(side="left", padx=2)
        
        Label(pag_center, text="Page:", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D").pack(side="left", padx=(10, 2))
        self.ent_page_num = Entry(pag_center, width=5, justify="center", font=("Segoe UI", 9))
        self.ent_page_num.pack(side="left", padx=2)
        self.ent_page_num.bind("<Return>", self.on_page_num_entry)
        
        self.lbl_total_pages = Label(pag_center, text="of 1", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D")
        self.lbl_total_pages.pack(side="left", padx=(2, 10))
        
        self.btn_next = Button(pag_center, text=">", command=self.goto_next_page, width=4, bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2")
        self.btn_next.pack(side="left", padx=2)
        
        self.btn_last = Button(pag_center, text=">|", command=self.goto_last_page, width=4, bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2")
        self.btn_last.pack(side="left", padx=2)
        
        Label(pag_center, text="Page Size:", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D").pack(side="left", padx=(20, 2))
        self.cmb_page_size = Combobox(pag_center, values=["100", "500", "1000", "5000"], width=8, state="readonly")
        self.cmb_page_size.set(str(self.page_size))
        self.cmb_page_size.pack(side="left", padx=2)
        self.cmb_page_size.bind("<<ComboboxSelected>>", self.on_page_size_changed)

        self.lbl_total_records = Label(pag_frame, text=f"Total Records: {len(self.unique_assemblies)}", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D")
        self.lbl_total_records.pack(side="right", padx=15)

        # --- Middle Section: Grid (takes up remaining space) ---
        grid_frame = LabelFrame(center_frame, text="2. Assign MOQs to Assemblies", padx=10, pady=10, bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10, "bold"))
        grid_frame.pack(side="top", fill="both", expand=True, pady=10)

        columns = ("assy", "moqs", "status")
        self.tree = ttk.Treeview(grid_frame, columns=columns, show="headings", height=10, selectmode="extended")
        self.tree.heading("assy", text="Assembly #")
        self.tree.heading("moqs", text="Assigned MOQs")
        self.tree.heading("status", text="Status")
        
        self.tree.column("assy", width=150, anchor="center")
        self.tree.column("moqs", width=250, anchor="center")
        self.tree.column("status", width=100, anchor="center")
        
        scrollbar = Scrollbar(grid_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree.tag_configure("missing", foreground="#d9534f")
        self.tree.tag_configure("global", foreground="#28a745")
        self.tree.tag_configure("custom", foreground="#0056b3")

    def _format_moq_list(self, moqs):
        if not moqs:
            return "- None -"
        return "[" + ", ".join(str(m) for m in moqs) + "]"

    def _populate_tree(self):
        # Clear existing
        for item in self.tree.get_children():
            self.tree.delete(item)

        total_matches = len(self.unique_assemblies)
        total_pages = max(1, ((total_matches - 1) // self.page_size) + 1)
        
        if self.current_page >= total_pages:
            self.current_page = total_pages - 1
        if self.current_page < 0:
            self.current_page = 0

        # Update Pagination widgets if they exist
        if hasattr(self, 'lbl_total_pages'):
            self.lbl_total_pages.config(text=f"of {total_pages}")
        if hasattr(self, 'ent_page_num'):
            self.ent_page_num.delete(0, tk.END)
            self.ent_page_num.insert(0, str(self.current_page + 1))
        if hasattr(self, 'update_nav_buttons'):
            self.update_nav_buttons(total_pages)

        # Slice for current page
        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, total_matches)
        page_assemblies = self.unique_assemblies[start_idx:end_idx]

        for assy in page_assemblies:
            moqs = self.assembly_moqs.get(assy, list(self.global_moqs) if hasattr(self, 'global_moqs') and self.global_moqs else [])
            moq_str = self.format_moq_str(moqs)
            is_custom = self.assembly_is_custom.get(assy, False)
            
            if not moqs:
                status = "🔴 Missing"
                tag = "missing"
            elif is_custom:
                status = "🔵 Custom"
                tag = "custom"
            else:
                status = "🟢 Global"
                tag = "global"
                
            self.tree.insert("", "end", iid=assy, values=(assy, moq_str, status), tags=(tag,))

    def goto_first_page(self):
        if self.current_page != 0:
            self.current_page = 0
            self._populate_tree()

    def goto_prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self._populate_tree()

    def goto_next_page(self):
        total_pages = max(1, ((len(self.unique_assemblies) - 1) // self.page_size) + 1)
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self._populate_tree()

    def goto_last_page(self):
        total_pages = max(1, ((len(self.unique_assemblies) - 1) // self.page_size) + 1)
        if self.current_page != total_pages - 1:
            self.current_page = total_pages - 1
            self._populate_tree()

    def on_page_size_changed(self, event=None):
        try:
            new_size = int(self.cmb_page_size.get())
            if new_size > 0:
                self.page_size = new_size
                self.current_page = 0
                self._populate_tree()
        except:
            pass

    def on_page_num_entry(self, event=None):
        try:
            val = int(self.ent_page_num.get())
            total_pages = max(1, ((len(self.unique_assemblies) - 1) // self.page_size) + 1)
            val = max(1, min(val, total_pages))
            self.current_page = val - 1
            self._populate_tree()
        except:
            self.ent_page_num.delete(0, tk.END)
            self.ent_page_num.insert(0, str(self.current_page + 1))

    def update_nav_buttons(self, total_pages):
        if self.current_page == 0:
            self.btn_first.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
            self.btn_prev.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
        else:
            self.btn_first.config(state="normal", bg="#1A365D", fg="#FFFFFF", cursor="hand2")
            self.btn_prev.config(state="normal", bg="#1A365D", fg="#FFFFFF", cursor="hand2")

        if self.current_page >= total_pages - 1:
            self.btn_next.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
            self.btn_last.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
        else:
            self.btn_next.config(state="normal", bg="#1A365D", fg="#FFFFFF", cursor="hand2")
            self.btn_last.config(state="normal", bg="#1A365D", fg="#FFFFFF", cursor="hand2")

    def format_moq_str(self, moqs):
        if not moqs:
            return "- None -"
        return f"[{', '.join(str(m) for m in sorted(moqs))}]"

    def _set_global_moq(self):
        # Reusing CategoryInputDialog to define Global MOQs
        dialog = CategoryInputDialog(self, initial_categories=list(self.global_moqs))
        res = dialog.get_categories()
        if res is not None:
            try:
                self.global_moqs = sorted([int(x) for x in res])
                self.global_moq_lbl_var.set(f"Current Global MOQs: {self.format_moq_str(self.global_moqs)}")
                # Apply Global MOQs ONLY to assemblies with status == Global or Missing (do NOT touch Custom MOQs!)
                for assy in self.unique_assemblies:
                    if not self.assembly_is_custom.get(assy, False):
                        self.assembly_moqs[assy] = list(self.global_moqs)
                        self.assembly_is_custom[assy] = False
                self._populate_tree()
            except ValueError:
                show_error("Invalid Input", "MOQs must be integer numbers.", parent=self)

    def _on_back(self):
        self.result = "BACK"
        self._wait_var.set(1)
        self.destroy()

    def _apply_global_to_selected(self):
        if not self.global_moqs:
            show_error("No Global MOQs", "Please define Global MOQs first.", parent=self)
            return

        selected = self.tree.selection()
        if not selected:
            show_info("No Selection", "Please select assemblies from the table.", parent=self)
            return

        for iid in selected:
            self.assembly_moqs[iid] = list(self.global_moqs)
            self.assembly_is_custom[iid] = False
            
        self._populate_tree()

    def _set_custom_to_selected(self):
        selected = self.tree.selection()
        if not selected:
            show_info("No Selection", "Please select one or more assemblies from the table.", parent=self)
            return

        # Pre-populate dialog with current assigned MOQs so user can add/remove instead of re-entering everything
        if len(selected) == 1:
            initial_vals = list(self.assembly_moqs.get(selected[0], []))
        else:
            first_vals = list(self.assembly_moqs.get(selected[0], []))
            all_same = all(sorted(self.assembly_moqs.get(iid, [])) == sorted(first_vals) for iid in selected)
            if all_same:
                initial_vals = first_vals
            else:
                initial_vals = sorted(list(set(m for iid in selected for m in self.assembly_moqs.get(iid, []))))

        dialog = CategoryInputDialog(self, initial_categories=initial_vals)
        res = dialog.get_categories()
        
        if res is not None:
            try:
                custom_moqs = sorted([int(x) for x in res])
                for iid in selected:
                    self.assembly_moqs[iid] = list(custom_moqs)
                    self.assembly_is_custom[iid] = True
                self._populate_tree()
            except ValueError:
                show_error("Invalid Input", "MOQs must be integer numbers.", parent=self)

    def _on_display_changes(self):
        BOMHistoryDialog(self, self.raw_data, title="Display Changes - MOQ Assignation")

    def _on_confirm(self):
        if getattr(self, 'read_only', False):
            self.result = self.assembly_moqs
            self._wait_var.set(1)
            return

        # 1. Validation: Any missing?
        missing = [assy for assy, moqs in self.assembly_moqs.items() if not moqs]
        if missing:
            msg = "The following assemblies have NO MOQs assigned:\n\n" + "\n".join(f"• {a}" for a in missing[:10])
            if len(missing) > 10:
                msg += f"\n... and {len(missing) - 10} more assemblies."
            msg += "\n\nPlease assign MOQs to all assemblies before saving."
            messagebox.showwarning("Missing MOQs Required", msg, parent=self)
            return

        # 2. Extract Original MOQs & Check for existing ones
        original_moqs = {}
        has_existing_moqs = False
        for assy in self.raw_data.get("Assemblies", []):
            a_num = str(assy.get("Assy #", ""))
            old_vals = [float(x) for x in assy.get("Assigned MOQs", []) if str(x).strip()]
            original_moqs[a_num] = old_vals
            if old_vals:
                has_existing_moqs = True
            
        # 3. Detect changes
        changes = []
        for assy, new_moqs in self.assembly_moqs.items():
            old_moqs = original_moqs.get(assy, [])
            new_moqs_clean = [float(x) for x in new_moqs]
            if sorted(old_moqs) != sorted(new_moqs_clean):
                changes.append({
                    "assembly": assy,
                    "old_val": self.format_moq_str(old_moqs),
                    "new_val": self.format_moq_str(new_moqs)
                })

        # 4. If changes exist, confirm with user
        if has_existing_moqs and changes:
            msg = "You are updating existing assigned MOQs for this RFQ.\n\nChanges to be saved:\n"
            for c in changes[:5]:
                msg += f"• Assembly {c['assembly']}: {c['old_val']} ➔ {c['new_val']}\n"
            if len(changes) > 5:
                msg += f"... and {len(changes)-5} more assemblies.\n"
            msg += "\nDo you wish to proceed?"
            
            if not messagebox.askyesno("Confirm MOQ Changes", msg, parent=self):
                return # Cancel and stay on the MOQ assignation page
            
        # 6. Record history in the BOM JSON
        from datetime import datetime
        now = datetime.now()
        date_str = now.strftime("%d.%m.%Y")
        time_str = now.strftime("%H:%M:%S")
        
        if "history" not in self.raw_data:
            self.raw_data["history"] = []
            
        for c in changes:
            history_entry = {
                "Date": date_str,
                "Time": time_str,
                "Changed By": self.current_user,
                "Field Name": f"Assembly {c['assembly']} MOQ",
                "Old Value": c['old_val'],
                "New Value": c['new_val']
            }
            self.raw_data["history"].append(history_entry)
        
        # Show Green Status Bar
        self.status_bar.config(text="MOQ Assign successfully", bg="#28a745", fg="white")
        self.update_idletasks()
        
        self.result = self.assembly_moqs
        self._wait_var.set(1)

    def get_assembly_moqs(self):
        try:
            self.master.configure(bg="#EBF8FF")
        except:
            pass
        self.wait_variable(self._wait_var)
        return getattr(self, 'result', None), self.global_moqs, getattr(self, 'assembly_is_custom', {})

class CalculationModeDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Select Calculation Mode")
        self.geometry("540x440")
        self.transient(parent)
        self.grab_set()
        self.configure(bg="#f2f2f2")
        
        self.result = None
        self.protocol("WM_DELETE_WINDOW", self.on_cancel)
        
        # 1. Header Frame (Light grey/white with solid black border)
        header_frame = tk.Frame(self, bg="#EBF8FF", bd=1, relief="solid")
        header_frame.pack(fill="x", padx=15, pady=(15, 10))
        
        tk.Label(
            header_frame, 
            text="CHOOSE EXCESS COST CALCULATION METHOD", 
            font=("Arial", 11, "bold"), 
            fg="#2B71B9", 
            bg="#EBF8FF"
        ).pack(pady=12)
        
        # 2. Main content Frame (White background with solid black border)
        msg_frame = tk.Frame(self, bg="white", bd=1, relief="solid")
        msg_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        self.mode_var = tk.StringVar(value="total_usage")
        
        # Radio buttons & Descriptions inside white frame
        radio_frame = tk.Frame(msg_frame, bg="white")
        radio_frame.pack(fill="both", expand=True, padx=20, pady=15)
        
        # Mode 1: Total Usage Sourcing
        rb_total = tk.Radiobutton(
            radio_frame, 
            text="Total Usage Sourcing (RFQ-wide)", 
            variable=self.mode_var, 
            value="total_usage",
            font=("Arial", 10, "bold"),
            bg="white",
            selectcolor="white",
            activebackground="white"
        )
        rb_total.pack(anchor="w", pady=(0, 2))
        
        desc_total = (
            "Sourcing is performed on the entire RFQ project. Part usage is "
            "aggregated across all assemblies to determine total quotes and "
            "allocate excess costs proportionally."
        )
        tk.Label(
            radio_frame, 
            text=desc_total, 
            font=("Arial", 9), 
            fg="#555555", 
            bg="white", 
            justify="left", 
            anchor="w",
            wraplength=460
        ).pack(anchor="w", padx=25, pady=(0, 15))
        
        # Mode 2: Isolated Sourcing
        rb_isolated = tk.Radiobutton(
            radio_frame, 
            text="Independent Assembly Sourcing (Isolated)", 
            variable=self.mode_var, 
            value="individual",
            font=("Arial", 10, "bold"),
            bg="white",
            selectcolor="white",
            activebackground="white"
        )
        rb_isolated.pack(anchor="w", pady=(0, 2))
        
        desc_isolated = (
            "Assemblies are assumed to be independent of each other. Sourcing "
            "is evaluated for each assembly in isolation, ignoring other assemblies "
            "when matching supplier MOQs and calculating excess costs."
        )
        tk.Label(
            radio_frame, 
            text=desc_isolated, 
            font=("Arial", 9), 
            fg="#555555", 
            bg="white", 
            justify="left", 
            anchor="w",
            wraplength=460
        ).pack(anchor="w", padx=25)
        
        # 3. Action Buttons Frame
        btn_frame = tk.Frame(self, bg="#f2f2f2")
        btn_frame.pack(fill="x", side="bottom", padx=15, pady=(0, 15))
        
        # Cancel (bottom-left) - Gray button with solid black border
        btn_no = tk.Button(
            btn_frame, 
            text="Cancel", 
            font=("Arial", 10), 
            command=self.on_cancel, 
            bg="#EBF8FF", 
            fg="black", 
            activebackground="#e0e0e0",
            relief="solid", 
            bd=1,
            width=12
        )
        btn_no.pack(side="left")
        
        # Confirm (bottom-right) - Blue button with solid black border
        btn_yes = tk.Button(
            btn_frame, 
            text="Confirm", 
            font=("Arial", 10, "bold"), 
            command=self.on_confirm, 
            bg="#2B71B9", 
            fg="white", 
            activebackground="#1d5084",
            relief="solid", 
            bd=1,
            width=15
        )
        btn_yes.pack(side="right")
        
        # Center on master
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def on_confirm(self):
        self.result = self.mode_var.get()
        self.destroy()

    def on_cancel(self):
        dialog = SourcingCancelWarningDialog(self, msg_type="sourcing")
        self.wait_window(dialog)
        if dialog.result:
            self.result = None
            self.destroy()

class SourcedAssemblySelectionDialog(tk.Toplevel):
    def __init__(self, parent, assemblies_history, current_bom_data):
        super().__init__(parent)
        self.title("Previously Quoted Assemblies Found")
        self.geometry("700x500")
        self.transient(parent)
        self.grab_set()
        self.configure(bg="#f2f2f2")
        
        self.result = {}
        self.cancelled = False
        
        self.protocol("WM_DELETE_WINDOW", self.on_cancel)
        
        header_frame = tk.Frame(self, bg="#EBF8FF", bd=1, relief="solid")
        header_frame.pack(fill="x", padx=15, pady=(15, 10))
        
        tk.Label(
            header_frame, 
            text="PREVIOUSLY SOURCED ASSEMBLIES DETECTED", 
            font=("Arial", 11, "bold"), 
            fg="#2B71B9", 
            bg="#EBF8FF"
        ).pack(pady=12)
        
        tk.Label(
            self,
            text="Select which previously sourced data to load directly, or choose [Fresh Sourcing].",
            font=("Arial", 9, "italic"),
            bg="#f2f2f2",
            fg="#555555"
        ).pack(pady=(0, 10))
        
        container = tk.Frame(self, bg="white", bd=1, relief="solid")
        container.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        canvas = tk.Canvas(container, bg="white", highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="white")
        
        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw", width=660)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        self.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))
        
        self.selection_vars = {}
        
        from bomformatter import is_bom_content_identical
        from utils import INDIVIDUAL_BOM_DATA_DIR
        import os
        import json
        
        row_idx = 0
        for assy_num, history in assemblies_history.items():
            current_comp_list = current_bom_data.get(assy_num, [])
            
            row_frame = tk.Frame(scroll_frame, bg="white", pady=10)
            row_frame.pack(fill="x", padx=10, pady=5)
            
            if row_idx > 0:
                separator = tk.Frame(scroll_frame, height=1, bg="#e0e0e0")
                separator.pack(fill="x", padx=10)
                
            tk.Label(
                row_frame,
                text=f"Assembly: {assy_num}",
                font=("Arial", 10, "bold"),
                bg="white",
                anchor="w"
            ).pack(anchor="w", padx=5)
            
            options = []
            option_map = {}
            
            default_opt = None
            
            for rec in history:
                cust = rec.get("Customer")
                rfq = rec.get("RFQ")
                timestamp = rec.get("Timestamp", "")
                filename = rec.get("Filename")
                
                safe_rfq = rfq.replace(" ", "_").replace("/", "_").replace("\\", "_")
                file_path = os.path.join(INDIVIDUAL_BOM_DATA_DIR, safe_rfq, filename)
                
                if not os.path.exists(file_path):
                    safe_cust = cust.replace(" ", "_")
                    file_path = os.path.join(INDIVIDUAL_BOM_DATA_DIR, safe_cust, filename)
                    
                if not os.path.exists(file_path):
                    safe_cust = cust.replace(" ", "_")
                    safe_assy = assy_num.replace('/', '_').replace('\\', '_')
                    file_path = os.path.join(INDIVIDUAL_BOM_DATA_DIR, safe_cust, f"{safe_assy}_{safe_rfq}.json")
                
                is_identical = False
                if os.path.exists(file_path):
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            saved_assy = json.load(f)
                        is_identical = is_bom_content_identical(saved_assy.get("saved_model", []), current_comp_list)
                    except Exception as ex:
                        print(f"Error checking identical status for {filename}: {ex}")
                        
                status_label = "Identical BOM" if is_identical else "BOM Changed/Outdated"
                label = f"{cust} - RFQ: {rfq} ({timestamp}) [{status_label}]"
                options.append(label)
                option_map[label] = {
                    "record": rec,
                    "filepath": file_path,
                    "is_identical": is_identical
                }
                
            fresh_label = "[Fresh Sourcing] (Do not load, run fresh calculations)"
            options.append(fresh_label)
            option_map[fresh_label] = None
            
            default_opt = fresh_label
                
            from tkinter import ttk
            var = tk.StringVar(value=default_opt)
            self.selection_vars[assy_num] = (var, option_map)
            
            combo = ttk.Combobox(row_frame, textvariable=var, values=options, state="readonly", width=80)
            combo.pack(anchor="w", padx=15, pady=5)
            
            row_idx += 1
            
        btn_frame = tk.Frame(self, bg="#f2f2f2")
        btn_frame.pack(fill="x", side="bottom", pady=15)
        
        btn_confirm = tk.Button(
            btn_frame, 
            text="Confirm & Proceed", 
            font=("Arial", 10, "bold"), 
            bg="#2B71B9", 
            fg="white",
            relief="flat",
            command=self.on_confirm,
            width=20
        )
        btn_confirm.pack(side="right", padx=(10, 20))
        
        btn_cancel = tk.Button(
            btn_frame, 
            text="Cancel", 
            font=("Arial", 10), 
            bg="#8c8c8c", 
            fg="white",
            relief="flat",
            command=self.on_cancel,
            width=10
        )
        btn_cancel.pack(side="right", padx=10)
        
        # Center dialog on master
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")
        
        try:
            from dialogs import apply_panel_theme
            apply_panel_theme(self)
        except:
            pass
            
    def on_confirm(self):
        warnings = []
        for assy_num, (var, option_map) in self.selection_vars.items():
            selected_label = var.get()
            info = option_map[selected_label]
            if info and not info.get("is_identical"):
                warnings.append(assy_num)
                
        if warnings:
            msg = f"The selected records for the following assemblies have mismatched/outdated BOM structures:\n"
            for w in warnings:
                msg += f" - {w}\n"
            msg += "\nLoading mismatched sourcing data may cause parts to be skipped or incorrectly priced. Are you sure you want to proceed?"
            if not messagebox.askyesno("Mismatched BOM Warning", msg, parent=self):
                return
                
        for assy_num, (var, option_map) in self.selection_vars.items():
            selected_label = var.get()
            info = option_map[selected_label]
            if info:
                self.result[assy_num] = {
                    "filepath": info["filepath"],
                    "record": info["record"]
                }
            else:
                self.result[assy_num] = None
                
        self.grab_release()
        self.destroy()
        
    def on_cancel(self):
        self.cancelled = True
        self.grab_release()
        self.destroy()

class BOMVerificationStatusWindow(tk.Toplevel):
    def __init__(self, parent, unique_assemblies, assembly_status):
        super().__init__(parent)
        self.title("Assembly Verification Status")
        self.geometry("550x500")
        self.transient(parent)
        self.grab_set()

        self.result = "BACK"

        # Title Frame (Yellow Tone)
        header_frame = tk.Frame(self, bg="#fffde7", bd=1, relief="ridge")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="ASSEMBLY VERIFICATION STATUS", font=("Arial", 14, "bold"), fg="#856404", bg="#fffde7").pack(pady=12)

        main_frame = tk.Frame(self, padx=10, pady=10)
        main_frame.pack(fill="both", expand=True)

        tk.Label(main_frame, text="Current verification status of all assemblies:", font=('Arial', 10, 'bold')).pack(pady=(5,5), anchor="w")

        list_frame = tk.Frame(main_frame)
        list_frame.pack(fill="both", expand=True, pady=5)

        cols = ("Status", "Assembly")
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=8)
        self.tree.heading("Status", text="Status")
        self.tree.heading("Assembly", text="Assembly")
        self.tree.column("Status", width=120, anchor="center")
        self.tree.column("Assembly", width=300, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)

        self.tree.tag_configure('viewed', foreground='green')
        self.tree.tag_configure('pending', foreground='orange')
        self.tree.tag_configure('unviewed', foreground='red')

        for assy in unique_assemblies:
            status = assembly_status.get(assy, "Unviewed")
            if status == "Viewed":
                bullet = "🟢 Verified"
                tag = "viewed"
            elif status == "Pending":
                bullet = "🟠 Pending"
                tag = "pending"
            else:
                bullet = "🔴 Not Verified"
                tag = "unviewed"
            self.tree.insert("", "end", values=(bullet, assy), tags=(tag,))

        scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        scroll.pack(side="right", fill="y")
        self.tree.config(yscrollcommand=scroll.set)

        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=10, side="bottom")

        tk.Button(btn_frame, text="💾 Confirm", command=self.on_proceed, bg="#2ead4e", fg="white", font=("Arial", 10, "bold")).pack(side="right", padx=10)
        tk.Button(btn_frame, text="Go Back", command=self.on_back, bg="#e2e8f0", font=("Arial", 10)).pack(side="right")
        
        try:
            from dialogs import apply_panel_theme
            apply_panel_theme(self)
        except:
            pass

    def on_proceed(self):
        self.result = "PROCEED"
        self.destroy()

    def on_back(self):
        self.result = "BACK"
        self.destroy()

class BOMVerificationPanel(BasePanel):
    def __init__(self, master, df, customer_info, mapping, assembly_status=None, temp_file_path=None, username=None, is_edit_saved=False, read_only=False):
        super().__init__(master)
        self.df = df.copy()
        self.customer_info = customer_info # (special_results, cust_name, rfq_id, email)
        self.mapping = mapping # {actual_header: standard_col}
        self.temp_file_path = temp_file_path
        self.username = username
        self.is_edit_saved = is_edit_saved
        self.read_only = read_only
        
        # UI State
        self.row_widgets = []
        self.bulk_add_history = []
        self.assembly_status = assembly_status.copy() if assembly_status else {}
        
        # Track edited assemblies and parts for optimized saving
        self.edited_assemblies = set()
        self.edited_parts = set()
        
        self._init_ui()

    def _init_ui(self):
        # Top: Customer Info & Assembly Selector
        top_frame = Frame(self, bg="#EBF8FF", padx=10, pady=10)
        top_frame.pack(fill="x", padx=10, pady=5)
        
        _, cust_name, rfq_id, email, *_ = self.customer_info
        
        info_container = Frame(top_frame, bg="#EBF8FF")
        info_container.pack(fill="x", pady=(0, 10))
        
        # Category 1: Customer BOM Info
        cust_frame = LabelFrame(info_container, text="Customer BOM Info", bg="#EBF8FF", padx=10, pady=5, font=("Segoe UI", 9, "bold"), fg="#1A365D")
        cust_frame.pack(side="left", fill="x", expand=True, padx=(0, 5))
        
        Label(cust_frame, text=f"Customer: {cust_name}", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10, "bold")).pack(side="left", padx=10)
        Label(cust_frame, text=f"RFQ: {rfq_id}", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10, "bold")).pack(side="left", padx=10)
        Label(cust_frame, text=f"Project / Email Subject: {email}", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10, "bold")).pack(side="left", padx=10)

        # Category 2: BOM Extraction Summary
        self.unique_assemblies = sorted([str(a) for a in self.df['Assy #'].unique() if a])
        
        summary_frame = LabelFrame(info_container, text="BOM Extraction Summary", bg="#EBF8FF", padx=10, pady=5, font=("Segoe UI", 9, "bold"), fg="#1A365D")
        summary_frame.pack(side="left", fill="x", expand=True, padx=(5, 0))
        
        self.total_assy_lbl = Label(summary_frame, text=f"Total Assy: {len(self.unique_assemblies)}", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10, "bold"))
        self.total_assy_lbl.pack(side="left", padx=10)
        
        self.total_rows_lbl = Label(summary_frame, text=f"Total Data Rows: {len(self.df)}", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10, "bold"))
        self.total_rows_lbl.pack(side="left", padx=10)

        # Category 3: Excel Export
        export_frame = LabelFrame(info_container, text="Excel Export", bg="#EBF8FF", padx=10, pady=5, font=("Segoe UI", 9, "bold"), fg="#1A365D")
        export_frame.pack(side="left", fill="x", expand=True, padx=(5, 0))
        
        self.export_btn = Button(export_frame, text="Export Cleaned BOM Data", bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2", command=self._export_bom_data)
        self.export_btn.pack(side="left", padx=10, pady=5)
        self.export_btn.bind("<Enter>", lambda e: self.export_btn.config(bg="#0077B6"))
        self.export_btn.bind("<Leave>", lambda e: self.export_btn.config(bg="#1A365D"))

        divider = tk.Frame(top_frame, height=2, bd=1, relief="sunken", bg="#EBF8FF")
        divider.pack(fill="x", pady=10)

        sel_row = Frame(top_frame, bg="#EBF8FF")
        sel_row.pack(fill="x")
        Label(sel_row, text="Select Assembly to Verify:", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10)).pack(side="left", padx=5)
        
        self.assy_var = StringVar()
        self.assy_combo = Combobox(sel_row, textvariable=self.assy_var, values=self.unique_assemblies, state="normal", width=30)
        self.assy_combo.pack(side="left", padx=5)
        self.assy_combo.bind("<<ComboboxSelected>>", lambda e: self._populate_tree())
        self.assy_combo.bind("<KeyRelease>", self._on_assy_key_release)
        self.assy_combo.bind("<Return>", self._on_assy_return)
        self.assy_combo.bind("<FocusOut>", self._on_assy_focus_out)
        
        if not self.read_only:
            self.btn_mark_viewed = Button(sel_row, text="🟢 Mark as Verified", bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2", command=self._mark_current_viewed)
            self.btn_mark_viewed.pack(side="left", padx=10)
            self.btn_mark_viewed.bind("<Enter>", lambda e: self.btn_mark_viewed.config(bg="#0077B6") if str(self.btn_mark_viewed.cget("state")) != "disabled" else None)
            self.btn_mark_viewed.bind("<Leave>", lambda e: self.btn_mark_viewed.config(bg="#1A365D") if str(self.btn_mark_viewed.cget("state")) != "disabled" else None)
            
            self.btn_mark_pending = Button(sel_row, text="🟠 Keep in View", bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2", command=self._mark_current_pending)
            self.btn_mark_pending.pack(side="left", padx=10)
            self.btn_mark_pending.bind("<Enter>", lambda e: self.btn_mark_pending.config(bg="#0077B6") if str(self.btn_mark_pending.cget("state")) != "disabled" else None)
            self.btn_mark_pending.bind("<Leave>", lambda e: self.btn_mark_pending.config(bg="#1A365D") if str(self.btn_mark_pending.cget("state")) != "disabled" else None)
        
        if self.unique_assemblies:
            self.assy_var.set(self.unique_assemblies[0])

        # Bottom Buttons
        btn_frame = Frame(self, bg="#EBF8FF")
        btn_frame.pack(fill="x", side="bottom", pady=10)
        
        cancel_text = "Close" if self.read_only else "Cancel"
        cancel_btn = Button(btn_frame, text=cancel_text, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", command=self._on_cancel_verification, width=15, pady=6)
        cancel_btn.pack(side="left", padx=5)
        cancel_btn.bind("<Enter>", lambda e: cancel_btn.config(bg="#0077B6"))
        cancel_btn.bind("<Leave>", lambda e: cancel_btn.config(bg="#1A365D"))

        if not self.read_only:
            delete_btn = Button(btn_frame, text="🗑️ Delete Selected Rows", bg="#dc3545", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", command=self._delete_checked, width=22, pady=6)
            delete_btn.pack(side="left", padx=5)
            delete_btn.bind("<Enter>", lambda e: delete_btn.config(bg="#c82333"))
            delete_btn.bind("<Leave>", lambda e: delete_btn.config(bg="#dc3545"))

            add_btn = Button(btn_frame, text="➕ Add New Row", bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", command=lambda: self._add_row(self.assy_var.get()), width=15, pady=6)
            add_btn.pack(side="left", padx=5)
            add_btn.bind("<Enter>", lambda e: add_btn.config(bg="#0077B6"))
            add_btn.bind("<Leave>", lambda e: add_btn.config(bg="#1A365D"))
            
            # New Bulk Add and Revert buttons
            bulk_btn = Button(btn_frame, text="✨ Bulk Add to All", bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", command=self._bulk_add_row, width=18, pady=6)
            bulk_btn.pack(side="left", padx=5)
            bulk_btn.bind("<Enter>", lambda e: bulk_btn.config(bg="#0077B6"))
            bulk_btn.bind("<Leave>", lambda e: bulk_btn.config(bg="#1A365D"))

            self.revert_btn = Button(btn_frame, text="↩️ Revert Bulk Add", bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", command=self._revert_bulk_add, width=18, state="disabled", pady=6)
            self.revert_btn.pack(side="left", padx=5)
            self.revert_btn.bind("<Enter>", lambda e: self.revert_btn.config(bg="#0077B6") if str(self.revert_btn.cget("state")) != "disabled" else None)
            self.revert_btn.bind("<Leave>", lambda e: self.revert_btn.config(bg="#1A365D") if str(self.revert_btn.cget("state")) != "disabled" else None)
        
        confirm_text = "Next ➔" if self.read_only else "💾 Confirm"
        confirm_btn = Button(btn_frame, text=confirm_text, bg="#2ead4e" if self.read_only else "#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", command=self._on_confirm, width=15 if self.read_only else 18, pady=6)
        confirm_btn.pack(side="right", padx=5)
        confirm_btn.bind("<Enter>", lambda e: confirm_btn.config(bg="#248a3e" if self.read_only else "#0077B6"))
        confirm_btn.bind("<Leave>", lambda e: confirm_btn.config(bg="#2ead4e" if self.read_only else "#1A365D"))

        history_btn = Button(btn_frame, text="View History", bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", command=self._on_display_changes, width=15, pady=6)
        history_btn.pack(side="right", padx=5)
        history_btn.bind("<Enter>", lambda e: history_btn.config(bg="#0077B6"))
        history_btn.bind("<Leave>", lambda e: history_btn.config(bg="#1A365D"))

        if not self.read_only:
            save_btn = Button(btn_frame, text="💾 Save Progress", bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), bd=0, relief="flat", cursor="hand2", command=self._save_progress, width=18, pady=6)
            save_btn.pack(side="right", padx=5)
            save_btn.bind("<Enter>", lambda e: save_btn.config(bg="#0077B6"))
            save_btn.bind("<Leave>", lambda e: save_btn.config(bg="#1A365D"))
        
        self.assy_rows_lbl = Label(btn_frame, text="Data Row: 0/0", font=("Segoe UI", 10, "bold"), bg="#EBF8FF", fg="#1A365D")
        self.assy_rows_lbl.pack(side="right", padx=5)

        # Bottom Assembly Navigation Bar (Centered)
        nav_frame = Frame(self, bg="#EBF8FF", padx=10, pady=4)
        nav_frame.pack(fill="x", side="bottom", pady=(0, 5))

        cnt_nav = Frame(nav_frame, bg="#EBF8FF")
        cnt_nav.pack(anchor="center")

        self.btn_prev_assy = Button(
            cnt_nav, text="◀ Previous", bg="#1A365D", fg="white",
            font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2",
            command=self._prev_assembly, padx=12, pady=4
        )
        self.btn_prev_assy.pack(side="left", padx=5)
        self.btn_prev_assy.bind("<Enter>", lambda e: self.btn_prev_assy.config(bg="#0077B6") if str(self.btn_prev_assy.cget("state")) != "disabled" else None)
        self.btn_prev_assy.bind("<Leave>", lambda e: self.btn_prev_assy.config(bg="#1A365D") if str(self.btn_prev_assy.cget("state")) != "disabled" else None)

        self.btn_next_assy = Button(
            cnt_nav, text="Next ▶", bg="#1A365D", fg="white",
            font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2",
            command=self._next_assembly, padx=12, pady=4
        )
        self.btn_next_assy.pack(side="left", padx=5)
        self.btn_next_assy.bind("<Enter>", lambda e: self.btn_next_assy.config(bg="#0077B6") if str(self.btn_next_assy.cget("state")) != "disabled" else None)
        self.btn_next_assy.bind("<Leave>", lambda e: self.btn_next_assy.config(bg="#1A365D") if str(self.btn_next_assy.cget("state")) != "disabled" else None)

        self.lbl_assy_nav_status = Label(
            cnt_nav, text="  |  Assembly 1 of 1: -", bg="#EBF8FF", fg="#1A365D",
            font=("Segoe UI", 9, "bold")
        )
        self.lbl_assy_nav_status.pack(side="left", padx=15)

        Label(cnt_nav, text="  |  Go to Page:", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(10, 2))
        self.page_jump_entry = Entry(cnt_nav, width=6, font=("Segoe UI", 9))
        self.page_jump_entry.pack(side="left", padx=3)
        self.page_jump_entry.bind("<Return>", lambda e: self._on_page_jump_submit())

        btn_go_page = Button(
            cnt_nav, text="Go", bg="#1A365D", fg="white",
            font=("Segoe UI", 8, "bold"), bd=0, relief="flat", cursor="hand2",
            command=self._on_page_jump_submit, padx=8, pady=2
        )
        btn_go_page.pack(side="left", padx=3)
        btn_go_page.bind("<Enter>", lambda e: btn_go_page.config(bg="#0077B6"))
        btn_go_page.bind("<Leave>", lambda e: btn_go_page.config(bg="#1A365D"))

        # Bottom Detail Pane
        self.detail_frame = LabelFrame(self, text="Part Assignment Details", bg="#EBF8FF", fg="#1A365D", padx=10, pady=10, font=("Segoe UI", 10, "bold"))
        self.detail_frame.pack(fill="x", side="bottom", padx=10, pady=10)
        
        self.detail_label = Label(self.detail_frame, text="Part Number: -", bg="#EBF8FF", fg="#1A365D", font=("Segoe UI", 10, "bold"))
        self.detail_label.pack(side="left", padx=10)
        
        if not self.read_only:
            self.edit_btn = Button(self.detail_frame, text="✏️ Edit MPN & MFR", state="disabled", bg="#1A365D", fg="white", font=("Segoe UI", 9, "bold"), bd=0, relief="flat", cursor="hand2", command=self._open_mpn_mfr_editor)
            self.edit_btn.pack(side="left", padx=20)
            self.edit_btn.bind("<Enter>", lambda e: self.edit_btn.config(bg="#0077B6") if str(self.edit_btn.cget("state")) != "disabled" else None)
            self.edit_btn.bind("<Leave>", lambda e: self.edit_btn.config(bg="#1A365D") if str(self.edit_btn.cget("state")) != "disabled" else None)

        # Middle: Treeview
        tree_frame = Frame(self, bg="#EBF8FF")
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # UI State for selections
        self.checked_iids = set()
        
        columns = ("line", "part", "desc", "mfr", "mpn", "qty", "uom")
        self.tree = Treeview(tree_frame, columns=columns, show="headings", height=15, selectmode="extended")
        
        self.tree.heading("line", text="Line Item")
        self.tree.heading("part", text="Part Number")
        self.tree.heading("desc", text="Description")
        self.tree.heading("mfr", text="MFR")
        self.tree.heading("mpn", text="MPN")
        self.tree.heading("qty", text="Qty")
        self.tree.heading("uom", text="UOM")
        
        self.tree.column("line", width=80, anchor="center")
        self.tree.column("part", width=150, anchor="w")
        self.tree.column("desc", width=250, anchor="w")
        self.tree.column("mfr", width=150, anchor="w")
        self.tree.column("mpn", width=200, anchor="w")
        self.tree.column("qty", width=60, anchor="center")
        self.tree.column("uom", width=60, anchor="center")
        
        y_scroll = Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        x_scroll = Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        self.tree.pack(fill="both", expand=True)
        y_scroll.pack(side="right", fill="y")
        x_scroll.pack(side="bottom", fill="x")
        
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<Button-1>", self._on_tree_click)
        self.tree.bind("<Double-1>", self._on_double_click)

        self._populate_tree()

    def _on_assy_key_release(self, event):
        if event.keysym in ("Return", "Tab", "Up", "Down", "Escape", "Left", "Right"):
            return
        typed = self.assy_var.get().strip()
        if not typed:
            self.assy_combo['values'] = self.unique_assemblies
            return
        matching = [a for a in self.unique_assemblies if typed.lower() in a.lower()]
        self.assy_combo['values'] = matching if matching else self.unique_assemblies

    def _on_assy_return(self, event=None):
        if not hasattr(self, 'tree'):
            return
        typed = self.assy_var.get().strip()
        if not typed:
            return
        if typed in self.unique_assemblies:
            self._populate_tree()
            return
        matching = [a for a in self.unique_assemblies if typed.lower() in a.lower()]
        if matching:
            exact_ci = [a for a in matching if a.lower() == typed.lower()]
            target = exact_ci[0] if exact_ci else matching[0]
            self.assy_var.set(target)
            self.assy_combo.set(target)
            self._populate_tree()

    def _on_assy_focus_out(self, event=None):
        if not hasattr(self, 'tree'):
            return
        self._on_assy_return()

    def _prev_assembly(self):
        curr_assy = self.assy_var.get()
        if curr_assy in self.unique_assemblies:
            idx = self.unique_assemblies.index(curr_assy)
            if idx > 0:
                self._switch_to_assembly(self.unique_assemblies[idx - 1])

    def _next_assembly(self):
        curr_assy = self.assy_var.get()
        if curr_assy in self.unique_assemblies:
            idx = self.unique_assemblies.index(curr_assy)
            if idx < len(self.unique_assemblies) - 1:
                self._switch_to_assembly(self.unique_assemblies[idx + 1])

    def _on_page_jump_submit(self):
        val = getattr(self, 'page_jump_entry', None)
        if not val:
            return
        txt = val.get().strip()
        if not txt:
            return
        try:
            page = int(txt)
            if 1 <= page <= len(self.unique_assemblies):
                self._switch_to_assembly(self.unique_assemblies[page - 1])
        except ValueError:
            pass

    def _switch_to_assembly(self, target_assy):
        self.assy_var.set(target_assy)
        self.assy_combo.set(target_assy)
        self._populate_tree()

    def _update_nav_controls(self):
        if not hasattr(self, 'lbl_assy_nav_status'):
            return
        N = len(self.unique_assemblies)
        curr_assy = self.assy_var.get()
        curr_idx = 0
        if curr_assy in self.unique_assemblies:
            curr_idx = self.unique_assemblies.index(curr_assy)
        page = curr_idx + 1 if N > 0 else 0

        self.lbl_assy_nav_status.config(text=f"  |  Assembly {page} of {N}: {curr_assy}")
        if hasattr(self, 'page_jump_entry') and self.page_jump_entry.winfo_exists():
            self.page_jump_entry.delete(0, "end")
            self.page_jump_entry.insert(0, str(page))

        if hasattr(self, 'btn_prev_assy'):
            if curr_idx > 0:
                self.btn_prev_assy.config(state="normal", bg="#1A365D")
            else:
                self.btn_prev_assy.config(state="disabled", bg="#A0AEC0")

        if hasattr(self, 'btn_next_assy'):
            if curr_idx < N - 1:
                self.btn_next_assy.config(state="normal", bg="#1A365D")
            else:
                self.btn_next_assy.config(state="disabled", bg="#A0AEC0")

    def _on_tree_click(self, event):
        if getattr(self, 'read_only', False):
            return "break"
        region = self.tree.identify("region", event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            if column == "#1": # 'sel' column
                iid = self.tree.identify_row(event.y)
                if iid:
                    self._toggle_check(iid)
                    return "break" # Prevent selection change if clicking checkbox

    def _on_double_click(self, event):
        if getattr(self, 'read_only', False):
            return "break"
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
            
        column = self.tree.identify_column(event.x)
        column_id = self.tree.column(column, "id")
        
        if column_id not in ("qty", "uom"):
            return
            
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
            
        # Get current value
        current_vals = list(self.tree.item(iid, "values"))
        col_idx = self.tree["columns"].index(column_id)
        current_val = current_vals[col_idx]
        
        # Bounding box of the cell
        bbox = self.tree.bbox(iid, column)
        if not bbox:
            return
        x, y, w, h = bbox
        
        # Create an Entry widget over the cell
        entry = Entry(self.tree, font=("Segoe UI", 9))
        entry.insert(0, current_val)
        entry.select_range(0, tk.END)
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        
        self._editing_entry = entry
        self._is_saving = False
        
        def save_edit(event=None):
            if self._is_saving:
                return
            self._is_saving = True
            
            entry.unbind("<FocusOut>")
            
            new_val = entry.get().strip()
            if column_id == "qty":
                try:
                    new_val_float = float(new_val)
                    val_to_save = int(new_val_float) if new_val_float.is_integer() else new_val_float
                    new_val = str(val_to_save)
                except ValueError:
                    show_error("Invalid Input", "Qty must be a numeric value.", parent=self)
                    entry.destroy()
                    self._is_saving = False
                    return
            else:
                val_to_save = new_val
                
            # Update DataFrame
            idx = int(iid)
            if column_id == "qty":
                self.df.at[idx, 'Qty'] = float(val_to_save)
            else:
                self.df.at[idx, 'UOM'] = str(val_to_save)

            # Apply automatic UOM conversion if applicable
            conv_uom, conv_qty = apply_uom_conversion_to_row(self.df.at[idx, 'UOM'], self.df.at[idx, 'Qty'])
            self.df.at[idx, 'UOM'] = conv_uom
            self.df.at[idx, 'Qty'] = conv_qty
                
            # Track modification
            assy_val = self.df.at[idx, 'Assy #']
            self.edited_assemblies.add(str(assy_val))
                
            # Update Treeview item
            self._populate_tree()

            # Record log to centralized backlog if changed
            if new_val != current_val:
                try:
                    user = "Unknown User"
                    if getattr(self, 'username', None):
                        user = self.username
                    elif getattr(self, 'user_name', None):
                        user = self.user_name
                    else:
                        curr = self.master
                        while curr:
                            if hasattr(curr, 'user_name') and curr.user_name:
                                user = curr.user_name
                                break
                            curr = getattr(curr, 'master', None)
                            
                    cust_name = self.customer_info[1] if len(self.customer_info) > 1 else ""
                    rfq_id = self.customer_info[2] if len(self.customer_info) > 2 else ""
                    part_num = self.df.at[idx, 'Part']
                    
                    from backlog_api import log_backlog_event
                    details = {
                        "customer": cust_name,
                        "rfq_number": rfq_id,
                        "part_number": part_num,
                        "column": column_id,
                        "old_value": current_val,
                        "new_value": new_val,
                        "source": "BOM Verification Panel (Qty/UOM Inline Editor)"
                    }
                    log_backlog_event(
                        event_type="EDIT_QTY_UOM",
                        app_name="BOM App",
                        user_name=user,
                        details=details
                    )
                except Exception as ex:
                    print(f"Failed to record backlog event: {ex}")
            
            entry.destroy()
            self._is_saving = False
            
        def cancel_edit(event=None):
            entry.destroy()
            
        entry.bind("<Return>", save_edit)
        entry.bind("<Escape>", cancel_edit)
        entry.bind("<FocusOut>", lambda e: save_edit() if entry.winfo_exists() else None)

    def _update_row_count_label(self):
        if not hasattr(self, 'tree'):
            return
        total_rows = len(self.tree.get_children())
        selected = self.tree.selection()
        if selected:
            try:
                focused_idx = self.tree.index(selected[0]) + 1
            except:
                focused_idx = 0
        else:
            focused_idx = 0
        
        if hasattr(self, 'assy_rows_lbl'):
            self.assy_rows_lbl.config(text=f"Data Row: {focused_idx}/{total_rows}")

    def _populate_tree(self):
        if not hasattr(self, 'tree'):
            return
        # Clear existing
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        self.checked_iids.clear()
        
        # Update Total Rows at the top
        if hasattr(self, 'total_rows_lbl'):
            self.total_rows_lbl.config(text=f"Total Data Rows: {len(self.df)}")
            
        assy = self.assy_var.get()
        if not assy:
            self._update_row_count_label()
            return
        
        assy_df = self.df[self.df['Assy #'].astype(str) == assy]
        
        for idx, row in assy_df.iterrows():
            vals = (
                str(row.get('Line Item', '')),
                str(row.get('Part', '')),
                str(row.get('Description', '')),
                str(row.get('MFR', '')),
                str(row.get('MPN', '')),
                format_to_sig_figs(row.get('Qty', '')),
                str(row.get('UOM', ''))
            )
            self.tree.insert("", "end", iid=str(idx), values=vals)

        self._update_row_count_label()
        self._update_mark_buttons(assy)
        self._update_nav_controls()

    def _update_mark_buttons(self, assy):
        if getattr(self, 'read_only', False) or not hasattr(self, 'btn_mark_viewed'):
            return
        status = self.assembly_status.get(assy, "Unviewed")
        if status == "Viewed":
            self.btn_mark_viewed.config(text="✔️ Marked Verified", state="disabled", bg="#28A745", fg="white")
            self.btn_mark_pending.config(text="🟠 Keep in View", state="normal", bg="#1A365D", fg="white")
        elif status == "Pending":
            self.btn_mark_viewed.config(text="🟢 Mark as Verified", state="normal", bg="#1A365D", fg="white")
            self.btn_mark_pending.config(text="⏳ Marked Pending", state="disabled", bg="#DD6B20", fg="white")
        else:
            self.btn_mark_viewed.config(text="🟢 Mark as Verified", state="normal", bg="#1A365D", fg="white")
            self.btn_mark_pending.config(text="🟠 Keep in View", state="normal", bg="#1A365D", fg="white")

    def _mark_current_viewed(self):
        assy = self.assy_var.get()
        if assy:
            self.assembly_status[assy] = "Viewed"
            self._update_mark_buttons(assy)

    def _mark_current_pending(self):
        assy = self.assy_var.get()
        if assy:
            self.assembly_status[assy] = "Pending"
            self._update_mark_buttons(assy)

    def _on_tree_select(self, event):
        selected = self.tree.selection()
        self._update_row_count_label()
        if not selected:
            self.detail_label.config(text="Part Number: -")
            if hasattr(self, 'edit_btn'):
                self.edit_btn.config(state="disabled")
            return
            
        iid = selected[0]
        row = self.df.loc[int(iid)]
        self.detail_label.config(text=f"Part Number: {row['Part']}")
        if hasattr(self, 'edit_btn'):
            self.edit_btn.config(state="normal")

    def _open_mpn_mfr_editor(self):
        selected = self.tree.selection()
        if not selected: return
        
        iid = int(selected[0])
        row = self.df.loc[iid]
        part_num = row['Part']
        current_mpn = row['MPN']
        current_mfr = row['MFR']
        
        # Helper to normalize and split comma-separated MPN/MFRs
        def normalize_pairs(mpn_str, mfr_str):
            mpns = [m.strip() for m in str(mpn_str).split(",") if m.strip()]
            mfrs = [m.strip() for m in str(mfr_str).split(",") if m.strip()]
            while len(mfrs) < len(mpns): mfrs.append("")
            while len(mpns) < len(mfrs): mpns.append("")
            return [(mpn.strip(), mfr.strip()) for mpn, mfr in zip(mpns, mfrs)]
            
        # Load potential database alternatives
        cust_name = self.customer_info[1] if len(self.customer_info) > 1 else ""
        from utils import get_alternative_mpn_path, merge_mpn_mfr_pairs
        import os, json
        
        alt_json_path = get_alternative_mpn_path(cust_name)
        db_mpn, db_mfr = "", ""
        alt_data = {"Customer": cust_name, "Parts": {}}
        
        if os.path.exists(alt_json_path):
            try:
                with open(alt_json_path, 'r', encoding='utf-8') as f:
                    alt_data = json.load(f)
                alt_parts = alt_data.get("Parts", {})
                if part_num in alt_parts:
                    db_rec = alt_parts[part_num]
                    db_mpn = db_rec.get("MPN", "")
                    db_mfr = db_rec.get("MFR", "")
            except Exception as e:
                print(f"Error checking database alternatives: {e}")
                
        # Check if there are current BOM pairs not in database Alternative MPNs
        current_pairs = normalize_pairs(current_mpn, current_mfr)
        db_pairs = normalize_pairs(db_mpn, db_mfr)
        db_pairs_upper = {(p[0].upper(), p[1].upper()) for p in db_pairs}
        
        new_pairs_to_sync = []
        for mpn, mfr in current_pairs:
            if (mpn.upper(), mfr.upper()) not in db_pairs_upper:
                new_pairs_to_sync.append((mpn, mfr))
                
        if new_pairs_to_sync:
            sync_text = "\n".join([f"• MPN: {p[0]} (MFR: {p[1]})" for p in new_pairs_to_sync])
            msg = f"The following current BOM MPN/MFR pairs for part '{part_num}' are not in the Alternative MPNs database:\n\n{sync_text}\n\nDo you want to sync (add) these new pairs to the database?"
            if messagebox.askyesno("Sync New Alternatives", msg, parent=self):
                new_mpns_str = ", ".join([p[0] for p in new_pairs_to_sync])
                new_mfrs_str = ", ".join([p[1] for p in new_pairs_to_sync])
                db_mpn, db_mfr = merge_mpn_mfr_pairs(db_mpn, db_mfr, new_mpns_str, new_mfrs_str)
                
                if "Parts" not in alt_data:
                    alt_data["Parts"] = {}
                alt_data['Parts'][part_num] = {
                    "MPN": db_mpn,
                    "MFR": db_mfr
                }
                
                try:
                    with open(alt_json_path, 'w', encoding='utf-8') as f:
                        json.dump(alt_data, f, indent=4)
                    # Refresh db_pairs
                    db_pairs = normalize_pairs(db_mpn, db_mfr)
                except Exception as e:
                    show_error("Sync Error", f"Failed to sync to database: {e}", parent=self)
                    
        # Check if database has extra alternatives NOT currently in the BOM
        has_alternatives = False
        if db_mpn.strip() or db_mfr.strip():
            current_pairs_upper = {(p[0].upper(), p[1].upper()) for p in current_pairs}
            db_has_extra = False
            for mpn, mfr in db_pairs:
                if (mpn.upper(), mfr.upper()) not in current_pairs_upper:
                    db_has_extra = True
                    break
            if db_has_extra:
                has_alternatives = True
                
        target_mpn = current_mpn
        target_mfr = current_mfr
        
        if has_alternatives:
            # Prompt user
            option_dlg = AlternativeMPNOptionDialog(self, part_num, current_mpn, current_mfr, db_mpn, db_mfr)
            self.wait_window(option_dlg)
            
            if option_dlg.result == "merge":
                target_mpn, target_mfr = merge_mpn_mfr_pairs(current_mpn, current_mfr, db_mpn, db_mfr)
            elif option_dlg.result == "select":
                select_dlg = AlternativeMPNSelectionDialog(self, part_num, db_mpn, db_mfr, current_mpn, current_mfr)
                self.wait_window(select_dlg)
                if select_dlg.result is not None:
                    # Merge selected items
                    sel_mpns = ", ".join([p[0] for p in select_dlg.result])
                    sel_mfrs = ", ".join([p[1] for p in select_dlg.result])
                    target_mpn, target_mfr = merge_mpn_mfr_pairs(current_mpn, current_mfr, sel_mpns, sel_mfrs)
                else:
                    return # Cancelled in selection
            elif option_dlg.result == "proceed":
                target_mpn = current_mpn
                target_mfr = current_mfr
            else:
                return # Cancelled or closed option dialog
                
        # Open main Editor Dialog
        dialog = MPNMFRAssignmentDialog(self, part_num, target_mpn, target_mfr)
        self.wait_window(dialog)
        
        if dialog.result:
            new_mpns, new_mfrs = dialog.result
            
            # Find all rows with the same Part number and update them
            same_part_rows = self.df[self.df['Part'].astype(str) == str(part_num)]
            for idx in same_part_rows.index:
                self.df.at[idx, 'MPN'] = new_mpns
                self.df.at[idx, 'MFR'] = new_mfrs
                
                # Track modification for all these assemblies
                assy_val = self.df.at[idx, 'Assy #']
                self.edited_assemblies.add(str(assy_val))
                
            self.edited_parts.add(str(part_num))
            
            # Update Tree
            self._populate_tree()
            self.tree.selection_set(str(iid))
            
            # Sync to JSON
            self._sync_to_json(part_num, new_mpns, new_mfrs, old_mpn=current_mpn, old_mfr=current_mfr)

    def _sync_to_json(self, part_num, mpn_str, mfr_str, old_mpn="", old_mfr=""):
        if not mpn_str.strip() and not mfr_str.strip():
            return

        cust_name = self.customer_info[1] if len(self.customer_info) > 1 else ""
        rfq_id = self.customer_info[2] if len(self.customer_info) > 2 else ""
        from utils import get_alternative_mpn_path
        import os, json
        
        alt_json_path = get_alternative_mpn_path(cust_name)
        
        data = {"Customer": cust_name, "Parts": {}}
        db_mpn, db_mfr = "", ""
        if os.path.exists(alt_json_path):
            try:
                with open(alt_json_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if str(part_num) in data.get("Parts", {}):
                    db_mpn = data["Parts"][str(part_num)].get("MPN", "")
                    db_mfr = data["Parts"][str(part_num)].get("MFR", "")
            except: pass
            
        data['Parts'][str(part_num)] = {
            "MPN": mpn_str,
            "MFR": mfr_str
        }
        
        try:
            with open(alt_json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4)
            
            # Record log to centralized backlog if changed
            if mpn_str != old_mpn or mfr_str != old_mfr or mpn_str != db_mpn or mfr_str != db_mfr:
                user = "Unknown User"
                try:
                    if getattr(self, 'username', None):
                        user = self.username
                    elif getattr(self, 'user_name', None):
                        user = self.user_name
                    else:
                        curr = self.master
                        while curr:
                            if hasattr(curr, 'user_name') and curr.user_name:
                                user = curr.user_name
                                break
                            curr = getattr(curr, 'master', None)
                except Exception:
                    pass
                
                from backlog_api import log_backlog_event
                details = {
                    "customer": cust_name,
                    "rfq_number": rfq_id,
                    "part_number": part_num,
                    "old_mpn": old_mpn,
                    "old_mfr": old_mfr,
                    "new_mpn": mpn_str,
                    "new_mfr": mfr_str,
                    "source": "BOM Verification Panel (MPN/MFR Editor)"
                }
                log_backlog_event(
                    event_type="EDIT_MPN_MFR_PAIR",
                    app_name="Sourcing App" if "Sourcing" in self.winfo_toplevel().title() else "BOM App",
                    user_name=user,
                    details=details
                )
        except Exception as e:
            print(f"Failed to sync alternative pairs: {e}")

    def _delete_checked(self):
        selected_iids = self.tree.selection()
        if not selected_iids:
            messagebox.showwarning("Warning", "Please click to select one or more rows to delete (use Ctrl+Click to select multiple rows).", parent=self)
            return
        
        count = len(selected_iids)
        msg = f"Delete {count} selected row(s)?" if count > 1 else "Delete selected row?"
        if messagebox.askyesno("Confirm Delete", msg, parent=self):
            to_delete = [int(iid) for iid in selected_iids]
            
            # Track modification before dropping
            for idx in to_delete:
                assy_val = self.df.at[idx, 'Assy #']
                self.edited_assemblies.add(str(assy_val))
                
            self.df = self.df.drop(to_delete).reset_index(drop=True)
            self._populate_tree()
            self.detail_label.config(text="Part Number: -")
            if hasattr(self, 'edit_btn'):
                self.edit_btn.config(state="disabled")
 
    def _add_row(self, assy_num):
        # Find representative row for metadata
        rep_rows = self.df[self.df['Assy #'].astype(str) == assy_num]
        if rep_rows.empty: return
        
        rep = rep_rows.iloc[0]
        
        # Open Add Row Dialog
        dialog = AddRowDialog(self, assy_num)
        self.wait_window(dialog)
        
        if dialog.result:
            res = dialog.result
            raw_uom = res.get('UOM', '')
            raw_qty = res.get('Qty', 0)
            conv_uom, conv_qty = apply_uom_conversion_to_row(raw_uom, raw_qty)

            new_row = {col: "" for col in self.df.columns}
            new_row['Assy #'] = assy_num
            new_row['Assy Model'] = rep.get('Assy Model', '')
            new_row['Assy Rev'] = rep.get('Assy Rev', '')
            
            new_row['Line Item'] = res.get('Line Item', '')
            new_row['Part'] = res.get('Part', '')
            new_row['Description'] = res.get('Description', '')
            new_row['MFR'] = res.get('MFR', '')
            new_row['MPN'] = res.get('MPN', '')
            new_row['Qty'] = conv_qty
            new_row['UOM'] = conv_uom
            
            # Track modification
            self.edited_assemblies.add(str(assy_num))
            self.edited_parts.add(str(res.get('Part', '')))
            
            import pandas as pd
            self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)
            self._populate_tree()
 
    def _bulk_add_row(self):
        if not self.unique_assemblies:
            messagebox.showwarning("Warning", "No unique assemblies found to bulk add.", parent=self)
            return

        dialog = BulkAddRowDialog(self, assemblies=self.unique_assemblies)
        self.wait_window(dialog)
        
        if dialog.result:
            target_assemblies = dialog.result.get('target_assemblies', self.unique_assemblies)
            if not target_assemblies:
                return

            # 1. Save current state to history for revert
            self.bulk_add_history.append(self.df.copy())
            self.revert_btn.config(state="normal")
            
            res = dialog.result
            raw_uom = res.get('UOM', '')
            raw_qty = res.get('Qty', '0')
            conv_uom, conv_qty = apply_uom_conversion_to_row(raw_uom, raw_qty)

            new_rows = []
            
            # Track modification for selected target assemblies and the part
            for assy_num in target_assemblies:
                self.edited_assemblies.add(str(assy_num))
            self.edited_parts.add(str(res.get('Part', '')))
            
            import pandas as pd
            for assy_num in target_assemblies:
                # Find representative row for metadata
                rep_rows = self.df[self.df['Assy #'].astype(str) == str(assy_num)]
                if not rep_rows.empty:
                    rep = rep_rows.iloc[0]
                    model = rep.get('Assy Model', '')
                    rev = rep.get('Assy Rev', '')
                else:
                    model = ''
                    rev = ''
                
                new_row = {col: "" for col in self.df.columns}
                new_row['Assy #'] = assy_num
                new_row['Assy Model'] = model
                new_row['Assy Rev'] = rev
                
                new_row['Line Item'] = res.get('Line Item', '')
                new_row['Part'] = res.get('Part', '')
                new_row['Description'] = res.get('Description', '')
                new_row['MFR'] = res.get('MFR', '')
                new_row['MPN'] = res.get('MPN', '')
                new_row['Qty'] = conv_qty
                new_row['UOM'] = conv_uom
                
                new_rows.append(new_row)
            
            if new_rows:
                self.df = pd.concat([self.df, pd.DataFrame(new_rows)], ignore_index=True)
                self._populate_tree()
                
                messagebox.showinfo(
                    "Success",
                    f"Successfully bulk added item '{res.get('Part')}' to {len(target_assemblies)} selected assembly(ies).\n"
                    "You can revert this action using the 'Revert Bulk Add' button.",
                    parent=self
                )

    def _revert_bulk_add(self):
        if not self.bulk_add_history:
            messagebox.showwarning("Warning", "No bulk add action to revert.", parent=self)
            return
            
        if messagebox.askyesno("Confirm Revert", "Are you sure you want to revert the last bulk add action?", parent=self):
            self.df = self.bulk_add_history.pop()
            if not self.bulk_add_history:
                self.revert_btn.config(state="disabled")
            
            self._populate_tree()
            messagebox.showinfo("Reverted", "The last bulk add action has been reverted successfully.", parent=self)

    def _export_bom_data(self):
        from tkinter import filedialog
        import pandas as pd
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import PatternFill, Border, Side, Font
        import os

        _, cust_name, rfq_id, email, *_ = self.customer_info
        
        now_str = datetime.now().strftime("%d-%m-%Y_%I%M%p")
        default_filename = f"Unverified BOM Data for Customer {cust_name}_{rfq_id}_{now_str}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Select Folder to Save Excel File"
        )
        
        if not file_path:
            return
            
        try:
            df_export = self.df.copy()
            if 'Qty' in df_export.columns:
                df_export['Qty'] = df_export['Qty'].apply(lambda x: float(format_to_sig_figs(x)) if pd.notna(x) and str(x).strip() != "" else x)
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_export.to_excel(writer, sheet_name="BOM Data", index=False)
                
            wb = openpyxl.load_workbook(file_path)
            ws = wb["BOM Data"]
            ws.sheet_view.showGridLines = False
            
            # light blue background for header
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") 
            header_font = Font(bold=True)
            
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for col_idx, cell in enumerate(ws[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
                
            wb.save(file_path)
            os.startfile(file_path)
            
        except Exception as e:
            from utils import show_error
            show_error("Export Error", f"Failed to export BOM data:\n{e}", parent=self)

    def _save_progress(self):
        from utils import TEMP_DIR, show_info, show_error
        import json
        import time
        import secrets

        # Time-ordered UUIDv7 generator
        def generate_uuid7():
            msec = int(time.time() * 1000)
            rand_a = secrets.randbits(12) | (0x7 << 12)
            rand_b = secrets.randbits(62) | (0x2 << 62)
            uuid_int = (msec << 80) | (rand_a << 64) | rand_b
            hex_str = f"{uuid_int:032x}"
            return f"{hex_str[:8]}-{hex_str[8:12]}-{hex_str[12:16]}-{hex_str[16:20]}-{hex_str[20:]}"

        if not self.temp_file_path:
            filename = f"{generate_uuid7()}.json"
            self.temp_file_path = os.path.join(TEMP_DIR, filename)

        # Serialize DataFrame
        df_records = self.df.to_dict(orient='records')
        
        session_data = {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "created_by": self.username or "Unknown",
            "customer_info": self.customer_info,
            "mapping": self.mapping,
            "assembly_status": self.assembly_status,
            "df_data": df_records,
            "temp_file_path": self.temp_file_path,
            "is_edit_saved": getattr(self, "is_edit_saved", False)
        }
        
        try:
            with open(self.temp_file_path, 'w', encoding='utf-8') as f:
                json.dump(session_data, f, indent=4)
            show_info("Success", f"Progress saved successfully.\nFile: {os.path.basename(self.temp_file_path)}", parent=self)
            self.result = None
            self._wait_var.set(1)
            self.destroy()
        except Exception as e:
            show_error("Save Error", f"Failed to save progress:\n{e}", parent=self)

    def _on_display_changes(self):
        if hasattr(self, 'raw_data') and self.raw_data:
            BOMHistoryDialog(self, self.raw_data, title="Display Changes - BOM Verification")
        elif hasattr(self, 'customer_info') and self.customer_info:
            _, cust_name, rfq_id, *rest = self.customer_info
            from utils import BOM_DATA_DIR, show_info
            import json, os
            cust_folder = str(cust_name).replace(" ", "_")
            filepath = os.path.normpath(os.path.join(BOM_DATA_DIR, cust_folder, f"{str(rfq_id).replace(' ', '_')}.json"))
            if os.path.exists(filepath):
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        raw_data = json.load(f)
                    BOMHistoryDialog(self, raw_data, title="Display Changes - BOM Verification")
                    return
                except Exception as e:
                    print(f"Error reading history file: {e}")
            show_info("No History", "No history logs available for this RFQ.", parent=self)
        else:
            from utils import show_info
            show_info("No History", "No history logs available.", parent=self)

    def _on_cancel_verification(self):
        if getattr(self, 'read_only', False):
            self.result = None
            self._wait_var.set(1)
            return
        dialog = SourcingCancelWarningDialog(self.master, msg_type="verification")
        self.master.wait_window(dialog)
        if dialog.result:
            self.result = None
            self._wait_var.set(1)

    def _on_confirm(self):
        if getattr(self, 'read_only', False):
            self.result = self.df
            self._wait_var.set(1)
            return
        if self.checked_iids:
            if not messagebox.askyesno("Unprocessed Checked Rows", f"You have {len(self.checked_iids)} row(s) checked but not deleted.\n\nAre you sure you want to proceed with verification without deleting them?", parent=self):
                return
                
        # Show Assembly Verification Status Window
        dialog = BOMVerificationStatusWindow(self.winfo_toplevel(), self.unique_assemblies, self.assembly_status)
        self.winfo_toplevel().wait_window(dialog)
        if dialog.result != "PROCEED":
            return
            
        self.result = self.df
        self._wait_var.set(1)

class AddRowDialog(tk.Toplevel):
    def __init__(self, parent, assy_num):
        super().__init__(parent)
        self.title(f"Add New Line Item for Assy: {assy_num}")
        self.geometry("700x750")
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()
        self.result = None

        self.fields = {
            'Line Item': StringVar(),
            'Part': StringVar(),
            'Description': StringVar(),
            'Qty': StringVar(),
            'UOM': StringVar()
        }
        
        self.pairs = [{
            'mpn_var': StringVar(),
            'mfr_var': StringVar(),
            'del_var': IntVar(value=0)
        }]
        self.entry_grid = []

        self.configure(bg="#EBF8FF")
        self.setup_ui()
        apply_panel_theme(self)

    def setup_ui(self):
        main_frame = Frame(self, padx=20, pady=10)
        main_frame.pack(fill="both", expand=True)

        Label(main_frame, text="Enter New Row Details", font=("Arial", 12, "bold")).pack(pady=(0, 10))

        # Basic Info Frame
        basic_f = LabelFrame(main_frame, text="Basic Information", padx=10, pady=10)
        basic_f.pack(fill="x", pady=5)

        self.basic_entry_widgets = []
        for i, (label, var) in enumerate(self.fields.items()):
            Label(basic_f, text=f"{label}:", font=("Arial", 10)).grid(row=i, column=0, sticky="w", pady=4)
            en = Entry(basic_f, textvariable=var, width=40, font=("Arial", 10))
            en.grid(row=i, column=1, sticky="w", padx=10, pady=4)
            self.basic_entry_widgets.append(en)
            
            en.bind("<FocusIn>", lambda e: e.widget.selection_range(0, 'end'))
            en.bind("<Up>", lambda e, idx=i: self._navigate_basic(idx, "up"))
            en.bind("<Down>", lambda e, idx=i: self._navigate_basic(idx, "down"))

        # MPN/MFR Pairs Frame
        pair_f = LabelFrame(main_frame, text="MPN & MFR Assignment", padx=10, pady=10)
        pair_f.pack(fill="both", expand=True, pady=10)

        # Scrollable area for pairs
        container = Frame(pair_f)
        container.pack(fill="both", expand=True)
        
        self.canvas = Canvas(container, height=250, bg="#EBF8FF", highlightthickness=0)
        scrollbar = Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = Frame(self.canvas, bg="#EBF8FF")
        
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0,0), window=self.scrollable_frame, anchor="nw")
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width))
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self._render_pair_table()

        # Bottom Buttons
        btn_f = Frame(main_frame, pady=10)
        btn_f.pack(fill="x", side="bottom")
        
        Button(btn_f, text="✅ Add Row", command=self._on_add, bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"), width=15, pady=4).pack(side="right", padx=5)
        Button(btn_f, text="Cancel", command=self.destroy, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), width=15, pady=4).pack(side="right", padx=5)

    def _render_pair_table(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
            
        self.scrollable_frame.columnconfigure(0, weight=0)
        self.scrollable_frame.columnconfigure(1, weight=1)
        self.scrollable_frame.columnconfigure(2, weight=1)

        # Headers
        Label(self.scrollable_frame, text="Del", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, padx=5)
        Label(self.scrollable_frame, text="MPN (Required if row used)", font=("Segoe UI", 9, "bold")).grid(row=0, column=1, padx=5, sticky="w")
        Label(self.scrollable_frame, text="MFR (Forced if MPN typed)", font=("Segoe UI", 9, "bold")).grid(row=0, column=2, padx=5, sticky="w")
        
        self.entry_grid = []
        for i, pair in enumerate(self.pairs):
            Checkbutton(self.scrollable_frame, variable=pair['del_var']).grid(row=i+1, column=0, padx=5)
            
            ent_mpn = Entry(self.scrollable_frame, textvariable=pair['mpn_var'], font=("Segoe UI", 9))
            ent_mpn.grid(row=i+1, column=1, padx=5, pady=2, sticky="ew")
            
            ent_mfr = Entry(self.scrollable_frame, textvariable=pair['mfr_var'], font=("Segoe UI", 9))
            ent_mfr.grid(row=i+1, column=2, padx=5, pady=2, sticky="ew")
            
            self.entry_grid.append((ent_mpn, ent_mfr))
            
            # Highlight on focus
            ent_mpn.bind("<FocusIn>", lambda e: e.widget.selection_range(0, 'end'))
            ent_mfr.bind("<FocusIn>", lambda e: e.widget.selection_range(0, 'end'))
            
            # Navigation
            ent_mpn.bind("<Up>", lambda e, r=i, c=0: self._navigate_pairs(r, c, "up"))
            ent_mpn.bind("<Down>", lambda e, r=i, c=0: self._navigate_pairs(r, c, "down"))
            ent_mpn.bind("<Left>", lambda e, r=i, c=0: self._navigate_pairs(r, c, "left"))
            ent_mpn.bind("<Right>", lambda e, r=i, c=0: self._navigate_pairs(r, c, "right"))
            
            ent_mfr.bind("<Up>", lambda e, r=i, c=1: self._navigate_pairs(r, c, "up"))
            ent_mfr.bind("<Down>", lambda e, r=i, c=1: self._navigate_pairs(r, c, "down"))
            ent_mfr.bind("<Left>", lambda e, r=i, c=1: self._navigate_pairs(r, c, "left"))
            ent_mfr.bind("<Right>", lambda e, r=i, c=1: self._navigate_pairs(r, c, "right"))

        # Pair control buttons
        next_row = len(self.pairs) + 1
        ctrl_frame = Frame(self.scrollable_frame, pady=5)
        ctrl_frame.grid(row=next_row, column=1, columnspan=2, sticky="w")
        
        btn_add = Button(ctrl_frame, text="➕ Add New Pair", command=self._add_pair, font=("Segoe UI", 10, "bold"), bg="#1A365D", fg="white", padx=12, pady=6)
        btn_add.pack(side="left", padx=5)
        
        btn_del = Button(ctrl_frame, text="🗑️ Delete Selected", command=self._delete_selected_pairs, font=("Segoe UI", 10, "bold"), bg="#dc3545", fg="white", padx=12, pady=6)
        btn_del.pack(side="left", padx=5)

        apply_panel_theme(self.scrollable_frame)

    def _navigate_basic(self, idx, direction):
        if direction == "up" and idx > 0:
            self.basic_entry_widgets[idx-1].focus_set()
        elif direction == "down":
            if idx < len(self.basic_entry_widgets) - 1:
                self.basic_entry_widgets[idx+1].focus_set()
            elif self.entry_grid:
                self.entry_grid[0][0].focus_set()

    def _navigate_pairs(self, row, col, direction):
        if direction == "up":
            if row > 0:
                self.entry_grid[row-1][col].focus_set()
            else:
                self.basic_entry_widgets[-1].focus_set()
        elif direction == "down":
            if row < len(self.entry_grid) - 1:
                self.entry_grid[row+1][col].focus_set()
        elif direction == "left":
            if col == 1:
                self.entry_grid[row][0].focus_set()
        elif direction == "right":
            if col == 0:
                self.entry_grid[row][1].focus_set()

    def _add_pair(self):
        if self.pairs:
            last_pair = self.pairs[-1]
            if not last_pair['mpn_var'].get().strip() or not last_pair['mfr_var'].get().strip():
                show_error("Incomplete Pair", "Please fill in both MPN and MFR for the current row before adding a new one.", parent=self)
                return

        self.pairs.append({
            'mpn_var': StringVar(),
            'mfr_var': StringVar(),
            'del_var': IntVar(value=0)
        })
        self._render_pair_table()
        if self.entry_grid:
            self.entry_grid[-1][0].focus_set()

    def _delete_selected_pairs(self):
        self.pairs = [p for p in self.pairs if p['del_var'].get() == 0]
        if not self.pairs: # Always keep at least one empty
             self.pairs.append({'mpn_var': StringVar(), 'mfr_var': StringVar(), 'del_var': IntVar(value=0)})
        self._render_pair_table()

    def _on_add(self):
        # Validation
        part = self.fields['Part'].get().strip()
        qty = self.fields['Qty'].get().strip()

        if not part:
            show_error("Validation Error", "Part Number is required.")
            return

        final_mpns = []
        final_mfrs = []
        for i, pair in enumerate(self.pairs):
            mpn = pair['mpn_var'].get().strip()
            mfr = pair['mfr_var'].get().strip()
            
            if mpn and not mfr:
                show_error("Validation Error", f"Row {i+1}: MPN and MFR must be maintained as a pair.\nSince MPN is entered, MFR is required.")
                return
            if mfr and not mpn:
                show_error("Validation Error", f"Row {i+1}: MPN and MFR must be maintained as a pair.\nSince MFR is entered, MPN is required.")
                return
                
            if mpn:
                final_mpns.append(mpn)
                final_mfrs.append(mfr)

        try:
            if qty:
                float(qty)
            else:
                self.fields['Qty'].set("0")
        except ValueError:
            show_error("Validation Error", "Qty must be a number.")
            return

        self.result = {k: v.get().strip() for k, v in self.fields.items()}
        self.result['MPN'] = ", ".join(final_mpns)
        self.result['MFR'] = ", ".join(final_mfrs)
        self.destroy()

class MPNMFRAssignmentDialog(tk.Toplevel):
    def __init__(self, parent, part_number, mpn_str, mfr_str):
        super().__init__(parent)
        self.title(f"Edit MPN & MFR for {part_number}")
        self.geometry("620x500")
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        
        # Split pairs
        mpns = [m.strip() for m in str(mpn_str).split(",") if m.strip()]
        mfrs = [m.strip() for m in str(mfr_str).split(",") if m.strip()]
        
        # Pad mfrs
        while len(mfrs) < len(mpns): mfrs.append("")
        while len(mpns) < len(mfrs): mpns.append("")
        
        self.pairs = []
        for mpn, mfr in zip(mpns, mfrs):
            self.pairs.append({
                'mpn_var': StringVar(value=mpn),
                'mfr_var': StringVar(value=mfr),
                'del_var': IntVar(value=0)
            })
            
        self.configure(bg="#EBF8FF")
        self.setup_ui()
        apply_panel_theme(self)
        
    def setup_ui(self):
        main_frame = Frame(self, padx=15, pady=15)
        main_frame.pack(fill="both", expand=True)
        
        Label(main_frame, text="Manage MPN and MFR Pairs", font=("Segoe UI", 11, "bold")).pack(pady=(0, 10))
        
        # Scrollable area for the "Excel-like" table
        container = Frame(main_frame)
        container.pack(fill="both", expand=True)
        
        self.canvas = Canvas(container, bg="#EBF8FF", highlightthickness=0)
        scrollbar = Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = Frame(self.canvas, bg="#EBF8FF")
        
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0,0), window=self.scrollable_frame, anchor="nw")
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width))
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self._render_table()
        
        btn_frame = Frame(main_frame)
        btn_frame.pack(fill="x", side="bottom", pady=(10, 0))
        Button(btn_frame, text="✅ Save & Apply", command=self._on_save, bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"), width=20, pady=4).pack(side="right", padx=5)
        Button(btn_frame, text="Cancel", command=self.destroy, bg="#1A365D", fg="white", font=("Segoe UI", 10, "bold"), width=15, pady=4).pack(side="right", padx=5)

    def _render_table(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
            
        self.scrollable_frame.columnconfigure(0, weight=0)
        self.scrollable_frame.columnconfigure(1, weight=1)
        self.scrollable_frame.columnconfigure(2, weight=1)

        # Headers
        Label(self.scrollable_frame, text="Del", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, padx=5)
        Label(self.scrollable_frame, text="MPN (Required if row used)", font=("Segoe UI", 9, "bold")).grid(row=0, column=1, padx=5, sticky="w")
        Label(self.scrollable_frame, text="MFR (Forced if MPN typed)", font=("Segoe UI", 9, "bold")).grid(row=0, column=2, padx=5, sticky="w")
        
        self.entry_grid = []
        for i, pair in enumerate(self.pairs):
            Checkbutton(self.scrollable_frame, variable=pair['del_var']).grid(row=i+1, column=0, padx=5)
            
            ent_mpn = Entry(self.scrollable_frame, textvariable=pair['mpn_var'], font=("Segoe UI", 9))
            ent_mpn.grid(row=i+1, column=1, padx=5, pady=2, sticky="ew")
            
            ent_mfr = Entry(self.scrollable_frame, textvariable=pair['mfr_var'], font=("Segoe UI", 9))
            ent_mfr.grid(row=i+1, column=2, padx=5, pady=2, sticky="ew")
            
            self.entry_grid.append((ent_mpn, ent_mfr))
            
            # Highlight on focus
            ent_mpn.bind("<FocusIn>", lambda e: e.widget.selection_range(0, 'end'))
            ent_mfr.bind("<FocusIn>", lambda e: e.widget.selection_range(0, 'end'))
            
            # Navigation
            ent_mpn.bind("<Up>", lambda e, r=i, c=0: self._navigate(r, c, "up"))
            ent_mpn.bind("<Down>", lambda e, r=i, c=0: self._navigate(r, c, "down"))
            ent_mpn.bind("<Left>", lambda e, r=i, c=0: self._navigate(r, c, "left"))
            ent_mpn.bind("<Right>", lambda e, r=i, c=0: self._navigate(r, c, "right"))
            
            ent_mfr.bind("<Up>", lambda e, r=i, c=1: self._navigate(r, c, "up"))
            ent_mfr.bind("<Down>", lambda e, r=i, c=1: self._navigate(r, c, "down"))
            ent_mfr.bind("<Left>", lambda e, r=i, c=1: self._navigate(r, c, "left"))
            ent_mfr.bind("<Right>", lambda e, r=i, c=1: self._navigate(r, c, "right"))

            # Validation logic: If MPN entered, force MFR
            ent_mpn.bind("<FocusOut>", lambda e, p=pair: self._validate_pairing(p))
            ent_mfr.bind("<FocusOut>", lambda e, p=pair: self._validate_pairing(p))

        # Add buttons right after the last row in the scrollable area
        next_row = len(self.pairs) + 1
        ctrl_frame = Frame(self.scrollable_frame, pady=10)
        ctrl_frame.grid(row=next_row, column=1, columnspan=2, sticky="w")
        
        btn_add = Button(ctrl_frame, text="➕ Add New Pair", command=self._add_pair, font=("Segoe UI", 10, "bold"), bg="#1A365D", fg="white", padx=12, pady=6)
        btn_add.pack(side="left", padx=5)
        
        btn_del = Button(ctrl_frame, text="🗑️ Delete Selected", command=self._delete_selected, font=("Segoe UI", 10, "bold"), bg="#dc3545", fg="white", padx=12, pady=6)
        btn_del.pack(side="left", padx=5)

        apply_panel_theme(self.scrollable_frame)

    def _navigate(self, row, col, direction):
        if direction == "up":
            if row > 0:
                self.entry_grid[row-1][col].focus_set()
        elif direction == "down":
            if row < len(self.entry_grid) - 1:
                self.entry_grid[row+1][col].focus_set()
        elif direction == "left":
            if col == 1:
                self.entry_grid[row][0].focus_set()
            elif row > 0:
                self.entry_grid[row-1][1].focus_set()
        elif direction == "right":
            if col == 0:
                self.entry_grid[row][1].focus_set()
            elif row < len(self.entry_grid) - 1:
                self.entry_grid[row+1][0].focus_set()

    def _validate_pairing(self, pair):
        mpn = pair['mpn_var'].get().strip()
        mfr = pair['mfr_var'].get().strip()
        
        # We'll do the actual "force" check on Save, but we can highlight here
        pass

    def _add_pair(self):
        if self.pairs:
            last_pair = self.pairs[-1]
            if not last_pair['mpn_var'].get().strip() or not last_pair['mfr_var'].get().strip():
                show_error("Incomplete Pair", "Please fill in both MPN and MFR for the current row before adding a new one.", parent=self)
                return

        self.pairs.append({
            'mpn_var': StringVar(),
            'mfr_var': StringVar(),
            'del_var': IntVar(value=0)
        })
        self._render_table()
        # Focus on the new row
        if self.entry_grid:
            self.entry_grid[-1][0].focus_set()

    def _delete_selected(self):
        self.pairs = [p for p in self.pairs if p['del_var'].get() == 0]
        self._render_table()

    def _on_save(self):
        final_mpns = []
        final_mfrs = []
        
        for i, pair in enumerate(self.pairs):
            mpn = pair['mpn_var'].get().strip()
            mfr = pair['mfr_var'].get().strip()
            
            if mpn and not mfr:
                show_error("Validation Error", f"Row {i+1}: MPN and MFR must be maintained as a pair.\nSince MPN is entered, MFR is required.")
                return
            if mfr and not mpn:
                show_error("Validation Error", f"Row {i+1}: MPN and MFR must be maintained as a pair.\nSince MFR is entered, MPN is required.")
                return
                
            if mpn:
                final_mpns.append(mpn)
                final_mfrs.append(mfr)
                
        self.result = (", ".join(final_mpns), ", ".join(final_mfrs))
        self.destroy()

ADDONS_FILE = os.path.join(MASTER_DATA_DIR, "Inhouse Addons.json") if 'MASTER_DATA_DIR' in globals() else os.path.join(os.path.dirname(os.path.abspath(__file__)), "Master Data", "Inhouse Addons.json")

def load_addons():
    if os.path.exists(ADDONS_FILE):
        try:
            with open(ADDONS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
        except Exception as e:
            print(f"Error loading addons: {e}")
    # Default initial presets
    return [
        {
            "Line Item": 1,
            "Part": "IH-TAPE-01",
            "Description": "In-house Adhesive Tape",
            "MFR": "In-house",
            "MPN": "IH-TAPE-01",
            "Qty": "1",
            "UOM": "PCS"
        },
        {
            "Line Item": 1,
            "Part": "IH-LABEL-02",
            "Description": "In-house Label",
            "MFR": "In-house",
            "MPN": "IH-LABEL-02",
            "Qty": "1",
            "UOM": "PCS"
        }
    ]

def save_addons(addons):
    try:
        # Ensure directory exists
        os.makedirs(os.path.dirname(ADDONS_FILE), exist_ok=True)
        with open(ADDONS_FILE, "w", encoding="utf-8") as f:
            json.dump(addons, f, indent=4)
    except Exception as e:
        print(f"Error saving addons: {e}")

class BulkAddRowDialog(tk.Toplevel):
    def __init__(self, parent, assemblies=None):
        super().__init__(parent)
        self._skip_autofit = True
        self.title("Bulk Add New Line Item to Selected Assemblies")
        
        # Fixed compact dialog geometry centered on screen
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w, h = 820, 750
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(750, 680)
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()
        self.result = None

        self.all_assemblies = [str(a) for a in (assemblies or [])]
        self.target_assemblies = list(self.all_assemblies)

        # Load master data list of inhouse addons
        self.addons = load_addons()

        self.fields = {
            'Line Item': StringVar(),
            'Part': StringVar(),
            'Description': StringVar(),
            'MFR': StringVar(),
            'MPN': StringVar(),
            'Qty': StringVar(value="1"),
            'UOM': StringVar(value="PCS")
        }

        self.configure(bg="#EBF8FF")
        self.setup_ui()
        apply_panel_theme(self)

    def setup_ui(self):
        # ── ALWAYS-VISIBLE BOTTOM BUTTON BAR (packed first to anchor it) ──
        btn_bar = Frame(self, bg="#EBF8FF", pady=8, padx=20)
        btn_bar.pack(side="bottom", fill="x")

        Button(btn_bar, text="✅ Bulk Add to Selected", command=self._on_add,
               bg="#28a745", fg="white", font=("Arial", 10, "bold"), width=24).pack(side="right", padx=5)
        Button(btn_bar, text="Cancel", command=self.destroy, width=15).pack(side="right", padx=5)

        # Thin separator above buttons
        Frame(self, bg="#CBD5E0", height=1).pack(side="bottom", fill="x")

        # ── SCROLLABLE CONTENT AREA ──
        canvas = tk.Canvas(self, bg="#EBF8FF", highlightthickness=0, borderwidth=0)
        v_scroll = Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=v_scroll.set)
        v_scroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        main_frame = Frame(canvas, bg="#EBF8FF", padx=20, pady=10)
        canvas_window = canvas.create_window((0, 0), window=main_frame, anchor="nw")

        def _on_frame_resize(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        main_frame.bind("<Configure>", _on_frame_resize)

        def _on_canvas_resize(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind("<Configure>", _on_canvas_resize)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # ── TITLE ──
        Label(main_frame, text="Enter Details to Bulk Add to Selected Assemblies",
              font=("Segoe UI", 12, "bold"), fg="#1A365D", bg="#EBF8FF").pack(pady=(0, 8))

        # ── DUAL LISTBOX ASSEMBLY SELECTION ──
        assy_sec = LabelFrame(main_frame, text=" 🎯 Target Assemblies Selection ",
                              font=("Segoe UI", 10, "bold"), padx=10, pady=8, fg="#1A365D", bg="#ffffff")
        assy_sec.pack(fill="x", pady=(0, 8))

        assy_grid = Frame(assy_sec, bg="#ffffff")
        assy_grid.pack(fill="x", expand=True)

        # Left Column: Available Assemblies with Search Filter
        left_col = Frame(assy_grid, bg="#ffffff")
        left_col.grid(row=0, column=0, sticky="nsew", padx=5)
        assy_grid.columnconfigure(0, weight=1)

        search_hdr = Frame(left_col, bg="#ffffff")
        search_hdr.pack(fill="x", pady=(0, 4))
        Label(search_hdr, text="🔍 Search Available:", font=("Segoe UI", 9, "bold"), bg="#ffffff", fg="#2D3748").pack(side="left")

        self.assy_search_var = StringVar()
        self.assy_search_var.trace_add("write", self._filter_available_assemblies)
        search_entry = Entry(search_hdr, textvariable=self.assy_search_var, font=("Segoe UI", 9), width=20)
        search_entry.pack(side="left", padx=5)

        self.lbl_avail_count = Label(search_hdr, text=f"({len(self.all_assemblies)})", font=("Segoe UI", 9, "italic"), bg="#ffffff", fg="#718096")
        self.lbl_avail_count.pack(side="right")

        left_box_frame = Frame(left_col, bg="#ffffff")
        left_box_frame.pack(fill="both", expand=True)

        self.lb_available = tk.Listbox(left_box_frame, selectmode="extended", font=("Segoe UI", 9), height=6, exportselection=False)
        sb_avail = Scrollbar(left_box_frame, orient="vertical", command=self.lb_available.yview)
        self.lb_available.configure(yscrollcommand=sb_avail.set)
        self.lb_available.pack(side="left", fill="both", expand=True)
        sb_avail.pack(side="right", fill="y")
        self.lb_available.bind("<Double-Button-1>", lambda e: self._add_selected_assemblies())

        # Center Column: Action Transfer Buttons
        mid_col = Frame(assy_grid, bg="#ffffff", padx=8)
        mid_col.grid(row=0, column=1, sticky="ns", padx=2)

        btn_add_sel = Button(mid_col, text="Add ▶", font=("Segoe UI", 9, "bold"), bg="#2B6CB0", fg="white", bd=0, relief="flat", cursor="hand2", command=self._add_selected_assemblies, width=12, pady=3)
        btn_add_sel.pack(pady=(10, 3))

        btn_add_all = Button(mid_col, text="Add All ⏩", font=("Segoe UI", 9, "bold"), bg="#2B6CB0", fg="white", bd=0, relief="flat", cursor="hand2", command=self._add_all_assemblies, width=12, pady=3)
        btn_add_all.pack(pady=3)

        btn_rem_sel = Button(mid_col, text="◀ Remove", font=("Segoe UI", 9, "bold"), bg="#718096", fg="white", bd=0, relief="flat", cursor="hand2", command=self._remove_selected_assemblies, width=12, pady=3)
        btn_rem_sel.pack(pady=3)

        btn_rem_all = Button(mid_col, text="⏪ Remove All", font=("Segoe UI", 9, "bold"), bg="#E53E3E", fg="white", bd=0, relief="flat", cursor="hand2", command=self._remove_all_assemblies, width=12, pady=3)
        btn_rem_all.pack(pady=3)

        # Right Column: Target Selected Assemblies
        right_col = Frame(assy_grid, bg="#ffffff")
        right_col.grid(row=0, column=2, sticky="nsew", padx=5)
        assy_grid.columnconfigure(2, weight=1)

        right_hdr = Frame(right_col, bg="#ffffff")
        right_hdr.pack(fill="x", pady=(0, 4))

        Label(right_hdr, text="🎯 Selected Target Assemblies:", font=("Segoe UI", 9, "bold"), bg="#ffffff", fg="#2F855A").pack(side="left")

        self.lbl_target_count = Label(right_hdr, text=f"({len(self.target_assemblies)} / {len(self.all_assemblies)})", font=("Segoe UI", 9, "bold"), bg="#ffffff", fg="#2F855A")
        self.lbl_target_count.pack(side="right")

        right_box_frame = Frame(right_col, bg="#ffffff")
        right_box_frame.pack(fill="both", expand=True)

        self.lb_target = tk.Listbox(right_box_frame, selectmode="extended", font=("Segoe UI", 9), height=6, exportselection=False)
        sb_target = Scrollbar(right_box_frame, orient="vertical", command=self.lb_target.yview)
        self.lb_target.configure(yscrollcommand=sb_target.set)
        self.lb_target.pack(side="left", fill="both", expand=True)
        sb_target.pack(side="right", fill="y")
        self.lb_target.bind("<Double-Button-1>", lambda e: self._remove_selected_assemblies())

        self._refresh_assembly_lists()

        # Preset Frame
        preset_f = LabelFrame(main_frame, text="Select In-house Add-on Preset", padx=10, pady=6)
        preset_f.pack(fill="x", pady=4)

        Label(preset_f, text="Select Preset:", font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=4)
        
        # Populate Combobox values
        self.preset_display_values = ["-- Custom (None) --"] + [f"{a['Part']} - {a['Description']}" for a in self.addons]
        self.preset_combo = Combobox(preset_f, values=self.preset_display_values, state="readonly", width=50, font=("Arial", 10))
        self.preset_combo.grid(row=0, column=1, sticky="w", padx=10, pady=4)
        self.preset_combo.set("-- Custom (None) --")
        self.preset_combo.bind("<<ComboboxSelected>>", self._on_preset_select)

        # Basic Info Frame
        basic_f = LabelFrame(main_frame, text="Item Details", padx=10, pady=6)
        basic_f.pack(fill="both", expand=True, pady=4)

        self.entry_widgets = {}
        for i, (label, var) in enumerate(self.fields.items()):
            Label(basic_f, text=f"{label}:", font=("Arial", 10)).grid(row=i, column=0, sticky="w", pady=4)
            en = Entry(basic_f, textvariable=var, width=50, font=("Arial", 10))
            en.grid(row=i, column=1, sticky="w", padx=10, pady=4)
            self.entry_widgets[label] = en
            
            en.bind("<FocusIn>", lambda e: e.widget.selection_range(0, 'end'))
            en.bind("<Up>", lambda e, idx=i: self._navigate_fields(idx, "up"))
            en.bind("<Down>", lambda e, idx=i: self._navigate_fields(idx, "down"))

        # Preset Management Frame
        maint_f = LabelFrame(main_frame, text="Preset Maintenance", padx=10, pady=6)
        maint_f.pack(fill="x", pady=4)

        Button(maint_f, text="💾 Save/Update as Preset", command=self._save_preset, bg="#3182ce", fg="white", font=("Arial", 9, "bold"), width=25).pack(side="left", padx=10)
        Button(maint_f, text="🗑️ Delete Selected Preset", command=self._delete_preset, bg="#dc3545", fg="white", font=("Arial", 9, "bold"), width=25).pack(side="left", padx=10)

        # Extra bottom padding inside scroll area
        Frame(main_frame, bg="#EBF8FF", height=10).pack()

    def _filter_available_assemblies(self, *args):
        query = self.assy_search_var.get().strip().lower()
        self.lb_available.delete(0, 'end')
        filtered = [a for a in self.all_assemblies if not query or query in a.lower()]
        for a in filtered:
            self.lb_available.insert('end', a)
        self.lbl_avail_count.config(text=f"({len(filtered)} / {len(self.all_assemblies)})")

    def _refresh_assembly_lists(self):
        self._filter_available_assemblies()
        self.lb_target.delete(0, 'end')
        for a in self.target_assemblies:
            self.lb_target.insert('end', a)
        self.lbl_target_count.config(text=f"({len(self.target_assemblies)} / {len(self.all_assemblies)})")

    def _add_selected_assemblies(self):
        sel_indices = self.lb_available.curselection()
        if not sel_indices: return
        added_any = False
        target_set = set(self.target_assemblies)
        for idx in sel_indices:
            val = self.lb_available.get(idx)
            if val not in target_set:
                self.target_assemblies.append(val)
                target_set.add(val)
                added_any = True
        if added_any:
            self._refresh_assembly_lists()

    def _add_all_assemblies(self):
        query = self.assy_search_var.get().strip().lower()
        filtered = [a for a in self.all_assemblies if not query or query in a.lower()]
        target_set = set(self.target_assemblies)
        added_any = False
        for val in filtered:
            if val not in target_set:
                self.target_assemblies.append(val)
                target_set.add(val)
                added_any = True
        if added_any:
            self._refresh_assembly_lists()

    def _remove_selected_assemblies(self):
        sel_indices = self.lb_target.curselection()
        if not sel_indices: return
        vals_to_remove = set(self.lb_target.get(i) for i in sel_indices)
        self.target_assemblies = [a for a in self.target_assemblies if a not in vals_to_remove]
        self._refresh_assembly_lists()

    def _remove_all_assemblies(self):
        self.target_assemblies.clear()
        self._refresh_assembly_lists()

    def _on_preset_select(self, event=None):
        val = self.preset_combo.get()
        if val == "-- Custom (None) --":
            for label, var in self.fields.items():
                if label not in ('Qty', 'UOM'):
                    var.set("")
                elif label == 'Qty':
                    var.set("1")
                elif label == 'UOM':
                    var.set("PCS")
            return

        selected_idx = self.preset_combo.current() - 1
        if 0 <= selected_idx < len(self.addons):
            addon = self.addons[selected_idx]
            self.fields['Line Item'].set(addon.get('Line Item', ''))
            self.fields['Part'].set(addon.get('Part', ''))
            self.fields['Description'].set(addon.get('Description', ''))
            self.fields['MFR'].set(addon.get('MFR', ''))
            self.fields['MPN'].set(addon.get('MPN', ''))
            self.fields['Qty'].set(addon.get('Qty', '1'))
            self.fields['UOM'].set(addon.get('UOM', 'PCS'))

    def _navigate_fields(self, idx, direction):
        keys = list(self.fields.keys())
        if direction == "up" and idx > 0:
            self.entry_widgets[keys[idx-1]].focus_set()
        elif direction == "down" and idx < len(keys) - 1:
            self.entry_widgets[keys[idx+1]].focus_set()

    def _save_preset(self):
        line = self.fields['Line Item'].get().strip()
        part = self.fields['Part'].get().strip()
        desc = self.fields['Description'].get().strip()
        mfr = self.fields['MFR'].get().strip()
        mpn = self.fields['MPN'].get().strip()
        qty = self.fields['Qty'].get().strip()
        uom = self.fields['UOM'].get().strip()

        if not part:
            show_error("Validation Error", "Part Number is required to save a preset.", parent=self)
            return

        existing_idx = None
        for i, a in enumerate(self.addons):
            if a.get('Part') == part:
                existing_idx = i
                break

        addon_data = {
            "Line Item": line,
            "Part": part,
            "Description": desc,
            "MFR": mfr,
            "MPN": mpn,
            "Qty": qty,
            "UOM": uom
        }

        if existing_idx is not None:
            self.addons[existing_idx] = addon_data
            msg = f"Preset for part '{part}' updated successfully!"
        else:
            self.addons.append(addon_data)
            msg = f"Preset for part '{part}' saved successfully!"

        save_addons(self.addons)

        self.preset_display_values = ["-- Custom (None) --"] + [f"{a['Part']} - {a['Description']}" for a in self.addons]
        self.preset_combo.config(values=self.preset_display_values)
        self.preset_combo.set(f"{part} - {desc}")

        show_info("Success", msg, parent=self)

    def _delete_preset(self):
        val = self.preset_combo.get()
        if val == "-- Custom (None) --":
            show_error("Selection Error", "Please select a valid preset to delete.", parent=self)
            return

        selected_idx = self.preset_combo.current() - 1
        if 0 <= selected_idx < len(self.addons):
            part = self.addons[selected_idx].get('Part', '')
            if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete preset '{part}'?", parent=self):
                self.addons.pop(selected_idx)
                save_addons(self.addons)

                self.preset_display_values = ["-- Custom (None) --"] + [f"{a['Part']} - {a['Description']}" for a in self.addons]
                self.preset_combo.config(values=self.preset_display_values)
                self.preset_combo.set("-- Custom (None) --")
                self._on_preset_select()

                show_info("Deleted", f"Preset '{part}' deleted successfully.", parent=self)

    def _on_add(self):
        part = self.fields['Part'].get().strip()
        qty = self.fields['Qty'].get().strip()

        if not part:
            show_error("Validation Error", "Part Number is required.", parent=self)
            return

        if not self.target_assemblies:
            show_error("Validation Error", "Please select at least one target assembly from the list.", parent=self)
            return

        try:
            if qty:
                float(qty)
            else:
                self.fields['Qty'].set("0")
        except ValueError:
            show_error("Validation Error", "Qty must be a number.", parent=self)
            return

        self.result = {k: v.get().strip() for k, v in self.fields.items()}
        self.result["target_assemblies"] = list(self.target_assemblies)
        self.destroy()

class BOMDatabaseSearchPanel(BasePanel):
    def __init__(self, master, title="BOM Sourcing", only_assigned_moqs=False, is_target_price=False, is_dispatch=False):
        super().__init__(master)
        self.result = None
        self.bom_records = []
        self.panel_title = title
        self.only_assigned_moqs = only_assigned_moqs
        self.is_target_price = is_target_price
        self.is_dispatch = is_dispatch
        
        self.current_page = 0
        self.page_size = 100
        self.filtered_indices = []
        self._sort_rules = []
        
        self._load_data()
        self._create_widgets()
        
        import tkinter as tk
        self.status_bar = tk.Label(self, text="", font=("Arial", 10, "bold"), anchor="center", pady=5)
        self.status_bar.pack(side="bottom", fill="x")
        
    def _load_data(self):
        from utils import BOM_DATA_DIR
        import os, sys, json
        from datetime import datetime
        
        try:
            from revert_workflow import resolve_assigned_pics
        except ImportError:
            try:
                pm_dir = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Project Management"))
                if pm_dir not in sys.path:
                    sys.path.insert(0, pm_dir)
                from revert_workflow import resolve_assigned_pics
            except Exception:
                def resolve_assigned_pics(data, json_status):
                    return "-"

        if not os.path.exists(BOM_DATA_DIR):
            return
            
        assigned_pairs = {}
        meta_path = os.path.join(BOM_DATA_DIR, "assigned_moqs_metadata.json")
        if os.path.exists(meta_path):
            try:
                with open(meta_path, 'r', encoding='utf-8') as f:
                    meta = json.load(f)
                    for item in meta.get("completed_moqs", []):
                        assigned_pairs[(item.get("Customer"), item.get("RFQ"))] = item
            except Exception as e:
                print(f"Error reading assigned MOQs metadata: {e}")
            
        for root_dir, dirs, files in os.walk(BOM_DATA_DIR):
            for file in files:
                if file.endswith('.json') and not file.endswith('metadata.json'):
                    filepath = os.path.join(root_dir, file)
                    try:
                        with open(filepath, 'r', encoding='utf-8-sig') as f:
                            data = json.load(f)
                            
                        cust_name = data.get("Customer", "Unknown")
                        rfq_num = data.get("RFQ", "Unknown")
                        
                        # Skip records that have already been dispatched to next stage
                        json_status = data.get("status", "")
                        is_dispatched = json_status in ("pending_sourcing_and_cycle_time", "pending_costing", "completed")
                        if is_dispatched:
                            continue
                        
                        # Inspect actual Assigned MOQs in assemblies
                        has_moq = False
                        all_moq = True
                        any_assy = False
                        for assy in data.get("Assemblies", []):
                            any_assy = True
                            moqs = assy.get("Assigned MOQs", [])
                            if moqs and len(moqs) > 0:
                                has_moq = True
                            else:
                                all_moq = False
                        
                        if not has_moq:
                            g_moqs = data.get("Global MOQs", [])
                            if g_moqs and len(g_moqs) > 0:
                                has_moq = True

                        is_confirmed_moq = (cust_name, rfq_num) in assigned_pairs
                        if is_confirmed_moq:
                            has_moq = True

                        if not any_assy:
                            moq_status = "Pending"
                        elif all_moq or is_confirmed_moq:
                            moq_status = "Completed"
                        elif has_moq:
                            moq_status = "Partial"
                        else:
                            moq_status = "Pending"

                        # If panel requires assigned MOQs (Target Price & Dispatch), skip RFQs that do not have assigned MOQs yet
                        if self.only_assigned_moqs and not has_moq:
                            continue

                        mtime = os.path.getmtime(filepath)
                        dt_obj = datetime.fromtimestamp(mtime)
                        date_str = dt_obj.strftime("%d.%m.%Y")
                        time_str = dt_obj.strftime("%I:%M %p")
                        
                        assigned_by = resolve_assigned_pics(data, json_status)
                        if (not assigned_by or assigned_by in ("Unassigned", "-")) and (cust_name, rfq_num) in assigned_pairs:
                            assigned_by = assigned_pairs[(cust_name, rfq_num)].get("AssignedBy", "-")
                        if not assigned_by or assigned_by in ("Unassigned", "-"):
                            assigned_by = data.get("bom_assigned_by") or data.get("dispatched_by") or "-"

                        status = moq_status if (not self.is_target_price and not self.is_dispatch) else "Pending"
                                
                        if self.is_target_price or self.is_dispatch:
                            has_tp = False
                            all_tp = True
                            has_eau = False
                            all_eau = True
                            for assy in data.get("Assemblies", []):
                                assigned_moqs = assy.get("Assigned MOQs", [])
                                tp_dict = assy.get("Target Prices", {})
                                for moq in assigned_moqs:
                                    val = tp_dict.get(str(moq))
                                    try:
                                        if val is not None and str(val).strip() != "" and float(val) > 0.0:
                                            has_tp = True
                                        else:
                                            all_tp = False
                                    except ValueError:
                                        all_tp = False

                                eau_data = assy.get("EAU", {})
                                if isinstance(eau_data, (int, float, str)):
                                    try:
                                        if float(eau_data) > 0:
                                            has_eau = True
                                        else:
                                            all_eau = False
                                    except ValueError:
                                        all_eau = False
                                elif isinstance(eau_data, dict):
                                    for moq in assigned_moqs:
                                        val = eau_data.get(str(moq))
                                        try:
                                            if val is not None and str(val).strip() != "" and float(val) > 0.0:
                                                has_eau = True
                                            else:
                                                all_eau = False
                                        except ValueError:
                                            all_eau = False
                                else:
                                    all_eau = False

                            if not any_assy:
                                tp_status = "Pending"
                                eau_status = "Pending"
                            else:
                                tp_status = "Completed" if all_tp else ("Partial" if has_tp else "Pending")
                                eau_status = "Completed" if all_eau else ("Partial" if has_eau else "Pending")
                            status = tp_status
                        else:
                            tp_status = "Pending"
                            eau_status = "Pending"

                        self.bom_records.append({
                            "RFQ": rfq_num,
                            "Customer": cust_name,
                            "Date": date_str,
                            "Time": time_str,
                            "Status": status,
                            "MOQStatus": moq_status,
                            "TPStatus": tp_status,
                            "EAUStatus": eau_status,
                            "AssignedBy": assigned_by,
                            "raw_data": data,
                            "filepath": filepath,
                            "mtime": mtime
                        })
                    except:
                        pass

        def get_status_priority(r):
            moq_st = r.get("MOQStatus", "Pending")
            tp_st = r.get("TPStatus", "Pending")
            eau_st = r.get("EAUStatus", "Pending")

            if self.is_dispatch:
                if moq_st == "Pending" or tp_st == "Pending" or eau_st == "Pending":
                    return 0
                elif moq_st == "Partial" or tp_st == "Partial" or eau_st == "Partial":
                    return 1
                else:
                    return 2
            elif self.is_target_price:
                if tp_st == "Pending" and eau_st == "Pending":
                    return 0
                elif tp_st in ("Completed", "Assigned") and eau_st in ("Completed", "Assigned"):
                    return 2
                else:
                    return 1
            else:
                if moq_st in ("Completed", "Assigned"):
                    return 2
                elif moq_st == "Partial":
                    return 1
                else:
                    return 0

        self.bom_records.sort(key=lambda r: (get_status_priority(r), -r.get("mtime", 0)))
                        
    def _create_widgets(self):
        import tkinter as tk
        from tkinter import ttk
        
        # Header
        header = tk.Frame(self, bg="#1A365D")
        header.pack(fill="x", side="top")
        tk.Label(header, text=self.panel_title, font=("Segoe UI", 14, "bold"), fg="white", bg="#1A365D", pady=12).pack(side="left", padx=20)

        # Search Filters
        search_frame = tk.LabelFrame(self, text="Search History Database", padx=20, pady=10)
        search_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        filter_f = tk.Frame(search_frame)
        filter_f.pack(side="top", fill="x", pady=5)
        
        self.search_rfq_var = tk.StringVar()
        self.search_cust_var = tk.StringVar()
        self.search_date_var = tk.StringVar()
        self.search_moq_status_var = tk.StringVar(value="-- All --")
        self.search_tp_status_var = tk.StringVar(value="-- All --")
        self.search_eau_status_var = tk.StringVar(value="-- All --")
        self.search_pic_var = tk.StringVar()
        
        unique_rfqs = sorted(list(set(r["RFQ"] for r in self.bom_records if r.get("RFQ"))))
        unique_custs = sorted(list(set(r["Customer"] for r in self.bom_records if r.get("Customer"))))
        unique_pics = sorted(list(set(r.get("AssignedBy", "-") for r in self.bom_records if r.get("AssignedBy"))))

        tk.Label(filter_f, text="Search by RFQ Number:").grid(row=0, column=0, sticky="w", pady=4)
        cb_rfq = ttk.Combobox(filter_f, textvariable=self.search_rfq_var, values=[""] + unique_rfqs, width=38)
        cb_rfq.grid(row=0, column=1, sticky="w", padx=(5, 15), pady=4)
        
        tk.Label(filter_f, text="Search by Customer:").grid(row=1, column=0, sticky="w", pady=4)
        cb_cust = ttk.Combobox(filter_f, textvariable=self.search_cust_var, values=[""] + unique_custs, width=38)
        cb_cust.grid(row=1, column=1, sticky="w", padx=(5, 15), pady=4)
        
        tk.Label(filter_f, text="Search by Date (DD.MM.YYYY):").grid(row=2, column=0, sticky="w", pady=4)
        ent_date = tk.Entry(filter_f, textvariable=self.search_date_var, width=38)
        ent_date.grid(row=2, column=1, sticky="w", padx=(5, 15), pady=4)

        status_options = ["-- All --", "Pending", "Partial", "Completed"]
        status_bind_widgets = []

        if self.is_dispatch:
            tk.Label(filter_f, text="Filter by MOQ Status:").grid(row=0, column=2, sticky="w", pady=4)
            cb_moq_st = ttk.Combobox(filter_f, textvariable=self.search_moq_status_var, values=status_options, state="readonly", width=18)
            cb_moq_st.grid(row=0, column=3, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_moq_st)

            tk.Label(filter_f, text="Filter by Target Price Status:").grid(row=1, column=2, sticky="w", pady=4)
            cb_tp_st = ttk.Combobox(filter_f, textvariable=self.search_tp_status_var, values=status_options, state="readonly", width=18)
            cb_tp_st.grid(row=1, column=3, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_tp_st)

            tk.Label(filter_f, text="Filter by EAU Status:").grid(row=2, column=2, sticky="w", pady=4)
            cb_eau_st = ttk.Combobox(filter_f, textvariable=self.search_eau_status_var, values=status_options, state="readonly", width=18)
            cb_eau_st.grid(row=2, column=3, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_eau_st)

            tk.Label(filter_f, text="Filter by PIC:").grid(row=0, column=4, sticky="w", padx=(15, 5), pady=4)
            cb_pic = ttk.Combobox(filter_f, textvariable=self.search_pic_var, values=[""] + unique_pics, width=22)
            cb_pic.grid(row=0, column=5, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_pic)

            sort_row = 1
            sort_btn_col = 4
            sort_colspan = 2
        elif self.is_target_price:
            tk.Label(filter_f, text="Filter by Target Price Status:").grid(row=0, column=2, sticky="w", pady=4)
            cb_tp_st = ttk.Combobox(filter_f, textvariable=self.search_tp_status_var, values=status_options, state="readonly", width=18)
            cb_tp_st.grid(row=0, column=3, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_tp_st)

            tk.Label(filter_f, text="Filter by EAU Status:").grid(row=1, column=2, sticky="w", pady=4)
            cb_eau_st = ttk.Combobox(filter_f, textvariable=self.search_eau_status_var, values=status_options, state="readonly", width=18)
            cb_eau_st.grid(row=1, column=3, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_eau_st)

            tk.Label(filter_f, text="Filter by PIC:").grid(row=2, column=2, sticky="w", pady=4)
            cb_pic = ttk.Combobox(filter_f, textvariable=self.search_pic_var, values=[""] + unique_pics, width=22)
            cb_pic.grid(row=2, column=3, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_pic)

            sort_row = 0
            sort_btn_col = 4
            sort_colspan = 1
        else:
            tk.Label(filter_f, text="Filter by MOQ Status:").grid(row=0, column=2, sticky="w", pady=4)
            cb_moq_st = ttk.Combobox(filter_f, textvariable=self.search_moq_status_var, values=status_options, state="readonly", width=18)
            cb_moq_st.grid(row=0, column=3, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_moq_st)

            tk.Label(filter_f, text="Filter by PIC:").grid(row=1, column=2, sticky="w", pady=4)
            cb_pic = ttk.Combobox(filter_f, textvariable=self.search_pic_var, values=[""] + unique_pics, width=22)
            cb_pic.grid(row=1, column=3, sticky="w", padx=(5, 15), pady=4)
            status_bind_widgets.append(cb_pic)

            sort_row = 0
            sort_btn_col = 4
            sort_colspan = 1

        for cb in (cb_rfq, cb_cust) + tuple(status_bind_widgets):
            cb.bind("<<ComboboxSelected>>", lambda e: self.reset_page_and_filter())
            cb.bind("<KeyRelease>", lambda e: self.reset_page_and_filter())
        ent_date.bind("<KeyRelease>", lambda e: self.reset_page_and_filter())

        btn_sort = tk.Button(
            filter_f, text="Sort Records ⇅", command=self._open_sort_dialog,
            font=("Segoe UI", 9, "bold"), bg="#1A365D", fg="white",
            activebackground="#0077B6", activeforeground="white",
            padx=14, pady=5, bd=0, relief="flat", cursor="hand2"
        )
        btn_sort.grid(row=sort_row, column=sort_btn_col, columnspan=sort_colspan, padx=(15, 5), pady=(4, 2), sticky="ew")

        btn_clear = tk.Button(
            filter_f, text="🧹 Clear Filters", command=self.clear_filters,
            font=("Segoe UI", 9, "bold"), bg="#4A5568", fg="white",
            activebackground="#2D3748", activeforeground="white",
            padx=14, pady=5, bd=0, relief="flat", cursor="hand2"
        )
        btn_clear.grid(row=sort_row+1, column=sort_btn_col, rowspan=(2 if sort_row == 0 else 1), columnspan=sort_colspan, padx=(15, 5), pady=(2, 4), sticky="ew")

        # Pinned bottom actions frame
        btn_frame = tk.Frame(search_frame, pady=10)
        btn_frame.pack(side="bottom", fill="x")
        
        if self.only_assigned_moqs:
            if self.is_dispatch:
                self.btn_start = tk.Button(btn_frame, text="Dispatch RFQ", bg="#2ead4e", fg="white", font=("Arial", 11, "bold"), command=lambda: self._start_sourcing("dispatch_rfq"), width=20)
                self.btn_start.pack(side="right", padx=5)
                note_text = "Note: Please select a verified BOM from the list above to dispatch to Sourcing & Cycle Time."
            else:
                start_lbl = "Start" if self.is_target_price else "Start Sourcing"
                load_lbl = "Load" if self.is_target_price else "Load Sourcing"
                
                self.btn_start = tk.Button(btn_frame, text=start_lbl, bg="#2ead4e", fg="white", font=("Arial", 11, "bold"), command=lambda: self._start_sourcing("start_sourcing"), width=15)
                self.btn_start.pack(side="right", padx=5)
                
                if not self.is_target_price:
                    self.btn_load = tk.Button(btn_frame, text=load_lbl, bg="#bee3f8", fg="#1A365D", font=("Arial", 11, "bold"), command=lambda: self._start_sourcing("load_sourcing"), width=15, state="disabled")
                    self.btn_load.pack(side="right", padx=5)
                
                if not self.is_target_price:
                    self.btn_download = tk.Button(btn_frame, text="Download Excel File", bg="#fffde7", fg="#1A365D", font=("Arial", 11, "bold"), command=lambda: self._start_sourcing("download_excel"), width=20, state="disabled")
                    self.btn_download.pack(side="right", padx=5)
                
                note_text = "Note: Only verified BOMs that have completed MOQ assignment are shown here."
        else:
            self.btn_assign = tk.Button(btn_frame, text="Assign MOQs", bg="#2ead4e", fg="white", font=("Arial", 11, "bold"), command=lambda: self._start_sourcing("assign_moqs"), width=25)
            self.btn_assign.pack(side="right", padx=10)
            
            note_text = "Note: Please select a record from the list above to assign or edit MOQs."
            
        tk.Button(btn_frame, text="Cancel", command=self._on_cancel, width=15).pack(side="left", padx=10)
        
        # Instruction/Note label
        note_lbl = tk.Label(search_frame, text=note_text, font=("Arial", 9, "italic"), fg="gray")
        note_lbl.pack(side="bottom", pady=(5, 0))

        # Total Records Count Header
        pag_frame = tk.Frame(search_frame, bg="#EBF8FF", pady=5)
        pag_frame.pack(side="bottom", fill="x")
        
        self.lbl_total_records = tk.Label(pag_frame, text="Total Records: 0", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D")
        self.lbl_total_records.pack(side="right", padx=15)

        # Treeview
        tree_frame = tk.Frame(search_frame)
        tree_frame.pack(side="top", fill="both", expand=True, pady=10)
        
        if self.is_dispatch:
            cols = ("Customer", "RFQ", "Date", "Time", "MOQ_Status", "TP_Status", "EAU_Status", "AssignedBy")
            self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=10)
            self.tree.heading("Customer", text="Customer", command=lambda: self._on_header_click("Customer"))
            self.tree.heading("RFQ", text="RFQ Number", command=lambda: self._on_header_click("RFQ"))
            self.tree.heading("Date", text="Last Update (Date)", command=lambda: self._on_header_click("Date"))
            self.tree.heading("Time", text="Last Update (Time)", command=lambda: self._on_header_click("Time"))
            self.tree.heading("MOQ_Status", text="MOQ Assignation Status", command=lambda: self._on_header_click("MOQ_Status"))
            self.tree.heading("TP_Status", text="Target Price Status", command=lambda: self._on_header_click("TP_Status"))
            self.tree.heading("EAU_Status", text="EAU Status", command=lambda: self._on_header_click("EAU_Status"))
            self.tree.heading("AssignedBy", text="PIC", command=lambda: self._on_header_click("AssignedBy"))
            
            self.tree.column("Customer", width=140, anchor="center")
            self.tree.column("RFQ", width=110, anchor="center")
            self.tree.column("Date", width=120, anchor="center")
            self.tree.column("Time", width=120, anchor="center")
            self.tree.column("MOQ_Status", width=140, anchor="center")
            self.tree.column("TP_Status", width=130, anchor="center")
            self.tree.column("EAU_Status", width=110, anchor="center")
            self.tree.column("AssignedBy", width=100, anchor="center")
        elif self.is_target_price:
            cols = ("Customer", "RFQ", "Date", "Time", "TP_Status", "EAU_Status", "AssignedBy")
            self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=10)
            self.tree.heading("Customer", text="Customer", command=lambda: self._on_header_click("Customer"))
            self.tree.heading("RFQ", text="RFQ Number", command=lambda: self._on_header_click("RFQ"))
            self.tree.heading("Date", text="Last Update (Date)", command=lambda: self._on_header_click("Date"))
            self.tree.heading("Time", text="Last Update (Time)", command=lambda: self._on_header_click("Time"))
            self.tree.heading("TP_Status", text="Target Price Status", command=lambda: self._on_header_click("TP_Status"))
            self.tree.heading("EAU_Status", text="EAU Status", command=lambda: self._on_header_click("EAU_Status"))
            self.tree.heading("AssignedBy", text="PIC", command=lambda: self._on_header_click("AssignedBy"))
            
            self.tree.column("Customer", width=160, anchor="center")
            self.tree.column("RFQ", width=120, anchor="center")
            self.tree.column("Date", width=130, anchor="center")
            self.tree.column("Time", width=130, anchor="center")
            self.tree.column("TP_Status", width=140, anchor="center")
            self.tree.column("EAU_Status", width=120, anchor="center")
            self.tree.column("AssignedBy", width=110, anchor="center")
        else:
            cols = ("Customer", "RFQ", "Date", "Time", "Status", "AssignedBy")
            self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=10)
            self.tree.heading("Customer", text="Customer", command=lambda: self._on_header_click("Customer"))
            self.tree.heading("RFQ", text="RFQ Number", command=lambda: self._on_header_click("RFQ"))
            self.tree.heading("Date", text="Last Update (Date)", command=lambda: self._on_header_click("Date"))
            self.tree.heading("Time", text="Last Update (Time)", command=lambda: self._on_header_click("Time"))
            status_text = "MOQ Status"
            self.tree.heading("Status", text=status_text, command=lambda: self._on_header_click("Status"))
            self.tree.heading("AssignedBy", text="PIC", command=lambda: self._on_header_click("AssignedBy"))
            
            self.tree.column("Customer", width=180, anchor="center")
            self.tree.column("RFQ", width=130, anchor="center")
            self.tree.column("Date", width=140, anchor="center")
            self.tree.column("Time", width=140, anchor="center")
            self.tree.column("Status", width=120, anchor="center")
            self.tree.column("AssignedBy", width=120, anchor="center")
        
        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.config(yscrollcommand=sb.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        import sys
        _shared_dir = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "shared"))
        if _shared_dir not in sys.path:
            sys.path.insert(0, _shared_dir)
        try:
            from treeview_sort import attach_treeview_sort
            attach_treeview_sort(self.tree)
        except Exception:
            pass

        # Configure row highlight tags for status
        self.tree.tag_configure("assigned",   background="#d4edda", foreground="#155724")   # green
        self.tree.tag_configure("partial",    background="#fff3cd", foreground="#856404")   # amber
        self.tree.tag_configure("pending",    background="#cce5ff", foreground="#1A365D")   # blue
        self.tree.tag_configure("dispatched", background="#e2e8f0", foreground="#4a5568")   # grey

        self._filter_tree()

    def clear_filters(self):
        if hasattr(self, 'search_rfq_var'): self.search_rfq_var.set("")
        if hasattr(self, 'search_cust_var'): self.search_cust_var.set("")
        if hasattr(self, 'search_date_var'): self.search_date_var.set("")
        if hasattr(self, 'search_moq_status_var'): self.search_moq_status_var.set("-- All --")
        if hasattr(self, 'search_tp_status_var'): self.search_tp_status_var.set("-- All --")
        if hasattr(self, 'search_eau_status_var'): self.search_eau_status_var.set("-- All --")
        if hasattr(self, 'search_pic_var'): self.search_pic_var.set("")
        self.reset_page_and_filter()

    def _open_sort_dialog(self):
        try:
            from treeview_sort import SortRecordsDialog
        except ImportError:
            import sys
            _shared_dir = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "shared"))
            if _shared_dir not in sys.path:
                sys.path.insert(0, _shared_dir)
            from treeview_sort import SortRecordsDialog

        if self.is_dispatch:
            col_map = {
                "Customer": "Customer", "RFQ Number": "RFQ", "Last Update (Date)": "Date",
                "Last Update (Time)": "Time", "MOQ Assignation Status": "MOQ_Status",
                "Target Price Status": "TP_Status", "EAU Status": "EAU_Status", "PIC": "AssignedBy"
            }
        elif self.is_target_price:
            col_map = {
                "Customer": "Customer", "RFQ Number": "RFQ", "Last Update (Date)": "Date",
                "Last Update (Time)": "Time", "Target Price Status": "TP_Status",
                "EAU Status": "EAU_Status", "PIC": "AssignedBy"
            }
        else:
            col_map = {
                "Customer": "Customer", "RFQ Number": "RFQ", "Last Update (Date)": "Date",
                "Last Update (Time)": "Time", "MOQ Status": "Status", "PIC": "AssignedBy"
            }

        rev_map = {v: k for k, v in col_map.items()}
        curr_rules = []
        if getattr(self, "_sort_rules", None):
            for k, d in self._sort_rules:
                disp = rev_map.get(k)
                if disp: curr_rules.append((disp, d))

        dlg = SortRecordsDialog(self, list(col_map.keys()), current_sort_rules=curr_rules)
        self.wait_window(dlg)
        if dlg.result_sort_rules:
            mapped_rules = []
            for disp, d in dlg.result_sort_rules:
                k = col_map.get(disp)
                if k: mapped_rules.append((k, d))
            self._sort_rules = mapped_rules
        else:
            self._sort_rules = []

        self.current_page = 0
        self._filter_tree()

    def _on_header_click(self, col_key):
        """Toggle multi-level column sort rule when clicking table header."""
        if not hasattr(self, "_sort_rules") or self._sort_rules is None:
            self._sort_rules = []

        existing_idx = next((i for i, r in enumerate(self._sort_rules) if r[0] == col_key), -1)

        if existing_idx == 0:
            col, state = self._sort_rules[0]
            if state == 'asc':
                self._sort_rules[0] = (col, 'desc')
            else:
                self._sort_rules.pop(0)
        elif existing_idx > 0:
            col, state = self._sort_rules.pop(existing_idx)
            new_state = 'desc' if state == 'asc' else 'asc'
            self._sort_rules.insert(0, (col, new_state))
        else:
            self._sort_rules.insert(0, (col_key, 'asc'))

        self.current_page = 0
        self._filter_tree()

    def reset_page_and_filter(self, event=None):
        self.current_page = 0
        self._filter_tree()

    def _filter_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        rfq_q = self.search_rfq_var.get().lower()
        cust_q = self.search_cust_var.get().lower()
        date_q = self.search_date_var.get().lower()
        pic_q = getattr(self, "search_pic_var", None)
        pic_val = pic_q.get().strip().lower() if pic_q else ""

        moq_st_q = getattr(self, "search_moq_status_var", None)
        tp_st_q = getattr(self, "search_tp_status_var", None)
        eau_st_q = getattr(self, "search_eau_status_var", None)

        moq_val = moq_st_q.get() if moq_st_q else "-- All --"
        tp_val = tp_st_q.get() if tp_st_q else "-- All --"
        eau_val = eau_st_q.get() if eau_st_q else "-- All --"
        
        self.filtered_indices = []
        for i, r in enumerate(self.bom_records):
            if rfq_q and rfq_q not in r["RFQ"].lower(): continue
            if cust_q and cust_q not in r["Customer"].lower(): continue
            if date_q and date_q not in r["Date"].lower(): continue
            if pic_val and pic_val != "-- all --" and pic_val not in str(r.get("AssignedBy", "")).lower(): continue
            
            if moq_val and moq_val != "-- All --":
                st = r.get("MOQStatus", r.get("Status", "Pending"))
                if str(st).lower() != moq_val.lower():
                    continue

            if tp_val and tp_val != "-- All --":
                st = r.get("TPStatus", "Pending")
                if str(st).lower() != tp_val.lower():
                    continue

            if eau_val and eau_val != "-- All --":
                st = r.get("EAUStatus", "Pending")
                if str(st).lower() != eau_val.lower():
                    continue

            self.filtered_indices.append(i)

        if getattr(self, "_sort_rules", None):
            def _get_val(idx, k):
                v = self.bom_records[idx].get(k, "")
                if v is None: v = ""
                try: return (0, float(str(v).replace(',', '')))
                except (ValueError, TypeError): return (1, str(v).lower())

            for col_key, direction in reversed(self._sort_rules):
                rev = (direction == 'desc')
                self.filtered_indices.sort(key=lambda idx: _get_val(idx, col_key), reverse=rev)

        # Update Treeview header text with sort arrows ▲/▼ and level numbers
        if hasattr(self, 'tree') and self.tree.winfo_exists():
            if self.is_dispatch:
                col_map = {
                    "Customer": "Customer", "RFQ": "RFQ Number", "Date": "Last Update (Date)",
                    "Time": "Last Update (Time)", "MOQ_Status": "MOQ Assignation Status",
                    "TP_Status": "Target Price Status", "EAU_Status": "EAU Status", "AssignedBy": "PIC"
                }
            elif self.is_target_price:
                col_map = {
                    "Customer": "Customer", "RFQ": "RFQ Number", "Date": "Last Update (Date)",
                    "Time": "Last Update (Time)", "TP_Status": "Target Price Status",
                    "EAU_Status": "EAU Status", "AssignedBy": "PIC"
                }
            else:
                col_map = {
                    "Customer": "Customer", "RFQ": "RFQ Number", "Date": "Last Update (Date)",
                    "Time": "Last Update (Time)", "Status": "MOQ Status", "AssignedBy": "PIC"
                }

            sort_rules = getattr(self, '_sort_rules', None) or []
            rule_map = {col: (direction, idx + 1) for idx, (col, direction) in enumerate(sort_rules)}
            cols = self.tree['columns']
            for c in cols:
                base_title = col_map.get(c, c)
                if c in rule_map:
                    direction, priority = rule_map[c]
                    arrow = " ▲" if direction == 'asc' else " ▼"
                    if len(sort_rules) > 1:
                        title_text = f"{base_title}{arrow} ({priority})"
                    else:
                        title_text = f"{base_title}{arrow}"
                else:
                    title_text = base_title
                self.tree.heading(c, text=title_text)

        total_matches = len(self.filtered_indices)
        total_pages = max(1, ((total_matches - 1) // self.page_size) + 1)
        
        if self.current_page >= total_pages:
            self.current_page = total_pages - 1
        if self.current_page < 0:
            self.current_page = 0

        # Update Pagination widgets if they exist
        if hasattr(self, 'lbl_total_pages'):
            self.lbl_total_pages.config(text=f"of {total_pages}")
        if hasattr(self, 'lbl_total_records'):
            self.lbl_total_records.config(text=f"Total Records: {total_matches}")
        if hasattr(self, 'ent_page_num'):
            self.ent_page_num.delete(0, tk.END)
            self.ent_page_num.insert(0, str(self.current_page + 1))
        if hasattr(self, 'update_nav_buttons'):
            self.update_nav_buttons(total_pages)

        # Slice for current page
        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, total_matches)
        page_indices = self.filtered_indices[start_idx:end_idx]

        for i in page_indices:
            r = self.bom_records[i]
            status = r.get("Status", "Pending")
            moq_st = r.get("MOQStatus", "Pending")
            tp_st = r.get("TPStatus", "Pending")
            eau_st = r.get("EAUStatus", "Pending")

            if self.is_dispatch:
                if moq_st == "Pending" or tp_st == "Pending" or eau_st == "Pending":
                    tag = "pending"
                elif moq_st == "Partial" or tp_st == "Partial" or eau_st == "Partial":
                    tag = "partial"
                else:
                    tag = "assigned"
            elif self.is_target_price:
                if tp_st == "Pending" and eau_st == "Pending":
                    tag = "pending"
                elif tp_st in ("Completed", "Assigned") and eau_st in ("Completed", "Assigned"):
                    tag = "assigned"
                else:
                    tag = "partial"
            else:
                if moq_st in ("Completed", "Assigned"):
                    tag = "assigned"
                elif moq_st == "Partial":
                    tag = "partial"
                else:
                    tag = "pending"

            if self.is_dispatch:
                self.tree.insert("", "end", iid=str(i), values=(r["Customer"], r["RFQ"], r["Date"], r.get("Time", "-"), moq_st, tp_st, eau_st, r.get("AssignedBy", "-")), tags=(tag,))
            elif self.is_target_price:
                self.tree.insert("", "end", iid=str(i), values=(r["Customer"], r["RFQ"], r["Date"], r.get("Time", "-"), tp_st, eau_st, r.get("AssignedBy", "-")), tags=(tag,))
            else:
                self.tree.insert("", "end", iid=str(i), values=(r["Customer"], r["RFQ"], r["Date"], r.get("Time", "-"), status, r.get("AssignedBy", "-")), tags=(tag,))

    def goto_first_page(self):
        if self.current_page != 0:
            self.current_page = 0
            self._filter_tree()

    def goto_prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self._filter_tree()

    def goto_next_page(self):
        total_pages = max(1, ((len(self.filtered_indices) - 1) // self.page_size) + 1)
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self._filter_tree()

    def goto_last_page(self):
        total_pages = max(1, ((len(self.filtered_indices) - 1) // self.page_size) + 1)
        if self.current_page != total_pages - 1:
            self.current_page = total_pages - 1
            self._filter_tree()

    def on_page_size_changed(self, event=None):
        try:
            new_size = int(self.cmb_page_size.get())
            if new_size > 0:
                self.page_size = new_size
                self.current_page = 0
                self._filter_tree()
        except:
            pass

    def on_page_num_entry(self, event=None):
        try:
            val = int(self.ent_page_num.get())
            total_pages = max(1, ((len(self.filtered_indices) - 1) // self.page_size) + 1)
            val = max(1, min(val, total_pages))
            self.current_page = val - 1
            self._filter_tree()
        except:
            self.ent_page_num.delete(0, tk.END)
            self.ent_page_num.insert(0, str(self.current_page + 1))

    def update_nav_buttons(self, total_pages):
        if not hasattr(self, 'btn_first'):
            return
        if self.current_page == 0:
            self.btn_first.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
            self.btn_prev.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
        else:
            self.btn_first.config(state="normal", bg="#2C5282", fg="#FFFFFF", cursor="hand2")
            self.btn_prev.config(state="normal", bg="#2C5282", fg="#FFFFFF", cursor="hand2")

        if self.current_page >= total_pages - 1:
            self.btn_next.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
            self.btn_last.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
        else:
            self.btn_next.config(state="normal", bg="#2C5282", fg="#FFFFFF", cursor="hand2")
            self.btn_last.config(state="normal", bg="#2C5282", fg="#FFFFFF", cursor="hand2")

    def _on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            if self.only_assigned_moqs:
                if hasattr(self, 'btn_load'):
                    self.btn_load.config(state="disabled")
                if hasattr(self, 'btn_download'):
                    self.btn_download.config(state="disabled")
            return
            
        idx = int(selected[0])
        record = self.bom_records[idx]
        raw_data = record["raw_data"]
        cust_name = record["Customer"]
        rfq_num = record["RFQ"]
        
        # Check if individual assembly JSON files exist in Individual BOM Data (any RFQ or customer)
        from utils import INDIVIDUAL_BOM_DATA_DIR, macro_file
        import os
        import glob
        
        has_sourcing = False
        if os.path.exists(INDIVIDUAL_BOM_DATA_DIR):
            for cust_folder in os.listdir(INDIVIDUAL_BOM_DATA_DIR):
                cust_path = os.path.join(INDIVIDUAL_BOM_DATA_DIR, cust_folder)
                if os.path.isdir(cust_path):
                    for assy in raw_data.get("Assemblies", []):
                        assy_num = str(assy.get("Assy #", ""))
                        safe_assy = assy_num.replace('/', '_').replace('\\', '_')
                        
                        pattern = os.path.join(cust_path, f"{safe_assy}_*.json")
                        files = glob.glob(pattern)
                        legacy_path = os.path.join(cust_path, f"{safe_assy}.json")
                        
                        if files or os.path.exists(legacy_path):
                            has_sourcing = True
                            break
                    if has_sourcing:
                        break
                        
        # Check if Excel report exists
        excel_exists = False
        target_dir = os.path.join(os.path.dirname(macro_file), cust_name.replace(" ", "_"))
        if os.path.exists(target_dir):
            excel_pattern = os.path.join(target_dir, f"{rfq_num}_{cust_name}_BOM Report_*.xlsx")
            excel_files = glob.glob(excel_pattern)
            if excel_files:
                excel_exists = True
                
        if self.only_assigned_moqs:
            if hasattr(self, 'btn_load'):
                if has_sourcing:
                    self.btn_load.config(state="normal")
                else:
                    self.btn_load.config(state="disabled")
                
            if hasattr(self, 'btn_download'):
                if excel_exists:
                    self.btn_download.config(state="normal")
                else:
                    self.btn_download.config(state="disabled")

    def _start_sourcing(self, action="start_sourcing"):
        selected = self.tree.selection()
        if not selected:
            from utils import show_info
            show_info("No Selection", "Please select a BOM record.", parent=self)
            return
            
        idx = int(selected[0])
        record = self.bom_records[idx]
        raw_data = record["raw_data"]
        cust_name = record["Customer"]
        rfq_num = record["RFQ"]
        filepath = record.get("filepath", "")
        date_str = record.get("Date", "")
        
        import pandas as pd
        from bomprocessor import BOMProcessor
        
        rows = []
        for assy in raw_data.get("Assemblies", []):
            assy_num = str(assy.get("Assy #", ""))
            model = str(assy.get("Assy Model", ""))
            rev = str(assy.get("Assy Rev", ""))
            
            for comp in assy.get("Components", []):
                rows.append({
                    'Assy #': assy_num,
                    'Assy Model': model,
                    'Assy Rev': rev,
                    'Part': str(comp.get('Part', '')),
                    'Description': str(comp.get('Description', '')),
                    'MFR': str(comp.get('MFR', '')),
                    'MPN': str(comp.get('MPN', '')),
                    'Qty': float(comp.get('Qty', 0.0)),
                    'UOM': str(comp.get('UOM', '')),
                    'Line Item': str(comp.get('Line Item', ''))
                })
                
        expected_cols = ['Assy #', 'Assy Model', 'Assy Rev', 'Part', 'Description', 'MFR', 'MPN', 'Qty', 'UOM', 'Line Item']
        df = pd.DataFrame(rows, columns=expected_cols)
        if action != "edit_saved":
            df = BOMProcessor._AssignLevel(df=df)
        
        self.result = (action, df, cust_name, rfq_num, filepath, raw_data, date_str)
        self._wait_var.set(1)

class AlternativeMPNOptionDialog(tk.Toplevel):
    def __init__(self, parent, part_number, current_mpn, current_mfr, db_mpn, db_mfr):
        super().__init__(parent)
        self.title("Alternative MPNs Available")
        self.geometry("600x380")
        self.transient(parent)
        self.grab_set()
        self.result = None # "merge", "select", "proceed", "cancel"
        
        self.configure(bg="#EBF8FF")
        
        main_frame = Frame(self, padx=20, pady=20, bg="#EBF8FF")
        main_frame.pack(fill="both", expand=True)
        
        Label(
            main_frame,
            text=f"Alternative MPNs Found for Part: {part_number}",
            font=("Arial", 12, "bold"),
            fg="#1a365d",
            bg="#EBF8FF"
        ).pack(anchor="w", pady=(0, 10))
        
        desc = (
            "We found matching alternative MPN/MFR records for this part in the database.\n"
            "Please select how you would like to proceed:"
        )
        Label(main_frame, text=desc, font=("Arial", 10), justify="left", bg="#EBF8FF").pack(anchor="w", pady=(0, 15))
        
        # Comparison box
        comp_frame = LabelFrame(main_frame, text="Comparison details", padx=10, pady=10, bg="#EBF8FF", fg="#1a365d", font=("Arial", 10, "bold"))
        comp_frame.pack(fill="x", pady=(0, 20))
        
        Label(comp_frame, text="Current in BOM:", font=("Arial", 9, "bold"), bg="#EBF8FF", fg="#4a5568").grid(row=0, column=0, sticky="w", pady=2)
        Label(comp_frame, text=f"MPN: {current_mpn or '(none)'}  |  MFR: {current_mfr or '(none)'}", font=("Arial", 9), bg="#EBF8FF").grid(row=0, column=1, sticky="w", padx=10, pady=2)
        
        Label(comp_frame, text="In Database:", font=("Arial", 9, "bold"), bg="#EBF8FF", fg="#4a5568").grid(row=1, column=0, sticky="w", pady=2)
        Label(comp_frame, text=f"MPN: {db_mpn}  |  MFR: {db_mfr}", font=("Arial", 9), bg="#EBF8FF", fg="#2b6cb0").grid(row=1, column=1, sticky="w", padx=10, pady=2)
        
        # Options Frame
        btn_frame = Frame(main_frame, bg="#EBF8FF")
        btn_frame.pack(fill="x", pady=10)
        
        Button(
            btn_frame, 
            text="✨ Merge Database Alternatives with Current", 
            command=lambda: self._set_result("merge"),
            bg="#2b6cb0", 
            fg="white", 
            font=("Arial", 10, "bold"),
            pady=8
        ).pack(fill="x", pady=4)
        
        Button(
            btn_frame, 
            text="🔍 Select Specific Alternatives to Add", 
            command=lambda: self._set_result("select"),
            bg="#3182ce", 
            fg="white", 
            font=("Arial", 10, "bold"),
            pady=8
        ).pack(fill="x", pady=4)
        
        Button(
            btn_frame, 
            text="➡️ Proceed to Edit without Alternatives", 
            command=lambda: self._set_result("proceed"),
            bg="#718096", 
            fg="white", 
            font=("Arial", 10, "bold"),
            pady=8
        ).pack(fill="x", pady=4)
        
        # Bottom Cancel Button
        cancel_frame = Frame(main_frame, bg="#EBF8FF")
        cancel_frame.pack(fill="x", side="bottom")
        Button(
            cancel_frame, 
            text="Cancel", 
            command=lambda: self._set_result("cancel"),
            width=15
        ).pack(side="right")
        
def _apply_bom_main_styling(win):
    try:
        import sys
        for mod_name in ('__main__', 'main'):
            mod = sys.modules.get(mod_name)
            if mod and hasattr(mod, '_apply_premium_global_styling'):
                mod._apply_premium_global_styling(win)
                if hasattr(mod, '_center_window_on_master_or_screen'):
                    mod._center_window_on_master_or_screen(win)
                return
    except Exception:
        pass

        # Apply premium styling
        _apply_bom_main_styling(self)
        
    def _set_result(self, res):
        self.result = res
        self.destroy()

class AlternativeMPNSelectionDialog(tk.Toplevel):
    def __init__(self, parent, part_number, db_mpn, db_mfr, current_mpn="", current_mfr=""):
        super().__init__(parent)
        self.title("Select Alternative MPNs")
        self.geometry("500x400")
        self.transient(parent)
        self.grab_set()
        self.result = None # Will contain list of checked (mpn, mfr) pairs
        
        self.configure(bg="#EBF8FF")
        
        # Split database pairs
        mpns = [m.strip() for m in str(db_mpn).split(",") if m.strip()]
        mfrs = [m.strip() for m in str(db_mfr).split(",") if m.strip()]
        while len(mfrs) < len(mpns): mfrs.append("")
        while len(mpns) < len(mfrs): mpns.append("")
        
        # Split current pairs to filter out from checklist
        curr_mpns = [m.strip().upper() for m in str(current_mpn).split(",") if m.strip()]
        curr_mfrs = [m.strip().upper() for m in str(current_mfr).split(",") if m.strip()]
        while len(curr_mfrs) < len(curr_mpns): curr_mfrs.append("")
        while len(curr_mpns) < len(curr_mfrs): curr_mpns.append("")
        curr_set = set(zip(curr_mpns, curr_mfrs))
        
        self.db_pairs = []
        for mpn, mfr in zip(mpns, mfrs):
            if (mpn.strip().upper(), mfr.strip().upper()) not in curr_set:
                self.db_pairs.append((mpn, mfr))
                
        self.chk_vars = []
        
        main_frame = Frame(self, padx=15, pady=15, bg="#EBF8FF")
        main_frame.pack(fill="both", expand=True)
        
        Label(
            main_frame,
            text=f"Select Database Alternatives for Part {part_number}",
            font=("Arial", 11, "bold"),
            fg="#1a365d",
            bg="#EBF8FF"
        ).pack(anchor="w", pady=(0, 10))
        
        # Scrollable Frame for checkbox list
        container = Frame(main_frame, bg="#EBF8FF")
        container.pack(fill="both", expand=True, pady=10)
        
        canvas = Canvas(container, bg="white", bd=1, relief="solid")
        scrollbar = Scrollbar(container, orient="vertical", command=canvas.yview)
        self.list_frame = Frame(canvas, bg="white")
        
        self.list_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=self.list_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Render check buttons
        for idx, (mpn, mfr) in enumerate(self.db_pairs):
            var = IntVar(value=1) # Checked by default
            self.chk_vars.append(var)
            
            row_f = Frame(self.list_frame, bg="white", pady=4)
            row_f.pack(fill="x", anchor="w", padx=10)
            
            # Checkbutton
            cb = Checkbutton(row_f, variable=var, bg="white", activebackground="white")
            cb.pack(side="left")
            
            lbl_text = f"MPN: {mpn}  |  MFR: {mfr}"
            Label(row_f, text=lbl_text, font=("Arial", 10), bg="white", anchor="w").pack(side="left", padx=5)
            
        btn_frame = Frame(main_frame, bg="#EBF8FF")
        btn_frame.pack(fill="x", side="bottom", pady=(10, 0))
        Button(
            btn_frame, 
            text="Confirm Selection", 
            command=self._on_confirm, 
            bg="#28a745", 
            fg="white", 
            font=("Arial", 10, "bold"),
            width=20
        ).pack(side="right", padx=5)
        Button(
            btn_frame, 
            text="Cancel", 
            command=self.destroy,
            width=12
        ).pack(side="right", padx=5)
        
        _apply_bom_main_styling(self)
        
    def _on_confirm(self):
        self.result = []
        for idx, var in enumerate(self.chk_vars):
            if var.get() == 1:
                self.result.append(self.db_pairs[idx])
        self.destroy()

class AddAlternativePartDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Add Alternative Part Mapping")
        self.geometry("500x300")
        self.transient(parent)
        self.grab_set()
        self.result = None
        
        self.configure(bg="#EBF8FF")
        
        main_frame = Frame(self, padx=15, pady=15, bg="#EBF8FF")
        main_frame.pack(fill="both", expand=True)
        
        Label(main_frame, text="Add New Alternative Part Mapping", font=("Arial", 11, "bold"), bg="#EBF8FF", fg="#1a365d").pack(anchor="w", pady=(0, 15))
        
        # Fields
        f_frame = Frame(main_frame, bg="#EBF8FF")
        f_frame.pack(fill="x", pady=5)
        
        Label(f_frame, text="Part Number:", font=("Arial", 10, "bold"), bg="#EBF8FF", width=12, anchor="w").grid(row=0, column=0, pady=5)
        self.part_entry = Entry(f_frame, width=35)
        self.part_entry.grid(row=0, column=1, pady=5)
        self.part_entry.focus_set()
        
        Label(f_frame, text="Alternative MPN:", font=("Arial", 10, "bold"), bg="#EBF8FF", width=12, anchor="w").grid(row=1, column=0, pady=5)
        self.mpn_entry = Entry(f_frame, width=35)
        self.mpn_entry.grid(row=1, column=1, pady=5)
        
        Label(f_frame, text="MFR:", font=("Arial", 10, "bold"), bg="#EBF8FF", width=12, anchor="w").grid(row=2, column=0, pady=5)
        self.mfr_entry = Entry(f_frame, width=35)
        self.mfr_entry.grid(row=2, column=1, pady=5)
        
        Label(main_frame, text="Note: Comma-separated values for multiple alternatives (e.g. MPN1, MPN2).", font=("Arial", 8, "italic"), bg="#EBF8FF", fg="#718096").pack(anchor="w", pady=5)
        
        btn_frame = Frame(main_frame, bg="#EBF8FF")
        btn_frame.pack(fill="x", side="bottom", pady=(10, 0))
        
        Button(btn_frame, text="Add Mapping", command=self._on_add, bg="#28a745", fg="white", font=("Arial", 10, "bold"), width=15).pack(side="right", padx=5)
        Button(btn_frame, text="Cancel", command=self.destroy, width=10).pack(side="right", padx=5)
        
        _apply_bom_main_styling(self)
        
    def _on_add(self):
        part = self.part_entry.get().strip()
        mpn = self.mpn_entry.get().strip()
        mfr = self.mfr_entry.get().strip()
        
        if not part:
            messagebox.showwarning("Validation Error", "Part Number cannot be empty.", parent=self)
            return
            
        self.result = {'part': part, 'mpn': mpn, 'mfr': mfr}
        self.destroy()

class SimpleTextInputDialog(tk.Toplevel):
    def __init__(self, parent, title, prompt):
        super().__init__(parent)
        self.title(title)
        self.geometry("380x150")
        self.transient(parent)
        self.grab_set()
        self.result = None
        
        self.configure(bg="#EBF8FF")
        
        main_frame = Frame(self, padx=15, pady=15, bg="#EBF8FF")
        main_frame.pack(fill="both", expand=True)
        
        Label(main_frame, text=prompt, font=("Arial", 10, "bold"), bg="#EBF8FF", anchor="w").pack(fill="x", pady=(0, 10))
        
        self.entry = Entry(main_frame, width=40)
        self.entry.pack(fill="x", pady=5)
        self.entry.focus_set()
        
        btn_frame = Frame(main_frame, bg="#EBF8FF")
        Button(btn_frame, text="OK", command=self._on_ok, bg="#28a745", fg="white", font=("Arial", 9, "bold"), width=10).pack(side="right", padx=5)
        Button(btn_frame, text="Cancel", command=self.destroy, width=10).pack(side="right", padx=5)
        
        _apply_bom_main_styling(self)
        
    def _on_ok(self):
        self.result = self.entry.get().strip()
        self.destroy()

class SearchableCustomerDropdown(Frame):
    """
    Smooth, non-blocking searchable combobox widget.
    Allows continuous fluid typing while displaying a live-filtered dropdown list below.
    """
    def __init__(self, master, values=None, width=28, on_select=None, bg="#EBF8FF", **kwargs):
        super().__init__(master, bg=bg, **kwargs)
        self.all_values = list(values) if values else []
        self.on_select = on_select
        self.popup = None
        self.listbox = None
        
        # Entry Widget for fluid typing
        self.entry_var = StringVar()
        self.entry = Entry(self, textvariable=self.entry_var, width=width, font=("Segoe UI", 10), bg="white", fg="#1A365D", relief="solid", bd=1)
        self.entry.pack(side="left", fill="x", expand=True)
        
        # Dropdown Arrow Button ▼ (Matching Entry height perfectly)
        self.btn = Label(self, text="▼", font=("Segoe UI", 7), width=3, bg="#e2e8f0", fg="#1A365D", relief="solid", bd=1, cursor="hand2")
        self.btn.pack(side="right", fill="y", padx=(1, 0))
        self.btn.bind("<Button-1>", lambda e: self.toggle_popup())
        
        # Event Bindings
        self.entry.bind("<KeyRelease>", self._on_keyrelease)
        self.entry.bind("<FocusIn>", self._on_focus_in)
        self.entry.bind("<Down>", self._on_arrow_down)
        self.entry.bind("<Return>", self._on_return)
        self.entry.bind("<Escape>", lambda e: self.hide_popup())
        self.entry.bind("<FocusOut>", self._on_focus_out)

    def set_all_values(self, values):
        self.all_values = list(values) if values else []

    def set_values(self, values, default_val=None):
        self.all_values = list(values) if values else []
        if default_val and default_val in self.all_values:
            self.entry_var.set(default_val)
        elif self.all_values and not self.entry_var.get():
            self.entry_var.set(self.all_values[0])

    def get(self):
        return self.entry_var.get().strip()

    def set(self, val):
        self.entry_var.set(val)

    def toggle_popup(self):
        if self.popup and self.popup.winfo_exists():
            self.hide_popup()
        else:
            self.show_popup(self.all_values)

    def show_popup(self, items):
        if not items:
            self.hide_popup()
            return

        # Create or update popup
        if not (self.popup and self.popup.winfo_exists()):
            self.update_idletasks()
            x = self.entry.winfo_rootx()
            y = self.entry.winfo_rooty() + self.entry.winfo_height() + 2
            w = max(self.entry.winfo_width() + self.btn.winfo_width(), 240)

            self.popup = tk.Toplevel(self)
            self.popup._is_autocomplete_popup = True
            self.popup._skip_autofit = True
            self.popup.wm_overrideredirect(True)
            self.popup.wm_geometry(f"{w}x{min(180, len(items)*22 + 8)}+{x}+{y}")
            self.popup.configure(bg="#3182CE", padx=1, pady=1)

            frame = Frame(self.popup, bg="white")
            frame.pack(fill="both", expand=True)

            scrollbar = Scrollbar(frame, orient="vertical")
            self.listbox = Listbox(
                frame, 
                selectmode="single", 
                yscrollcommand=scrollbar.set,
                font=("Segoe UI", 9),
                bg="white",
                fg="#1A365D",
                selectbackground="#3182CE",
                selectforeground="white",
                relief="flat",
                bd=0,
                activestyle="none"
            )
            scrollbar.config(command=self.listbox.yview)

            self.listbox.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            self.listbox.bind("<ButtonRelease-1>", self._on_listbox_click)
            self.listbox.bind("<Return>", self._on_listbox_select)
            self.listbox.bind("<Escape>", lambda e: self.hide_popup())

        # Populate items
        if self.listbox:
            self.listbox.delete(0, "end")
            for item in items:
                self.listbox.insert("end", item)
            
            # Reposition in case window moved
            x = self.entry.winfo_rootx()
            y = self.entry.winfo_rooty() + self.entry.winfo_height() + 2
            w = max(self.entry.winfo_width() + self.btn.winfo_width(), 240)
            self.popup.wm_geometry(f"{w}x{min(180, len(items)*22 + 8)}+{x}+{y}")

        # ALWAYS MAINTAIN FOCUS ON ENTRY WIDGET FOR FLUID TYPING
        self.entry.focus_set()

    def hide_popup(self):
        if self.popup and self.popup.winfo_exists():
            self.popup.destroy()
            self.popup = None
            self.listbox = None

    def _on_keyrelease(self, event):
        if event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            return
        typed = self.entry_var.get().strip().lower()
        if not typed:
            matching = self.all_values
        else:
            matching = [v for v in self.all_values if typed in str(v).lower()]
        self.show_popup(matching)

    def _on_focus_in(self, event):
        pass

    def _on_focus_out(self, event):
        self.after(200, self._check_focus_out)

    def _check_focus_out(self):
        try:
            focus_widget = self.focus_get()
            if self.popup and self.popup.winfo_exists():
                if focus_widget != self.entry and focus_widget != self.listbox and focus_widget != self.btn:
                    self.hide_popup()
        except Exception:
            self.hide_popup()

    def _on_arrow_down(self, event):
        if self.popup and self.listbox and self.listbox.size() > 0:
            self.listbox.focus_set()
            self.listbox.selection_clear(0, "end")
            self.listbox.selection_set(0)
            self.listbox.activate(0)
        else:
            self.toggle_popup()

    def _on_listbox_click(self, event):
        if self.listbox and self.listbox.curselection():
            idx = self.listbox.curselection()[0]
            val = self.listbox.get(idx)
            self.entry_var.set(val)
            self.hide_popup()
            if self.on_select:
                self.on_select(val)

    def _on_listbox_select(self, event):
        if self.listbox and self.listbox.curselection():
            idx = self.listbox.curselection()[0]
            val = self.listbox.get(idx)
            self.entry_var.set(val)
            self.hide_popup()
            self.entry.focus_set()
            if self.on_select:
                self.on_select(val)

    def _on_return(self, event):
        val = self.entry_var.get().strip()
        self.hide_popup()
        if self.on_select:
            self.on_select(val)

class CustomerAlternativeMPNMaintenanceDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Customer Alternative MPNs")
        self._skip_autofit = True
        self.geometry("1200x750")
        self.grab_set()
        
        self.current_customer = None
        self.alt_data = {"Customer": "", "Parts": {}}
        self.has_unsaved_changes = False
        self.current_page = 0
        self.page_size = 100
        self.filtered_parts = []
        
        self.configure(bg="#EBF8FF")
        self.setup_ui()
        
        # Load customer list
        self._refresh_customer_list()
        
        _apply_bom_main_styling(self)
        
        # Intercept closing to prompt for unsaved changes
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        
        self._populate_tree()

    def setup_ui(self):
        # Header banner
        header = Frame(self, bg="#1A365D")
        header.pack(fill="x", side="top")
        Label(header, text="Customer Alternative MPNs", font=("Segoe UI", 14, "bold"), fg="white", bg="#1A365D", pady=12).pack(side="left", padx=20)

        # Main layout
        top_frame = Frame(self, padx=10, pady=10, bg="#EBF8FF")
        top_frame.pack(fill="x")
        
        Label(top_frame, text="Customer:", font=("Arial", 10, "bold"), bg="#EBF8FF").pack(side="left", padx=5)
        self.cust_combo = SearchableCustomerDropdown(top_frame, width=28, on_select=self._on_customer_change)
        self.cust_combo.pack(side="left", padx=5)
        
        Button(top_frame, text="📁 Create Customer Folder", command=self._create_new_customer, bg="#1A365D", fg="white").pack(side="left", padx=10)
        
        # Search Filter (Right aligned)
        search_frame = Frame(top_frame, bg="#EBF8FF")
        search_frame.pack(side="right", padx=5)
        Label(search_frame, text="Filter Parts:", font=("Arial", 10, "bold"), bg="#EBF8FF").pack(side="left", padx=5)
        self.search_var = StringVar()
        self.search_var.trace_add("write", lambda *a: self._populate_tree())
        self.search_entry = Entry(search_frame, textvariable=self.search_var, width=25)
        self.search_entry.pack(side="left", padx=5)
        
        # Treeview Area
        mid_frame = Frame(self, padx=15, pady=5)
        mid_frame.pack(fill="both", expand=True)
        
        self.tree = Treeview(mid_frame, columns=("part", "mpn", "mfr"), show="headings")
        self.tree.heading("part", text="Part Number", anchor="w")
        self.tree.heading("mpn", text="Alternative MPNs", anchor="w")
        self.tree.heading("mfr", text="Alternative MFRs", anchor="w")
        
        self.tree.column("part", width=180, anchor="w")
        self.tree.column("mpn", width=300, anchor="w")
        self.tree.column("mfr", width=250, anchor="w")
        
        scrollbar = Scrollbar(mid_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.tree.bind("<<TreeviewSelect>>", self._on_select_item)
        self.tree.bind("<Double-1>", lambda e: self._edit_alternative())
        
        # Pagination Controls Frame
        pag_frame = Frame(self, bg="#EBF8FF", pady=5)
        pag_frame.pack(fill="x")
        
        pag_center = Frame(pag_frame, bg="#EBF8FF")
        pag_center.pack(anchor="center")

        self.btn_first = Button(pag_center, text="|<", command=self.goto_first_page, width=4)
        self.btn_first.pack(side="left", padx=2)

        self.btn_prev = Button(pag_center, text="<", command=self.goto_prev_page, width=4)
        self.btn_prev.pack(side="left", padx=2)

        Label(pag_center, text="Page:", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D").pack(side="left", padx=(10, 2))
        self.ent_page_num = Entry(pag_center, width=5, justify="center")
        self.ent_page_num.pack(side="left", padx=2)
        self.ent_page_num.bind("<Return>", self.on_page_num_entry)

        self.lbl_total_pages = Label(pag_center, text="of 1", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D")
        self.lbl_total_pages.pack(side="left", padx=(2, 10))

        self.btn_next = Button(pag_center, text=">", command=self.goto_next_page, width=4)
        self.btn_next.pack(side="left", padx=2)

        self.btn_last = Button(pag_center, text=">|", command=self.goto_last_page, width=4)
        self.btn_last.pack(side="left", padx=2)

        Label(pag_center, text="Page Size:", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D").pack(side="left", padx=(20, 2))
        self.cmb_page_size = Combobox(pag_center, values=["100", "500", "1000", "5000"], width=8, state="readonly")
        self.cmb_page_size.set(str(self.page_size))
        self.cmb_page_size.pack(side="left", padx=2)
        self.cmb_page_size.bind("<<ComboboxSelected>>", self.on_page_size_changed)

        self.lbl_total_records = Label(pag_frame, text="Total Records: 0", font=("Segoe UI", 9, "bold"), bg="#EBF8FF", fg="#1A365D")
        self.lbl_total_records.pack(side="right", padx=15)

        # Button Area
        btn_frame = Frame(self, padx=15, pady=10, bg="#EBF8FF")
        btn_frame.pack(fill="x")
        
        Button(btn_frame, text="➕ Add Mapping", command=self._add_alternative, bg="#3182ce", fg="white", width=15).pack(side="left", padx=5)
        Button(btn_frame, text="✏️ Edit Selected", command=self._edit_alternative, bg="#3182ce", fg="white", width=15).pack(side="left", padx=5)
        Button(btn_frame, text="🗑️ Delete Mapping", command=self._delete_alternative, bg="#dc3545", fg="white", width=15).pack(side="left", padx=5)
        
        Button(btn_frame, text="💾 Save Changes", command=self._save_changes, bg="#28a745", fg="white", font=("Arial", 10, "bold"), width=18).pack(side="right", padx=5)
        
        self.status_lbl = Label(btn_frame, text="Total alternative parts: 0", font=("Arial", 9, "italic"), bg="#EBF8FF")
        self.status_lbl.pack(side="right", padx=15)
        
    def _on_select_item(self, event=None):
        pass
        
    def _update_status_label(self):
        count = len(self.filtered_parts)
        self.status_lbl.config(text=f"Total alternative parts: {count}")
        if hasattr(self, 'lbl_total_records'):
            self.lbl_total_records.config(text=f"Total Records: {count}")

    def _refresh_customer_list(self):
        from utils import ALT_MPN_DIR
        alt_mpn_dir = ALT_MPN_DIR
        customers = set()
        
        if os.path.exists(alt_mpn_dir):
            for item in os.listdir(alt_mpn_dir):
                item_path = os.path.join(alt_mpn_dir, item)
                if os.path.isdir(item_path):
                    customers.add(item)
                elif item.endswith(".json"):
                    customers.add(item[:-5]) # filename without .json
                    
        sorted_custs = sorted(list(customers))
        self.cust_combo.set_all_values(sorted_custs)
        
        if sorted_custs:
            if not self.current_customer or self.current_customer not in sorted_custs:
                self.cust_combo.set(sorted_custs[0])
                self._on_customer_change()

    def _on_customer_change(self, event=None):
        cust = self.cust_combo.get().strip()
        if not cust:
            return
        if self.has_unsaved_changes:
            if not messagebox.askyesno("Unsaved Changes", "You have unsaved changes for the current customer. Do you want to discard them and load the new customer?", parent=self):
                if self.current_customer:
                    self.cust_combo.set(self.current_customer)
                return
                
        self.current_customer = cust
        self.has_unsaved_changes = False
        self._load_customer_data(cust)

    def _load_customer_data(self, cust):
        from utils import get_alternative_mpn_path
        import json
        
        alt_json_path = get_alternative_mpn_path(cust)
        self.alt_data = {"Customer": cust, "Parts": {}}
        
        if os.path.exists(alt_json_path):
            try:
                with open(alt_json_path, 'r', encoding='utf-8') as f:
                    self.alt_data = json.load(f)
            except Exception as e:
                print(f"Error loading alternative MPNs: {e}")
                
        if "Parts" not in self.alt_data:
            self.alt_data["Parts"] = {}
            
        self._populate_tree()

    def _populate_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        filter_text = self.search_var.get().strip().lower()
        parts = self.alt_data.get("Parts", {})
        
        # Reset page to 0 if filter changed
        if not hasattr(self, '_last_filter_text') or self._last_filter_text != filter_text:
            self._last_filter_text = filter_text
            self.current_page = 0

        self.filtered_parts = []
        for part in sorted(parts.keys()):
            rec = parts[part]
            mpn = rec.get("MPN", "")
            mfr = rec.get("MFR", "")
            
            if filter_text:
                if filter_text not in part.lower() and filter_text not in mpn.lower() and filter_text not in mfr.lower():
                    continue
            self.filtered_parts.append(part)
            
        total_matches = len(self.filtered_parts)
        total_pages = max(1, ((total_matches - 1) // self.page_size) + 1)
        
        if self.current_page >= total_pages:
            self.current_page = total_pages - 1
        if self.current_page < 0:
            self.current_page = 0

        # Update Pagination widgets if they exist
        if hasattr(self, 'lbl_total_pages'):
            self.lbl_total_pages.config(text=f"of {total_pages}")
        if hasattr(self, 'ent_page_num'):
            self.ent_page_num.delete(0, tk.END)
            self.ent_page_num.insert(0, str(self.current_page + 1))
        if hasattr(self, 'update_nav_buttons'):
            self.update_nav_buttons(total_pages)

        # Slice for current page
        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, total_matches)
        page_parts = self.filtered_parts[start_idx:end_idx]

        for part in page_parts:
            rec = parts[part]
            mpn = rec.get("MPN", "")
            mfr = rec.get("MFR", "")
            self.tree.insert("", "end", iid=part, values=(part, mpn, mfr))
            
        self._update_status_label()

    def goto_first_page(self):
        if self.current_page != 0:
            self.current_page = 0
            self._populate_tree()

    def goto_prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self._populate_tree()

    def goto_next_page(self):
        total_pages = max(1, ((len(self.filtered_parts) - 1) // self.page_size) + 1)
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self._populate_tree()

    def goto_last_page(self):
        total_pages = max(1, ((len(self.filtered_parts) - 1) // self.page_size) + 1)
        if self.current_page != total_pages - 1:
            self.current_page = total_pages - 1
            self._populate_tree()

    def on_page_size_changed(self, event=None):
        try:
            new_size = int(self.cmb_page_size.get())
            if new_size > 0:
                self.page_size = new_size
                self.current_page = 0
                self._populate_tree()
        except:
            pass

    def on_page_num_entry(self, event=None):
        try:
            val = int(self.ent_page_num.get())
            total_pages = max(1, ((len(self.filtered_parts) - 1) // self.page_size) + 1)
            val = max(1, min(val, total_pages))
            self.current_page = val - 1
            self._populate_tree()
        except:
            self.ent_page_num.delete(0, tk.END)
            self.ent_page_num.insert(0, str(self.current_page + 1))

    def update_nav_buttons(self, total_pages):
        if self.current_page == 0:
            self.btn_first.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
            self.btn_prev.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
        else:
            self.btn_first.config(state="normal", bg="#2C5282", fg="#FFFFFF", cursor="hand2")
            self.btn_prev.config(state="normal", bg="#2C5282", fg="#FFFFFF", cursor="hand2")

        if self.current_page >= total_pages - 1:
            self.btn_next.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
            self.btn_last.config(state="disabled", bg="#CBD5E0", fg="#718096", cursor="arrow")
        else:
            self.btn_next.config(state="normal", bg="#2C5282", fg="#FFFFFF", cursor="hand2")
            self.btn_last.config(state="normal", bg="#2C5282", fg="#FFFFFF", cursor="hand2")

    def _add_alternative(self):
        if not self.current_customer:
            messagebox.showwarning("No Customer", "Please select or create a customer first.", parent=self)
            return
            
        dialog = AddAlternativePartDialog(self)
        self.wait_window(dialog)
        
        if dialog.result:
            part = dialog.result['part']
            mpn = dialog.result['mpn']
            mfr = dialog.result['mfr']
            
            parts = self.alt_data["Parts"]
            if part in parts:
                if messagebox.askyesno("Part Exists", f"Part {part} already has alternative MPNs in the database. Do you want to merge the new alternatives with the existing ones?\n\nSelecting 'No' will overwrite the existing alternatives.", parent=self):
                    from utils import merge_mpn_mfr_pairs
                    existing = parts[part]
                    merged_mpn, merged_mfr = merge_mpn_mfr_pairs(existing.get("MPN", ""), existing.get("MFR", ""), mpn, mfr)
                    parts[part] = {"MPN": merged_mpn, "MFR": merged_mfr}
                else:
                    parts[part] = {"MPN": mpn, "MFR": mfr}
            else:
                parts[part] = {"MPN": mpn, "MFR": mfr}
                
            self.has_unsaved_changes = True
            
            # Find the page containing the part
            filter_text = self.search_var.get().strip().lower()
            matching_keys = sorted([p for p in parts.keys() if not filter_text or filter_text in p.lower() or filter_text in parts[p].get("MPN","").lower() or filter_text in parts[p].get("MFR","").lower()])
            if part in matching_keys:
                self.current_page = matching_keys.index(part) // self.page_size
                
            self._populate_tree()
            if self.tree.exists(part):
                self.tree.selection_set(part)
                self.tree.see(part)

    def _edit_alternative(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a part mapping to edit.", parent=self)
            return
            
        part = selected[0]
        rec = self.alt_data["Parts"][part]
        
        dialog = MPNMFRAssignmentDialog(self, part, rec.get("MPN", ""), rec.get("MFR", ""))
        self.wait_window(dialog)
        
        if dialog.result:
            new_mpn, new_mfr = dialog.result
            self.alt_data["Parts"][part] = {"MPN": new_mpn, "MFR": new_mfr}
            self.has_unsaved_changes = True
            self._populate_tree()
            self.tree.selection_set(part)

    def _delete_alternative(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a part mapping to delete.", parent=self)
            return
            
        part = selected[0]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete alternative mapping for part: {part}?", parent=self):
            del self.alt_data["Parts"][part]
            self.has_unsaved_changes = True
            self._populate_tree()

    def _save_changes(self):
        if not self.current_customer:
            return
            
        from utils import get_alternative_mpn_path
        alt_json_path = get_alternative_mpn_path(self.current_customer)
        
        try:
            with open(alt_json_path, 'w', encoding='utf-8') as f:
                json.dump(self.alt_data, f, indent=4)
            self.has_unsaved_changes = False
            messagebox.showinfo("Success", f"Alternative MPNs successfully saved for customer: {self.current_customer}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Alternative MPNs: {e}", parent=self)

    def _create_new_customer(self):
        dialog = SimpleTextInputDialog(self, "Create Customer Folder", "Enter New Customer Name:")
        self.wait_window(dialog)
        
        if dialog.result:
            new_cust = dialog.result.strip()
            if not new_cust:
                return
            
            clean_cust = new_cust.replace(' ', '_').replace('/', '_').replace('\\', '_')
            
            vals = list(self.cust_combo['values'])
            if clean_cust not in vals:
                vals.append(clean_cust)
                vals.sort()
                self.cust_combo['values'] = vals
                
            self.cust_combo.set(clean_cust)
            self._on_customer_change()

    def _on_close(self):
        if self.has_unsaved_changes:
            if not messagebox.askyesno("Unsaved Changes", "You have unsaved changes. Do you want to discard them and close?", parent=self):
                return
        self.destroy()


class UOMConversionMaintenanceDialog(tk.Toplevel):
    """Full maintenance dialog for UOM conversion rules stored in uom_conversions.json."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Unit of Measurement (UOM) Conversion")
        self._skip_autofit = True
        self.geometry("860x560")
        self.resizable(True, True)
        self.minsize(700, 450)
        self.grab_set()
        self.configure(bg="#f0f4f8")

        from utils import load_uom_conversions, save_uom_conversions
        self._load_fn = load_uom_conversions
        self._save_fn = save_uom_conversions

        self._config = self._load_fn()
        self._rules = dict(self._config.get("rules", {}))   # mutable copy
        self._tol_pct = float(self._config.get("tolerance_pct", 5.0))

        self._build_ui()
        self._refresh_tree()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    # ------------------------------------------------------------------
    def _build_ui(self):
        import tkinter as tk
        from tkinter import ttk

        header = tk.Frame(self, bg="#1A365D")
        header.pack(fill="x")
        tk.Label(header, text="Unit of Measurement (UOM) Conversion Rules",
                 font=("Segoe UI", 14, "bold"), fg="white", bg="#1A365D",
                 pady=12).pack(side="left", padx=20)

        # Global tolerance row
        tol_frame = tk.Frame(self, bg="#f0f4f8", pady=6)
        tol_frame.pack(fill="x", padx=20)
        tk.Label(tol_frame, text="Global Tolerance %:", font=("Segoe UI", 10, "bold"),
                 bg="#f0f4f8").pack(side="left")
        self._tol_var = tk.StringVar(value=str(self._tol_pct))
        tk.Entry(tol_frame, textvariable=self._tol_var, width=8,
                 font=("Segoe UI", 10)).pack(side="left", padx=8)
        tk.Label(tol_frame, text="(applied to rules where 'Apply Tolerance' is enabled)",
                 font=("Segoe UI", 9, "italic"), fg="gray", bg="#f0f4f8").pack(side="left")

        # Button row (packed at bottom so it remains visible regardless of tree height or banners)
        btn_frame = tk.Frame(self, bg="#f0f4f8", pady=8)
        btn_frame.pack(side="bottom", fill="x", padx=20)

        def btn(text, cmd, bg="#2C5282", fg="white"):
            return tk.Button(btn_frame, text=text, command=cmd,
                             bg=bg, fg=fg, font=("Segoe UI", 10, "bold"),
                             padx=12, pady=4, relief="flat", cursor="hand2")

        btn("+ Add Rule",    self._add_rule).pack(side="left", padx=4)
        btn("✏ Edit",        self._edit_rule, bg="#2563EB").pack(side="left", padx=4)
        btn("🗑 Delete",     self._delete_rule, bg="#C53030").pack(side="left", padx=4)
        btn("Close",         self.destroy, bg="#718096").pack(side="right", padx=4)
        btn("💾 Save All",   self._save_all, bg="#276749").pack(side="right", padx=4)

        self.lbl_total_records = tk.Label(btn_frame, text="Total Records: 0", font=("Segoe UI", 9, "bold"), bg="#f0f4f8", fg="#1A365D")
        self.lbl_total_records.pack(side="right", padx=15)

        # Treeview (fills remaining middle space)
        tree_frame = tk.Frame(self, bg="#f0f4f8")
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 5))

        cols = ("from_uom", "to_uom", "factor", "apply_tol")
        self._tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=10)
        self._tree.heading("from_uom", text="From UOM")
        self._tree.heading("to_uom",   text="To UOM")
        self._tree.heading("factor",   text="Conversion Factor")
        self._tree.heading("apply_tol", text="Apply Tolerance")
        self._tree.column("from_uom",  width=130, anchor="center")
        self._tree.column("to_uom",    width=130, anchor="center")
        self._tree.column("factor",    width=180, anchor="center")
        self._tree.column("apply_tol", width=140, anchor="center")

        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=self._tree.yview)
        self._tree.config(yscrollcommand=sb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    # ------------------------------------------------------------------
    def _refresh_tree(self):
        for row in self._tree.get_children():
            self._tree.delete(row)
        for uom, rule in sorted(self._rules.items()):
            self._tree.insert("", "end", iid=uom, values=(
                uom,
                rule.get("to_uom", ""),
                rule.get("factor", 1.0),
                "Yes" if rule.get("apply_tolerance", False) else "No"
            ))
        if hasattr(self, 'lbl_total_records'):
            self.lbl_total_records.config(text=f"Total Records: {len(self._rules)}")

    # ------------------------------------------------------------------
    def _add_rule(self):
        self._open_rule_editor(mode="add")

    def _edit_rule(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("No Selection", "Please select a rule to edit.", parent=self)
            return
        self._open_rule_editor(mode="edit", from_uom=sel[0])

    def _delete_rule(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("No Selection", "Please select a rule to delete.", parent=self)
            return
        from_uom = sel[0]
        if messagebox.askyesno("Delete Rule", f"Delete conversion rule for '{from_uom}'?", parent=self):
            self._rules.pop(from_uom, None)
            self._refresh_tree()

    # ------------------------------------------------------------------
    def _open_rule_editor(self, mode="add", from_uom=None):
        import tkinter as tk
        dlg = tk.Toplevel(self)
        dlg.title("Add Rule" if mode == "add" else f"Edit Rule — {from_uom}")
        dlg._skip_autofit = True
        dlg.geometry("380x260")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(bg="#f0f4f8")

        existing = self._rules.get(from_uom, {}) if from_uom else {}

        vars_ = {
            "from": tk.StringVar(value=from_uom or ""),
            "to":   tk.StringVar(value=existing.get("to_uom", "")),
            "fac":  tk.StringVar(value=str(existing.get("factor", 1.0))),
            "tol":  tk.BooleanVar(value=existing.get("apply_tolerance", False)),
        }

        grid = tk.Frame(dlg, bg="#f0f4f8", padx=20, pady=15)
        grid.pack(fill="both", expand=True)

        def row(label, widget, r):
            tk.Label(grid, text=label, font=("Segoe UI", 10), bg="#f0f4f8",
                     anchor="w").grid(row=r, column=0, sticky="w", pady=5)
            widget.grid(row=r, column=1, sticky="ew", padx=10, pady=5)
        grid.columnconfigure(1, weight=1)

        from_ent = tk.Entry(grid, textvariable=vars_["from"], font=("Segoe UI", 10))
        if mode == "edit":
            from_ent.config(state="disabled")
        to_ent  = tk.Entry(grid, textvariable=vars_["to"],  font=("Segoe UI", 10))
        fac_ent = tk.Entry(grid, textvariable=vars_["fac"], font=("Segoe UI", 10))
        tol_chk = tk.Checkbutton(grid, variable=vars_["tol"], bg="#f0f4f8",
                                  text="Enable", font=("Segoe UI", 10))

        row("From UOM:",           from_ent, 0)
        row("To UOM:",             to_ent,   1)
        row("Conversion Factor:",  fac_ent,  2)
        row("Apply Tolerance:",    tol_chk,  3)

        def _ok():
            fu = vars_["from"].get().strip().upper()
            tu = vars_["to"].get().strip().upper()
            try:
                fac = float(vars_["fac"].get())
            except ValueError:
                messagebox.showerror("Invalid", "Conversion factor must be a number.", parent=dlg)
                return
            if not fu or not tu:
                messagebox.showerror("Required", "From UOM and To UOM are required.", parent=dlg)
                return
            self._rules[fu] = {
                "to_uom": tu,
                "factor": fac,
                "apply_tolerance": vars_["tol"].get()
            }
            self._refresh_tree()
            dlg.destroy()

        btn_r = tk.Frame(dlg, bg="#f0f4f8", pady=8)
        btn_r.pack(fill="x", padx=20)
        tk.Button(btn_r, text="OK", command=_ok, bg="#276749", fg="white",
                  font=("Segoe UI", 10, "bold"), padx=16).pack(side="right", padx=4)
        tk.Button(btn_r, text="Cancel", command=dlg.destroy, bg="#718096", fg="white",
                  font=("Segoe UI", 10, "bold"), padx=12).pack(side="right", padx=4)

    # ------------------------------------------------------------------
    def _save_all(self):
        try:
            tol = float(self._tol_var.get())
        except ValueError:
            messagebox.showerror("Invalid", "Global tolerance must be a number.", parent=self)
            return
        data = {"tolerance_pct": tol, "rules": self._rules}
        if self._save_fn(data):
            messagebox.showinfo("Saved", "UOM conversion rules saved successfully.", parent=self)
        else:
            messagebox.showerror("Error", "Failed to save UOM conversion rules.", parent=self)


class PromptMissingUOMDialog(UOMConversionMaintenanceDialog):
    """Variant of UOMConversionMaintenanceDialog that is triggered during BOM
    verification when one or more UOMs in the consolidated BOM have no
    conversion rule.  Displays a prominent warning banner listing the
    unmapped UOMs so the user knows exactly what needs to be added before
    they can proceed.

    Signature:  PromptMissingUOMDialog(parent, missing_uoms)
        parent        – tk parent window (wizard_window)
        missing_uoms  – list[str] of UOM strings that lack a rule
    """

    def __init__(self, parent, missing_uoms):
        self._missing_uoms = list(missing_uoms)
        super().__init__(parent)
        self.title("UOM Mapping Required – BOM Verification")
        self.geometry("860x650")
        self.minsize(750, 500)
        self._inject_warning_banner()

    # ------------------------------------------------------------------
    def _inject_warning_banner(self):
        """Insert a warning frame above the main content area."""
        import tkinter as tk

        uom_list = ", ".join(self._missing_uoms) if self._missing_uoms else "—"

        banner = tk.Frame(self, bg="#7B2D00", pady=6)
        # Place it just below the header (index 1 in pack order)
        banner.pack(fill="x", after=self.winfo_children()[0])

        tk.Label(
            banner,
            text=f"⚠  The following UOMs have no conversion rule and must be mapped before verification can continue:",
            font=("Segoe UI", 9, "bold"),
            fg="#FFE4B5",
            bg="#7B2D00",
            anchor="w",
        ).pack(fill="x", padx=16, pady=(4, 0))

        tk.Label(
            banner,
            text=uom_list,
            font=("Segoe UI", 9),
            fg="#FFFFFF",
            bg="#7B2D00",
            anchor="w",
            wraplength=820,
            justify="left",
        ).pack(fill="x", padx=16, pady=(0, 4))
