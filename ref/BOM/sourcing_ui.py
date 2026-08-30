import os
import tkinter as tk
from tkinter import ttk, filedialog
from utils import SemanticMessageBox as messagebox, BASE_DIR, USAGE_SUMMARY_DIR
from dialogs import BaseDialog, BasePanel, apply_panel_theme
import pandas as pd
try:
    from treeview_sort import attach_treeview_sort
except ImportError:
    attach_treeview_sort = lambda tree: None

def format_to_sig_figs(val, sig_figs=2):
    import math
    import pandas as pd
    if val is None or str(val).strip() == "":
        return ""
    try:
        f_val = float(val)
        if f_val == 0.0:
            return "0"
        if f_val == int(f_val):
            return str(int(f_val))
        dec = -int(math.floor(math.log10(abs(f_val))))
        places = max(0, dec + sig_figs - 1)
        fmt = f"{{:.{places}f}}"
        res = fmt.format(f_val)
        if "." in res:
            res = res.rstrip('0').rstrip('.')
        return res
    except:
        return str(val)

class MultiValueInputDialog(BaseDialog):
    def __init__(self, master, title, initial_values=None):
        super().__init__(master, title)
        self.geometry("400x500")
        self.initial_values = initial_values or []
        self._create_widgets()
        
    def _create_widgets(self):
        btn_frame = tk.Frame(self)
        btn_frame.pack(side="bottom", fill="x", padx=15, pady=(0, 15))
        
        btn_cancel = tk.Button(btn_frame, text="Cancel", command=self._on_cancel, width=12)
        btn_cancel.pack(side="left")
        
        btn_confirm = tk.Button(btn_frame, text="Confirm", command=self._on_confirm, width=12)
        btn_confirm.pack(side="right")
        
        from dialogs import style_premium_button
        style_premium_button(btn_cancel)
        style_premium_button(btn_confirm)

        main_frame = tk.Frame(self)
        main_frame.pack(side="top", fill="both", expand=True, padx=15, pady=(15, 0))
        
        tk.Label(main_frame, text="Enter values (one per line):", font=("Arial", 10, "bold")).pack(anchor="w", pady=(0, 5))
        
        self.text_area = tk.Text(main_frame, wrap="none", font=("Courier", 10), height=15)
        self.text_area.pack(fill="both", expand=True, pady=5)
        
        if self.initial_values:
            self.text_area.insert("1.0", "\n".join(self.initial_values))
        
    def _on_confirm(self):
        text = self.text_area.get("1.0", "end-1c")
        self.result = [line.strip() for line in text.split("\n") if line.strip()]
        self.destroy()
        
    def _on_cancel(self):
        self.result = None
        self.destroy()

class BOMFilterDialog(BaseDialog):
    def __init__(self, master, current_filters=None):
        super().__init__(master, "BOM Calculation Filter")
        # self.geometry("500x200")
        self.current_filters = current_filters or {"Part": [], "MPN": []}
        self.result = None
        self._create_widgets()
        
    def _create_widgets(self):
        main_frame = tk.Frame(self, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Grid layout for inputs
        grid_frame = tk.Frame(main_frame)
        grid_frame.pack(fill="x", expand=False)
        
        tk.Label(grid_frame, text="Part #:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", pady=10)
        self.part_entry = ttk.Entry(grid_frame, width=35)
        self.part_entry.grid(row=0, column=1, sticky="ew", padx=10, pady=10)
        btn_part_multi = tk.Button(grid_frame, text="⫘", font=("Arial", 10, "bold"), command=self._edit_part_multi, bg="#e2e8f0", relief="groove")
        btn_part_multi.grid(row=0, column=2, padx=5, pady=10)
        
        tk.Label(grid_frame, text="MPN:", font=("Arial", 10, "bold")).grid(row=1, column=0, sticky="w", pady=10)
        self.mpn_entry = ttk.Entry(grid_frame, width=35)
        self.mpn_entry.grid(row=1, column=1, sticky="ew", padx=10, pady=10)
        btn_mpn_multi = tk.Button(grid_frame, text="⫘", font=("Arial", 10, "bold"), command=self._edit_mpn_multi, bg="#e2e8f0", relief="groove")
        btn_mpn_multi.grid(row=1, column=2, padx=5, pady=10)
        
        grid_frame.columnconfigure(1, weight=1)
        
        # Populate initial values
        self.part_list = list(self.current_filters.get("Part", []))
        self.mpn_list = list(self.current_filters.get("MPN", []))
        
        self._update_entry_fields()
        
        # Action Buttons
        btn_frame = tk.Frame(main_frame, padx=0, pady=15)
        btn_frame.pack(fill="x", side="bottom")
        
        tk.Button(btn_frame, text="Execute Filter", command=self._on_execute, width=15).pack(side="right", padx=5)
        tk.Button(btn_frame, text="Clear Filters", command=self._on_clear, width=15).pack(side="right", padx=5)
        tk.Button(btn_frame, text="Cancel", command=self._on_cancel, width=12).pack(side="left")
        
    def _update_entry_fields(self):
        # Update part entry
        self.part_entry.delete(0, "end")
        if self.part_list:
            if len(self.part_list) == 1:
                self.part_entry.insert(0, self.part_list[0])
            else:
                self.part_entry.insert(0, f"[{len(self.part_list)} Values Selected]")
                self.part_entry.config(state="readonly")
        else:
            self.part_entry.config(state="normal")
            
        # Update mpn entry
        self.mpn_entry.delete(0, "end")
        if self.mpn_list:
            if len(self.mpn_list) == 1:
                self.mpn_entry.insert(0, self.mpn_list[0])
            else:
                self.mpn_entry.insert(0, f"[{len(self.mpn_list)} Values Selected]")
                self.mpn_entry.config(state="readonly")
        else:
            self.mpn_entry.config(state="normal")
            
    def _edit_part_multi(self):
        dialog = MultiValueInputDialog(self, "Multi Part Search", self.part_list)
        self.wait_window(dialog)
        if dialog.result is not None:
            self.part_list = dialog.result
            self._update_entry_fields()
            
        else:
            self.filter_criteria = None
            if hasattr(self, 'populate_tree'):
                self.populate_tree()
            elif hasattr(self, 'refresh_projects'):
                self.refresh_projects()
    def _edit_mpn_multi(self):
        dialog = MultiValueInputDialog(self, "Multi MPN Search", self.mpn_list)
        self.wait_window(dialog)
        if dialog.result is not None:
            self.mpn_list = dialog.result
            self._update_entry_fields()
            
        else:
            self.filter_criteria = None
            if hasattr(self, 'populate_tree'):
                self.populate_tree()
            elif hasattr(self, 'refresh_projects'):
                self.refresh_projects()
    def _on_execute(self):
        if self.part_entry.cget("state") == "normal":
            val = self.part_entry.get().strip()
            self.part_list = [val] if val else []
            
        if self.mpn_entry.cget("state") == "normal":
            val = self.mpn_entry.get().strip()
            self.mpn_list = [val] if val else []
            
        self.result = {"Part": self.part_list, "MPN": self.mpn_list}
        self.destroy()
        
    def _on_clear(self):
        self.result = {"Part": [], "MPN": []}
        self.destroy()
        
    def _on_cancel(self):
        self.result = None
        self.destroy()

class SourcingStatusWindow(tk.Toplevel):
    def __init__(self, parent, unique_assemblies, assembly_status):
        super().__init__(parent)
        self.title("Assembly Review Status")
        self.geometry("500x400")
        self.transient(parent)
        self.grab_set()

        self.result = "BACK"

        # Title Frame (Yellow Tone)
        header_frame = tk.Frame(self, bg="#fffde7", bd=1, relief="ridge")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="ASSEMBLY REVIEW STATUS", font=("Arial", 14, "bold"), fg="#856404", bg="#fffde7").pack(pady=12)

        main_frame = tk.Frame(self, padx=10, pady=10)
        main_frame.pack(fill="both", expand=True)

        tk.Label(main_frame, text="Current status of all assemblies in the preview:", font=('Arial', 10, 'bold')).pack(pady=(5,5), anchor="w")

        # Pack btn_frame FIRST (before list_frame) so it always stays visible at the bottom
        btn_frame = tk.Frame(main_frame, pady=8)
        btn_frame.pack(fill="x", side="bottom")

        tk.Button(btn_frame, text="Proceed", command=self.on_proceed, bg="#2ead4e", fg="white", font=("Arial", 10, "bold"), width=12).pack(side="right", padx=10)
        tk.Button(btn_frame, text="Go Back", command=self.on_back, bg="#e2e8f0", font=("Arial", 10), width=10).pack(side="right")

        list_frame = tk.Frame(main_frame)
        list_frame.pack(fill="both", expand=True, pady=(5, 0))

        cols = ("Status", "Assembly")
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=8)
        self.tree.heading("Status", text="Status")
        self.tree.heading("Assembly", text="Assembly")
        self.tree.column("Status", width=120, anchor="center")
        self.tree.column("Assembly", width=300, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)
        attach_treeview_sort(self.tree)

        self.tree.tag_configure('viewed', foreground='green')
        self.tree.tag_configure('pending', foreground='orange')
        self.tree.tag_configure('unviewed', foreground='red')
        
        _st = ttk.Style()
        _st.theme_use("clam")
        _st.map("Treeview", background=[("selected", "#0078D7")], foreground=[("selected", "white")])

        for assy in unique_assemblies:
            status = assembly_status.get(assy, "Unviewed")
            if status == "Viewed":
                bullet = "🟢 Viewed"
                tag = "viewed"
            elif status == "Pending":
                bullet = "🟠 Pending"
                tag = "pending"
            else:
                bullet = "🔴 Not Viewed"
                tag = "unviewed"
            self.tree.insert("", "end", values=(bullet, assy), tags=(tag,))

        scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        scroll.pack(side="right", fill="y")
        self.tree.config(yscrollcommand=scroll.set)

    def on_proceed(self):
        self.result = "PROCEED"
        self.destroy()

    def on_back(self):
        self.result = "BACK"
        self.destroy()

class SourcingPreviewPanel(BasePanel):
    def __init__(self, parent, gui_model, engine, assembly_moqs, on_approve_callback, on_export_callback, calc_mode='total_usage', cust_name="", rfq_num="", user_name="", bom_filepath=None, read_only=False, raw_data=None):
        super().__init__(parent)
        self.window = self
        self.user_name = user_name
        self.bom_filepath = bom_filepath
        self.read_only = read_only
        self.raw_data = raw_data or {}
        self.load_target_price_data()
        
        self.gui_model = gui_model
        self.engine = engine
        self.assembly_moqs = assembly_moqs
        self.calc_mode = calc_mode
        self.cust_name = cust_name
        self.rfq_num = rfq_num
        self.unique_assemblies = sorted(list(self.assembly_moqs.keys()))
        self.current_filters = {"Part": [], "MPN": []}
        
        # Assign Global Ref # for easy user identification
        for i, item in enumerate(self.gui_model):
            item['ref_id'] = i + 1
            item['is_manual'] = item.get('is_manual', False) # Track manual changes
        
        # Recalculate ExMat costs for all items based on usage summary
        for item in self.gui_model:
            self._recalculate_exmat_for_item(item)
        
        self.selected_assembly = tk.StringVar()
        if self.unique_assemblies:
            self.selected_assembly.set(self.unique_assemblies[0])
            
        self.on_approve = on_approve_callback
        self.on_export = on_export_callback
        
        # Save original state for revert functionality
        for item in self.gui_model:
            moqs = self.assembly_moqs.get(item['Assy'].strip(), [])
            for q in moqs:
                if 'original_winner' not in item['moq_results'][q]:
                    item['moq_results'][q]['original_winner'] = item['moq_results'][q].get('winner')
        
        self.current_selected_idx = None
        self.active_edit_ref_id = None
        self.assy_popup = None
        self.show_details = False
        self.active_categories = []
        
        # Base Columns
        self.base_cols = ["Comp Level", "Assy", "Part", "Description", "MFR", "MPN", "BOM Qty", "UOM"]
        
        # Detailed columns for each MOQ
        self.moq_col_defs = [
            ("Winner", "Winning Supplier", 120),
            ("UnitP", "Unit Price", 80),
            ("MOQ", "MOQ", 60),
            ("SrcDate", "Source Date", 90),
            ("MPN", "MPN", 120),
            ("Cur", "Currency", 70),
            ("ConvUSD", "Conv USD", 80),
            ("Markup", "Markup", 60),
            ("BOMCost", "BOM Cost", 80),
            ("ExCost", "ExMat Cost", 80),
            ("TotCost", "Total Cost", 80)
        ]
        
        self.all_cols = []
        self.compact_cols = []
        self.assembly_status = {}
        
        self.setup_ui()
        self.rebuild_grouped_model()
        self.on_assembly_changed(None)
        self.show_green_bar()

    def rebuild_grouped_model(self):
        """Re-indexes the gui_model by assembly to ensure fast retrieval and correct row counts."""
        self.grouped_model = {}
        for idx, item in enumerate(self.gui_model):
            assy = str(item.get('Assy', '')).strip()
            if assy not in self.grouped_model:
                self.grouped_model[assy] = []
            self.grouped_model[assy].append((idx, item))

    def show_green_bar(self):
        self.preview_status.config(text="☑ Preview is ready", bg="green", fg="white")
        self.window.update_idletasks()
        self.window.after(5000, lambda: self.preview_status.config(text="", bg=self.window.cget("bg")))

    def setup_ui(self):
        # --- Bottom Buttons ---
        self.preview_status = tk.Label(self.window, text="", font=('Arial', 10, 'bold'), anchor="center", pady=4)
        self.preview_status.pack(side="bottom", fill="x")
        
        btn_frame = tk.Frame(self.window, padx=10, pady=10)
        btn_frame.pack(side="bottom", fill="x")
        
        back_text = "Close" if self.read_only else "Cancel"
        self.btn_back = tk.Button(btn_frame, text=back_text, command=self.go_back)
        self.btn_back.pack(side="right" if self.read_only else "left", padx=5)
        
        if not self.read_only:
            # Approve Calculations button (Green)
            self.btn_approve = tk.Button(btn_frame, text="✅ Approve Calculations", command=self.approve, bg="#2ead4e", fg="white", font=("Arial", 10, "bold"))
            self.btn_approve.pack(side="right", padx=5)
            
            # Save Progress button (Slate Gray)
            self.btn_save_progress = tk.Button(btn_frame, text="💾 Save Progress", command=self.save_progress, bg="#4A5568", fg="white", font=("Arial", 10, "bold"))
            self.btn_save_progress.pack(side="right", padx=5)
            
        # Export Excel button (Theme-styled dark navy by default)
        self.btn_export_excel = tk.Button(btn_frame, text="📊 Export Excel File", command=self.export_excel)
        self.btn_export_excel.pack(side="right", padx=5)

        if not self.read_only:
            self.btn_export_missing = tk.Button(btn_frame, text="📋 Export Missing Sourcing Report", command=self.export_missing_sourcing_report)
            self.btn_export_missing.pack(side="right", padx=5)

        self.lbl_row_count = tk.Label(btn_frame, text="Data Rows: 0 / 0", font=("Arial", 9, "bold"))
        self.lbl_row_count.pack(side="right", padx=15)
        
        # --- Top Bar (Row 1: Customer, RFQ & Assembly Selection & Info) ---
        top_bar = tk.Frame(self.window, padx=5, pady=5)
        top_bar.pack(fill="x")

        cust_display = self.cust_name if self.cust_name else "-"
        rfq_display = self.rfq_num if self.rfq_num else "-"
        from utils import get_bom_creation_date, BOM_DATA_DIR
        raw = getattr(self, 'raw_data', {})
        proj_title = raw.get("description", "") or raw.get("project_title", "") or raw.get("email_subject", "")
        created_at_display = get_bom_creation_date(raw)

        if not proj_title or not created_at_display or created_at_display.startswith("("):
            safe_cust = str(cust_display).replace('/', '_').replace('\\', '_').replace(' ', '_')
            safe_rfq = str(rfq_display).replace('/', '_').replace('\\', '_').replace(' ', '_')
            bom_file = os.path.join(BOM_DATA_DIR, safe_cust, f"{safe_rfq}.json")
            if os.path.exists(bom_file):
                try:
                    with open(bom_file, 'r', encoding='utf-8-sig') as f:
                        b_data = json.load(f)
                    if not proj_title:
                        proj_title = b_data.get("description", "") or b_data.get("project_title", "") or b_data.get("email_subject", "")
                    created_at_display = get_bom_creation_date(b_data, bom_file)
                except Exception:
                    pass

        tk.Label(top_bar, text="RFQ:", font=('Segoe UI', 10, 'bold'), fg="#1A365D").pack(side="left", padx=(5, 2))
        self.lbl_rfq = tk.Label(top_bar, text=f"{rfq_display}", font=('Segoe UI', 10, 'bold'), fg="#cc0000")
        self.lbl_rfq.pack(side="left", padx=(0, 14))

        tk.Label(top_bar, text="Customer:", font=('Segoe UI', 10, 'bold'), fg="#1A365D").pack(side="left", padx=(0, 2))
        self.lbl_cust = tk.Label(top_bar, text=f"{cust_display}", font=('Segoe UI', 10), fg="#1A365D")
        self.lbl_cust.pack(side="left", padx=(0, 14))

        tk.Label(top_bar, text="Project / Email Subject:", font=('Segoe UI', 10, 'bold'), fg="#1A365D").pack(side="left", padx=(0, 2))
        self.lbl_proj = tk.Label(top_bar, text=f"{proj_title or '-'}", font=('Segoe UI', 10), fg="#1A365D")
        self.lbl_proj.pack(side="left", padx=(0, 14))

        tk.Label(top_bar, text="BOM Created:", font=('Segoe UI', 10, 'bold'), fg="#1A365D").pack(side="left", padx=(0, 2))
        self.lbl_created = tk.Label(top_bar, text=f"{created_at_display}", font=('Segoe UI', 10), fg="#1A365D")
        self.lbl_created.pack(side="left", padx=(0, 15))
        
        tk.Label(top_bar, text="Select Assembly to Preview:").pack(side="left", padx=5)
        self.assy_combo = ttk.Combobox(top_bar, textvariable=self.selected_assembly, values=self.unique_assemblies, state="normal", width=30)
        self.assy_combo.pack(side="left", padx=5)
        self.assy_combo.bind("<<ComboboxSelected>>", self.on_assembly_changed)
        self.assy_combo.bind("<KeyRelease>", self.on_assy_combo_keyrelease)
        self.assy_combo.bind("<Return>", lambda e: self.on_assembly_changed())
        
        self.lbl_assy_model = tk.Label(top_bar, text="Model: -", font=('Arial', 9, 'bold'))
        self.lbl_assy_model.pack(side="left", padx=10)
        self.lbl_assy_rev = tk.Label(top_bar, text="Rev: -", font=('Arial', 9, 'bold'))
        self.lbl_assy_rev.pack(side="left", padx=10)
        
        # --- Action Bar (Row 2: Action Buttons) ---
        action_bar = tk.Frame(self.window, padx=10, pady=5)
        action_bar.pack(fill="x")
        
        self.btn_toggle = tk.Button(action_bar, text="[+] Expand Sourcing Details in Table", command=self.toggle_details)
        self.btn_toggle.pack(side="left", padx=(0, 10))
        
        #self.btn_filter = tk.Button(action_bar, text="🔍 Filter Records", command=self.open_filter_dialog)
        #self.btn_filter.pack(side="left", padx=10)
        
        if not self.read_only:
            self.btn_mark_viewed = tk.Button(action_bar, text="🟢 Mark as Viewed", command=self.mark_current_viewed)
            self.btn_mark_viewed.pack(side="left", padx=10)
            
            self.btn_mark_pending = tk.Button(action_bar, text="🟠 Keep in View", command=self.mark_current_pending)
            self.btn_mark_pending.pack(side="left", padx=10)

        # Create Vertical PanedWindow to let users dynamically adjust table vs card heights
        self.paned = tk.PanedWindow(self.window, orient="vertical", bd=0, bg="#EBF8FF", sashwidth=4, opaqueresize=True)
        self.paned.pack(fill="both", expand=True, padx=10, pady=5)

        # --- Top Section: Treeview (Master Sheet) ---
        self.tree_frame = tk.Frame(self.paned, padx=10, pady=10)
        
        self.header_canvas = tk.Canvas(self.tree_frame, height=90, highlightthickness=0, bg="#EBF8FF")
        self.header_canvas.pack(fill="x", side="top")
        
        style = ttk.Style(self)
        try:
            style.theme_use('clam')
        except:
            pass
        style.configure("SourcingUI.Treeview", font=("Segoe UI", 9), rowheight=26, background="white", fieldbackground="white")
        style.configure("SourcingUI.Treeview.Heading", font=("Segoe UI", 9, "bold"), background="#dcebfa", foreground="#1A365D")
        style.map("SourcingUI.Treeview", background=[("selected", "#FFF2B2")], foreground=[("selected", "#000000")])
        self.tree = ttk.Treeview(self.tree_frame, show="headings", height=10, selectmode="extended", style="SourcingUI.Treeview")
        self.tree.bind("<ButtonRelease-1>", lambda event: self.redraw_header_canvas(), add="+")
        self.tree.bind("<B1-Motion>", lambda event: self.redraw_header_canvas(), add="+")
        self.tree.bind("<Configure>", lambda event: self.redraw_header_canvas(), add="+")
        self.tree.tag_configure("assy_separator", background="#dcebfa", font=('Arial', 9, 'bold'))
        self.tree.tag_configure("manual_override", background="#e8f5e9") # Light Green
        self.tree.tag_configure("synced_override", background="#e3f2fd") # Light Blue
    
        top_y_scroll = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        
        def sync_xview(*args):
            self.tree.xview(*args)
            self.header_canvas.xview(*args)
            
        top_x_scroll = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=sync_xview)
        
        def update_scroll(first, last):
            top_x_scroll.set(first, last)
            self.header_canvas.xview_moveto(first)
            
        self.tree.configure(yscrollcommand=top_y_scroll.set, xscrollcommand=update_scroll)
        
        top_y_scroll.pack(side="right", fill="y")
        top_x_scroll.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.paned.add(self.tree_frame, minsize=150)
    
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.tree.bind("<Button-1>", self.on_tree_click)
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.bind("<Control-c>", self.copy_selected_range)
        self.tree.bind("<Control-C>", self.copy_selected_range)
        self.tree.bind("<Button-3>", self.show_context_menu)
        
        # Setup copy context menu
        self.context_menu = tk.Menu(self.window, tearoff=0)
        self.context_menu.add_command(label="📋 Copy Selected BOM Info (Excel Format)", command=self.copy_selected_range)
        self.context_menu.add_command(label="📋 Copy All Assembly BOM Info (Excel Format)", command=self.copy_all_assembly_bom_info)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="📋 Copy Selected Part Numbers", command=self.copy_selected_parts)
        self.context_menu.add_command(label="📋 Copy Selected MPNs", command=self.copy_selected_mpns)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="📦 Toggle Consign Item (Mark / Remove)", command=self.toggle_consign_item)
        self.context_menu.add_command(label="✏️ Edit MPN/MFR Pairs", command=self.open_pair_editor)

        # Container frame for pagination and details cards (only when not read_only)
        if self.read_only:
            self.bottom_pane = None
            self.pagination_frame = tk.Frame(self.tree_frame)
            self.pagination_frame.pack(side="bottom", fill="x", padx=10, pady=(5, 5))
        else:
            self.bottom_pane = tk.Frame(self.paned)
            self.pagination_frame = tk.Frame(self.bottom_pane)
            self.pagination_frame.pack(fill="x", padx=10, pady=(5, 5))
            
            # --- Middle Section: Detail Pane (Edit Mode) ---
            self.detail_frame = tk.LabelFrame(self.bottom_pane, text="Editing Sourcing for Part (Click a row above to edit)", padx=10, pady=10)
            self.detail_frame.pack(fill="both", expand=True, padx=10, pady=5)
            
            top_ctrl_frame = tk.Frame(self.detail_frame)
            top_ctrl_frame.pack(fill="x", pady=(0, 5))
            
            self.btn_revert = tk.Button(top_ctrl_frame, text="⏪ Revert to Original Quotes", command=self.revert_part, state="disabled")
            self.btn_revert.pack(side="left", padx=5)
            
            self.btn_edit_pairs = tk.Button(top_ctrl_frame, text="✏️ Edit MPN/MFR Pairs", command=self.open_pair_editor, state="disabled")
            self.btn_edit_pairs.pack(side="left", padx=5)
            
            self.btn_delete_rows = tk.Button(top_ctrl_frame, text="🗑️ Delete Selected Rows", command=self.delete_selected_rows)
            self.btn_delete_rows.pack(side="left", padx=5)

            self.btn_toggle_consign = tk.Button(top_ctrl_frame, text="📦 Mark as Consign", command=self.toggle_consign_item, state="disabled")
            self.btn_toggle_consign.pack(side="left", padx=5)

            ttk.Separator(top_ctrl_frame, orient="vertical").pack(side="left", padx=10, fill="y")

            self.btn_apply_manual = tk.Button(top_ctrl_frame, text="💾 Apply Changes to Row(s)", command=self.apply_manual_changes, state="disabled")
            self.btn_apply_manual.pack(side="left", padx=5)

            self.linked_parts_frame = tk.Frame(self.detail_frame)
            self.linked_parts_frame.pack(fill="x", pady=(5, 0), anchor="w")

            self.lbl_linked_parts = tk.Label(self.linked_parts_frame, text="", font=('Arial', 11, 'bold'), foreground="#003399", anchor="w")
            self.lbl_linked_parts.pack(side="left", pady=0)

            self.btn_show_shared = tk.Button(
                self.linked_parts_frame,
                text="🔍 View Shared Details",
                command=self.show_shared_details_dialog,
                state="disabled"
            )
            self.btn_show_shared.pack(side="left", padx=10)

            scroll_outer_frame = tk.Frame(self.detail_frame)
            scroll_outer_frame.pack(fill="both", expand=True)
            
            cards_canvas = tk.Canvas(scroll_outer_frame, highlightthickness=0, height=260)
            
            bottom_x_scroll = ttk.Scrollbar(scroll_outer_frame, orient="horizontal", command=cards_canvas.xview)
            bottom_y_scroll = ttk.Scrollbar(scroll_outer_frame, orient="vertical", command=cards_canvas.yview)
            
            cards_canvas.configure(xscrollcommand=bottom_x_scroll.set, yscrollcommand=bottom_y_scroll.set)
            
            bottom_y_scroll.pack(side="right", fill="y")
            bottom_x_scroll.pack(side="bottom", fill="x")
            cards_canvas.pack(side="left", fill="both", expand=True)
            
            self.cards_frame = tk.Frame(cards_canvas)
            cards_canvas.create_window((0, 0), window=self.cards_frame, anchor="nw")
            self.cards_frame.bind("<Configure>", lambda e: cards_canvas.configure(scrollregion=cards_canvas.bbox("all")))
            
            def _on_mousewheel(event):
                cards_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                
            self.cards_frame.bind("<MouseWheel>", _on_mousewheel)
            cards_canvas.bind("<MouseWheel>", _on_mousewheel)

            # Add the bottom container pane to the PanedWindow
            self.paned.add(self.bottom_pane, minsize=340)

        # Apply global theme styling!
        from dialogs import apply_panel_theme
        apply_panel_theme(self)
        
    def load_target_price_data(self):
        self.target_prices_by_assy = {}
        self.eau_by_assy = {}
        self.target_currency = "USD"
        self.target_markdown_pct = 20.0
        if self.bom_filepath and os.path.exists(self.bom_filepath):
            try:
                import json
                with open(self.bom_filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.target_currency = data.get("Target Currency", data.get("target_currency", "USD"))
                self.target_markdown_pct = float(data.get("Target Markdown %", data.get("target_markdown_pct", 20.0)))
                for assy in data.get("Assemblies", []):
                    assy_num = assy.get("Assy #")
                    if assy_num:
                        self.target_prices_by_assy[assy_num] = assy.get("Target Prices", {})
                        
                        # Load EAU as dict or legacy single value
                        self.eau_by_assy[assy_num] = {}
                        eau_raw = assy.get("EAU")
                        if isinstance(eau_raw, dict):
                            for mq, ev in eau_raw.items():
                                self.eau_by_assy[assy_num][str(mq)] = ev
                        elif eau_raw is not None and str(eau_raw).strip():
                            # Legacy fallback: copy to all assigned MOQs
                            moqs = assy.get("Assigned MOQs", [])
                            for mq in moqs:
                                self.eau_by_assy[assy_num][str(mq)] = eau_raw
            except Exception as e:
                print(f"Error loading target price data: {e}")

    def open_filter_dialog(self):
        dialog = BOMFilterDialog(self.window, self.current_filters)
        dialog.wait_for_close()
        if dialog.result is not None:
            self.current_filters = dialog.result
            self.apply_filters()

        else:
            self.filter_criteria = None
            if hasattr(self, 'populate_tree'):
                self.populate_tree()
            elif hasattr(self, 'refresh_projects'):
                self.refresh_projects()
    def apply_filters(self):
        part_filters = [p.lower() for p in self.current_filters.get("Part", []) if p.strip()]
        mpn_filters = [m.lower() for m in self.current_filters.get("MPN", []) if m.strip()]
        
        filtered_assemblies = []
        for assy in self.unique_assemblies:
            items = self.grouped_model.get(assy, [])
            match_found = False
            
            if not part_filters and not mpn_filters:
                match_found = True
            else:
                for idx, item in items:
                    item_part = str(item.get("Part", "")).lower()
                    item_mpns = [m.strip().lower() for m in str(item.get("MPN", "")).split(",")]
                    
                    part_match = False
                    if part_filters:
                        for pf in part_filters:
                            if pf in item_part:
                                part_match = True
                                break
                    else:
                        part_match = True
                        
                    mpn_match = False
                    if mpn_filters:
                        for mf in mpn_filters:
                            for im in item_mpns:
                                if mf in im:
                                    mpn_match = True
                                    break
                            if mpn_match:
                                break
                    else:
                        mpn_match = True
                        
                    if part_match and mpn_match:
                        match_found = True
                        break
                        
            if match_found:
                filtered_assemblies.append(assy)
                
        self.assy_combo.config(values=filtered_assemblies)
        
        if filtered_assemblies:
            current = self.selected_assembly.get()
            if current not in filtered_assemblies:
                self.selected_assembly.set(filtered_assemblies[0])
            self.on_assembly_changed(None)
        else:
            self.selected_assembly.set("")
            self.tree.delete(*self.tree.get_children())
            self.lbl_row_count.config(text="Data Rows: 0 / 0")
            self.clear_detail_pane()

    def go_back(self):
        if getattr(self, 'read_only', False):
            if hasattr(self, '_wait_var'):
                self._wait_var.set(1)
            return
        dialog = SourcingCancelWarningDialog(self.winfo_toplevel())
        self.winfo_toplevel().wait_window(dialog)
        if dialog.result:
            self._wait_var.set(1)
            if hasattr(self, 'on_back_callback') and self.on_back_callback:
                self.on_back_callback()

    def refresh_moq_widgets(self):
        if not hasattr(self, 'cards_frame') or self.cards_frame is None:
            return
        for widget in self.cards_frame.winfo_children():
            widget.destroy()

        self.moq_widgets = {}
        
        for i, q in enumerate(self.active_categories):
            frame = tk.Frame(self.cards_frame, highlightbackground="#1A365D", highlightthickness=2, bd=0, padx=12, pady=12)
            frame.grid(row=0, column=i, padx=8, pady=8, sticky="n")
            
            tk.Label(frame, text=f"MOQ {q} Quotes", font=('Segoe UI', 9, 'bold'), wraplength=230).pack(anchor='w')
            
            lbl_sel = tk.Label(frame, text="SYSTEM BEST CHOICE", font=('Segoe UI', 8, 'bold'), fg="#2b6cb0", bg="#ebf8ff", bd=1, relief="solid", padx=5, pady=3, wraplength=230)
            lbl_sel.pack(fill="x", pady=(5, 8))
            
            btn_sel = tk.Button(frame, text="🔍 Change Winning Supplier", font=('Segoe UI', 8, 'bold'), 
                               command=lambda current_q=q: self.open_sourcing_detail_dialog(current_q), 
                               bg="#EBF8FF", fg="#475569", relief="flat", bd=1)
            btn_sel.pack(fill="x", pady=(0, 10))

            lbl_supplier = tk.Label(frame, text="Supplier: -", font=('Segoe UI', 8, 'bold'), wraplength=230)
            lbl_supplier.pack(anchor='w', pady=(5,0))
            
            lbl_mpn = tk.Label(frame, text="MPN: -", font=('Segoe UI', 9), wraplength=230)
            lbl_mpn.pack(anchor='w', pady=(2,0))
            
            lbl_moq = tk.Label(frame, text="MOQ: -", font=('Segoe UI', 9), wraplength=230)
            lbl_moq.pack(anchor='w', pady=(2,0))
            
            lbl_date = tk.Label(frame, text="Date: -", font=('Segoe UI', 9), wraplength=230)
            lbl_date.pack(anchor='w', pady=(2,0))
            
            lbl_ltime = tk.Label(frame, text="L/Time: -", font=('Segoe UI', 9), wraplength=230)
            lbl_ltime.pack(anchor='w', pady=(2,0))
            
            lbl_unitp = tk.Label(frame, text="Unit Price: -", font=('Segoe UI', 9), wraplength=230)
            lbl_unitp.pack(anchor='w', pady=(2,0))
            
            ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=8)
            
            lbl_bom = tk.Label(frame, text="BOM: $0.0000", font=('Segoe UI', 9), foreground="blue", wraplength=230)
            lbl_bom.pack(anchor='w')
            
            lbl_exc = tk.Label(frame, text="Excess: $0.0000", font=('Segoe UI', 9), foreground="red", wraplength=230)
            lbl_exc.pack(anchor='w')
            
            lbl_tot = tk.Label(frame, text="Total: $0.0000", font=('Segoe UI', 8, 'bold'), wraplength=230)
            lbl_tot.pack(anchor='w')

            def on_double_click(event, current_q=q):
                self.open_sourcing_detail_dialog(current_q)

            frame.bind("<Double-1>", on_double_click)
            def bind_recursively(w):
                w.bind("<Double-1>", on_double_click)
                for child in w.winfo_children():
                    bind_recursively(child)
            bind_recursively(frame)
            
            # Apply panel theme styling with bg="#FFFFFF" and fg="#1A365D"
            # to make cards stand out as white panels on the light-blue background.
            apply_panel_theme(frame, bg="#FFFFFF", fg="#1A365D")
            
            # Re-apply state specific foreground/background colors for specialized labels to prevent override
            lbl_sel.config(text="SYSTEM BEST CHOICE", font=('Segoe UI', 8, 'bold'), fg="#2b6cb0", bg="#ebf8ff", bd=1, relief="solid")
            lbl_bom.config(foreground="blue")
            lbl_exc.config(foreground="red")
            
            self.moq_widgets[q] = {
                'lbl_sel': lbl_sel,
                'lbl_supplier': lbl_supplier,
                'lbl_mpn': lbl_mpn,
                'lbl_moq': lbl_moq,
                'lbl_date': lbl_date,
                'lbl_ltime': lbl_ltime,
                'lbl_unitp': lbl_unitp,
                'lbl_bom': lbl_bom,
                'lbl_exc': lbl_exc,
                'lbl_tot': lbl_tot
            }

    def mark_current_viewed(self):
        assy = self.selected_assembly.get()
        if assy:
            self.assembly_status[assy] = "Viewed"
            self.update_mark_buttons(assy)

    def mark_current_pending(self):
        assy = self.selected_assembly.get()
        if assy:
            self.assembly_status[assy] = "Pending"
            self.update_mark_buttons(assy)

    def update_mark_buttons(self, assy):
        if not hasattr(self, 'btn_mark_viewed') or self.btn_mark_viewed is None:
            return
        status = self.assembly_status.get(assy, "Unviewed")
        if status == "Viewed":
            self.btn_mark_viewed.config(text="✔️ Marked Viewed", state="disabled", bg="#28A745", fg="white")
            self.btn_mark_pending.config(text="🟠 Keep in View", state="normal", bg="#1A365D", fg="white")
        elif status == "Pending":
            self.btn_mark_viewed.config(text="🟢 Mark as Viewed", state="normal", bg="#1A365D", fg="white")
            self.btn_mark_pending.config(text="⏳ Marked Pending", state="disabled", bg="#DD6B20", fg="white")
        else:
            self.btn_mark_viewed.config(text="🟢 Mark as Viewed", state="normal", bg="#1A365D", fg="white")
            self.btn_mark_pending.config(text="🟠 Keep in View", state="normal", bg="#1A365D", fg="white")

    def on_assembly_changed(self, event=None):
        self.assy_combo.config(values=self.unique_assemblies)
        self.dismiss_suggestions_popup()
        
        assy = self.assy_combo.get().strip()
        if not assy:
            if self.unique_assemblies:
                assy = self.unique_assemblies[0]
                self.selected_assembly.set(assy)
                self.assy_combo.set(assy)
            else:
                return
        
        if assy not in self.unique_assemblies:
            matches = [a for a in self.unique_assemblies if assy.lower() in a.lower()]
            if matches:
                assy = matches[0]
                self.selected_assembly.set(assy)
                self.assy_combo.set(assy)
            else:
                self.selected_assembly.set("")
                self.assy_combo.set("")
                return
        else:
            self.selected_assembly.set(assy)
            
        self.active_categories = sorted(self.assembly_moqs.get(assy, []))
        self.update_mark_buttons(assy)
        
        self.all_cols = list(self.base_cols)
        self.compact_cols = list(self.base_cols)
        
        for q in self.active_categories:
            self.compact_cols.append(f"Winner_{q}")
            for col_id, col_name, width in self.moq_col_defs:
                self.all_cols.append(f"{col_id}_{q}")
                
        self.tree["displaycolumns"] = "#all"
        self.tree["columns"] = self.all_cols
        self._setup_tree_columns()
        
        if self.show_details:
            self.tree["displaycolumns"] = self.all_cols
        else:
            self.tree["displaycolumns"] = self.compact_cols
            
        self.redraw_header_canvas()
        self.refresh_moq_widgets()
        
        # Select the first row of the new assembly as active edit row by default
        items_for_assy = self.grouped_model.get(assy, [])
        if items_for_assy:
            self.active_edit_ref_id = items_for_assy[0][1].get('ref_id')
        else:
            self.active_edit_ref_id = None
            
        self.current_selected_idx = None
        self.populate_tree()
        self.load_active_edit_row()

        # Update Model and Rev labels
        found_model = "-"
        found_rev = "-"

        def _extract_from_dict(d, *keys):
            for k in keys:
                v = d.get(k)
                if v is not None and str(v).strip() != "" and str(v).strip() != "-":
                    return str(v).strip()
            return "-"

        model_keys = ('Assy Model', 'Assy_Model', 'Model', 'Model #', 'Model Description', 'Model Name', 'model')
        rev_keys = ('Assy Rev', 'Assy_Rev', 'Rev', 'Revision', 'rev')

        # 1. Check item dict keys in gui_model
        for item in self.gui_model:
            item_assy = str(item.get('Assy', '') or item.get('Assembly #', '')).strip()
            if item_assy == assy:
                if found_model == "-":
                    found_model = _extract_from_dict(item, *model_keys)
                if found_rev == "-":
                    found_rev = _extract_from_dict(item, *rev_keys)
                if found_model != "-" and found_rev != "-":
                    break

        # 2. Check raw_data Assemblies list
        if found_model == "-" or found_rev == "-":
            raw_d = getattr(self, 'raw_data', {})
            for a_info in raw_d.get('Assemblies', []):
                if isinstance(a_info, dict):
                    a_num = str(a_info.get('Assy #') or a_info.get('Assy') or a_info.get('Assembly #')).strip()
                    if a_num == assy:
                        if found_model == "-":
                            found_model = _extract_from_dict(a_info, *model_keys)
                        if found_rev == "-":
                            found_rev = _extract_from_dict(a_info, *rev_keys)
                        break

        # 3. Fallback: check raw BOM file in BOM_DATA_DIR
        if found_model == "-" or found_rev == "-":
            from utils import BOM_DATA_DIR
            safe_c = str(getattr(self, 'cust_name', '')).replace('/', '_').replace('\\', '_').replace(' ', '_')
            safe_r = str(getattr(self, 'rfq_num', '')).replace('/', '_').replace('\\', '_').replace(' ', '_')
            bom_file = os.path.join(BOM_DATA_DIR, safe_c, f"{safe_r}.json")
            if os.path.exists(bom_file):
                try:
                    with open(bom_file, 'r', encoding='utf-8-sig') as f:
                        b_data = json.load(f)
                    for a_info in b_data.get('Assemblies', []):
                        if isinstance(a_info, dict):
                            a_num = str(a_info.get('Assy #') or a_info.get('Assy') or a_info.get('Assembly #')).strip()
                            if a_num == assy:
                                if found_model == "-":
                                    found_model = _extract_from_dict(a_info, *model_keys)
                                if found_rev == "-":
                                    found_rev = _extract_from_dict(a_info, *rev_keys)
                                break
                except Exception:
                    pass

        self.lbl_assy_model.config(text=f"Model: {found_model}")
        self.lbl_assy_rev.config(text=f"Rev: {found_rev}")

        # Update pagination UI
        self.update_pagination_ui()

    def update_pagination_ui(self):
        # Clear existing widgets in pagination_frame
        for widget in self.pagination_frame.winfo_children():
            widget.destroy()
            
        N = len(self.unique_assemblies)
        if N <= 0: return
        
        current_assy = self.selected_assembly.get()
        current_idx = self.unique_assemblies.index(current_assy) if current_assy in self.unique_assemblies else 0
        page = current_idx + 1 # 1-based page index
        
        # Outer container to center the pagination widgets
        cnt_frame = tk.Frame(self.pagination_frame, bg="#EBF8FF")
        cnt_frame.pack(anchor="center")
        
        # Prev button
        btn_prev = tk.Button(cnt_frame, text="◀ Prev", font=("Arial", 9, "bold"),
                             command=lambda: self.change_assembly_by_page(page - 1),
                             state="normal" if page > 1 else "disabled")
        btn_prev.pack(side="left", padx=5)
        
        # Determine range of pages to show
        start_page = max(1, page - 2)
        end_page = min(N, start_page + 4)
        if end_page - start_page < 4:
            start_page = max(1, end_page - 4)
            
        # First page with ellipsis if needed
        if start_page > 1:
            btn_first = tk.Button(cnt_frame, text="1", font=("Arial", 9),
                                  command=lambda: self.change_assembly_by_page(1))
            btn_first.pack(side="left", padx=2)
            if start_page > 2:
                tk.Label(cnt_frame, text="...", bg="#EBF8FF", font=("Arial", 9)).pack(side="left", padx=2)
                
        # Middle pages
        for p in range(start_page, end_page + 1):
            is_active = (p == page)
            btn_p = tk.Button(cnt_frame, text=str(p), font=("Arial", 9, "bold" if is_active else "normal"))
            btn_p.pack(side="left", padx=2)
            btn_p.config(command=lambda target_p=p: self.change_assembly_by_page(target_p))
            
        # Last page with ellipsis if needed
        if end_page < N:
            if end_page < N - 1:
                tk.Label(cnt_frame, text="...", bg="#EBF8FF", font=("Arial", 9)).pack(side="left", padx=2)
            btn_last = tk.Button(cnt_frame, text=str(N), font=("Arial", 9),
                                 command=lambda: self.change_assembly_by_page(N))
            btn_last.pack(side="left", padx=2)
            
        # Next button
        btn_next = tk.Button(cnt_frame, text="Next ▶", font=("Arial", 9, "bold"),
                             command=lambda: self.change_assembly_by_page(page + 1),
                             state="normal" if page < N else "disabled")
        btn_next.pack(side="left", padx=5)
        
        # Status Label: "Assembly 5 of 20 (12-E56253-N200)"
        lbl_status = tk.Label(cnt_frame, text=f"  |  Assembly {page} of {N}: {current_assy}", font=("Arial", 9, "bold"), fg="#1A365D", bg="#EBF8FF")
        lbl_status.pack(side="left", padx=10)
        
        # Direct page jump box
        tk.Label(cnt_frame, text="  |  Go to Page:", bg="#EBF8FF", font=("Arial", 9, "bold"), fg="#1A365D").pack(side="left", padx=(10, 2))
        self.page_jump_entry = tk.Entry(cnt_frame, width=5, font=("Arial", 9))
        self.page_jump_entry.pack(side="left", padx=2)
        # Pre-fill with current page number
        self.page_jump_entry.insert(0, str(page))
        self.page_jump_entry.bind("<Return>", lambda e: self.on_page_jump_submit())
        self.page_jump_entry.bind("<FocusOut>", lambda e: self.on_page_jump_submit())
        # Auto-apply after short delay while typing
        self._page_jump_after_id = None
        def _on_page_key(event):
            if self._page_jump_after_id:
                try: self.page_jump_entry.after_cancel(self._page_jump_after_id)
                except: pass
            self._page_jump_after_id = self.page_jump_entry.after(600, self.on_page_jump_submit)
        self.page_jump_entry.bind("<KeyRelease>", _on_page_key)
        
        btn_go = tk.Button(cnt_frame, text="Go", font=("Arial", 8, "bold"), command=self.on_page_jump_submit)
        btn_go.pack(side="left", padx=2)
        
        # Apply premium button styling
        from dialogs import style_premium_button
        style_premium_button(btn_go, bg_color="#1A365D", fg_color="#ffffff", hover_bg="#0077B6")
        
        for child in cnt_frame.winfo_children():
            if isinstance(child, tk.Button) and child != btn_go:
                txt = child.cget("text")
                if txt.isdigit():
                    p_num = int(txt)
                    if p_num == page:
                        style_premium_button(child, bg_color="#0077B6", fg_color="#ffffff", hover_bg="#1A365D")
                    else:
                        style_premium_button(child, bg_color="#ebf8ff", fg_color="#1A365D", hover_bg="#bee3f8")
                elif "Prev" in txt:
                    style_premium_button(child, bg_color="#1A365D", fg_color="#ffffff", hover_bg="#0077B6")
                elif "Next" in txt:
                    style_premium_button(child, bg_color="#1A365D", fg_color="#ffffff", hover_bg="#0077B6")

    def on_page_jump_submit(self):
        if not hasattr(self, 'page_jump_entry'): return
        val = self.page_jump_entry.get().strip()
        try:
            p = int(val)
            if 1 <= p <= len(self.unique_assemblies):
                self.change_assembly_by_page(p)
            else:
                messagebox.showwarning("Invalid Page", f"Please enter a page number between 1 and {len(self.unique_assemblies)}.", parent=self.winfo_toplevel())
        except ValueError:
            messagebox.showwarning("Invalid Input", "Please enter a valid page number.", parent=self.winfo_toplevel())

    def change_assembly_by_page(self, target_page):
        if 1 <= target_page <= len(self.unique_assemblies):
            assy = self.unique_assemblies[target_page - 1]
            self.selected_assembly.set(assy)
            self.assy_combo.set(assy)
            self.on_assembly_changed()

    def on_assy_combo_keyrelease(self, event):
        if event.keysym in ("UpDown", "Down", "Up", "Left", "Right", "Return", "Tab", "Escape", "Shift_L", "Shift_R", "Control_L", "Control_R"):
            return
            
        typed_text = self.assy_combo.get().strip().lower()
        if not typed_text:
            self.dismiss_suggestions_popup()
            return
            
        filtered = [a for a in self.unique_assemblies if typed_text in a.lower()]
        
        if filtered:
            self.show_assy_suggestions_popup(filtered)
        else:
            self.dismiss_suggestions_popup()

    def on_suggestion_select(self, event=None):
        if not hasattr(self, 'assy_listbox') or not self.assy_listbox.winfo_exists():
            return
        selection = self.assy_listbox.curselection()
        if selection:
            idx = selection[0]
            assy = self.assy_listbox.get(idx)
            self.selected_assembly.set(assy)
            self.assy_combo.set(assy)
            self.dismiss_suggestions_popup()
            self.on_assembly_changed()

    def _setup_tree_columns(self):
        base_widths = {"Comp Level": 80, "Assy": 150, "Part": 120, "Description": 200, "MFR": 120, "MPN": 200, "BOM Qty": 60, "UOM": 50}
        for col in self.base_cols:
            text = col if col not in ["Assy", "Part", "BOM Qty"] else {"Assy":"Assy #", "Part":"Part #", "BOM Qty":"BOM Qty"}[col]
            self.tree.heading(col, text=text, anchor="w")
            self.tree.column(col, width=base_widths.get(col, 100), minwidth=40, stretch=False)
            
        for q in self.active_categories:
            for col_id, col_name, width in self.moq_col_defs:
                full_col = f"{col_id}_{q}"
                self.tree.heading(full_col, text=col_name, anchor="w")
                self.tree.column(full_col, width=width, minwidth=40, stretch=False)

    def toggle_details(self):
        self.show_details = not self.show_details
        if self.show_details:
            self.tree["displaycolumns"] = self.all_cols
            self.btn_toggle.config(text="[-] Collapse Sourcing Details")
        else:
            self.tree["displaycolumns"] = self.compact_cols
            self.btn_toggle.config(text="[+] Expand Sourcing Details in Table")
        
        self.redraw_header_canvas()

    def redraw_header_canvas(self):
        self.header_canvas.delete("all")
        
        # Get actual column widths dynamically from Treeview columns to ensure perfect alignment
        actual_base_widths = {}
        for col in self.base_cols:
            try:
                actual_base_widths[col] = self.tree.column(col, "width")
            except:
                actual_base_widths[col] = {"Comp Level": 80, "Assy": 150, "Part": 120, "Description": 200, "MFR": 120, "MPN": 200, "BOM Qty": 60, "UOM": 50}.get(col, 100)
                
        base_width = sum(actual_base_widths.values())
        x_offset = base_width
        
        self.header_canvas.create_rectangle(0, 0, base_width, 90, fill="#e8eaed", outline="lightgray")
        self.header_canvas.create_text(base_width/2, 45, text="Base BOM Information", font=('Segoe UI', 11, 'bold'), fill="#333333")
        
        assy = self.selected_assembly.get().strip()
            
        for q in self.active_categories:
            # 1. Calculate active assembly totals for this MOQ q
            has_any_winner = False
            has_missing_quote = False
            total_bom = 0.0
            total_exc = 0.0
            
            items_for_assy = [it for idx, it in self.grouped_model.get(assy, [])]
            for it in items_for_assy:
                res = it.get('moq_results', {}).get(q, {})
                winner = res.get('pending_winner') or res.get('winner')
                if winner:
                    has_any_winner = True
                    total_bom += winner.get('BOM Cost', 0.0)
                    total_exc += winner.get('Excess Cost', 0.0)
                else:
                    has_mpn = bool(str(it.get('MPN', '')).strip())
                    has_mfr = bool(str(it.get('MFR', '')).strip())
                    has_qty = pd.to_numeric(it.get('BOM Qty', 0), errors='coerce') > 0
                    if has_mpn and has_mfr and has_qty and not it.get('is_consign'):
                        has_missing_quote = True

            if has_any_winner:
                status_suffix = " (Pending)" if has_missing_quote else ""
                bom_str = f"BOM: ${total_bom:.4f}{status_suffix}"
                exc_str = f"Exc: ${total_exc:.4f}{status_suffix}"
            else:
                bom_str = "BOM: Pending"
                exc_str = "Exc: Pending"

            # Parse Target Price
            target_price_val = None
            if hasattr(self, 'target_prices_by_assy') and self.target_prices_by_assy:
                target_price_val = self.target_prices_by_assy.get(assy, {}).get(str(q))
                
            eau_val = self.eau_by_assy.get(assy, {}).get(str(q))
            eau_str = f"EAU: {int(float(eau_val)):,}" if eau_val is not None and str(eau_val).strip() else "EAU: None"

            if target_price_val is not None:
                try:
                    T = float(target_price_val)
                    t_curr = getattr(self, 'target_currency', 'USD')
                    rate_t = self.engine._get_rate_to_usd(t_curr)
                    usd_target = T / rate_t if rate_t > 0 else T
                    markdown_pct = getattr(self, 'target_markdown_pct', 20.0)
                    marked_down_usd = usd_target * (1.0 - markdown_pct / 100.0)
                    tgt_str = f"Tgt: USD ${marked_down_usd:.4f}"
                except Exception:
                    tgt_str = "Tgt: Error"
            else:
                tgt_str = "Tgt: None"

            if self.show_details:
                # Sum of actual columns in the first group
                g1_actual_w = 0
                for col_id, _, _ in self.moq_col_defs[0:6]:
                    full_col = f"{col_id}_{q}"
                    try:
                        g1_actual_w += self.tree.column(full_col, "width")
                    except:
                        g1_actual_w += {"Winner": 120, "UnitPrice": 80, "MOQ": 60, "SourceDate": 100, "MPN": 150, "Currency": 80}.get(col_id, 100)
                
                # Sum of actual columns in the third group (Costing & Calculation)
                g3_actual_w = 0
                for col_id, _, _ in self.moq_col_defs[6:11]:
                    full_col = f"{col_id}_{q}"
                    try:
                        g3_actual_w += self.tree.column(full_col, "width")
                    except:
                        g3_actual_w += {"ConvUSD": 90, "Markup": 90, "BOMCost": 90, "ExCost": 90, "TotCost": 90}.get(col_id, 100)
                
                moq_width = g1_actual_w + g3_actual_w
                
                # Top MOQ Category Title (height 25)
                self.header_canvas.create_rectangle(x_offset, 0, x_offset + moq_width, 25, fill="#cfe2f3", outline="gray")
                self.header_canvas.create_text(x_offset + (moq_width / 2), 12, text=f"MOQ {q}", font=('Segoe UI', 10, 'bold'), fill="#000000")
                
                # Active Assembly summary line (height 45)
                self.header_canvas.create_rectangle(x_offset, 25, x_offset + moq_width, 70, fill="#f8fafc", outline="gray")
                info_text = f"{eau_str}   |   {tgt_str}   |   {bom_str}   |   {exc_str}"
                self.header_canvas.create_text(x_offset + (moq_width / 2), 47, text=info_text, font=('Segoe UI', 9, 'bold'), fill="#1e293b")
                
                # Sub-headers (height 20)
                self.header_canvas.create_rectangle(x_offset, 70, x_offset + g1_actual_w, 90, fill="#d9ead3", outline="gray")
                self.header_canvas.create_text(x_offset + (g1_actual_w / 2), 80, text="Supplier Source Data", font=('Segoe UI', 8, 'bold'))
                x_sub = x_offset + g1_actual_w
                
                self.header_canvas.create_rectangle(x_sub, 70, x_sub + g3_actual_w, 90, fill="#f4cccc", outline="gray")
                self.header_canvas.create_text(x_sub + (g3_actual_w / 2), 80, text="Costing & Calculation", font=('Segoe UI', 8, 'bold'))
            else:
                try:
                    moq_width = self.tree.column(f"Winner_{q}", "width")
                except:
                    moq_width = 120
                    
                # Top MOQ Category Title (height 25)
                self.header_canvas.create_rectangle(x_offset, 0, x_offset + moq_width, 25, fill="#cfe2f3", outline="gray")
                self.header_canvas.create_text(x_offset + (moq_width / 2), 12, text=f"MOQ {q}", font=('Segoe UI', 10, 'bold'), fill="#000000")
                
                # Active Assembly Info Panel (height 65) - Stacked vertically
                self.header_canvas.create_rectangle(x_offset, 25, x_offset + moq_width, 90, fill="#f8fafc", outline="gray")
                self.header_canvas.create_text(x_offset + (moq_width / 2), 35, text=eau_str, font=('Segoe UI', 8, 'bold'), fill="#1e293b")
                self.header_canvas.create_text(x_offset + (moq_width / 2), 49, text=tgt_str, font=('Segoe UI', 8, 'bold'), fill="#0d9488")
                self.header_canvas.create_text(x_offset + (moq_width / 2), 63, text=bom_str, font=('Segoe UI', 8, 'bold'), fill="blue")
                self.header_canvas.create_text(x_offset + (moq_width / 2), 77, text=exc_str, font=('Segoe UI', 8, 'bold'), fill="red")

            x_offset += moq_width
            
        self.header_canvas.config(scrollregion=(0, 0, x_offset, 90))

    def _get_row_values(self, item_data):
        comp_level = str(item_data.get('Comp Level', ''))
        is_checked = (item_data.get('ref_id') == getattr(self, 'active_edit_ref_id', None))
        cb_char = "☑ " if is_checked else "☐ "
        
        if item_data.get('manual_moqs'):
            if not comp_level.startswith("✏️"): comp_level = "✏️ " + comp_level
        elif item_data.get('synced_moqs'):
            if not comp_level.startswith("🔗"): comp_level = "🔗 " + comp_level
            
        comp_level = cb_char + comp_level
            
        desc = str(item_data.get('Description', ''))
        mpn = str(item_data.get('MPN', ''))

        val_map = {
            "Comp Level": comp_level,
            "Assy": item_data.get('Assy', ''),
            "Part": item_data.get('Part', ''),
            "Description": desc,
            "MFR": item_data.get('MFR', ''),
            "MPN": mpn,
            "BOM Qty": format_to_sig_figs(item_data.get('BOM Qty', '')),
            "UOM": item_data.get('UOM', '')
        }
        
        for q in self.active_categories:
            moq_data = item_data.get('moq_results', {}).get(q, {})
            winner = moq_data.get('pending_winner') or moq_data.get('winner')
            is_consign_win = item_data.get('is_consign') or (winner and str(winner.get('Supplier', '')).strip().upper() in ('CONSIGN', 'CONSIGN ITEM'))
            if is_consign_win:
                val_map[f"Winner_{q}"] = "CONSIGN ITEM"
                val_map[f"UnitP_{q}"] = "$0.0000"
                val_map[f"MOQ_{q}"] = "1"
                val_map[f"SrcDate_{q}"] = "-"
                val_map[f"MPN_{q}"] = "CONSIGN ITEM"
                val_map[f"Cur_{q}"] = "USD"
                val_map[f"UOM_{q}"] = str(item_data.get('UOM', ''))
                val_map[f"StdPack_{q}"] = "-"
                val_map[f"Stock_{q}"] = "-"
                val_map[f"LTime_{q}"] = "0"
                val_map[f"Remark_{q}"] = ""  # Leave blank for CONSIGN ITEM
                val_map[f"ConvUSD_{q}"] = "$0.0000"
                val_map[f"Markup_{q}"] = "$0.0000"
                val_map[f"BOMCost_{q}"] = "$0.0000"
                val_map[f"ExCost_{q}"] = "$0.0000"
                val_map[f"TotCost_{q}"] = "$0.0000"
            elif winner:
                supp_name = winner.get('Supplier', '')
                val_map[f"Winner_{q}"] = "CONSIGN ITEM" if str(supp_name).strip().upper() in ('CONSIGN', 'CONSIGN ITEM') else supp_name
                val_map[f"UnitP_{q}"] = f"${winner.get('Unit Price', 0):.4f}"
                val_map[f"MOQ_{q}"] = winner.get('MOQ', '')
                
                sd = winner.get('Source Date', '')
                date_clean = "-"
                if sd and pd.notna(sd):
                    try:
                        if hasattr(sd, 'strftime'):
                            date_clean = sd.strftime("%d/%m/%Y")
                        else:
                            date_clean = str(sd).split(' ')[0]
                    except:
                        date_clean = str(sd).split(' ')[0]
                val_map[f"SrcDate_{q}"] = date_clean
                
                val_map[f"MPN_{q}"] = winner.get('MPN', '')
                val_map[f"Cur_{q}"] = winner.get('Currency', '')
                val_map[f"UOM_{q}"] = winner.get('uom', '')
                val_map[f"StdPack_{q}"] = winner.get('Std Pack', '')
                val_map[f"Stock_{q}"] = winner.get('Stock', '')
                val_map[f"LTime_{q}"] = winner.get('L/Time (weeks)', winner.get('L/Time', ''))
                val_map[f"Remark_{q}"] = winner.get('Remark', '')
                
                val_map[f"ConvUSD_{q}"] = f"${winner.get('Convert to USD', 0):.4f}"
                marked_up_val = winner.get('Convert to USD', 0) * winner.get('Markup Rate', 1.0)
                val_map[f"Markup_{q}"] = f"${marked_up_val:.4f}"
                val_map[f"BOMCost_{q}"] = f"${winner.get('BOM Cost', 0):.4f}"
                val_map[f"ExCost_{q}"] = f"${winner.get('Excess Cost', 0):.4f}"
                val_map[f"TotCost_{q}"] = f"${winner.get('Total Cost', 0):.4f}"
            else:
                for col_id, _, _ in self.moq_col_defs:
                    val_map[f"{col_id}_{q}"] = "-"
                has_mpn = bool(str(item_data.get('MPN', '')).strip())
                has_mfr = bool(str(item_data.get('MFR', '')).strip())
                has_qty = pd.to_numeric(item_data.get('BOM Qty', 0), errors='coerce') > 0
                if not (has_mpn and has_mfr and has_qty):
                    val_map[f"Winner_{q}"] = ""
                else:
                    val_map[f"Winner_{q}"] = "❌ No Valid Quote"
                
        return [val_map.get(col, '') for col in self.all_cols]

    def populate_tree(self):
        self.tree.delete(*self.tree.get_children())
        
        selected_assy = self.selected_assembly.get()
        if not selected_assy:
            self.lbl_row_count.config(text="Data Rows: 0 / 0")
            return
            
        items_for_assy = self.grouped_model.get(selected_assy, [])
        
        part_filters = [p.lower() for p in self.current_filters.get("Part", []) if p.strip()]
        mpn_filters = [m.lower() for m in self.current_filters.get("MPN", []) if m.strip()]
        
        for idx, item in items_for_assy:
            part_match = False
            item_part = str(item.get("Part", "")).lower()
            if part_filters:
                for pf in part_filters:
                    if pf in item_part:
                        part_match = True
                        break
            else:
                part_match = True
                
            mpn_match = False
            item_mpns = [m.strip().lower() for m in str(item.get("MPN", "")).split(",")]
            if mpn_filters:
                for mf in mpn_filters:
                    for im in item_mpns:
                        if mf in im:
                            mpn_match = True
                            break
                    if mpn_match:
                        break
            else:
                mpn_match = True
                
            if part_match and mpn_match:
                vals = self._get_row_values(item)
                
                tags = ()
                if item.get('manual_moqs'):
                    tags = ("manual_override",)
                elif item.get('synced_moqs'):
                    tags = ("synced_override",)
                
                self.tree.insert("", "end", iid=str(idx), values=vals, tags=tags)
                
        total = len(self.tree.get_children())
        self.lbl_row_count.config(text=f"Data Rows: 0 / {total}")
        self.redraw_header_canvas()

    def on_tree_select(self, event=None):
        selected = self.tree.selection()
        total = len(self.tree.get_children())
        if selected:
            children = self.tree.get_children()
            try:
                idx = children.index(selected[0]) + 1
                self.lbl_row_count.config(text=f"Data Rows: {idx} / {total}")
            except:
                self.lbl_row_count.config(text=f"Data Rows: ? / {total}")
        else:
            self.lbl_row_count.config(text=f"Data Rows: 0 / {total}")

        if not selected: return
        
        # If exactly one row is selected, make it the active edit row automatically!
        if len(selected) == 1:
            iid = selected[0]
            if iid.startswith("sep_"):
                self.detail_frame.config(text="Editing Sourcing for Part (Please select a Part row, not an Assembly header)")
                self.current_selected_idx = None
                self.clear_detail_pane()
                return
            idx = int(iid)
            item = self.gui_model[idx]
            if getattr(self, 'active_edit_ref_id', None) != item.get('ref_id'):
                self.active_edit_ref_id = item.get('ref_id')
                self.populate_tree_preserving_selection()
                self.load_active_edit_row()

    def show_shared_details_dialog(self):
        if self.current_selected_idx is None: return
        item = self.gui_model[self.current_selected_idx]
        dialog = SharedPartsDialog(self, item['Part'], item['Assy'], self.current_shared_items, self.assembly_moqs)
        self.wait_window(dialog)

    def open_pair_editor(self):
        if self.current_selected_idx is None: return
        item = self.gui_model[self.current_selected_idx]
        part_num = item.get('Part', '')
        original_bom_mpn = item.get('MPN', '')
        original_bom_mfr = item.get('MFR', '')
        
        dialog = MPNMFRPairDialog(self.window, item['MPN'], item['MFR'])
        self.window.wait_window(dialog)
        
        if dialog.result:
            new_mpn, new_mfr = dialog.result
            
            # --- Sync to Alternative MPNs Database ---
            from utils import get_alternative_mpn_path, merge_mpn_mfr_pairs
            import os, json
            
            alt_json_path = get_alternative_mpn_path(self.cust_name)
            
            db_mpn, db_mfr = "", ""
            alt_data = {"Customer": self.cust_name, "Parts": {}}
            
            if os.path.exists(alt_json_path):
                try:
                    with open(alt_json_path, 'r', encoding='utf-8') as f:
                        alt_data = json.load(f)
                    alt_parts = alt_data.get("Parts", {})
                    if part_num in alt_parts:
                        db_rec = alt_parts[part_num]
                        db_mpn = db_rec.get("MPN", "")
                        db_mfr = db_rec.get("MFR", "")
                except Exception as e:
                    print(f"Error checking database alternatives: {e}")
            
            # Merge new pairs with existing database alternatives to preserve database history
            merged_mpn, merged_mfr = merge_mpn_mfr_pairs(db_mpn, db_mfr, new_mpn, new_mfr)
            
            # Update DB JSON file if changed
            db_changed = merged_mpn != db_mpn or merged_mfr != db_mfr
            if db_changed:
                if "Parts" not in alt_data:
                    alt_data["Parts"] = {}
                alt_data['Parts'][part_num] = {
                    "MPN": merged_mpn,
                    "MFR": merged_mfr
                }
                
                try:
                    with open(alt_json_path, 'w', encoding='utf-8') as f:
                        json.dump(alt_data, f, indent=4)
                except Exception as e:
                    print(f"Failed to sync alternative pairs: {e}")

            # Update verified BOM JSON file on disk if changed (globally for all assemblies)
            bom_changed = new_mpn != original_bom_mpn or new_mfr != original_bom_mfr
            if getattr(self, 'bom_filepath', None) and os.path.exists(self.bom_filepath) and bom_changed:
                try:
                    with open(self.bom_filepath, 'r', encoding='utf-8') as f:
                        bom_data = json.load(f)
                    
                    updated = False
                    for assy in bom_data.get("Assemblies", []):
                        for comp in assy.get("Components", []):
                            if str(comp.get("Part", "")).strip() == str(part_num).strip():
                                comp["MPN"] = new_mpn
                                comp["MFR"] = new_mfr
                                updated = True
                    
                    if updated:
                        with open(self.bom_filepath, 'w', encoding='utf-8') as f:
                            json.dump(bom_data, f, indent=4)
                except Exception as e:
                    print(f"Failed to update BOM Data file: {e}")
            
            # Record to the centralized backlog if there was any change
            if db_changed or bom_changed:
                try:
                    from backlog_api import log_backlog_event
                    details = {
                        "customer": self.cust_name,
                        "rfq_number": self.rfq_num,
                        "part_number": part_num,
                        "old_mpn": original_bom_mpn,
                        "old_mfr": original_bom_mfr,
                        "new_mpn": new_mpn,
                        "new_mfr": new_mfr,
                        "source": "Sourcing Preview Panel (MPN/MFR Editor)"
                    }
                    log_backlog_event(
                        event_type="EDIT_MPN_MFR_PAIR",
                        app_name="Sourcing App",
                        user_name=self.user_name or "Unknown User",
                        details=details
                    )
                except Exception as e:
                    print(f"Failed to record backlog event: {e}")
            
            # Reprocess all rows with the same customer part number in this RFQ
            for it in self.gui_model:
                if str(it.get('Part', '')).strip().upper() == str(part_num).strip().upper():
                    it['MPN'] = new_mpn
                    it['MFR'] = new_mfr
                    self.engine.reprocess_item(it)
                    self._recalculate_exmat_for_item(it)
            
            self.refresh_detail_pane()
            self.populate_tree()
            self.tree.selection_set(str(self.current_selected_idx))

    def delete_selected_rows(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Selection", "Please select one or more rows to delete.", parent=self.winfo_toplevel())
            return
        
        msg_list = []
        for iid in selected:
            item_data = self.tree.item(iid, "values")
            comp_level = item_data[0] if item_data else "Unknown"
            part = item_data[2] if len(item_data) > 2 else "Unknown"
            msg_list.append(f"Comp Level {comp_level}: Part {part}")

        dialog = ConfirmationDialog(self, "Confirm Delete", f"Are you sure you want to delete {len(selected)} selected row(s)?", msg_list, action_text="Delete Rows")
        self.wait_window(dialog)
        
        if dialog.result == "PROCEED":
            indices_to_delete = sorted([int(iid) for iid in selected], reverse=True)
            for idx in indices_to_delete:
                self.gui_model.pop(idx)
            
            self.rebuild_grouped_model()
            self.current_selected_idx = None
            self.clear_detail_pane()
            self.populate_tree()

    def open_add_row_dialog(self):
        assy = self.selected_assembly.get()
        if not assy:
            messagebox.showwarning("Warning", "Please select an assembly first.", parent=self.winfo_toplevel())
            return
            
        model = ""
        rev = ""
        for item in self.gui_model:
            if str(item.get('Assy', '')).strip() == assy:
                model = item.get('Assy Model', '')
                rev = item.get('Assy Rev', '')
                break
                
        dialog = AddNewRowDialog(self, assy, model, rev)
        self.wait_window(dialog)
        
        if dialog.result:
            new_item = dialog.result
            self.engine.reprocess_item(new_item)
            self._recalculate_exmat_for_item(new_item)
            self.gui_model.append(new_item)
            self.rebuild_grouped_model()
            self.populate_tree()

    def get_valid_options_filtered(self, item, current_moq):
        assy = str(item.get('Assy', '')).strip()
        item_moqs = sorted(self.assembly_moqs.get(assy, []))
        
        if current_moq not in item_moqs:
            return item.get('moq_results', {}).get(current_moq, {}).get('options', [])

        idx = item_moqs.index(current_moq)
        if idx == 0:
            return item['moq_results'][current_moq]['options']
            
        prev_moq = item_moqs[idx-1]
        prev_winner = item['moq_results'][prev_moq].get('pending_winner') or item['moq_results'][prev_moq].get('winner')
        
        if not prev_winner:
            return item['moq_results'][current_moq]['options']
            
        limit = prev_winner['BOM Cost']
        return [opt for opt in item['moq_results'][current_moq]['options'] if opt['BOM Cost'] <= (limit + 0.0001)]

    def toggle_consign_item(self):
        selected = self.tree.selection()
        if not selected and self.current_selected_idx is None:
            messagebox.showinfo("Selection", "Please select a row to mark or remove CONSIGN ITEM status.", parent=self.window)
            return

        indices = [int(iid) for iid in selected if not iid.startswith("sep_")] if selected else [self.current_selected_idx]
        if not indices:
            return

        first_item = self.gui_model[indices[0]]
        will_consign = not first_item.get('is_consign', False)

        action_name = "Mark as CONSIGN ITEM" if will_consign else "Remove CONSIGN ITEM Status"
        confirm_msg = f"Are you sure you want to {action_name} for the selected {len(indices)} item(s)?"

        dialog = ConfirmationDialog(
            self.window, 
            f"Confirm {action_name}", 
            confirm_msg, 
            [f"Part: {self.gui_model[i].get('Part', '')} (Comp {self.gui_model[i].get('Comp Level', '')})" for i in indices[:15]], 
            action_text=action_name
        )
        self.wait_window(dialog)

        if dialog.result != "PROCEED":
            return

        affected_items = []
        # If user explicitly selected multiple rows, apply ONLY to those selected rows!
        if len(indices) > 1:
            for idx in indices:
                item = self.gui_model[idx]
                if item not in affected_items:
                    affected_items.append(item)
        else:
            # Single row selected: link to matching part number across assemblies if part is specific
            item = self.gui_model[indices[0]]
            affected_items.append(item)
            if self.calc_mode == 'total_usage':
                part_id = str(item.get('Part', '')).strip().upper()
                mpn_id = str(item.get('MPN', '')).strip().upper()
                uom_id = str(item.get('UOM', '')).strip().upper()
                if part_id and part_id not in ("PN", "RES", "CAP", "MISC", "PART"):
                    for other_item in self.gui_model:
                        if (str(other_item.get('Part', '')).strip().upper() == part_id and 
                            str(other_item.get('UOM', '')).strip().upper() == uom_id and
                            (not mpn_id or str(other_item.get('MPN', '')).strip().upper() == mpn_id)):
                            if other_item not in affected_items:
                                affected_items.append(other_item)

        for item in affected_items:
            item['is_consign'] = will_consign
            item['is_manual'] = True if will_consign else False

            consign_winner = {
                'Supplier': 'CONSIGN ITEM',
                'MPN': 'CONSIGN ITEM',
                'MFR': 'CONSIGN ITEM',
                'MOQ': 1,
                'Unit Price': 0.0,
                'Convert to USD': 0.0,
                'Markup Rate': 1.0,
                'BOM Cost': 0.0,
                'Excess Cost': 0.0,
                'Total Cost': 0.0,
                'Currency': 'USD',
                'uom': str(item.get('UOM', '')),
                'Std Pack': '-',
                'Stock': '-',
                'L/Time': 0,
                'Shipping Terms': '-',
                'Remark': ''
            }

            for q in list(item.get('moq_results', {}).keys()):
                if will_consign:
                    item['moq_results'][q]['winner'] = consign_winner
                    if 'pending_winner' in item['moq_results'][q]:
                        del item['moq_results'][q]['pending_winner']
                else:
                    orig_w = item['moq_results'][q].get('original_winner', item['moq_results'][q].get('winner'))
                    item['moq_results'][q]['winner'] = orig_w
                    if 'pending_winner' in item['moq_results'][q]:
                        del item['moq_results'][q]['pending_winner']

            if not will_consign:
                item['is_manual'] = False

            self._recalculate_exmat_for_item(item)

        self.rebuild_grouped_model()
        self.populate_tree()
        self.refresh_detail_pane()
        status_txt = "marked as CONSIGN ITEM" if will_consign else "restored from CONSIGN ITEM"
        messagebox.showinfo("Success", f"{len(affected_items)} row(s) {status_txt}.", parent=self.window)

    def clear_detail_pane(self):
        self.btn_revert.config(state="disabled")
        self.btn_apply_manual.config(state="disabled")
        self.btn_edit_pairs.config(state="disabled")
        if hasattr(self, 'btn_toggle_consign'):
            self.btn_toggle_consign.config(text="📦 Mark as CONSIGN ITEM", state="disabled")
        self.lbl_linked_parts.config(text="")
        self.btn_show_shared.config(state="disabled")
        self.current_shared_items = []
        for q in self.active_categories:
            w = self.moq_widgets[q]
            w['lbl_sel'].config(text="NO PART SELECTED", bg="#EBF8FF", fg="#a0aec0", font=('Segoe UI', 8, 'bold'))
            w['lbl_supplier'].config(text="Supplier: -", font=('Segoe UI', 8, 'bold'))
            w['lbl_mpn'].config(text="MPN: -", font=('Segoe UI', 9))
            w['lbl_moq'].config(text="MOQ: -", font=('Segoe UI', 9))
            w['lbl_date'].config(text="Date: -", font=('Segoe UI', 9))
            w['lbl_unitp'].config(text="Unit Price: -", font=('Segoe UI', 9))
            w['lbl_bom'].config(text="BOM: -", font=('Segoe UI', 9), fg="#1A365D")
            w['lbl_exc'].config(text="Excess: -", font=('Segoe UI', 9), fg="#1A365D")
            w['lbl_tot'].config(text="Total: -", font=('Segoe UI', 9))

    def refresh_detail_pane(self):
        if self.current_selected_idx is None: return
        item = self.gui_model[self.current_selected_idx]
        
        has_pending = any('pending_winner' in it['moq_results'][m] for it in self.gui_model for m in it['moq_results'])
        self.btn_apply_manual.config(state="normal" if has_pending else "disabled")
        
        if hasattr(self, 'btn_toggle_consign'):
            if item.get('is_consign'):
                self.btn_toggle_consign.config(text="📦 Remove CONSIGN ITEM", state="normal")
            else:
                self.btn_toggle_consign.config(text="📦 Mark as CONSIGN ITEM", state="normal")

        for q in self.active_categories:
            self.refresh_moq_card(q)

    def refresh_moq_card(self, moq_qty):
        if self.current_selected_idx is None: return
        item = self.gui_model[self.current_selected_idx]
        moq_data = item['moq_results'][moq_qty]
        w = self.moq_widgets[moq_qty]

        if item.get('is_consign'):
            w['lbl_sel'].config(text="CONSIGN ITEM", bg="#e2e8f0", fg="#2d3748", font=('Segoe UI', 8, 'bold'))
            w['lbl_supplier'].config(text="Supplier: CONSIGN ITEM", font=('Segoe UI', 8, 'bold'))
            w['lbl_mpn'].config(text="MPN: CONSIGN ITEM", font=('Segoe UI', 9))
            w['lbl_moq'].config(text="MOQ: 1", font=('Segoe UI', 9))
            w['lbl_date'].config(text="Date: -", font=('Segoe UI', 9))
            w['lbl_ltime'].config(text="L/Time: 0", font=('Segoe UI', 9))
            w['lbl_unitp'].config(text="Unit Price: USD $0.0000", font=('Segoe UI', 9))
            w['lbl_bom'].config(text="BOM: $0.0000", font=('Segoe UI', 9), fg="blue")
            w['lbl_exc'].config(text="Excess: $0.0000", font=('Segoe UI', 9), fg="red")
            w['lbl_tot'].config(text="Total: $0.0000", font=('Segoe UI', 8, 'bold'))
            return

        is_pending = 'pending_winner' in moq_data
        is_manual_saved = moq_qty in item.get('manual_moqs', set())
        is_synced_saved = moq_qty in item.get('synced_moqs', set())
        
        if is_pending:
            w['lbl_sel'].config(text="MANUAL OVERRIDE", bg="#fff5f5", fg="#c53030", font=('Segoe UI', 8, 'bold'))
        elif is_manual_saved:
            w['lbl_sel'].config(text="MANUAL (SAVED)", bg="#f0fff4", fg="#2f855a", font=('Segoe UI', 8, 'bold'))
        elif is_synced_saved:
            w['lbl_sel'].config(text="SYNCED (SAVED)", bg="#e3f2fd", fg="#2980b9", font=('Segoe UI', 8, 'bold'))
        else:
            w['lbl_sel'].config(text="SYSTEM BEST CHOICE", bg="#ebf8ff", fg="#2b6cb0", font=('Segoe UI', 8, 'bold'))

        display_winner = moq_data.get('pending_winner') or moq_data.get('winner')
        
        if display_winner:
            supp_text = f"Supplier: {display_winner.get('Supplier', 'Unknown')}"
            if is_pending: supp_text += " (Pending)"
            w['lbl_supplier'].config(text=supp_text, font=('Segoe UI', 8, 'bold'))
            
            w['lbl_mpn'].config(text=f"MPN: {display_winner.get('MPN', '')}", font=('Segoe UI', 9))
            w['lbl_moq'].config(text=f"MOQ: {display_winner.get('MOQ', '')}", font=('Segoe UI', 9))
            
            # Safe and clean date parsing
            sd = display_winner.get('Source Date', '')
            date_str = "-"
            if sd and pd.notna(sd):
                try:
                    if hasattr(sd, 'strftime'):
                        date_str = sd.strftime("%d %b %Y")
                    else:
                        try:
                            parsed_dt = pd.to_datetime(sd)
                        except:
                            parsed_dt = sd
                        if hasattr(parsed_dt, 'strftime'):
                            date_str = parsed_dt.strftime("%d %b %Y")
                        else:
                            date_str = str(sd).split(' ')[0]
                except:
                    date_str = str(sd).split(' ')[0]
            w['lbl_date'].config(text=f"Date: {date_str}", font=('Segoe UI', 9))
            
            lt_val = display_winner.get('L/Time (weeks)', display_winner.get('L/Time'))
            lt_str = f"{lt_val} wks" if lt_val is not None and str(lt_val).strip() != "" else "-"
            w['lbl_ltime'].config(text=f"L/Time: {lt_str}", font=('Segoe UI', 9))
            
            w['lbl_unitp'].config(text=f"Unit Price: {display_winner.get('Currency', 'USD')} ${display_winner.get('Unit Price', 0):.4f}", font=('Segoe UI', 9))
            bom_cost = display_winner.get('BOM Cost', 0.0)
            total_excess_cost = display_winner.get('Total Excess Cost', display_winner.get('Excess Cost', 0.0))
            card_total = bom_cost + total_excess_cost
            w['lbl_bom'].config(text=f"BOM: ${bom_cost:.4f}", font=('Segoe UI', 9), fg="blue")
            w['lbl_exc'].config(text=f"Excess: ${total_excess_cost:.4f}", font=('Segoe UI', 9), fg="red")
            w['lbl_tot'].config(text=f"Total: ${card_total:.4f}", font=('Segoe UI', 8, 'bold'))
        else:
            has_mpn = bool(str(item.get('MPN', '')).strip())
            has_mfr = bool(str(item.get('MFR', '')).strip())
            has_qty = pd.to_numeric(item.get('BOM Qty', 0), errors='coerce') > 0
            if not (has_mpn and has_mfr and has_qty):
                w['lbl_sel'].config(text="", bg="#FFFFFF", fg="#7f8c8d")
            else:
                w['lbl_sel'].config(text="- NO VALID OPTION -", bg="#f1f3f5", fg="#7f8c8d", font=('Segoe UI', 8, 'bold'))
            w['lbl_supplier'].config(text="Supplier: -", font=('Segoe UI', 8, 'bold'))
            w['lbl_mpn'].config(text="MPN: -", font=('Segoe UI', 9))
            w['lbl_moq'].config(text="MOQ: -", font=('Segoe UI', 9))
            w['lbl_date'].config(text="Date: -", font=('Segoe UI', 9))
            w['lbl_ltime'].config(text="L/Time: -", font=('Segoe UI', 9))
            w['lbl_unitp'].config(text="Unit Price: -", font=('Segoe UI', 9))
            w['lbl_bom'].config(text="BOM: -", font=('Segoe UI', 9), fg="#1A365D")
            w['lbl_exc'].config(text="Excess: -", font=('Segoe UI', 9), fg="#1A365D")
            w['lbl_tot'].config(text="Total: -", font=('Segoe UI', 9))

    def apply_manual_changes(self):
        pending_changes_data = []
        for idx, it in enumerate(self.gui_model):
            for q, res in it['moq_results'].items():
                if 'pending_winner' in res:
                    pending_changes_data.append({
                        'idx': idx,
                        'moq': q,
                        'assy': it['Assy'],
                        'comp_level': it.get('Comp Level', ''),
                        'part': it.get('Part', ''),
                        'uom': it.get('UOM', ''),
                        'supplier': res['pending_winner']['Supplier'],
                        'winner': res['pending_winner'],
                        'is_primary': res.get('is_pending_primary', True)
                    })

        if not pending_changes_data:
            return

        all_staged = list(pending_changes_data)
        if self.calc_mode == 'total_usage':
            for p in pending_changes_data:
                for other_idx, other_item in enumerate(self.gui_model):
                    if other_idx == p['idx']: continue
                    part_id = str(p.get('part', '')).strip()
                    if not part_id: continue
                    
                    if str(other_item.get('Part', '')).strip() == part_id and str(other_item.get('UOM', '')).strip() == str(p.get('uom', '')).strip():
                        q = p['moq']
                        if q in other_item['moq_results']:
                            for opt in other_item['moq_results'][q]['options']:
                                if (str(opt['Supplier']).strip() == str(p['winner']['Supplier']).strip() and 
                                    str(opt['MPN']).strip() == str(p['winner']['MPN']).strip() and
                                    abs(float(opt.get('Unit Price', 0)) - float(p['winner'].get('Unit Price', 0))) < 0.00001 and
                                    float(opt.get('MOQ', 0)) == float(p['winner'].get('MOQ', 0))):
                                    
                                    if not any(s['idx'] == other_idx and s['moq'] == q for s in all_staged):
                                        all_staged.append({
                                            'idx': other_idx,
                                            'moq': q,
                                            'assy': other_item['Assy'],
                                            'comp_level': other_item.get('Comp Level', ''),
                                            'supplier': opt['Supplier'],
                                            'winner': opt,
                                            'is_primary': False
                                        })
                                    break

        dialog = ManualOverrideConfirmationDialog(self, "Confirm Sourcing Overrides", all_staged)
        self.window.wait_window(dialog)
        
        if not dialog.result:
            return

        affected_indices = set()
        for change in dialog.result:
            idx = change['idx']
            q = change['moq']
            target_item = self.gui_model[idx]
            
            target_item['moq_results'][q]['winner'] = change['winner']
            
            if change['is_primary']:
                if 'manual_moqs' not in target_item: target_item['manual_moqs'] = set()
                target_item['manual_moqs'].add(q)
                if 'synced_moqs' in target_item and q in target_item['synced_moqs']:
                    target_item['synced_moqs'].remove(q)
            else:
                if 'synced_moqs' not in target_item: target_item['synced_moqs'] = set()
                target_item['synced_moqs'].add(q)
                if 'manual_moqs' in target_item and q in target_item['manual_moqs']:
                    target_item['manual_moqs'].remove(q)
            
            if 'pending_winner' in target_item['moq_results'][q]:
                del target_item['moq_results'][q]['pending_winner']
            if 'is_pending_primary' in target_item['moq_results'][q]:
                del target_item['moq_results'][q]['is_pending_primary']
                
            affected_indices.add(idx)
            self._cascade_validation(target_item, q)
            self._recalculate_exmat_for_item(target_item)

        self.rebuild_grouped_model()
        self.populate_tree()
        self.refresh_detail_pane()
        
        current_assy = self.selected_assembly.get()
        to_select = []
        for idx in affected_indices:
            if str(self.gui_model[idx].get('Assy', '')).strip() == current_assy:
                to_select.append(str(idx))
        if to_select:
            self.tree.selection_set(to_select)
            self.tree.see(to_select[0])
            
        messagebox.showinfo("Success", "Sourcing changes applied successfully.", parent=self.window)

    def _is_same_option(self, opt1, opt2):
        if opt1 is None or opt2 is None: return opt1 is opt2
        try:
            core_keys = [
                'Supplier', 'MPN', 'MOQ', 'Unit Price', 'Source Date', 
                'uom', 'Std Pack', 'Stock', 'L/Time', 'Remark', 
                'Total Cost', 'BOM Cost', 'Excess Cost', 'Currency'
            ]
            for key in core_keys:
                v1 = opt1.get(key)
                v2 = opt2.get(key)
                if isinstance(v1, (float, int)) or isinstance(v2, (float, int)):
                    if abs(float(v1 or 0) - float(v2 or 0)) > 0.00001: return False
                else:
                    if str(v1).strip() != str(v2).strip(): return False
            return True
        except: return False

    def revert_part(self):
        if self.current_selected_idx is None: return
        item = self.gui_model[self.current_selected_idx]
        
        revert_summary = []
        affected_items = [item]
        revert_summary.append(f"Current (Comp {item.get('Comp Level', '')})")

        if self.calc_mode == 'total_usage':
            part = item['Part']
            uom = item['UOM']
            for other_item in self.gui_model:
                if other_item['Part'] == part and other_item['UOM'] == uom and other_item['ref_id'] != item['ref_id']:
                    affected_items.append(other_item)
                    revert_summary.append(f"Linked (Comp {other_item.get('Comp Level', '')})")

        dialog = ConfirmationDialog(self, "Confirm Revert", "Revert manual changes to original system-selected winners for:", revert_summary, action_text="Revert Now")
        self.wait_window(dialog)
        
        if dialog.result != "PROCEED":
            return

        for aff_item in affected_items:
            for q in aff_item['moq_results']:
                orig_w = aff_item['moq_results'][q].get('original_winner', aff_item['moq_results'][q].get('winner'))
                aff_item['moq_results'][q]['winner'] = orig_w
                if 'pending_winner' in aff_item['moq_results'][q]:
                    del aff_item['moq_results'][q]['pending_winner']
                if 'is_pending_primary' in aff_item['moq_results'][q]:
                    del aff_item['moq_results'][q]['is_pending_primary']
            
            aff_item['is_manual'] = False
            aff_item['is_synced'] = False
            if 'manual_moqs' in aff_item: aff_item['manual_moqs'] = set()
            if 'synced_moqs' in aff_item: aff_item['synced_moqs'] = set()
            self._recalculate_exmat_for_item(aff_item)
            
        self.rebuild_grouped_model()
        self.populate_tree()
        self.refresh_detail_pane()
        messagebox.showinfo("Revert Complete", f"Reverted {len(affected_items)} row(s) to original system choices.", parent=self.window)
        self.tree.selection_set(str(self.current_selected_idx))

    def _cascade_validation(self, item, current_moq):
        assy = str(item.get('Assy', '')).strip()
        item_moqs = sorted(self.assembly_moqs.get(assy, []))
        
        if current_moq not in item_moqs:
            return
            
        start_idx = item_moqs.index(current_moq)
        current_winner = item['moq_results'][current_moq].get('winner')
        current_bom_cost = current_winner['BOM Cost'] if current_winner else None
        
        for i in range(start_idx + 1, len(item_moqs)):
            next_moq = item_moqs[i]
            next_data = item['moq_results'][next_moq]
            
            valid_for_next = self.get_valid_options_filtered(item, next_moq)
            
            if valid_for_next:
                new_best = valid_for_next[0]
                next_data['winner'] = new_best
                current_bom_cost = new_best['BOM Cost']
            else:
                next_data['winner'] = None
                current_bom_cost = None

    def open_sourcing_detail_dialog(self, moq_qty):
        if self.current_selected_idx is None: return
        item = self.gui_model[self.current_selected_idx]
        
        options = self.get_valid_options_filtered(item, moq_qty)
        if not options:
            messagebox.showinfo("No Data", f"No sourcing options available for MOQ {moq_qty}.", parent=self.master)
            return
            
        system_winner_idx = 0
        moq_data = item['moq_results'][moq_qty]
        current_active = moq_data.get('pending_winner') or moq_data.get('winner')
        
        dialog = SourcingOptionDetailDialog(self.window, item, moq_qty, options, system_winner_idx, current_active)
        self.window.wait_window(dialog)
        try: self.window.update()
        except: pass
        
        if dialog.result is not None:
            new_winner, is_different_row = dialog.result
            
            original_winner = moq_data.get('original_winner')
            if not original_winner: 
                original_winner = moq_data.get('winner')
            
            is_same_as_system = self._is_same_option(new_winner, original_winner)
            
            if not is_different_row:
                messagebox.showinfo("Selection Unchanged", "No changes made. Winning supplier selection remains unchanged.", parent=self.master)
            else:
                affected_items = [item]
                if self.calc_mode == 'total_usage':
                    part_id = str(item.get('Part', '')).strip()
                    uom_id = str(item.get('UOM', '')).strip()
                    for other_item in self.gui_model:
                        if other_item.get('ref_id') != item.get('ref_id'):
                            if str(other_item.get('Part', '')).strip() == part_id and str(other_item.get('UOM', '')).strip() == uom_id:
                                affected_items.append(other_item)

                if is_same_as_system:
                    for aff in affected_items:
                        if moq_qty in aff['moq_results']:
                            if 'pending_winner' in aff['moq_results'][moq_qty]:
                                del aff['moq_results'][moq_qty]['pending_winner']
                            if 'is_pending_primary' in aff['moq_results'][moq_qty]:
                                del aff['moq_results'][moq_qty]['is_pending_primary']
                    messagebox.showinfo("Selection Reverted", "Winning supplier reverted to system choice. \n\nPlease confirm by clicking 'Apply Changes'.", parent=self.master)
                else:
                    item['moq_results'][moq_qty]['pending_winner'] = new_winner
                    item['moq_results'][moq_qty]['is_pending_primary'] = True
                    
                    for aff in affected_items:
                        if aff.get('ref_id') == item.get('ref_id'):
                            continue
                        
                        # Find the matching option in this other item
                        found_opt = None
                        if moq_qty in aff['moq_results']:
                            for opt in aff['moq_results'][moq_qty].get('options', []):
                                if (str(opt.get('Supplier')).strip() == str(new_winner.get('Supplier')).strip() and 
                                    str(opt.get('MPN')).strip() == str(new_winner.get('MPN')).strip() and
                                    abs(float(opt.get('Unit Price', 0)) - float(new_winner.get('Unit Price', 0))) < 0.00001 and
                                    float(opt.get('MOQ', 0)) == float(new_winner.get('MOQ', 0))):
                                    found_opt = opt
                                    break
                        if found_opt:
                            aff['moq_results'][moq_qty]['pending_winner'] = found_opt
                            aff['moq_results'][moq_qty]['is_pending_primary'] = False
                            
                    messagebox.showinfo("Selection Changed", "Winning supplier selection changed. \n\nPlease confirm by clicking the 'Apply Changes' button.", parent=self.master)
            
                for aff in affected_items:
                    self._cascade_pending_validation(aff)
                    self._recalculate_exmat_for_item(aff)
            
            self.populate_tree()
            self.refresh_detail_pane()
            self.tree.selection_set(str(self.current_selected_idx))

        else:
            self.filter_criteria = None
            if hasattr(self, 'populate_tree'):
                self.populate_tree()
            elif hasattr(self, 'refresh_projects'):
                self.refresh_projects()
    def approve(self):
        dialog = SourcingStatusWindow(self.window, self.unique_assemblies, self.assembly_status)
        self.window.wait_window(dialog)
        if dialog.result != "PROCEED":
            return

        self.btn_approve.config(state="disabled")
        from dialogs import ProgressWindow
        prog_win = ProgressWindow(self.window, title="Applying BOM Calculation", message="Applying BOM Calculation... Please wait...")
        prog_win.update_progress(50, 100, "Saving sourcing calculations...")
        prog_win.update()
        try:
            success, msg = self.on_approve(self.gui_model)
            try: prog_win.destroy()
            except: pass
            if success:
                self.preview_status.config(text=f"☑ {msg}", bg="green", fg="white")
                self.window.update_idletasks()
                
                # Ask to export Excel
                from utils import messagebox
                if messagebox.askyesno("Export Excel", "Sourcing calculations saved successfully!\nDo you want to proceed and export the Excel file now?", parent=self.window):
                    self.export_excel()
                
                if hasattr(self, 'on_back_callback') and callable(self.on_back_callback):
                    self.on_back_callback()
                self._wait_var.set(1)
            else:
                self.preview_status.config(text=f"⚠ {msg}", bg="red", fg="white")
        except Exception as e:
            try: prog_win.destroy()
            except: pass
            self.preview_status.config(text=f"⚠ Unexpected Error: {str(e)}", bg="red", fg="white")
        finally:
            try:
                if self.btn_approve.winfo_exists():
                    self.btn_approve.config(state="normal")
            except:
                pass

    def save_progress(self):
        self.btn_save_progress.config(state="disabled")
        from dialogs import ProgressWindow
        prog_win = ProgressWindow(self.window, title="Saving Sourcing Draft", message="Saving draft calculations. Please wait...")
        prog_win.update_progress(50, 100, "Saving draft data...")
        prog_win.update()
        try:
            success, msg = self.on_approve(self.gui_model, is_final=False)
            try: prog_win.destroy()
            except: pass
            if success:
                self.preview_status.config(text=f"☑ {msg}", bg="green", fg="white")
                self.window.update_idletasks()
                
                if hasattr(self, 'on_back_callback') and callable(self.on_back_callback):
                    self.on_back_callback()
                self._wait_var.set(1)
            else:
                self.preview_status.config(text=f"⚠ {msg}", bg="red", fg="white")
        except Exception as e:
            try: prog_win.destroy()
            except: pass
            self.preview_status.config(text=f"⚠ Unexpected Error: {str(e)}", bg="red", fg="white")
        finally:
            try:
                if self.btn_save_progress.winfo_exists():
                    self.btn_save_progress.config(state="normal")
            except:
                pass

    def export_excel(self):
        self.btn_export_excel.config(state="disabled")
        self.window.update_idletasks()
        try:
            if not self.on_export:
                messagebox.showinfo("Export Excel", "Export function is not available.", parent=self.window)
                return
            success, msg = self.on_export(self.gui_model)
            if success:
                self.preview_status.config(text=f"☑ {msg}", bg="green", fg="white")
                self.window.update_idletasks()
                
                self.window.after(5000, lambda: self.preview_status.config(text="", bg=self.window.cget("bg")))
                
                if hasattr(self.window.master, 'status_bar'):
                    self.window.master.status_bar.config(text=f"☑ {msg}", bg="green", fg="white")
                    self.window.master.update_idletasks()
                    self.window.master.after(5000, lambda: self.window.master.status_bar.config(text="", bg=self.window.master.cget("bg")))
            else:
                self.preview_status.config(text=f"⚠ {msg}", bg="red", fg="white")
        except Exception as e:
            self.preview_status.config(text=f"⚠ Unexpected Error: {str(e)}", bg="red", fg="white")
        finally:
            self.btn_export_excel.config(state="normal")

    def export_missing_sourcing_report(self):
        missing_items = []
        for item in self.gui_model:
            all_missing = True
            for q in self.assembly_moqs.get(item['Assy'].strip(), []):
                if item['moq_results'].get(q, {}).get('winner') is not None:
                    all_missing = False
                    break
            if all_missing:
                missing_items.append(item)

        if not missing_items:
            messagebox.showinfo("Report", "No items with missing sourcing data found.", parent=self.window)
            return

        from datetime import datetime
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%I.%M%p")
        default_name = f"Missing Sourcing Data in {self.cust_name}_ {self.rfq_num}_{date_str}_{time_str}"
        
        save_path = filedialog.asksaveasfilename(
            parent=self.window,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_name,
            title="Save Missing Sourcing Report"
        )
        if not save_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            usage_data = []
            part_usage_map = {}
            for item in missing_items:
                part = item['Part']
                if part not in part_usage_map:
                    part_usage_map[part] = {
                        'Description': item['Description'],
                        'MPNs': set(),
                        'MFRs': set(),
                        'Assy Models': set(),
                        'UOM': item['UOM'],
                        'Total Usage': 0.0
                    }
                
                if item.get('MPN'):
                    for m in str(item['MPN']).split(","):
                        if m.strip(): part_usage_map[part]['MPNs'].add(m.strip())
                if item.get('MFR'):
                    for m in str(item['MFR']).split(","):
                        if m.strip(): part_usage_map[part]['MFRs'].add(m.strip())
                if item.get('Assy Model'):
                    part_usage_map[part]['Assy Models'].add(str(item['Assy Model']).strip())
                
                part_usage_map[part]['Total Usage'] += item['BOM Qty']
                
                assy_num = item['Assy'].strip()
                assy_moqs = self.assembly_moqs.get(assy_num, [])
                if 'Quote_Qtys' not in part_usage_map[part]:
                    part_usage_map[part]['Quote_Qtys'] = {}
                
                for q in assy_moqs:
                    q_int = int(q)
                    part_usage_map[part]['Quote_Qtys'][q_int] = part_usage_map[part]['Quote_Qtys'].get(q_int, 0.0) + (item['BOM Qty'] * q_int)

            all_moqs = sorted(list({int(q) for moqs in self.assembly_moqs.values() for q in moqs}))
            
            for part, info in part_usage_map.items():
                row = {
                    'Part Number': part,
                    'Description': info['Description'],
                    'MPN': ", ".join(sorted(list(info['MPNs']))),
                    'MFR': ", ".join(sorted(list(info['MFRs']))),
                    'Assembly Model(s)': ", ".join(sorted(list(info['Assy Models']))),
                    'UOM': info['UOM'],
                    'Total Usage': info['Total Usage']
                }
                for q in all_moqs:
                    val = info.get('Quote_Qtys', {}).get(q)
                    row[f'Quote Qty @ {q}'] = val if val is not None else ""
                usage_data.append(row)

            df_usage = pd.DataFrame(usage_data)

            sourcing_template_data = []
            headers_sourcing = [
                "Customer Part", "Source Date", "Description", "Supplier", "MFR", "MPN", "Currency", "UOM", 
                "Std Pack", "Stock", "L/Time (weeks)", "Shipping Terms", "Remark (Attachment)", 
                "Unit Price", "Supplier Quote", "Unit Price Validity Duration (Days)", "Stock Duration (Days)"
            ]

            for item in missing_items:
                mpns = [m.strip() for m in str(item['MPN']).split(",") if m.strip()]
                mfrs = [m.strip() for m in str(item['MFR']).split(",") if m.strip()]
                while len(mfrs) < len(mpns): mfrs.append("")
                
                for mpn, mfr in zip(mpns, mfrs):
                    row = {h: "" for h in headers_sourcing}
                    row['Customer Part'] = item['Part']
                    row['Description'] = item['Description']
                    row['MPN'] = mpn
                    row['MFR'] = mfr
                    sourcing_template_data.append(row)

            df_sourcing = pd.DataFrame(sourcing_template_data, columns=headers_sourcing)

            wb = Workbook()
            
            ws1 = wb.active
            ws1.title = "Usage Summary"
            ws1.sheet_properties.tabColor = "9BBB59"
            for col_idx, column_title in enumerate(df_usage.columns, 1):
                cell = ws1.cell(row=1, column=col_idx, value=column_title)
                cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

            for r_idx, row in enumerate(df_usage.values, 2):
                for c_idx, value in enumerate(row, 1):
                    ws1.cell(row=r_idx, column=c_idx, value=value)

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            last_row_ws1 = len(df_usage) + 1
            for row in ws1.iter_rows(min_row=1, max_row=last_row_ws1, min_col=1, max_col=len(df_usage.columns)):
                for cell in row:
                    cell.border = thin_border
            ws1.sheet_view.showGridLines = False

            ws2 = wb.create_sheet("Update Sourcing Data")
            ws2.sheet_properties.tabColor = "ADD8E6"
            for col_idx, column_title in enumerate(df_sourcing.columns, 1):
                cell = ws2.cell(row=1, column=col_idx, value=column_title)
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for r_idx, row in enumerate(df_sourcing.values, 2):
                for c_idx, value in enumerate(row, 1):
                    ws2.cell(row=r_idx, column=c_idx, value=value)

            last_row_ws2 = len(df_sourcing) + 1
            for row in ws2.iter_rows(min_row=1, max_row=last_row_ws2, min_col=1, max_col=len(df_sourcing.columns)):
                for cell in row:
                    cell.border = thin_border
            ws2.sheet_view.showGridLines = False

            # --- Sheet 3: BOM Data (Missing Items Only) ---
            ws3 = wb.create_sheet("BOM Data")
            ws3.sheet_properties.tabColor = "B1A0C7"
            
            # Merge and write Qty to Quote in Row 1
            if len(all_moqs) > 1:
                for c_idx in range(7, 7 + len(all_moqs)):
                    cell = ws3.cell(row=1, column=c_idx)
                    cell.border = thin_border
                ws3.merge_cells(start_row=1, start_column=7, end_row=1, end_column=7 + len(all_moqs) - 1)
                for c_idx in range(7, 7 + len(all_moqs)):
                    cell = ws3.cell(row=1, column=c_idx)
                    cell.border = thin_border
            cell_q2q = ws3.cell(row=1, column=7, value="Qty to Quote")
            cell_q2q.font = Font(bold=True, name="Segoe UI", size=10)
            cell_q2q.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write Headers in Row 2
            bom_headers = ["Comp Level", "Assy #", "Assy Model", "Part", "Qty", "UOM"]
            for idx, h in enumerate(bom_headers, 1):
                ws3.cell(row=2, column=idx, value=h)
                
            for idx, moq_val in enumerate(all_moqs, 7):
                ws3.cell(row=2, column=idx, value=moq_val)
                
            # Style Row 2 Headers with #CCC0DA background
            header_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")
            header_font = Font(bold=True, name="Segoe UI", size=10)
            header_align = Alignment(horizontal="center", vertical="center")
            
            for col_idx in range(1, 7 + len(all_moqs)):
                cell = ws3.cell(row=2, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                if col_idx in (3, 4):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = header_align
                    
            # Write Data Rows (Missing Items Only)
            current_row = 3
            for item in missing_items:
                comp_level = str(item.get('Comp Level', ''))
                assy_num = str(item.get('Assy', '')).strip()
                assy_model = str(item.get('Assy Model', ''))
                part = str(item.get('Part', ''))
                qty = item.get('BOM Qty', 0.0)
                uom = str(item.get('UOM', ''))
                
                try:
                    if qty.is_integer():
                        qty = int(qty)
                except:
                    pass
                
                ws3.cell(row=current_row, column=1, value=comp_level)
                ws3.cell(row=current_row, column=2, value=assy_num)
                ws3.cell(row=current_row, column=3, value=assy_model)
                ws3.cell(row=current_row, column=4, value=part)
                ws3.cell(row=current_row, column=5, value=qty)
                ws3.cell(row=current_row, column=6, value=uom)
                
                # Write dynamic MOQ quantities to quote
                for idx, moq_val in enumerate(all_moqs, 7):
                    q_val = qty * moq_val
                    ws3.cell(row=current_row, column=idx, value=q_val)
                    
                # Format Data Cell Styles
                for col_idx in range(1, 7 + len(all_moqs)):
                    cell = ws3.cell(row=current_row, column=col_idx)
                    cell.font = Font(name="Segoe UI", size=10)
                    if col_idx in (1, 2, 5, 6) or col_idx >= 7:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                current_row += 1
                
            # Apply Thin Borders recursively matching your exact table design
            last_row_ws3 = current_row - 1
            
            # Columns A-F border starts from Row 2
            for r in range(2, last_row_ws3 + 1):
                for c in range(1, 7):
                    ws3.cell(row=r, column=c).border = thin_border
                    
            # Columns G+ border starts from Row 1
            for r in range(1, last_row_ws3 + 1):
                for c in range(7, 7 + len(all_moqs)):
                    ws3.cell(row=r, column=c).border = thin_border
                    
            ws3.sheet_view.showGridLines = False

            # Add premium guidelines sheet as the first sheet
            try:
                from sourcing_master_ui import add_guideline_sheet
                add_guideline_sheet(wb)
                wb.active = 0
            except Exception as e:
                print(f"Failed to add guideline sheet: {e}")

            for ws in [ws1, ws2, ws3]:
                for column in ws.columns:
                    max_length = 0
                    from openpyxl.utils import get_column_letter
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column_letter].width = max_length + 2

            wb.save(save_path)
            success_msg = (
                f"Missing Sourcing Report saved successfully to:\n{save_path}\n\n"
                "Please review and proceed to update sourcing data to ensure Sourcing Completeness."
            )
            messagebox.showinfo("Success", success_msg, parent=self.window)
            os.startfile(save_path)

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to generate report: {str(e)}", parent=self.window)

    def _get_safe_part_filename(self, part_name):
        safe_part = "".join([c for c in str(part_name) if c.isalnum() or c in (' ', '-', '_')]).strip().replace(" ", "_")
        if not safe_part:
            safe_part = "Unknown_Part"
        return safe_part

    def _load_part_usage_summary(self, part_name):
        import json
        safe_part = self._get_safe_part_filename(part_name)
        project_name = f"{self.cust_name}_{self.rfq_num}".replace(" ", "_")
        json_path = os.path.join(USAGE_SUMMARY_DIR, project_name, f"{safe_part}.json")
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading usage summary JSON for {part_name}: {e}")
        return None

    def _calculate_option_exmat(self, item, q, opt):
        if not opt:
            return
        
        bom_qty = float(item.get('BOM Qty', 0))
        q_val = float(q)
        qty_to_quote = bom_qty * q_val
        
        if self.calc_mode == 'total_usage':
            # Load Usage Summary JSON for project-wide aggregation
            usage_summary = self._load_part_usage_summary(item.get('Part', ''))
            total_quantity_to_quote = None
            if usage_summary:
                tqq_dict = usage_summary.get("Total_Quote_Quantities", {})
                for key in [str(q), str(int(q_val)), f"{q_val:.1f}", f"{q_val:.0f}"]:
                    if key in tqq_dict:
                        total_quantity_to_quote = float(tqq_dict[key])
                        break
                
                # If still not found, try matching by float value
                if total_quantity_to_quote is None:
                    for k, v in tqq_dict.items():
                        try:
                            if abs(float(k) - q_val) < 0.0001:
                                total_quantity_to_quote = float(v)
                                break
                        except ValueError:
                            continue
            
            if total_quantity_to_quote is None:
                # Fallback to qty_to_quote if JSON or key is missing
                total_quantity_to_quote = qty_to_quote
        else:
            # Isolated assembly sourcing: Sum the quantities of this part within this specific assembly only
            assy = str(item.get('Assy', '')).strip()
            part = str(item.get('Part', '')).strip()
            total_quantity_to_quote = 0.0
            for it in self.gui_model:
                if str(it.get('Assy', '')).strip() == assy and str(it.get('Part', '')).strip() == part:
                    total_quantity_to_quote += float(it.get('BOM Qty', 0)) * q_val
            
            if total_quantity_to_quote == 0.0:
                total_quantity_to_quote = qty_to_quote
            
        if total_quantity_to_quote > 0:
            proportion = qty_to_quote / total_quantity_to_quote
        else:
            proportion = 1.0
            
        moq = float(opt.get('MOQ', 0))
        unit_usd = float(opt.get('Convert to USD', 0))
        
        # Total excess cost (non-allocated, showing on canvas card)
        total_excess_cost = max(0.0, moq - total_quantity_to_quote) * unit_usd
        
        # Proportional allocated excess cost (showing in Treeview and Excel)
        allocated_excess_cost = proportion * total_excess_cost
        
        opt['Total Excess Cost'] = total_excess_cost
        opt['Excess Cost'] = allocated_excess_cost
        opt['Total Cost'] = opt.get('BOM Cost', 0.0) + allocated_excess_cost

    def _cascade_pending_validation(self, item):
        assy = str(item.get('Assy', '')).strip()
        item_moqs = sorted(self.assembly_moqs.get(assy, []))
        if not item_moqs:
            return
            
        for i in range(1, len(item_moqs)):
            current_moq = item_moqs[i]
            prev_moq = item_moqs[i-1]
            
            prev_winner = item['moq_results'][prev_moq].get('pending_winner') or item['moq_results'][prev_moq].get('winner')
            current_data = item['moq_results'][current_moq]
            current_active = current_data.get('pending_winner') or current_data.get('winner')
            
            if not prev_winner:
                continue
                
            is_valid = False
            if current_active:
                if current_active.get('BOM Cost', 0.0) <= (prev_winner.get('BOM Cost', 0.0) + 0.0001):
                    is_valid = True
                    
            if not is_valid:
                # Violates golden rule! We must automatically find the best valid quote for current_moq
                valid_options = self.get_valid_options_filtered(item, current_moq)
                if valid_options:
                    current_data['pending_winner'] = valid_options[0]
                    current_data['is_pending_primary'] = False
                else:
                    current_data['pending_winner'] = None
                    if 'is_pending_primary' in current_data:
                        del current_data['is_pending_primary']

    def _recalculate_exmat_for_item(self, item):
        if not item or 'moq_results' not in item:
            return
        
        assy = str(item.get('Assy', '')).strip()
        moqs = self.assembly_moqs.get(assy, [])
        for q in moqs:
            res = item['moq_results'].get(q)
            if not res:
                continue
            
            # Recalculate for all options in this MOQ
            for opt in res.get('options', []):
                self._calculate_option_exmat(item, q, opt)
                
            # Recalculate for current winner
            if res.get('winner'):
                self._calculate_option_exmat(item, q, res['winner'])
                
            # Recalculate for pending_winner (if any)
            if res.get('pending_winner'):
                self._calculate_option_exmat(item, q, res['pending_winner'])
                
            # Recalculate for original_winner (if any)
            if res.get('original_winner'):
                self._calculate_option_exmat(item, q, res['original_winner'])

    def populate_tree_preserving_selection(self):
        sel = self.tree.selection()
        self.populate_tree()
        existing_iids = [iid for iid in sel if self.tree.exists(iid)]
        if existing_iids:
            self.tree.selection_set(existing_iids)

    def on_tree_double_click(self, event):
        if getattr(self, 'read_only', False):
            return "break"
        iid = self.tree.identify_row(event.y)
        if iid and not iid.startswith("sep_"):
            idx = int(iid)
            item = self.gui_model[idx]
            self.active_edit_ref_id = item.get('ref_id')
            self.populate_tree_preserving_selection()
            self.load_active_edit_row()

    def on_tree_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            if column == "#1":  # First column (Comp Level)
                iid = self.tree.identify_row(event.y)
                if iid and not iid.startswith("sep_"):
                    idx = int(iid)
                    item = self.gui_model[idx]
                    self.active_edit_ref_id = item.get('ref_id')
                    self.populate_tree_preserving_selection()
                    if not getattr(self, 'read_only', False):
                        self.load_active_edit_row()
                    return "break"

    def load_active_edit_row(self):
        if self.read_only or not hasattr(self, 'detail_frame') or self.detail_frame is None:
            return
        if not hasattr(self, 'active_edit_ref_id') or self.active_edit_ref_id is None:
            return
            
        item = None
        for x in self.gui_model:
            if x.get('ref_id') == self.active_edit_ref_id:
                item = x
                break
        if not item: return
        
        idx = self.gui_model.index(item)
        self.detail_frame.config(text=f"Editing Sourcing for Part: {item['Part']} (Comp Level {item.get('Comp Level', '')}) | MPNs: {item['MPN']}")
        self.btn_revert.config(state="normal")
        self.btn_edit_pairs.config(state="normal")
        
        self.current_shared_items = []
        if self.calc_mode == 'total_usage':
            part = item['Part']
            uom = item['UOM']
            for other_item in self.gui_model:
                if other_item['Part'] == part and other_item['UOM'] == uom and other_item.get('ref_id') != item.get('ref_id'):
                    self.current_shared_items.append(other_item)

        if self.current_shared_items:
            self.lbl_linked_parts.config(text="Shared commonly with other assemblies.")
            self.btn_show_shared.config(state="normal")
        else:
            self.lbl_linked_parts.config(text="Not shared with other assemblies.")
            self.btn_show_shared.config(state="disabled")

        self.current_selected_idx = idx
        self.refresh_detail_pane()

    def show_context_menu(self, event):
        if getattr(self, 'read_only', False):
            return "break"
        iid = self.tree.identify_row(event.y)
        if iid:
            if iid not in self.tree.selection():
                self.tree.selection_set(iid)
            self.context_menu.post(event.x_root, event.y_root)

    def copy_selected_range(self, event=None):
        selected_iids = self.tree.selection()
        if not selected_iids:
            return
        
        lines = []
        # Only copy base BOM columns
        copy_cols = self.base_cols
            
        headers = []
        for col in copy_cols:
            headers.append(self.tree.heading(col, "text"))
        lines.append("\t".join(headers))
        
        for iid in selected_iids:
            row_vals = self.tree.item(iid, "values")
            filtered_vals = []
            for col in copy_cols:
                try:
                    idx = self.all_cols.index(col)
                    val = str(row_vals[idx])
                except:
                    val = ""
                # Strip checkbox symbol if copying Comp Level
                if col == "Comp Level" and (val.startswith("☐") or val.startswith("☑")):
                    val = val[2:]
                filtered_vals.append(val)
            lines.append("\t".join(filtered_vals))
            
        clipboard_text = "\n".join(lines)
        self.window.clipboard_clear()
        self.window.clipboard_append(clipboard_text)

    def copy_all_assembly_bom_info(self):
        children = self.tree.get_children()
        if not children:
            return
            
        lines = []
        copy_cols = self.base_cols
            
        headers = []
        for col in copy_cols:
            headers.append(self.tree.heading(col, "text"))
        lines.append("\t".join(headers))
        
        for iid in children:
            if iid.startswith("sep_"):
                continue
            row_vals = self.tree.item(iid, "values")
            filtered_vals = []
            for col in copy_cols:
                try:
                    idx = self.all_cols.index(col)
                    val = str(row_vals[idx])
                except:
                    val = ""
                # Strip checkbox symbol if copying Comp Level
                if col == "Comp Level" and (val.startswith("☐") or val.startswith("☑")):
                    val = val[2:]
                filtered_vals.append(val)
            lines.append("\t".join(filtered_vals))
            
        clipboard_text = "\n".join(lines)
        self.window.clipboard_clear()
        self.window.clipboard_append(clipboard_text)

    def copy_selected_parts(self):
        selected_iids = self.tree.selection()
        if not selected_iids: return
        parts = [self.tree.heading("Part", "text")]
        for iid in selected_iids:
            try:
                row_vals = self.tree.item(iid, "values")
                idx = self.all_cols.index("Part")
                parts.append(str(row_vals[idx]).strip())
            except:
                pass
        self.window.clipboard_clear()
        self.window.clipboard_append("\n".join(parts))
        
    def copy_selected_mpns(self):
        selected_iids = self.tree.selection()
        if not selected_iids: return
        mpns = [self.tree.heading("MPN", "text")]
        for iid in selected_iids:
            try:
                row_vals = self.tree.item(iid, "values")
                idx = self.all_cols.index("MPN")
                mpns.append(str(row_vals[idx]).strip())
            except:
                pass
        self.window.clipboard_clear()
        self.window.clipboard_append("\n".join(mpns))

    def show_assy_suggestions_popup(self, filtered_list):
        if not hasattr(self, 'assy_popup') or self.assy_popup is None or not self.assy_popup.winfo_exists():
            self.assy_popup = tk.Toplevel(self.window)
            self.assy_popup._is_autocomplete_popup = True
            self.assy_popup.overrideredirect(True)
            self.assy_popup.configure(bg="#1A365D")
            
            container = tk.Frame(self.assy_popup, bg="#ffffff", bd=1, relief="flat")
            container.pack(fill="both", expand=True, padx=1, pady=1)
            
            self.assy_listbox = tk.Listbox(container, height=6, font=("Arial", 9), selectbackground="#0077B6", selectforeground="white", bd=0, highlightthickness=0)
            self.assy_listbox.pack(side="left", fill="both", expand=True)
            
            sb = ttk.Scrollbar(container, orient="vertical", command=self.assy_listbox.yview)
            sb.pack(side="right", fill="y")
            self.assy_listbox.config(yscrollcommand=sb.set)
            
            self.assy_listbox.bind("<Double-1>", self.on_suggestion_select)
            self.assy_listbox.bind("<Return>", self.on_suggestion_select)
            
            self.assy_combo.bind("<FocusOut>", self.on_assy_focus_out)
            self.assy_listbox.bind("<FocusOut>", self.on_assy_focus_out)
            
        self.assy_listbox.delete(0, "end")
        for item in filtered_list:
            self.assy_listbox.insert("end", item)
            
        self.window.update_idletasks()
        x = self.assy_combo.winfo_rootx()
        y = self.assy_combo.winfo_rooty() + self.assy_combo.winfo_height()
        w = max(self.assy_combo.winfo_width(), 200)
        
        num_items = min(len(filtered_list), 6)
        h = num_items * 20 + 4
        
        def set_geom():
            if hasattr(self, 'assy_popup') and self.assy_popup is not None and self.assy_popup.winfo_exists():
                self.assy_popup.geometry(f"{w}x{h}+{x}+{y}")
                
        set_geom()
        self.window.after(80, set_geom)
        self.assy_popup.deiconify()
        self.assy_popup.lift()

    def on_assy_focus_out(self, event=None):
        self.window.after(100, self.check_focus_and_dismiss)
        
    def check_focus_and_dismiss(self):
        if not hasattr(self, 'assy_popup') or self.assy_popup is None or not self.assy_popup.winfo_exists():
            return
        focus_widget = self.window.focus_get()
        if focus_widget != self.assy_combo and focus_widget != self.assy_listbox:
            try:
                parent = focus_widget.nametowidget(focus_widget.winfo_parent())
                if parent == self.assy_popup or parent.winfo_parent() == str(self.assy_popup):
                    return
            except:
                pass
            self.dismiss_suggestions_popup()
            
    def dismiss_suggestions_popup(self):
        if hasattr(self, 'assy_popup') and self.assy_popup is not None and self.assy_popup.winfo_exists():
            self.assy_popup.destroy()
            self.assy_popup = None

class SourcingOptionDetailDialog(tk.Toplevel):
    def __init__(self, parent, item_data, moq_qty, options, system_winner_idx, current_ui_winner=None):
        self._skip_autofit = True
        super().__init__(parent)
        self.title("Candidate Suppliers Comparison")
        self.geometry("1250x700")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        self.pending_selection = None
        self.item_data = item_data
        self.moq_qty = moq_qty
        self.options = options
        self.system_winner_idx = system_winner_idx
        
        self.initial_winner_idx = -1
        self.current_winner_idx = -1
        if current_ui_winner:
            for i, opt in enumerate(options):
                if (str(opt['Supplier']).strip() == str(current_ui_winner['Supplier']).strip() and 
                    str(opt['MPN']).strip() == str(current_ui_winner['MPN']).strip() and 
                    opt.get('MOQ') == current_ui_winner.get('MOQ') and
                    str(opt.get('Source Date')) == str(current_ui_winner.get('Source Date')) and
                    abs(float(opt.get('Unit Price', 0)) - float(current_ui_winner.get('Unit Price', 0))) < 0.00001):
                    self.initial_winner_idx = i
                    self.current_winner_idx = i
                    self.pending_selection = opt
                    break
        
        if self.current_winner_idx == -1:
            self.initial_winner_idx = system_winner_idx
            self.current_winner_idx = system_winner_idx
            self.pending_selection = options[system_winner_idx]

        self.setup_ui()
        
        if self.current_winner_idx != -1:
            self.tree.selection_set(str(self.current_winner_idx))
            self.tree.see(str(self.current_winner_idx))
            
        self._center_on_master()

    def setup_ui(self):
        header_frame = tk.Frame(self, bg="#fffde7", bd=1, relief="ridge")
        header_frame.pack(fill="x")
        
        title_text = f"SOURCING OPTIONS: {self.item_data['Part']} | MOQ {self.moq_qty}"
        tk.Label(header_frame, text=title_text, font=("Arial", 16, "bold"), fg="#856404", bg="#fffde7").pack(pady=15)
        
        desc_text = f"Comp Level: {self.item_data.get('Comp Level', '')} | BOM Qty: {self.item_data['BOM Qty']} {self.item_data['UOM']}"
        tk.Label(header_frame, text=desc_text, font=("Arial", 10), fg="#555555", bg="#fffde7").pack(pady=(0, 10))

        main_frame = tk.Frame(self, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # 1. Pack Footer Frame FIRST at the bottom so it remains strictly visible
        footer = tk.Frame(main_frame, padx=0, pady=15)
        footer.pack(side="bottom", fill="x")
        
        tk.Label(footer, text="👑 = System's Best Choice | ✅ = Current Selection", font=("Arial", 9, "italic")).pack(side="left")

        self.lbl_pending = tk.Label(footer, text="Proposed: (None)", font=("Arial", 10, "bold"), foreground="#003399")
        if self.pending_selection:
            self.lbl_pending.config(text=f"Proposed: {self.pending_selection['Supplier']} - ${self.pending_selection['BOM Cost']:.4f}")
        self.lbl_pending.pack(side="left", padx=20)

        # Create a frame for the buttons on the right side of the footer to prevent layout squeezing
        btn_box = tk.Frame(footer)
        btn_box.pack(side="right")

        btn_cancel = tk.Button(btn_box, text="Cancel", command=self.destroy, width=12)
        btn_cancel.pack(side="left", padx=3)
        
        btn_change = tk.Button(btn_box, text="Change Winning Supplier", command=self.on_change_click)
        btn_change.pack(side="left", padx=3)
        
        btn_revert = tk.Button(btn_box, text="Revert to System Selection", command=self.on_revert_click)
        btn_revert.pack(side="left", padx=3)
        
        btn_confirm = tk.Button(btn_box, text="Confirm Selection", command=self.on_proceed, width=15)
        btn_confirm.pack(side="left", padx=3)
        
        from dialogs import style_premium_button
        style_premium_button(btn_cancel)
        style_premium_button(btn_change)
        style_premium_button(btn_revert)
        style_premium_button(btn_confirm)

        # 2. Create and pack horizontal scrollbar above footer
        sb_x = ttk.Scrollbar(main_frame, orient="horizontal")
        sb_x.pack(side="bottom", fill="x")

        # 3. Pack Table Frame THIRD to occupy all remaining space above the scrollbar
        table_frame = tk.Frame(main_frame)
        table_frame.pack(fill="both", expand=True)

        style = ttk.Style(self)
        try:
            style.theme_use('clam')
        except:
            pass
        style.configure("Options.Treeview", font=("Segoe UI", 9), rowheight=26, background="white", fieldbackground="white")
        style.configure("Options.Treeview.Heading", font=("Segoe UI", 9, "bold"), background="#dcebfa", foreground="#1A365D")
        style.map("Options.Treeview", background=[("selected", "#0078D7")], foreground=[("selected", "white")])

        cols = ("Status", "Supplier", "MPN", "Date", "Unit Price", "Cur", "USD Price", "MOQ", "BOM Cost", "Excess", "Total")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse", style="Options.Treeview")
        self.tree.config(xscrollcommand=sb_x.set)
        sb_x.config(command=self.tree.xview)
        
        headings = {
            "Status": "Status", "Supplier": "Supplier", "MPN": "MPN", 
            "Date": "Source Date", "Unit Price": "Price", "Cur": "Cur", 
            "USD Price": "USD Price", "MOQ": "MOQ", "BOM Cost": "BOM Cost", 
            "Excess": "Excess Cost", "Total": "Total Cost"
        }
        
        widths = {
            "Status": 70, "Supplier": 150, "MPN": 150, 
            "Date": 100, "Unit Price": 80, "Cur": 50, 
            "USD Price": 80, "MOQ": 60, "BOM Cost": 80, 
            "Excess": 80, "Total": 90
        }

        for c, h in headings.items():
            self.tree.heading(c, text=h)
            self.tree.column(c, width=widths.get(c, 100), anchor="center")
        attach_treeview_sort(self.tree)

        self.tree.pack(side="left", fill="both", expand=True)
        
        sb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.config(yscrollcommand=sb.set)

        self.tree.tag_configure('system_winner', background='#fff9c4')
        self.tree.tag_configure('current_winner', font=('Arial', 10, 'bold'))

        for i, opt in enumerate(self.options):
            status = ""
            tags = []
            if i == self.system_winner_idx:
                status = "👑 System"
                tags.append('system_winner')
            
            if i == self.current_winner_idx:
                if not status: status = "✅ Active"
                else: status = "👑✅ Active"
                tags.append('current_winner')

            sd = opt.get('Source Date', '')
            try: date_str = sd.strftime("%d %b %Y") if hasattr(sd, 'strftime') else str(sd)
            except: date_str = str(sd)

            supplier = str(opt.get('Supplier', ''))
            mpn = str(opt.get('MPN', ''))
            vals = (
                status,
                supplier,
                mpn,
                date_str[:12],
                f"{opt.get('Unit Price', 0):.4f}",
                opt.get('Currency', ''),
                f"{opt.get('Convert to USD', 0):.4f}",
                opt.get('MOQ', 0),
                f"{opt.get('BOM Cost', 0):.4f}",
                f"{opt.get('Excess Cost', 0):.4f}",
                f"{opt.get('Total Cost', 0):.4f}"
            )
            self.tree.insert("", "end", iid=str(i), values=vals, tags=tags)

    def on_change_click(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a row from the table first.", parent=self.winfo_toplevel())
            return
        
        idx = int(selected[0])
        self.pending_selection = self.options[idx]
        self.lbl_pending.config(text=f"Proposed: {self.pending_selection['Supplier']} - ${self.pending_selection['BOM Cost']:.4f}")
        
        if self.current_winner_idx != idx:
            self.current_winner_idx = idx
            self.refresh_tree_status()
        
        self.tree.see(str(idx))

    def on_revert_click(self):
        self.pending_selection = self.options[self.system_winner_idx]
        self.lbl_pending.config(text=f"Proposed: {self.pending_selection['Supplier']} - ${self.pending_selection['BOM Cost']:.4f}")
        
        self.current_winner_idx = self.system_winner_idx
        self.refresh_tree_status()
        
        self.tree.see(str(self.system_winner_idx))
        self.tree.selection_set(str(self.system_winner_idx))

    def refresh_tree_status(self):
        for i in range(len(self.options)):
            status = ""
            tags = []
            if i == self.system_winner_idx:
                status = "👑 System"
                tags.append('system_winner')
            
            if i == self.current_winner_idx:
                if not status: status = "✅ Active"
                else: status = "👑✅ Active"
                tags.append('current_winner')
            
            current_vals = list(self.tree.item(str(i), 'values'))
            current_vals[0] = status
            self.tree.item(str(i), values=tuple(current_vals), tags=tags)

    def on_proceed(self):
        if not self.pending_selection:
            messagebox.showwarning("Warning", "Please select a winning supplier using 'Confirm Selection' button first.", parent=self.winfo_toplevel())
            return
        
        is_different_row = (self.current_winner_idx != self.initial_winner_idx)
        self.result = (self.pending_selection, is_different_row)
        self.destroy()

    def destroy(self):
        try: self.grab_release()
        except: pass
        super().destroy()

    def _center_on_master(self):
        try: self.update_idletasks()
        except: pass
        if self.master:
            x = self.master.winfo_x() + (self.master.winfo_width() // 2) - (self.winfo_width() // 2)
            y = self.master.winfo_y() + (self.master.winfo_height() // 2) - (self.winfo_height() // 2)
            self.geometry(f"+{x}+{y}")

class MPNMFRPairDialog(tk.Toplevel):
    def __init__(self, parent, mpn_str, mfr_str):
        super().__init__(parent)
        self.title("Edit MPN/MFR Pairs")
        self.geometry("680x450")
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        
        mpns = [m.strip() for m in str(mpn_str).split(",") if m.strip()]
        mfrs = [m.strip() for m in str(mfr_str).split(",") if m.strip()]
        
        while len(mfrs) < len(mpns): mfrs.append("")
        while len(mpns) < len(mfrs): mpns.append("")
        
        self.pairs = []
        for mpn, mfr in zip(mpns, mfrs):
            self.pairs.append({'mpn': mpn, 'mfr': mfr, 'delete': tk.BooleanVar(value=False)})
            
        self.setup_ui()
        self._center_on_master()
        
    def setup_ui(self):
        header_frame = tk.Frame(self, bg="#fffde7", bd=1, relief="ridge")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="EDIT MPN/MFR PAIRS", font=("Arial", 14, "bold"), fg="#856404", bg="#fffde7").pack(pady=12)

        main_frame = tk.Frame(self, padx=10, pady=10)
        main_frame.pack(fill="both", expand=True)
        
        tk.Label(main_frame, text="Existing MPN and MFR Pairs", font=('Arial', 10, 'bold')).pack(pady=5)
        
        self.table_frame = tk.Frame(main_frame)
        self.table_frame.pack(fill="both", expand=True)
        
        self.render_pairs()
        
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=10)
        
        btn_cancel = tk.Button(btn_frame, text="Cancel", command=self.destroy, width=12)
        btn_cancel.pack(side="left", padx=5)
        
        btn_add = tk.Button(btn_frame, text="➕ Add New Pair", command=self.add_pair)
        btn_add.pack(side="left", padx=5)
        
        btn_delete = tk.Button(btn_frame, text="🗑️ Delete Selected", command=self.delete_pairs)
        btn_delete.pack(side="left", padx=5)
        
        btn_save = tk.Button(btn_frame, text="✅ Save & Join", command=self.save, width=15)
        btn_save.pack(side="right", padx=5)
        
        from dialogs import style_premium_button
        style_premium_button(btn_cancel)
        style_premium_button(btn_add)
        style_premium_button(btn_delete)
        style_premium_button(btn_save)
        
    def render_pairs(self):
        for widget in self.table_frame.winfo_children():
            widget.destroy()
            
        tk.Label(self.table_frame, text="MPN", font=('Arial', 8, 'bold')).grid(row=0, column=1, sticky="w", padx=5)
        tk.Label(self.table_frame, text="MFR", font=('Arial', 8, 'bold')).grid(row=0, column=2, sticky="w", padx=5)
        
        for i, pair in enumerate(self.pairs):
            ttk.Checkbutton(self.table_frame, variable=pair['delete']).grid(row=i+1, column=0, padx=5)
            
            mpn_var = tk.StringVar(value=pair['mpn'])
            mfr_var = tk.StringVar(value=pair['mfr'])
            
            ent_mpn = ttk.Entry(self.table_frame, textvariable=mpn_var, width=30)
            ent_mpn.grid(row=i+1, column=1, padx=5, pady=2)
            ent_mfr = ttk.Entry(self.table_frame, textvariable=mfr_var, width=30)
            ent_mfr.grid(row=i+1, column=2, padx=5, pady=2)
            
            pair['mpn_var'] = mpn_var
            pair['mfr_var'] = mfr_var
            
    def add_pair(self):
        self.sync_vars()
        self.pairs.append({'mpn': '', 'mfr': '', 'delete': tk.BooleanVar(value=False)})
        self.render_pairs()
        
    def delete_pairs(self):
        self.sync_vars()
        self.pairs = [p for p in self.pairs if not p['delete'].get()]
        self.render_pairs()
        
    def sync_vars(self):
        for p in self.pairs:
            if 'mpn_var' in p:
                p['mpn'] = p['mpn_var'].get().strip()
                p['mfr'] = p['mfr_var'].get().strip()
                
    def save(self):
        self.sync_vars()
        mpns = [p['mpn'] for p in self.pairs if p['mpn']]
        mfrs = [p['mfr'] for p in self.pairs if p['mpn']]
        
        self.result = (", ".join(mpns), ", ".join(mfrs))
        self.destroy()

    def destroy(self):
        try: self.grab_release()
        except: pass
        super().destroy()

    def _center_on_master(self):
        try: self.update_idletasks()
        except: pass
        if self.master:
            x = self.master.winfo_x() + (self.master.winfo_width() // 2) - (self.winfo_width() // 2)
            y = self.master.winfo_y() + (self.master.winfo_height() // 2) - (self.winfo_height() // 2)
            self.geometry(f"+{x}+{y}")

class AddNewRowDialog(tk.Toplevel):
    def __init__(self, parent, assy, model, rev):
        super().__init__(parent)
        self.title("Add New Row")
        self.geometry("450x400")
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        self.assy = assy
        self.model = model
        self.rev = rev
        
        self.vars = {
            'Part': tk.StringVar(),
            'Description': tk.StringVar(),
            'MFR': tk.StringVar(),
            'MPN': tk.StringVar(),
            'BOM Qty': tk.StringVar(value="1"),
            'UOM': tk.StringVar(value="EA")
        }
        
        self.setup_ui()
        
    def setup_ui(self):
        header_frame = tk.Frame(self, bg="#fffde7", bd=1, relief="ridge")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="ADD NEW ROW", font=("Arial", 14, "bold"), fg="#856404", bg="#fffde7").pack(pady=12)

        main_frame = tk.Frame(self, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        tk.Label(main_frame, text=f"Adding to Assembly: {self.assy}", font=('Arial', 10, 'bold')).pack(pady=(0, 15))
        
        for i, (label, var) in enumerate(self.vars.items()):
            f = tk.Frame(main_frame)
            f.pack(fill="x", pady=5)
            tk.Label(f, text=label, width=15).pack(side="left")
            ttk.Entry(f, textvariable=var).pack(side="left", fill="x", expand=True)
            
        btn_frame = tk.Frame(main_frame, padx=20, pady=20)
        btn_frame.pack(fill="x", side="bottom", pady=20)
        
        btn_cancel = tk.Button(btn_frame, text="Cancel", command=self.destroy, width=15)
        btn_cancel.pack(side="left", padx=5)
        
        btn_save = tk.Button(btn_frame, text="✅ Add Row", command=self.save, width=15)
        btn_save.pack(side="right", padx=5)
        
        from dialogs import style_premium_button
        style_premium_button(btn_cancel)
        style_premium_button(btn_save)
        
    def save(self):
        try:
            qty = float(self.vars['BOM Qty'].get())
        except ValueError:
            messagebox.showerror("Error", "BOM Qty must be a number.", parent=self.winfo_toplevel())
            return
            
        self.result = {
            'Assy': self.assy,
            'Assy Model': self.model,
            'Assy Rev': self.rev,
            'Part': self.vars['Part'].get().strip(),
            'Description': self.vars['Description'].get().strip(),
            'MFR': self.vars['MFR'].get().strip(),
            'MPN': self.vars['MPN'].get().strip(),
            'BOM Qty': qty,
            'UOM': self.vars['UOM'].get().strip(),
            'moq_results': {}
        }
        self.destroy()

class ConfirmationDialog(tk.Toplevel):
    def __init__(self, parent, title, heading, message_list, action_text="Proceed"):
        super().__init__(parent)
        self.title(title)
        self.geometry("540x480")
        self.minsize(500, 420)
        self.transient(parent)
        self.grab_set()
        
        self.result = "CANCEL"
        self.configure(bg="#F8FAFC")
        
        # 1. Header Banner (Top)
        header_frame = tk.Frame(self, bg="#1A365D", pady=12)
        header_frame.pack(fill="x", side="top")
        tk.Label(header_frame, text=title.upper(), font=("Segoe UI", 12, "bold"), fg="white", bg="#1A365D").pack()

        # 2. Footer Action Bar (Bottom - packed BEFORE main_frame so it never overflows or clips!)
        btn_frame = tk.Frame(self, bg="#F1F5F9", padx=20, pady=12, bd=1, relief="solid")
        btn_frame.pack(fill="x", side="bottom")
        
        btn_cancel = tk.Button(
            btn_frame, text="Cancel", command=self.destroy,
            font=('Segoe UI', 10, 'bold'), bg="#718096", fg="white",
            activebackground="#4A5568", activeforeground="white",
            relief="flat", bd=0, cursor="hand2", width=12, height=1
        )
        btn_cancel.pack(side="left", padx=5)
        btn_cancel.bind("<Enter>", lambda e: btn_cancel.config(bg="#4A5568"))
        btn_cancel.bind("<Leave>", lambda e: btn_cancel.config(bg="#718096"))
        
        btn_bg = "#C53030" if ("Delete" in action_text or "Remove" in action_text) else "#1A365D"
        btn_hover = "#9B2C2C" if ("Delete" in action_text or "Remove" in action_text) else "#0077B6"
        
        btn_proceed = tk.Button(
            btn_frame, text=action_text, command=self.on_proceed,
            font=('Segoe UI', 10, 'bold'), bg=btn_bg, fg="white",
            activebackground=btn_hover, activeforeground="white",
            relief="flat", bd=0, cursor="hand2", width=22, height=1
        )
        btn_proceed.pack(side="right", padx=5)
        btn_proceed.bind("<Enter>", lambda e: btn_proceed.config(bg=btn_hover))
        btn_proceed.bind("<Leave>", lambda e: btn_proceed.config(bg=btn_bg))

        # 3. Main Body Frame (Center)
        main_frame = tk.Frame(self, bg="#F8FAFC", padx=20, pady=15)
        main_frame.pack(fill="both", expand=True, side="top")
        
        tk.Label(main_frame, text=heading, font=('Segoe UI', 11, 'bold'), fg="#1A365D", bg="#F8FAFC", wraplength=480, justify="left").pack(pady=(0, 10), anchor="w")
        
        list_frame = tk.Frame(main_frame, bg="#FFFFFF", bd=1, relief="solid")
        list_frame.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(list_frame, highlightthickness=0, bg="#FFFFFF")
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#FFFFFF")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scrollbar.pack(side="right", fill="y")
        
        for msg in message_list:
            tk.Label(scrollable_frame, text=f"• {msg}", font=('Segoe UI', 9), fg="#2D3748", bg="#FFFFFF", wraplength=450, justify="left").pack(anchor="w", pady=3, padx=5)

    def on_proceed(self):
        self.result = "PROCEED"
        self.destroy()

class ManualOverrideConfirmationDialog(tk.Toplevel):
    def __init__(self, parent, title, changes_data):
        self._skip_autofit = True
        super().__init__(parent)
        self.title(title)
        self.geometry("850x650")
        self.resizable(False, False)
        self.configure(bg="#EBF8FF")
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        self.changes_data = changes_data
        
        self.vars = []
        
        # Premium Blue Header Banner
        header_frame = tk.Frame(self, bg="#1A365D")
        header_frame.pack(fill="x", side="top")
        tk.Label(header_frame, text="REVIEW & CONFIRM OVERRIDES", font=("Segoe UI", 14, "bold"), fg="white", bg="#1A365D").pack(pady=15)

        # 1. Pack Footer Frame FIRST at the bottom so it remains strictly visible
        btn_frame = tk.Frame(self, bg="#EBF8FF", padx=20, pady=15)
        btn_frame.pack(fill="x", side="bottom")
        
        btn_cancel = tk.Button(btn_frame, text="Cancel", command=self.destroy, width=10)
        btn_cancel.pack(side="left", padx=5)
        
        btn_apply = tk.Button(btn_frame, text="✅ Apply Selected Changes", command=self.on_apply, width=25)
        btn_apply.pack(side="right", padx=5)
        
        from dialogs import style_premium_button
        style_premium_button(btn_cancel)
        style_premium_button(btn_apply)

        # 2. Main Frame takes up remaining space
        main_frame = tk.Frame(self, bg="#EBF8FF", padx=20, pady=15)
        main_frame.pack(fill="both", expand=True)

        summary_f = tk.Frame(main_frame, bg="#dcebfa", bd=1, relief="solid", pady=10)
        summary_f.pack(fill="x", pady=(0, 15))
        
        total_rows = len(changes_data)
        total_assys = len(set(c['assy'] for c in changes_data))
        
        tk.Label(summary_f, text=f"Total Updates: {total_rows}", font=("Arial", 10, "bold"), bg="#dcebfa", fg="#1A365D").pack(side="left", padx=20)
        tk.Label(summary_f, text=f"Assemblies Affected: {total_assys}", font=("Arial", 10, "bold"), bg="#dcebfa", fg="#1A365D").pack(side="left", padx=20)
        
        tk.Label(main_frame, text="Select which manual overrides to apply:", font=("Arial", 11, "bold"), bg="#EBF8FF", fg="#1A365D").pack(anchor="w", pady=(0, 10))

        tbl_header = tk.Frame(main_frame, bg="#dcebfa")
        tbl_header.pack(fill="x")
        tk.Label(tbl_header, text="Apply", width=6, bg="#dcebfa", font=("Arial", 9, "bold")).pack(side="left")
        tk.Label(tbl_header, text="Comp Level", width=12, bg="#dcebfa", font=("Arial", 9, "bold")).pack(side="left")
        tk.Label(tbl_header, text="Assembly", width=25, bg="#dcebfa", anchor="w", font=("Arial", 9, "bold")).pack(side="left", padx=5)
        tk.Label(tbl_header, text="MOQ", width=8, bg="#dcebfa", font=("Arial", 9, "bold")).pack(side="left")
        tk.Label(tbl_header, text="New Winner", width=20, bg="#dcebfa", anchor="w", font=("Arial", 9, "bold")).pack(side="left")

        scroll_container = tk.Frame(main_frame, bg="#EBF8FF", bd=1, relief="solid")
        scroll_container.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(scroll_container, highlightthickness=0, bg="#EBF8FF")
        scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg="#EBF8FF")
        
        self.scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw", width=700)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for i, change in enumerate(changes_data):
            row_bg = "#EBF8FF" if i % 2 == 0 else "#DCEBFA"
            row_f = tk.Frame(self.scrollable_frame, bg=row_bg, pady=5)
            row_f.pack(fill="x")
            
            tk.Label(row_f, text="☑", font=("Arial", 12), fg="#2ead4e", bg=row_bg).pack(side="left", padx=(15, 5))
            
            var = tk.BooleanVar(value=True)
            self.vars.append(var)
            
            comp_txt = f"{change['comp_level']}"
            if change['is_primary']: comp_txt += "*"
            
            tk.Label(row_f, text=comp_txt, width=12, bg=row_bg, font=("Arial", 9, "bold" if change['is_primary'] else "normal")).pack(side="left")
            
            assy_txt = str(change['assy'])
            if len(assy_txt) > 30: assy_txt = assy_txt[:27] + "..."
            tk.Label(row_f, text=assy_txt, width=32, bg=row_bg, anchor="w", font=("Arial", 9)).pack(side="left", padx=5)
            
            tk.Label(row_f, text=change['moq'], width=8, bg=row_bg).pack(side="left")
            
            supplier_color = "#27ae60" if change['is_primary'] else "#2980b9"
            tk.Label(row_f, text=change['supplier'], width=25, bg=row_bg, anchor="w", fg=supplier_color, font=("Arial", 9, "bold")).pack(side="left")

        legend_f = tk.Frame(main_frame, bg="#EBF8FF")
        legend_f.pack(fill="x", pady=10)
        tk.Label(legend_f, text="* Primary Selection (Current Row)", font=("Arial", 8, "italic"), bg="#EBF8FF", fg="#7f8c8d").pack(side="left")
        
    def on_apply(self):
        selected = []
        for var, data in zip(self.vars, self.changes_data):
            if var.get():
                selected.append(data)
        
        if not selected:
            if messagebox.askyesno("No Selection", "You haven't selected any changes to apply. Cancel override?", parent=self):
                self.destroy()
            return
            
        self.result = selected
        self.destroy()

class SharedPartsDialog(tk.Toplevel):
    def __init__(self, parent, part_number, current_assembly, shared_items, assembly_moqs):
        super().__init__(parent)
        self.title(f"Shared Part Details - {part_number}")
        self.transient(parent)
        self.grab_set()
        self.part_number = part_number
        self.current_assembly = current_assembly
        self.all_shared_items = shared_items
        self.assembly_moqs = assembly_moqs
        self.current_filters = {"Assembly": [], "MOQ": [], "MOQ_Mode": "OR"}
        
        # Center the dialog on parent
        self.geometry("850x550")
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")
        
        header_frame = tk.Frame(self, bg="#fffde7", bd=1, relief="ridge")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="SHARED PART DETAILS", font=("Arial", 16, "bold"), fg="#856404", bg="#fffde7").pack(pady=15)
        
        # Close button at the bottom (packed first at the bottom of the window so it is never pushed off)
        btn_frame = tk.Frame(self, bg="#f1f3f5", padx=20, pady=15)
        btn_frame.pack(fill="x", side="bottom")
        
        btn_close = tk.Button(btn_frame, text="Close", command=self.destroy, width=12)
        btn_close.pack(side="right", padx=5)
        
        from dialogs import style_premium_button
        style_premium_button(btn_close)

        main_frame = tk.Frame(self, bg="#EBF8FF", padx=20, pady=15)
        main_frame.pack(fill="both", expand=True)
        
        # Info about current part
        info_frame = tk.Frame(main_frame, bg="white", bd=1, relief="solid", pady=10, padx=15)
        info_frame.pack(fill="x", pady=(0, 15))
        tk.Label(info_frame, text=f"Part Number: {part_number}", font=("Arial", 10, "bold"), bg="white", fg="#2c3e50").pack(anchor="w")
        tk.Label(info_frame, text=f"Current Assembly: {current_assembly}", font=("Arial", 10), bg="white", fg="#7f8c8d").pack(anchor="w", pady=(2, 0))
        
        # Control frame for filters
        ctrl_frame = tk.Frame(main_frame, bg="#EBF8FF")
        ctrl_frame.pack(fill="x", pady=(0, 10))
        
        #btn_filter = tk.Button(ctrl_frame, text="🔍 Filter Records", command=self.open_filter_dialog)
        #btn_filter.pack(side="left")
        
        #from dialogs import style_premium_button
        #style_premium_button(btn_filter)
        
        self.lbl_filter_status = tk.Label(ctrl_frame, text="", font=("Arial", 9, "bold"), bg="#EBF8FF", fg="#555555")
        self.lbl_filter_status.pack(side="left", padx=15)
        
        tk.Label(main_frame, text="This part is also shared with the following assemblies:", font=("Arial", 11, "bold"), bg="#EBF8FF", fg="#34495e").pack(anchor="w", pady=(0, 10))
        
        # Table Header
        tbl_header = tk.Frame(main_frame, bg="#e9ecef")
        tbl_header.pack(fill="x")
        tk.Label(tbl_header, text="Assembly", width=32, bg="#e9ecef", anchor="w", font=("Arial", 9, "bold")).pack(side="left", padx=(15, 5))
        tk.Label(tbl_header, text="Comp Level", width=15, bg="#e9ecef", anchor="w", font=("Arial", 9, "bold")).pack(side="left", padx=5)
        tk.Label(tbl_header, text="MOQs Assigned", width=25, bg="#e9ecef", anchor="w", font=("Arial", 9, "bold")).pack(side="left", padx=5)
        tk.Label(tbl_header, text="Qty to Quote", width=25, bg="#e9ecef", anchor="w", font=("Arial", 9, "bold")).pack(side="left", padx=5)
        
        # Scrollable container
        scroll_container = tk.Frame(main_frame, bg="white", bd=1, relief="solid")
        scroll_container.pack(fill="both", expand=True)
        
        self.canvas = tk.Canvas(scroll_container, highlightthickness=0, bg="white")
        scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg="white")
        
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw", width=800)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Mousewheel binding helpers
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            
        self.bind_mousewheel_recursively = lambda w: self._bind_mousewheel_recursively(w, _on_mousewheel)
        
        self.bind_mousewheel_recursively(self.canvas)
        self.bind_mousewheel_recursively(self.scrollable_frame)
        
        # Populate table initially
        self.populate_table()

    def _bind_mousewheel_recursively(self, w, on_mousewheel):
        w.bind("<MouseWheel>", on_mousewheel)
        for child in w.winfo_children():
            self._bind_mousewheel_recursively(child, on_mousewheel)

    def open_filter_dialog(self):
        dialog = SharedPartsFilterDialog(self, self.current_filters)
        self.wait_window(dialog)
        if dialog.result is not None:
            self.current_filters = dialog.result
            self.populate_table()

        else:
            self.filter_criteria = None
            if hasattr(self, 'populate_tree'):
                self.populate_tree()
            elif hasattr(self, 'refresh_projects'):
                self.refresh_projects()
    def populate_table(self):
        # Clear existing row widgets in scrollable_frame
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
            
        # Get active filters
        assy_filters = [a.lower().strip() for a in self.current_filters.get("Assembly", []) if a.strip()]
        moq_filters = [m.lower().strip() for m in self.current_filters.get("MOQ", []) if m.strip()]
        moq_mode = self.current_filters.get("MOQ_Mode", "OR")
        
        filtered_items = []
        for item in self.all_shared_items:
            assy_txt = str(item.get('Assy', '')).strip()
            moqs = self.assembly_moqs.get(assy_txt, [])
            
            # Check assembly match
            assy_match = False
            if not assy_filters:
                assy_match = True
            else:
                for af in assy_filters:
                    if af in assy_txt.lower():
                        assy_match = True
                        break
                        
            # Check moq match
            moq_match = False
            if not moq_filters:
                moq_match = True
            else:
                if moq_mode == "AND":
                    # All filter terms must match at least one MOQ in the assembly's moqs
                    matches_all = True
                    for mf in moq_filters:
                        found_match = False
                        for q in moqs:
                            if mf in str(q).lower():
                                found_match = True
                                break
                        if not found_match:
                            matches_all = False
                            break
                    moq_match = matches_all
                else: # OR mode
                    for mf in moq_filters:
                        for q in moqs:
                            if mf in str(q).lower():
                                moq_match = True
                                break
                        if moq_match:
                            break
                            
            if assy_match and moq_match:
                filtered_items.append(item)
                
        # Now render filtered_items
        for i, item in enumerate(filtered_items):
            row_bg = "white" if i % 2 == 0 else "#EBF8FF"
            row_f = tk.Frame(self.scrollable_frame, bg=row_bg, pady=5)
            row_f.pack(fill="x")
            
            # Assembly
            assy_txt = str(item.get('Assy', ''))
            if len(assy_txt) > 30:
                assy_txt = assy_txt[:27] + "..."
            assy_entry = tk.Entry(row_f, width=32, font=("Arial", 9), relief="flat", bd=0, bg=row_bg, readonlybackground=row_bg, fg="#333333")
            assy_entry.insert(0, assy_txt)
            assy_entry.config(state="readonly")
            assy_entry.pack(side="left", padx=(15, 5))
            
            # Comp Level
            comp_level = str(item.get('Comp Level', ''))
            comp_entry = tk.Entry(row_f, width=15, font=("Arial", 9), relief="flat", bd=0, bg=row_bg, readonlybackground=row_bg, fg="#333333")
            comp_entry.insert(0, comp_level)
            comp_entry.config(state="readonly")
            comp_entry.pack(side="left", padx=5)
            
            # MOQs Assigned (no "MOQ" repeat, e.g. "100, 200, 300")
            moqs = self.assembly_moqs.get(str(item.get('Assy', '')).strip(), [])
            moqs_txt = ", ".join(str(q) for q in sorted(moqs)) if moqs else "None"
            moq_entry = tk.Entry(row_f, width=25, font=("Arial", 9), relief="flat", bd=0, bg=row_bg, readonlybackground=row_bg, fg="#333333")
            moq_entry.insert(0, moqs_txt)
            moq_entry.config(state="readonly")
            moq_entry.pack(side="left", padx=5)
            
            # Qty to Quote
            try:
                bom_qty = float(item.get('BOM Qty', 0))
            except:
                bom_qty = 0.0
            qtys_txt = ", ".join(f"{bom_qty * float(q):.4f}".rstrip('0').rstrip('.') for q in sorted(moqs)) if moqs else "None"
            qty_entry = tk.Entry(row_f, width=25, font=("Arial", 9), relief="flat", bd=0, bg=row_bg, readonlybackground=row_bg, fg="#333333")
            qty_entry.insert(0, qtys_txt)
            qty_entry.config(state="readonly")
            qty_entry.pack(side="left", padx=5)
            
            # Bind scroll wheel
            self.bind_mousewheel_recursively(row_f)
            
        # Update scrollregion
        self.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        # Update filter status label
        total_all = len(self.all_shared_items)
        total_filtered = len(filtered_items)
        if assy_filters or moq_filters:
            self.lbl_filter_status.config(text=f"Filtered: {total_filtered} of {total_all} records")
        else:
            self.lbl_filter_status.config(text=f"Total: {total_all} records")

class SharedPartsFilterDialog(BaseDialog):
    def __init__(self, master, current_filters=None):
        super().__init__(master, "Filter Shared Parts")
        # self.geometry("500x250")
        self.current_filters = current_filters or {"Assembly": [], "MOQ": [], "MOQ_Mode": "OR"}
        
        self.assy_list = list(self.current_filters.get("Assembly", []))
        self.moq_list = list(self.current_filters.get("MOQ", []))
        self.assy_multi_selected = len(self.assy_list) > 1
        self.moq_multi_selected = len(self.moq_list) > 1
        
        self.result = None
        self._create_widgets()
        
    def _create_widgets(self):
        main_frame = tk.Frame(self, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        grid_frame = tk.Frame(main_frame)
        grid_frame.pack(fill="x", expand=False)
        
        # Assembly input
        tk.Label(grid_frame, text="Assembly #:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", pady=10)
        self.assy_entry = ttk.Entry(grid_frame, width=35)
        self.assy_entry.grid(row=0, column=1, sticky="ew", padx=10, pady=10)
        btn_assy_multi = tk.Button(grid_frame, text="⫘", font=("Arial", 10, "bold"), command=self._edit_assy_multi, bg="#e2e8f0", relief="groove")
        btn_assy_multi.grid(row=0, column=2, padx=5, pady=10)
        
        # MOQ input
        tk.Label(grid_frame, text="MOQ:", font=("Arial", 10, "bold")).grid(row=1, column=0, sticky="w", pady=10)
        self.moq_entry = ttk.Entry(grid_frame, width=35)
        self.moq_entry.grid(row=1, column=1, sticky="ew", padx=10, pady=10)
        btn_moq_multi = tk.Button(grid_frame, text="⫘", font=("Arial", 10, "bold"), command=self._edit_moq_multi, bg="#e2e8f0", relief="groove")
        btn_moq_multi.grid(row=1, column=2, padx=5, pady=10)
        
        # MOQ match mode
        tk.Label(grid_frame, text="MOQ Match Mode:", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky="w", pady=10)
        mode_frame = tk.Frame(grid_frame)
        mode_frame.grid(row=2, column=1, sticky="w", padx=10, pady=10)
        
        self.moq_mode_var = tk.StringVar(value=self.current_filters.get("MOQ_Mode", "OR"))
        tk.Radiobutton(mode_frame, text="Any (OR)", variable=self.moq_mode_var, value="OR").pack(side="left", padx=(0, 10))
        tk.Radiobutton(mode_frame, text="All (AND)", variable=self.moq_mode_var, value="AND").pack(side="left")
        
        grid_frame.columnconfigure(1, weight=1)
        
        self._update_entry_fields()
        
        btn_frame = tk.Frame(main_frame, padx=0, pady=15)
        btn_frame.pack(fill="x", side="bottom")
        
        btn_confirm = tk.Button(btn_frame, text="Execute Filter", command=self._on_execute, width=15)
        btn_confirm.pack(side="right", padx=5)
        btn_clear = tk.Button(btn_frame, text="Clear Filters", command=self._on_clear, width=15)
        btn_clear.pack(side="right", padx=5)
        btn_cancel = tk.Button(btn_frame, text="Cancel", command=self._on_cancel, width=12)
        btn_cancel.pack(side="left")
        
        from dialogs import style_premium_button
        style_premium_button(btn_confirm)
        style_premium_button(btn_clear)
        style_premium_button(btn_cancel)
        
    def _update_entry_fields(self):
        self.assy_entry.config(state="normal")
        self.assy_entry.delete(0, "end")
        if self.assy_list:
            if len(self.assy_list) == 1:
                self.assy_entry.insert(0, self.assy_list[0])
                self.assy_multi_selected = False
            else:
                self.assy_entry.insert(0, f"[{len(self.assy_list)} Values Selected]")
                self.assy_entry.config(state="readonly")
                self.assy_multi_selected = True
        else:
            self.assy_multi_selected = False
            
        self.moq_entry.config(state="normal")
        self.moq_entry.delete(0, "end")
        if self.moq_list:
            if len(self.moq_list) == 1:
                self.moq_entry.insert(0, self.moq_list[0])
                self.moq_multi_selected = False
            else:
                self.moq_entry.insert(0, f"[{len(self.moq_list)} Values Selected]")
                self.moq_entry.config(state="readonly")
                self.moq_multi_selected = True
        else:
            self.moq_multi_selected = False
            
    def _edit_assy_multi(self):
        dialog = MultiValueInputDialog(self, "Multi Assembly Search", self.assy_list)
        self.wait_window(dialog)
        if dialog.result is not None:
            self.assy_list = dialog.result
            self.assy_multi_selected = len(self.assy_list) > 1
            self._update_entry_fields()
            
        else:
            self.filter_criteria = None
            if hasattr(self, 'populate_tree'):
                self.populate_tree()
            elif hasattr(self, 'refresh_projects'):
                self.refresh_projects()
    def _edit_moq_multi(self):
        dialog = MultiValueInputDialog(self, "Multi MOQ Search", self.moq_list)
        self.wait_window(dialog)
        if dialog.result is not None:
            self.moq_list = dialog.result
            self.moq_multi_selected = len(self.moq_list) > 1
            self._update_entry_fields()
            
        else:
            self.filter_criteria = None
            if hasattr(self, 'populate_tree'):
                self.populate_tree()
            elif hasattr(self, 'refresh_projects'):
                self.refresh_projects()
    def _on_execute(self):
        if not self.assy_multi_selected:
            val = self.assy_entry.get().strip()
            self.assy_list = [val] if val else []
            
        if not self.moq_multi_selected:
            val = self.moq_entry.get().strip()
            self.moq_list = [val] if val else []
            
        self.result = {"Assembly": self.assy_list, "MOQ": self.moq_list, "MOQ_Mode": self.moq_mode_var.get()}
        self.destroy()
        
    def _on_clear(self):
        self.result = {"Assembly": [], "MOQ": [], "MOQ_Mode": "OR"}
        self.destroy()
        
    def _on_cancel(self):
        self.result = None
        self.destroy()

