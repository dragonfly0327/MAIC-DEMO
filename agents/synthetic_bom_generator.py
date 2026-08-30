# ==============================================================================
# --- ContinuumX Synthetic BOM Generator ---
# Converts extracted intermediate RFQ JSON into standardized Synthetic Excel workbooks
# with exact columns: Line, Assy#, Assy Model, Assy Rev, Part, Description, MPN, MFR, QTY, UOM, EAU, Target Price (USD).
# ==============================================================================

import os
import sys
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Standardized Base Directory Resolution Pattern
if getattr(sys, 'frozen', False) or hasattr(sys, '_MEIPASS') or "__compiled__" in globals():
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
elif '__file__' in globals():
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
elif len(sys.argv) > 0 and sys.argv[0] and sys.argv[0] != '-c':
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(sys.argv[0])))
else:
    BASE_DIR = os.getcwd()

if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)


def format_paired_mpn_mfr(mpn_val, mfr_val):
    """
    Ensures MPN and MFR are strictly 1-to-1 paired and comma-separated with ', '.
    If 1 MPN -> 1 MFR. If 2 MPNs -> 2 MFRs.
    """
    if not mpn_val and not mfr_val:
        return "", ""

    raw_mpns = [m.strip() for m in re.split(r'[/;|,\\]', str(mpn_val or "")) if m.strip()]
    raw_mfrs = [f.strip() for f in re.split(r'[/;|,\\]', str(mfr_val or "")) if f.strip()]

    if not raw_mpns:
        return "", ", ".join(raw_mfrs)

    if len(raw_mfrs) == 0:
        raw_mfrs = ["Unknown"] * len(raw_mpns)
    elif len(raw_mfrs) < len(raw_mpns):
        if len(raw_mfrs) == 1:
            raw_mfrs = [raw_mfrs[0]] * len(raw_mpns)
        else:
            raw_mfrs.extend([raw_mfrs[-1]] * (len(raw_mpns) - len(raw_mfrs)))
    elif len(raw_mfrs) > len(raw_mpns):
        raw_mfrs = raw_mfrs[:len(raw_mpns)]

    return ", ".join(raw_mpns), ", ".join(raw_mfrs)


class SyntheticBOMGenerator:
    """
    Transforms structured RFQ intermediate JSON objects into standard
    ContinuumX synthetic BOM Excel files (.xlsx).
    """
    STANDARD_COLUMNS = [
        "Line", "Assy#", "Assy Model", "Assy Rev", "Part", "Description", "MPN", "MFR", "QTY", "UOM", "EAU", "Target Price (USD)"
    ]

    def __init__(self, output_dir=None):
        if not output_dir:
            try:
                bom_ref = os.path.join(BASE_DIR, "ref", "BOM")
                if bom_ref not in sys.path:
                    sys.path.append(bom_ref)
                from utils import SYNTHETIC_BOM_DIR
                if SYNTHETIC_BOM_DIR and os.path.exists(os.path.dirname(SYNTHETIC_BOM_DIR)):
                    os.makedirs(SYNTHETIC_BOM_DIR, exist_ok=True)
                    output_dir = SYNTHETIC_BOM_DIR
            except Exception:
                pass

        if not output_dir:
            local_appdata = os.environ.get('LOCALAPPDATA', os.environ.get('TEMP', os.path.join(BASE_DIR, 'data')))
            output_dir = os.path.join(local_appdata, "ContXs", "SyntheticBOM")
            os.makedirs(output_dir, exist_ok=True)

        self.output_dir = output_dir

    def generate_synthetic_excel(self, rfq_json, custom_filename=None):
        """
        Creates an openpyxl workbook formatted with exact ContinuumX columns:
        Line | Assy# | Assy Model | Assy Rev | Part | Description | MPN | MFR | QTY | UOM | EAU | Target Price (USD)
        """
        try:
            meta = rfq_json.get("rfq_metadata", {})
            customer = meta.get("customer_name") or "Customer"
            rfq_num = meta.get("rfq_number") or f"RFQ{datetime.now().strftime('%m%d%H%M')}"
            proj_title = meta.get("project_title") or f"{customer} Assembly"
            commodity = meta.get("commodity") or "Wire Harness"
            target_price = meta.get("target_price") or "Not Specified"
            eau = meta.get("eau") if meta.get("eau") is not None else ""
            moqs = meta.get("default_moqs") or [100, 200, 300, 500]

            # Generate Clean Target Filename
            clean_cust = re.sub(r'[^a-zA-Z0-9_-]', '_', customer)
            clean_rfq = re.sub(r'[^a-zA-Z0-9_-]', '_', rfq_num)
            if not custom_filename:
                filename = f"synthetic_{clean_cust}_{clean_rfq}.xlsx"
            else:
                filename = custom_filename if custom_filename.endswith(".xlsx") else f"{custom_filename}.xlsx"

            file_path = os.path.join(self.output_dir, filename)

            # Create Workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM"

            # Header Styling
            header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
            header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            border_thin = Side(border_style="thin", color="CBD5E0")
            cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

            # Write Headers
            for col_idx, col_name in enumerate(self.STANDARD_COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = cell_border

            # Populate Rows from Assemblies (Deduplicate assemblies and items strictly)
            raw_assemblies = rfq_json.get("assemblies", [])
            assemblies = []
            seen_assy_set = set()
            for assy in raw_assemblies:
                a_no = str(assy.get("assy_no", "")).strip()
                if a_no and a_no not in seen_assy_set:
                    seen_assy_set.add(a_no)
                    # Auto-sum quantities for repeated parts with matching UOM
                    consolidated_map = {}
                    for it in assy.get("items", []):
                        p_no = str(it.get("part_number", "")).strip()
                        uom_val = str(it.get("uom", "EA")).strip().upper()
                        key = (p_no, uom_val)
                        
                        try:
                            cur_q = float(it.get("qty", 1))
                        except Exception:
                            cur_q = 1.0

                        if key not in consolidated_map:
                            it_copy = dict(it)
                            it_copy["qty"] = cur_q
                            consolidated_map[key] = it_copy
                        else:
                            # Auto-Sum repeated occurrence
                            consolidated_map[key]["qty"] = round(consolidated_map[key]["qty"] + cur_q, 4)
                    
                    # Convert whole floats back to clean ints
                    for it in consolidated_map.values():
                        if isinstance(it.get("qty"), float) and it["qty"].is_integer():
                            it["qty"] = int(it["qty"])

                    assy["items"] = list(consolidated_map.values()) if consolidated_map else assy.get("items", [])
                    assemblies.append(assy)
                elif not a_no:
                    assemblies.append(assy)

            row_idx = 2
            total_items = 0

            for assy_idx, assy in enumerate(assemblies, start=1):
                assy_no = assy.get("assy_no") or f"ASSY-{assy_idx:02d}"
                assy_model = assy.get("assy_model") or proj_title
                assy_rev = str(assy.get("assy_rev") or "00").replace("Rev", "").strip()
                assy_eau = assy.get("eau") if assy.get("eau") is not None else (eau or "")
                assy_tp = str(assy.get("target_price", target_price)).replace('$', '').strip()
                try:
                    assy_tp_num = float(assy_tp)
                except Exception:
                    assy_tp_num = assy_tp

                items = assy.get("items", [])

                if not items:
                    items = [{
                        "line_item": 1,
                        "part_number": assy_no,
                        "description": assy_model,
                        "mfr": "",
                        "mpn": "",
                        "qty": 1,
                        "uom": "EA",
                        "eau": assy_eau,
                        "target_price": assy_tp_num
                    }]

                for comp_idx, item in enumerate(items, start=1):
                    part_num = item.get("part_number") or ""
                    desc = item.get("description") or "Component"
                    mfr_raw = item.get("mfr") or ""
                    mpn_raw = item.get("mpn") or ""
                    mpn, mfr = format_paired_mpn_mfr(mpn_raw, mfr_raw)
                    qty = item.get("qty", 1)
                    uom = item.get("uom") or "EA"
                    row_eau = item.get("eau") if item.get("eau") is not None else (assy_eau or "")
                    row_tp = str(item.get("target_price", assy_tp_num)).replace('$', '').strip()
                    try:
                        row_tp_val = float(row_tp)
                    except Exception:
                        row_tp_val = row_tp

                    # Line | Assy# | Assy Model | Assy Rev | Part | Description | MPN | MFR | QTY | UOM | EAU | Target Price (USD)
                    row_values = [
                        comp_idx,       # Line (1..N per assembly)
                        assy_no,        # Assy#
                        assy_model,     # Assy Model
                        assy_rev,       # Assy Rev
                        part_num,       # Part
                        desc,           # Description
                        mpn,            # MPN
                        mfr,            # MFR
                        qty,            # QTY
                        uom,            # UOM
                        row_eau,        # EAU
                        row_tp_val      # Target Price (USD)
                    ]

                    for col_idx, val in enumerate(row_values, start=1):
                        c = ws.cell(row=row_idx, column=col_idx, value=val)
                        c.font = Font(name="Segoe UI", size=9)
                        c.border = cell_border
                        if col_idx in [1, 4, 9, 10, 11, 12]:  # Line, Assy Rev, QTY, UOM, EAU, Target Price
                            c.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            c.alignment = Alignment(horizontal="left", vertical="center")

                    row_idx += 1
                    total_items += 1

            # Auto-fit Column Widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # --- Conflict Log Sheet (Phase 3) ---
            conflicts = rfq_json.get("conflict_candidates", [])
            if conflicts:
                wc = wb.create_sheet("Conflict Log")
                conflict_header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
                conflict_header_font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
                warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                conflict_cols = [
                    "Field", "Source A Doc", "Source A Value",
                    "Source B Doc", "Source B Value",
                    "Auto Resolution", "Reasoning"
                ]
                for ci, col_name in enumerate(conflict_cols, start=1):
                    cell = wc.cell(row=1, column=ci, value=col_name)
                    cell.fill = conflict_header_fill
                    cell.font = conflict_header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = cell_border
                for ri, cc in enumerate(conflicts, start=2):
                    sa = cc.get("source_a", {})
                    sb = cc.get("source_b", {})
                    row_vals = [
                        cc.get("field", ""),
                        sa.get("source_document", ""),
                        sa.get("value", ""),
                        sb.get("source_document", ""),
                        sb.get("value", ""),
                        cc.get("auto_resolution", "NEEDS_HUMAN"),
                        cc.get("auto_reasoning", "")
                    ]
                    for ci, val in enumerate(row_vals, start=1):
                        c = wc.cell(row=ri, column=ci, value=str(val))
                        c.font = Font(name="Segoe UI", size=9)
                        c.fill = warn_fill
                        c.border = cell_border
                        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                for col in wc.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    wc.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

            wb.save(file_path)

            # Special column config for CombinedMappingPanel
            suggested_special = {
                "Assy #": {"method": "map", "source_column": "Assy#", "value": ""},
                "Assy Model": {"method": "map", "source_column": "Assy Model", "value": ""},
                "Assy Rev": {"method": "map", "source_column": "Assy Rev", "value": ""}
            }

            # Standard and Multi-Source Mapping for CombinedMappingPanel
            suggested_mapping = {
                "Line": "Line Item",
                "Part": "Part",
                "Description": "Description",
                "QTY": "Qty",
                "UOM": "UOM",
                "MFR": ["MFR"],
                "MPN": ["MPN"]
            }

            return {
                "success": True,
                "file_path": file_path,
                "customer_name": customer,
                "rfq_number": rfq_num,
                "project_title": proj_title,
                "commodity": commodity,
                "target_price": str(target_price),
                "eau": eau,
                "assembly_count": len(assemblies),
                "total_items": total_items,
                "suggested_mapping": suggested_mapping,
                "suggested_special": suggested_special,
                "error": None
            }

        except Exception as e:
            return {
                "success": False,
                "file_path": None,
                "error": str(e)
            }


if __name__ == "__main__":
    generator = SyntheticBOMGenerator()
    from agents.multimodal_extractor import MultimodalExtractor
    extractor = MultimodalExtractor()
    staging_tecan = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'ContXs', 'EmailStaging', '47456_Enquiry___Cable___Tecan_-_RS25')
    attachments = []
    if os.path.exists(staging_tecan):
        for f in os.listdir(staging_tecan):
            attachments.append({"filename": f, "path": os.path.join(staging_tecan, f)})

    sample_email = {
        "id": "47456",
        "subject": "Enquiry ~ Cable _ Tecan - RS25-8099",
        "sender": "jessiekong@radysis-asia.com",
        "date": "2026-08-13 09:52:12",
        "body": """
Tecan Article No	Rev. Level	Description	Estimated annual volume quantity	TP USD
30079632	5	HARNESS 2 VISION	900	73.12
30078993	2	HARNESS 3 VISION (MAX)	1000	53.90
30059436	1	SENSOR LIGHT ASSY SICK WLG4S-3W1132	300	61.10
30059453	2	CABLE RSN 9*0.34MM2 1M DSUB9F-DSUB9M	500	10.30
30077977	0	CABLE RSN 10*0.5MM2 1.75M 2*MFT2*5F	200	7.23
        """,
        "attachments": attachments
    }

    rfq_json = extractor.extract_full_rfq(sample_email)
    res = generator.generate_synthetic_excel(rfq_json)
    print("Generated Synthetic Excel:", res["file_path"])
    print("Total rows written:", res["total_items"])
