# ==============================================================================
# --- ContinuumX Multimodal Drawing & Document Extractor ---
# Source-agnostic multi-document entity resolver with Evidence Graph generation,
# conflict candidate detection, and multi-assembly BOM merging.
# ==============================================================================

import os
import sys
import re
import json
import configparser
import urllib.request
from datetime import datetime

# Standardized Base Directory Resolution Pattern
if getattr(sys, 'frozen', False) or hasattr(sys, '_MEIPASS') or "__compiled__" in globals():
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
elif '__file__' in globals():
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
elif len(sys.argv) > 0 and sys.argv[0] and sys.argv[0] != '-c':
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(sys.argv[0])))
else:
    BASE_DIR = os.getcwd()

if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from agents.evidence_schema import ResolutionType, make_evidence, ConflictCandidate
from agents.drawing_agent import DrawingVisionAgent

class MultimodalExtractor:
    """
    Source-Agnostic Multi-Document RFQ Extractor.
    Resolves commercial information and engineering BOM items across emails, drawings, and spreadsheets.
    """
    def __init__(self, config_path=None):
        if not config_path:
            config_path = os.path.join(BASE_DIR, "config.ini")

        self.config_path = config_path
        self.config = configparser.ConfigParser()
        self._load_config()

    def _load_config(self):
        if os.path.exists(self.config_path):
            try:
                self.config.read(self.config_path, encoding='utf-8')
            except Exception as e:
                print(f"[MultimodalExtractor] Warning reading config.ini: {e}")

    @property
    def _gemini_api_key(self):
        """Reads Gemini API key from environment or config.ini."""
        key = os.environ.get("GEMINI_API_KEY", "").strip()
        if not key and "AGENTS_LLM" in self.config:
            key = self.config["AGENTS_LLM"].get("gemini_api_key", "").strip()
        return key

    def _llm_extract_components(self, raw_block, assy_context="", doc_hint=""):
        """
        Uses Gemini LLM at runtime to parse any raw component text block from
        an email body (any format, any manufacturers) into structured JSON.

        Phase 5: Loads past user corrections from CorrectionStore and injects
        them as few-shot examples in the prompt for improved accuracy.

        Returns a list of component dicts, or [] on failure.
        """
        api_key = self._gemini_api_key
        if not api_key:
            return []

        # Phase 5: Load correction context
        correction_block = ""
        try:
            from agents.correction_store import CorrectionStore
            cs = CorrectionStore()
            hints = [doc_hint] if doc_hint else []
            if assy_context:
                # Try extracting a part number hint from assy_context
                import re as _re
                pn_m = _re.search(r'\b([0-9]{6,10})\b', assy_context)
                if pn_m:
                    hints.append(pn_m.group(1))
            corrections = []
            for hint in hints:
                corrections = cs.get_relevant_corrections(hint, ["mpn", "mfr", "description", "qty", "uom"])
                if corrections:
                    break
            if corrections:
                lines = []
                for c in corrections:
                    lines.append(
                        f"  - Field '{c['field']}': was {c['wrong_value']!r} → correct is {c['correct_value']!r}"
                        + (f" (MFR: {c['mfr']})" if c.get('mfr') else "")
                        + (f" — Note: {c['note']}" if c.get('note') else "")
                    )
                correction_block = "\n\nPast user corrections for similar documents — apply these:\n" + "\n".join(lines) + "\n"
        except Exception:
            pass

        system_prompt = (
            "You are a Manufacturing BOM data extraction specialist. "
            "Given a raw text block from an email (may contain wire harness components, "
            "PCB parts, connectors, terminals, contacts, or any other electronic/mechanical parts), "
            "extract every distinct component line and return ONLY a JSON array. "
            "Each element must have these exact keys: "
            '"mpn" (manufacturer part number or part identifier, string), '
            '"mfr" (manufacturer name, string, empty string if unknown), '
            '"description" (part description, string), '
            '"qty" (quantity as a number), '
            '"uom" (unit of measure, e.g. EA, IN, M, PCS, default EA). '
            "Do NOT include any explanation. Return ONLY the JSON array, no markdown, no code fences."
        )

        user_prompt = (
            f"Assembly context: {assy_context}{correction_block}\n\n"
            f"Raw component text block:\n{raw_block}\n\n"
            "Extract all components as a JSON array."
        )

        from agents.llm_gateway import LLMGateway
        components, status = LLMGateway(self.config_path).generate_json(
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            doc_name=f"EmailBodyComponents:{doc_hint}"
        )
        if isinstance(components, list):
            return components
        return []


    @staticmethod
    def _parse_eau_value(raw_val):
        """Converts strings like '3-5k', '5k', '10,000', '2640', '900' to standard integer, or None if absent."""
        if not raw_val:
            return None
        raw_str = str(raw_val).strip().lower()
        try:
            range_m = re.search(r'(\d+)\s*[-~]\s*(\d+)\s*k', raw_str)
            if range_m:
                return int(range_m.group(2)) * 1000

            k_m = re.search(r'(\d+(?:\.\d+)?)\s*k', raw_str)
            if k_m:
                return int(float(k_m.group(1)) * 1000)

            digits = re.sub(r'[^0-9]', '', raw_str)
            if digits:
                return int(digits)
        except Exception:
            pass
        return None

    @classmethod
    def _create_evidence(cls, value, res_type, source_doc, page=1, zone="METADATA", snippet="", reasoning="", confidence=0.95):
        """Thin wrapper — delegates to shared make_evidence factory."""
        return make_evidence(value, res_type, source_doc, page=page, zone=zone,
                             snippet=snippet, reasoning=reasoning, confidence=confidence)

    def _extract_commercial_params_source_agnostic(self, subject, body, sender="", attachments=None):
        """
        Dynamically extracts Customer, RFQ Number, Project, EAU, and Target Price
        from subject, body text, or attached spreadsheets/documents.
        """
        full_text = f"{subject}\n{body}"
        params = {}
        evidence = {}
        doc_source = "email_header_and_body"

        # 1. Customer Name Resolution
        cust = ""
        cust_snip = ""
        end_cust_m = re.search(r'(?:end\s+customer|customer|client)[:\s]+([^\n\r,;]+)', body, re.IGNORECASE)
        if end_cust_m:
            cust = end_cust_m.group(1).strip()
            cust_snip = end_cust_m.group(0)

        if not cust:
            for known in ["Tecan", "Graco", "Honeywell", "Eastek", "Radysis", "Intel", "Plexus", "Celestica", "Jabil"]:
                if re.search(r'\b' + re.escape(known) + r'\b', full_text, re.I):
                    cust = known
                    cust_snip = f"Matched known customer keyword: {known}"
                    break

        if not cust:
            subj_parts = re.split(r'[-_~]', subject)
            if len(subj_parts) >= 2:
                for part in subj_parts:
                    p_clean = part.strip()
                    if p_clean and not re.search(r'\b(?:rfq|enquiry|cable|for\s+localization|rs25|rs26)\b', p_clean, re.I):
                        cust = p_clean
                        cust_snip = p_clean
                        break

        resolved_cust = cust or "Customer"
        params["customer_name"] = resolved_cust
        evidence["customer_name"] = self._create_evidence(resolved_cust, ResolutionType.DIRECT, doc_source, snippet=cust_snip)

        # 2. RFQ Number Resolution
        rfq_num = ""
        rfq_snip = ""
        rfq_code_m = re.search(r'\b(RS2[4567]-[\d]{4,5})\b', full_text, re.IGNORECASE)
        if rfq_code_m:
            rfq_num = rfq_code_m.group(1).upper()
            rfq_snip = rfq_code_m.group(0)
        else:
            rfq_gen_m = re.search(r'\brfq[-_\s#:]*([A-Za-z0-9_-]+)', subject, re.IGNORECASE)
            if rfq_gen_m:
                val = rfq_gen_m.group(1).strip().upper()
                if len(val) >= 3 and val not in ["FOR", "NEW", "THE", "PLEASE"]:
                    rfq_num = val
                    rfq_snip = rfq_gen_m.group(0)

        if not rfq_num:
            rfq_num = f"RFQ-{datetime.now().strftime('%m%d%H%M')}"
            rfq_snip = "Generated timestamp fallback"

        params["rfq_number"] = rfq_num
        evidence["rfq_number"] = self._create_evidence(rfq_num, ResolutionType.DIRECT, doc_source, snippet=rfq_snip)

        # 3. Project Title Resolution
        proj_title = ""
        proj_m = re.search(r'project\s+(?:name|title)[:\s]+([^\n\r]+)', body, re.IGNORECASE)
        if proj_m:
            proj_title = proj_m.group(1).strip()
        else:
            proj_title = subject[:50]
        params["project_title"] = proj_title
        evidence["project_title"] = self._create_evidence(proj_title, ResolutionType.DIRECT, doc_source)

        # 4. EAU Resolution & Ambiguity Detection
        eau_val = None
        eau_snip = ""
        is_ambiguous_eau = False
        eau_options = []
        raw_eau_text = ""

        eau_m = re.search(r'(?:annual\s+consumption(?:\s*\(eau\))?|eau|estimated\s+annual\s+volume\s+quantity|estimated\s+annual\s+usage)[:\s]+([^\n\r]+)', body, re.IGNORECASE)
        if eau_m:
            raw_eau_text = eau_m.group(1).strip()
            eau_snip = eau_m.group(0)
            
            # Check for range: e.g. 3-5k, 3000-5000
            range_m = re.search(r'(\d+(?:\.\d+)?)\s*[-~]\s*(\d+(?:\.\d+)?)\s*(k)?', raw_eau_text, re.I)
            if range_m:
                is_ambiguous_eau = True
                mult = 1000 if (range_m.group(3) or 'k' in raw_eau_text.lower()) else 1
                v_min = int(float(range_m.group(1)) * mult)
                v_max = int(float(range_m.group(2)) * mult)
                v_mid = int((v_min + v_max) / 2)
                eau_options = [v_min, v_mid, v_max]
                eau_val = v_max  # default max fallback
            else:
                eau_val = self._parse_eau_value(raw_eau_text)

        fc_m = re.search(r'forecast[:\s]+(\d+)', body, re.IGNORECASE)
        if fc_m:
            fc_val = int(fc_m.group(1))
            if fc_val not in eau_options:
                eau_options.insert(0, fc_val)
            if not eau_val:
                eau_val = fc_val
                eau_snip = fc_m.group(0)

        resolved_eau = eau_val  # Keep None if not specified in email or attachments
        params["eau"] = resolved_eau
        params["eau_ambiguity"] = {
            "is_ambiguous": is_ambiguous_eau,
            "raw_text": raw_eau_text or (str(resolved_eau) if resolved_eau is not None else ""),
            "suggested_options": eau_options if eau_options else ([resolved_eau] if resolved_eau is not None else [])
        }
        evidence["eau"] = self._create_evidence(resolved_eau, ResolutionType.DIRECT if eau_val else ResolutionType.NOT_AVAILABLE, doc_source, snippet=eau_snip)

        # 5. Target Price Resolution
        target_price = "Not Specified"
        tp_snip = ""
        dollar_m = re.search(r'\$\s*([0-9]+(?:\.[0-9]{2,4})?)', body)
        if dollar_m:
            target_price = f"${dollar_m.group(1)}"
            tp_snip = dollar_m.group(0)
        else:
            tp_m = re.search(r'(?:target\s+price|tp\s+usd)[^\n\r\d$]{0,15}[:=]?\s*(\$\s*[\d.,]+|\b\d+\.\d{2}\b)', body, re.IGNORECASE)
            if tp_m:
                val = tp_m.group(1).strip()
                target_price = val if val.startswith('$') else f"${val}"
                tp_snip = tp_m.group(0)

        params["target_price"] = target_price
        evidence["target_price"] = self._create_evidence(target_price, ResolutionType.DIRECT if target_price != "Not Specified" else ResolutionType.NOT_AVAILABLE, doc_source, snippet=tp_snip)

        # 6. Commodity
        params["commodity"] = "Wire Harness" if any(k in full_text.lower() for k in ["wire", "harness", "cable", "awg", "terminal", "housing", "jst", "molex"]) else ("PCBA" if "pcba" in full_text.lower() else "Wire Harness")

        return params, evidence

    def _parse_embedded_component_lines(self, mpn_block, desc_context="", source_doc="email_body"):
        """
        Source-agnostic component line parser.
        Step 1: Calls Gemini LLM at runtime to identify any manufacturer/part/description.
        Step 2: Falls back to a generic heuristic if LLM is unavailable.
        No manufacturer names, part numbers, or patterns are hardcoded.
        """
        items = []

        # --- Step 1: LLM-powered extraction (runtime, generalizes to any customer/manufacturer) ---
        llm_components = self._llm_extract_components(mpn_block, assy_context=desc_context)
        if llm_components:
            for idx, comp in enumerate(llm_components, start=1):
                mpn = str(comp.get("mpn", "")).strip()
                mfr = str(comp.get("mfr", "")).strip()
                desc = str(comp.get("description", "Component")).strip()
                try:
                    qty = float(comp.get("qty", 1))
                except (ValueError, TypeError):
                    qty = 1
                uom = str(comp.get("uom", "EA")).strip().upper() or "EA"
                if not mpn:
                    continue
                items.append({
                    "line_item": idx,
                    "part_number": mpn,
                    "description": desc,
                    "mfr": mfr,
                    "mpn": mpn,
                    "qty": qty,
                    "uom": uom,
                    "evidence": {
                        "part": self._create_evidence(mpn, ResolutionType.DIRECT, source_doc, snippet=mpn),
                        "mpn": self._create_evidence(mpn, ResolutionType.DIRECT, source_doc, snippet=mpn),
                        "mfr": self._create_evidence(mfr, ResolutionType.DIRECT if mfr else ResolutionType.NOT_AVAILABLE, source_doc),
                        "qty": self._create_evidence(qty, ResolutionType.DIRECT, source_doc, snippet=f"Qty: {qty} {uom}"),
                        "uom": self._create_evidence(uom, ResolutionType.DIRECT, source_doc)
                    }
                })
            return items

        # --- Step 2: Generic heuristic fallback (no hardcoded manufacturer names) ---
        print("[MultimodalExtractor] LLM unavailable — using generic heuristic fallback for component parsing.")
        skip_keywords = ["wire assembly", "total", "drawing", "please refer", "note:", "remark"]
        lines = [l.strip() for l in mpn_block.splitlines() if l.strip() and not l.strip().startswith("*")]
        for line in lines:
            if any(k in line.lower() for k in skip_keywords):
                continue

            mfr = ""
            mpn = line
            desc = "Component"
            qty = 1
            uom = "EA"

            # Generic: try to detect "MFR: MPN" or "MFR - MPN" separator patterns
            sep_m = re.match(r'^([A-Za-z][A-Za-z0-9 .&]{1,30})\s*[:–-]\s*([A-Z0-9][\w\-./]+)', line)
            if sep_m:
                mfr = sep_m.group(1).strip()
                mpn = sep_m.group(2).strip()
            else:
                # If line looks like a standalone part number (alphanumeric with dashes)
                pn_m = re.match(r'^([A-Z0-9][\w\-./]{3,})\s*(.*)', line, re.IGNORECASE)
                if pn_m:
                    mpn = pn_m.group(1).strip()
                    desc = pn_m.group(2).strip() or "Component"

            items.append({
                "line_item": len(items) + 1,
                "part_number": mpn,
                "description": desc,
                "mfr": mfr,
                "mpn": mpn,
                "qty": qty,
                "uom": uom,
                "evidence": {
                    "part": self._create_evidence(mpn, ResolutionType.DIRECT, source_doc, snippet=line),
                    "mpn": self._create_evidence(mpn, ResolutionType.DIRECT, source_doc, snippet=line),
                    "mfr": self._create_evidence(mfr, ResolutionType.DIRECT if mfr else ResolutionType.NOT_AVAILABLE, source_doc),
                    "qty": self._create_evidence(qty, ResolutionType.DIRECT, source_doc, snippet=f"Qty: {qty} {uom}"),
                    "uom": self._create_evidence(uom, ResolutionType.DIRECT, source_doc)
                }
            })
        return items

    def _parse_email_summary_table(self, body, source_doc="email_body"):
        """
        Parses RFQ tables in email bodies (supporting Graco, Tecan, and standard multi-assembly tables).
        Supports both embedded sub-component blocks and single-line / multi-line summary formats.
        """
        assemblies = []
        normalized_body = body.replace('\r\n', '\n').replace('\r', '\n')

        # 1. Graco Multi-Component Embedded Table Pattern
        # ID \n+ DESC \n+ (MULTILINE MPN BLOCK) \n+ FORECAST \n+ TARGET PRICE
        g_pattern = re.finditer(
            r'(\b\d{3}-\d{6}-\d{3}-\d{2}[A-Za-z0-9]?)\s*\n+([^\n\r]+?)\s*\n+((?:[^\n\r]+\n+)+?)\s*(\d{2,6})\s*\n+\$?\s*([\d.]+)',
            normalized_body
        )
        for m in g_pattern:
            assy_id, desc, mpn_block, forecast, tp = m.groups()
            if "FORECAST" in assy_id or "TARGET PRICE" in assy_id:
                continue
            items = self._parse_embedded_component_lines(mpn_block, desc, source_doc)
            if items:
                assemblies.append({
                    "assy_no": assy_id.strip(),
                    "assy_rev": assy_id.split("-")[-1] if "-" in assy_id else "00",
                    "assy_model": desc.strip(),
                    "eau": int(forecast.strip()),
                    "target_price": f"${float(tp.strip()):.2f}",
                    "items": items,
                    "evidence": {
                        "assy_no": self._create_evidence(assy_id.strip(), ResolutionType.DIRECT, source_doc, snippet=assy_id),
                        "assy_rev": self._create_evidence(assy_id.split("-")[-1] if "-" in assy_id else "00", ResolutionType.DIRECT, source_doc),
                        "eau": self._create_evidence(int(forecast.strip()), ResolutionType.DIRECT, source_doc, snippet=f"Forecast: {forecast}"),
                        "target_price": self._create_evidence(f"${float(tp.strip()):.2f}", ResolutionType.DIRECT, source_doc, snippet=f"TP: ${tp}")
                    }
                })

        # 2. Tecan Multi-line pattern: Article \n+ Rev \n+ Desc \n+ EAU \n+ TP
        if not assemblies:
            m_pattern = re.finditer(
                r'\b([0-9]{6,10}|[0-9]{3,4}-[0-9]{5,7}-[0-9]{3}-[0-9]{2}[A-Za-z0-9]?)\s*\n+([0-9]{1,2}|Rev\s*[A-Z0-9]+)\s*\n+([A-Za-z0-9\s()_*.-]+?)\s*\n+([0-9]{2,6})\s*\n+([$\d.,]+)',
                normalized_body
            )
            for m in m_pattern:
                art_no, rev, desc, eau, tp_val = m.groups()
                art_no = art_no.strip()
                rev = rev.replace("Rev", "").strip()
                desc = desc.strip()
                eau_num = int(eau.strip())
                tp_clean_str = tp_val.strip().replace('$', '')
                try:
                    tp_clean = f"${float(tp_clean_str):.2f}"
                except Exception:
                    tp_clean = f"${tp_clean_str}"

                assemblies.append({
                    "assy_no": art_no,
                    "assy_rev": rev,
                    "assy_model": desc,
                    "eau": eau_num,
                    "target_price": tp_clean,
                    "items": [],
                    "evidence": {
                        "assy_no": self._create_evidence(art_no, ResolutionType.DIRECT, source_doc, snippet=m.group(0)),
                        "assy_rev": self._create_evidence(rev, ResolutionType.DIRECT, source_doc, snippet=f"Rev: {rev}"),
                        "eau": self._create_evidence(eau_num, ResolutionType.DIRECT, source_doc, snippet=f"EAU: {eau_num}"),
                        "target_price": self._create_evidence(tp_clean, ResolutionType.DIRECT, source_doc, snippet=f"TP: {tp_clean}")
                    }
                })

        # 3. Single-line pattern fallback
        if not assemblies:
            for line in normalized_body.splitlines():
                row_m = re.search(r'\b([0-9]{6,10}|[0-9]{3,4}-[0-9]{5,7}-[0-9]{3}-[0-9]{2}[A-Za-z0-9]?)\s+(\d{1,2}|Rev\s*[A-Z0-9]+)\s+([A-Za-z0-9\s()_*.-]+?)\s+(\d{2,6})\s+([$\d.,]+)', line)
                if row_m:
                    art_no = row_m.group(1).strip()
                    rev = row_m.group(2).replace("Rev", "").strip()
                    desc = row_m.group(3).strip()
                    eau_num = int(row_m.group(4))
                    tp_val = row_m.group(5).strip().replace('$', '')
                    tp_clean = f"${float(tp_val):.2f}" if re.match(r'^\d+(\.\d+)?$', tp_val) else f"${tp_val}"

                    assemblies.append({
                        "assy_no": art_no,
                        "assy_rev": rev,
                        "assy_model": desc,
                        "eau": eau_num,
                        "target_price": tp_clean,
                        "items": [],
                        "evidence": {
                            "assy_no": self._create_evidence(art_no, ResolutionType.DIRECT, source_doc, snippet=row_m.group(0)),
                            "assy_rev": self._create_evidence(rev, ResolutionType.DIRECT, source_doc, snippet=f"Rev: {rev}"),
                            "eau": self._create_evidence(eau_num, ResolutionType.DIRECT, source_doc, snippet=f"EAU: {eau_num}"),
                            "target_price": self._create_evidence(tp_clean, ResolutionType.DIRECT, source_doc, snippet=f"TP: {tp_clean}")
                        }
                    })

        return assemblies

    def _parse_bom_xls_file(self, xls_path, assy_eau=None):
        """Extracts exact engineering components from an unzipped BOM_<assy>.xls sheet."""
        items = []
        try:
            import xlrd
            wb = xlrd.open_workbook(xls_path)
            ws = wb.sheet_by_index(0)
            
            header_map = {}
            for c in range(ws.ncols):
                h = str(ws.cell_value(0, c)).strip().lower()
                if h == "component number":
                    header_map["part_number"] = c
                elif h in ("comp. qty (cun)", "comp. qty", "quantity", "qty"):
                    header_map["qty"] = c
                elif h in ("component unit", "unit", "uom"):
                    header_map["uom"] = c
                elif h in ("object description", "description") and "desc" not in header_map:
                    header_map["desc"] = c
                elif "vendor" in h and "vendor material" not in h:
                    header_map["mfr"] = c

            doc_fn = os.path.basename(xls_path)
            for r in range(1, ws.nrows):
                exp = str(ws.cell_value(r, 0)).strip()
                p_no = str(ws.cell_value(r, header_map.get("part_number", 4))).strip() if "part_number" in header_map else ""
                raw_qty = ws.cell_value(r, header_map.get("qty", 2)) if "qty" in header_map else 1
                uom_val = str(ws.cell_value(r, header_map.get("uom", 3))).strip() if "uom" in header_map else "EA"
                desc_val = str(ws.cell_value(r, header_map.get("desc", 13))).strip() if "desc" in header_map else ""
                
                if p_no and p_no != "None" and exp.startswith('.'):
                    try:
                        q_num = float(raw_qty)
                    except Exception:
                        q_num = 1.0
                    
                    mfr_val = ""
                    for m in ["Molex", "JST", "Heiniger", "Alpha", "TE Connectivity", "Sick", "FCI", "3M"]:
                        if m.lower() in desc_val.lower():
                            mfr_val = m
                            break

                    items.append({
                        "line_item": len(items) + 1,
                        "part_number": p_no,
                        "description": desc_val or f"Component {p_no}",
                        "mfr": mfr_val,
                        "mpn": p_no,
                        "qty": q_num,
                        "uom": uom_val or "EA",
                        "eau": assy_eau,
                        "evidence": {
                            "part": self._create_evidence(p_no, ResolutionType.DIRECT, doc_fn),
                            "mpn": self._create_evidence(p_no, ResolutionType.DIRECT, doc_fn),
                            "mfr": self._create_evidence(mfr_val, ResolutionType.DIRECT if mfr_val else ResolutionType.NOT_AVAILABLE, doc_fn),
                            "qty": self._create_evidence(q_num, ResolutionType.DIRECT, doc_fn),
                            "uom": self._create_evidence(uom_val, ResolutionType.DIRECT, doc_fn)
                        }
                    })
        except Exception as e:
            print(f"[MultimodalExtractor] BOM xls parse error on {xls_path}: {e}")
        return items

    def _create_evidence(self, value, resolution_type, source_doc, zone="METADATA", snippet=None, reasoning=None):
        return make_evidence(value, resolution_type, source_doc, zone=zone, snippet=snippet, reasoning=reasoning)

    def _detect_conflicts(self, field, val_a, source_a_doc, val_b, source_b_doc, zone="METADATA"):
        """
        Detects a conflict between two source values for the same field.
        If both values are non-null and different -> returns a ConflictCandidate.
        If one is None/empty -> no conflict (accept the non-null).

        Args:
            field:       Field name (e.g. 'assy_model', 'assy_rev', 'eau')
            val_a:       Value from source A (e.g. email)
            source_a_doc: Source A document label
            val_b:       Value from source B (e.g. drawing)
            source_b_doc: Source B document label
            zone:        Evidence zone for both sources

        Returns:
            ConflictCandidate or None
        """
        a_clean = str(val_a).strip() if val_a is not None else ""
        b_clean = str(val_b).strip() if val_b is not None else ""

        if not a_clean or not b_clean:
            return None  # One side missing — no conflict
        if a_clean.lower() == b_clean.lower():
            return None  # Values agree

        # Drawing is usually more authoritative for technical fields
        if field in ("assy_no", "assy_rev", "assy_model"):
            auto_res = "USE_B"  # Prefer drawing
            reasoning = f"Drawing value '{b_clean}' preferred over email value '{a_clean}' for technical field '{field}'"
        else:
            auto_res = "USE_A"  # Prefer email for commercial fields
            reasoning = f"Email value '{a_clean}' preferred over drawing value '{b_clean}' for commercial field '{field}'"

        ev_a = make_evidence(a_clean, ResolutionType.DIRECT, source_a_doc, zone=zone, snippet=a_clean)
        ev_b = make_evidence(b_clean, ResolutionType.DIRECT, source_b_doc, zone=zone, snippet=b_clean)
        return ConflictCandidate(
            field=field,
            source_a=ev_a,
            source_b=ev_b,
            auto_resolution=auto_res,
            auto_reasoning=reasoning
        )

    def extract_full_rfq(self, email_data, progress_callback=None):
        """
        Consolidates email text, Excel spreadsheets, and PDF drawing blueprints
        into a standardized, complete intermediate RFQ JSON object with Evidence.
        """
        subject = email_data.get("subject", "")
        body = email_data.get("body", "")
        sender = email_data.get("sender", "")
        date_str = email_data.get("date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        attachments = email_data.get("attachments", [])

        # 1. Source-Agnostic Commercial Parameters
        params, meta_evidence = self._extract_commercial_params_source_agnostic(subject, body, sender, attachments)

        # 1.5 Customer Identification & Pattern Loading
        from agents.customer_profile_store import CustomerProfileStore
        cust_store = CustomerProfileStore()
        att_filenames = [att.get("filename", "") if isinstance(att, dict) else os.path.basename(str(att)) for att in attachments]
        cust_ident = cust_store.identify_customer(subject=subject, body=body, sender=sender, filenames=att_filenames)
        if cust_ident.get("customer_name") and cust_ident["customer_name"] not in ("Customer", "New Customer"):
            params["customer_name"] = cust_ident["customer_name"]
        resolved_cust = params.get("customer_name") or cust_ident.get("customer_name") or "Customer"

        # 2.5 Expand ZIP/archive attachments safely (never unpack .xlsx/.docx files)
        expanded_attachments = []
        import zipfile
        office_exts = ('.xlsx', '.xls', '.docx', '.pptx', '.doc', '.ppt')
        
        for att in attachments:
            fp = att.get("path", "") if isinstance(att, dict) else str(att)
            fn = att.get("filename", "") if isinstance(att, dict) else os.path.basename(fp)
            is_office = fn.lower().endswith(office_exts)
            
            # Preserve original attachment in list
            expanded_attachments.append(att if isinstance(att, dict) else {"filename": fn, "path": fp})
            
            if fp and os.path.exists(fp) and not is_office and (fn.lower().endswith((".zip", ".tar", ".gz", ".7z")) or zipfile.is_zipfile(fp)):
                try:
                    extract_dir = os.path.join(os.path.dirname(fp), f"unzipped_{os.path.splitext(fn)[0]}")
                    os.makedirs(extract_dir, exist_ok=True)
                    with zipfile.ZipFile(fp, 'r') as zf:
                        for member in zf.namelist():
                            target_out = os.path.join(extract_dir, member)
                            if not os.path.exists(target_out):
                                try:
                                    zf.extract(member, extract_dir)
                                except Exception:
                                    pass

                    for root, dirs, files in os.walk(extract_dir):
                        for f in files:
                            sub_fp = os.path.join(root, f)
                            expanded_attachments.append({
                                "filename": f,
                                "path": sub_fp,
                                "content_type": "application/pdf" if f.lower().endswith(".pdf") else "application/octet-stream",
                                "size_bytes": os.path.getsize(sub_fp),
                                "extracted_from_zip": fn
                            })
                except Exception as z_ex:
                    print(f"[MultimodalExtractor] Zip unpack notice for {fn}: {z_ex}")

        attachments = expanded_attachments

        # 2. Extract Assemblies: Prioritize Authoritative Excel RFQ Sheet (e.g. Cable_RFQ_AEI.xlsx)
        email_table_assemblies = []
        
        for att in attachments:
            fp = att.get("path", "") if isinstance(att, dict) else str(att)
            fn = att.get("filename", "") if isinstance(att, dict) else os.path.basename(fp)
            
            # Robust fallback: If recorded path does not exist, search staging directories
            if not fp or not os.path.exists(fp):
                stg_root = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'ContXs', 'EmailStaging')
                if os.path.exists(stg_root):
                    for root, _, files in os.walk(stg_root):
                        if fn in files:
                            cand_p = os.path.join(root, fn)
                            if os.path.isfile(cand_p) and os.path.getsize(cand_p) > 0:
                                fp = cand_p
                                break

            if fp and os.path.exists(fp) and fn.lower().endswith(('.xlsx', '.xls')) and not fn.lower().startswith('synthetic_') and not fn.lower().startswith('bom_'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(fp, data_only=True)
                    excel_assemblies = []
                    
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        headers = []
                        header_row_idx = None
                        
                        # Scan rows 1 through 10 for column header row
                        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                            if r_idx > 12:
                                break
                            row_strs = [" ".join(str(c).replace('\n', ' ').replace('\r', ' ').split()).strip().lower() for c in row if c is not None]
                            if any(k in s for s in row_strs for k in ["article", "part no", "item no", "part number", "assembly"]) and \
                               any(k in s for s in row_strs for k in ["rev", "status", "description", "volume", "quantity", "eau", "drawing", "bom"]):
                                headers = [" ".join(str(c).replace('\n', ' ').replace('\r', ' ').split()).strip() if c is not None else "" for c in row]
                                header_row_idx = r_idx
                                break

                        if header_row_idx:
                            col_map = {}
                            for c_idx, h in enumerate(headers):
                                h_low = h.lower()
                                if any(k in h_low for k in ["article", "item", "part no", "part number", "assembly"]):
                                    col_map["assy_no"] = c_idx
                                elif any(k in h_low for k in ["rev", "level"]):
                                    col_map["assy_rev"] = c_idx
                                elif any(k in h_low for k in ["desc", "description", "model", "title"]):
                                    col_map["assy_model"] = c_idx
                                elif any(k in h_low for k in ["annual", "volume", "quantity", "eau", "consumption", "forecast"]):
                                    col_map["eau"] = c_idx
                                elif "drawing" in h_low:
                                    col_map["drawing"] = c_idx

                            for row in list(ws.iter_rows(values_only=True))[header_row_idx:]:
                                if not any(row):
                                    continue
                                raw_no = row[col_map["assy_no"]] if "assy_no" in col_map and col_map["assy_no"] < len(row) else None
                                if not raw_no or str(raw_no).strip() in ("", "None"):
                                    continue
                                art_no = str(raw_no).strip()
                                rev = str(row[col_map["assy_rev"]]).replace("Rev", "").strip() if "assy_rev" in col_map and col_map["assy_rev"] < len(row) and row[col_map["assy_rev"]] is not None else "00"
                                desc = str(row[col_map["assy_model"]]).strip() if "assy_model" in col_map and col_map["assy_model"] < len(row) and row[col_map["assy_model"]] is not None else art_no
                                
                                # Parse pure numeric EAU per assembly
                                eau_raw = row[col_map["eau"]] if "eau" in col_map and col_map["eau"] < len(row) else None
                                eau_num = params["eau"]
                                if eau_raw is not None and str(eau_raw).strip() not in ("", "None"):
                                    try:
                                        clean_e = re.sub(r'[^\d.]', '', str(eau_raw))
                                        if clean_e:
                                            eau_num = int(float(clean_e))
                                    except Exception:
                                        pass

                                dwg_link = str(row[col_map["drawing"]]).strip() if "drawing" in col_map and col_map["drawing"] < len(row) and row[col_map["drawing"]] is not None else ""

                                # Step 2a: Check for child engineering BOM xls file in unzipped attachments
                                assy_items = []
                                for b_att in attachments:
                                    b_fp = b_att.get("path", "") if isinstance(b_att, dict) else str(b_att)
                                    b_fn = os.path.basename(b_fp).lower()
                                    if art_no.lower() in b_fn and b_fn.endswith(('.xls', '.xlsx')) and "cable_rfq" not in b_fn:
                                        assy_items = self._parse_bom_xls_file(b_fp, assy_eau=eau_num)
                                        if assy_items:
                                            break

                                # Step 2b: Multi-Source Drawing Reconciliation (Check for drawing callouts / extra parts)
                                drawing_extra_items = []
                                for d_att in attachments:
                                    d_fp = d_att.get("path", "") if isinstance(d_att, dict) else str(d_att)
                                    d_fn = os.path.basename(d_fp).lower()
                                    if art_no.lower() in d_fn and d_fn.endswith('.pdf') and not d_fn.startswith(('bb0_', 'datenblatt_')):
                                        try:
                                            dwg_parsed = DrawingVisionAgent().parse_drawing_file(d_fp, use_vision=False)
                                            if dwg_parsed and dwg_parsed.get("items"):
                                                existing_pns = {str(it.get("part_number", "")).strip() for it in assy_items}
                                                for d_it in dwg_parsed["items"]:
                                                    d_pn = str(d_it.get("part_number", "")).strip()
                                                    if d_pn and d_pn not in existing_pns:
                                                        d_it_copy = dict(d_it)
                                                        d_it_copy["eau"] = eau_num
                                                        d_it_copy["description"] = f"{d_it.get('description', 'Component')} [From Drawing]"
                                                        drawing_extra_items.append(d_it_copy)
                                                        existing_pns.add(d_pn)
                                        except Exception:
                                            pass

                                combined_items = assy_items + drawing_extra_items
                                for idx, it in enumerate(combined_items, start=1):
                                    it["line_item"] = idx
                                    it["eau"] = eau_num

                                excel_assemblies.append({
                                    "assy_no": art_no,
                                    "assy_rev": rev,
                                    "assy_model": desc,
                                    "eau": eau_num,
                                    "target_price": params["target_price"],
                                    "drawing_link": dwg_link,
                                    "items": combined_items,
                                    "evidence": {
                                        "assy_no": self._create_evidence(art_no, ResolutionType.DIRECT, fn, snippet=f"{art_no} Rev {rev}"),
                                        "assy_rev": self._create_evidence(rev, ResolutionType.DIRECT, fn, snippet=f"Rev: {rev}"),
                                        "eau": self._create_evidence(eau_num, ResolutionType.DIRECT if eau_num else ResolutionType.NOT_AVAILABLE, fn, snippet=f"EAU: {eau_num}"),
                                        "target_price": self._create_evidence(params["target_price"], ResolutionType.DIRECT, fn)
                                    }
                                })

                    if excel_assemblies:
                        # Deduplicate by assy_no
                        seen_a = set()
                        deduped = []
                        for ea in excel_assemblies:
                            if ea["assy_no"] not in seen_a:
                                seen_a.add(ea["assy_no"])
                                deduped.append(ea)
                        if deduped:
                            email_table_assemblies = deduped
                            break
                except Exception as ex_err:
                    print(f"[MultimodalExtractor] Excel RFQ sheet parser notice: {ex_err}")

            if email_table_assemblies:
                break

        # If no Excel sheet found, fallback to email body summary table
        if not email_table_assemblies:
            email_table_assemblies = self._parse_email_summary_table(body)

        # 3. Technical PDF Drawing Parsing & Source Linking
        drawing_assemblies = []
        drawings_metadata = []

        for att in attachments:
            fp = att.get("path", "") if isinstance(att, dict) else str(att)
            fn = att.get("filename", "") if isinstance(att, dict) else os.path.basename(fp)
            fn_lower = fn.lower()
            if not fp or not os.path.exists(fp):
                continue

            # Skip non-drawing datasheets (do not skip BB0_ as Tecan uses BB0 for cable blueprints)
            is_component_datasheet = fn_lower.startswith(("datenblatt_", "datasheet_", "spec_sheet_"))
            if fn_lower.endswith(".pdf") or (fn_lower.endswith((".png", ".jpg", ".jpeg")) and not any(k in fn_lower for k in ["image0", "image1", "photo", "logo", "icon"])):
                if progress_callback:
                    progress_callback(f"📄 Mapping {fn}...")
                parsed_dwg = DrawingVisionAgent().parse_drawing_file(fp, progress_callback=progress_callback, use_vision=True)
                if parsed_dwg:
                    drawings_metadata.append({
                        "filename": fn,
                        "drawing_number": parsed_dwg.get("assy_no", fn),
                        "revision": parsed_dwg.get("assy_rev", "Rev A"),
                        "customer_part_number": parsed_dwg.get("assy_no", fn),
                        "title": parsed_dwg.get("assy_model", fn)
                    })
                    drawing_assemblies.append(parsed_dwg)

        # 4. Multi-Document Consolidation & Evidence Merging + Conflict Detection
        final_assemblies = []
        all_conflicts = []

        if email_table_assemblies:
            for e_assy in email_table_assemblies:
                matched_dwg = None
                for d_assy in drawing_assemblies:
                    e_no = str(e_assy.get("assy_no", ""))
                    d_no = str(d_assy.get("assy_no", ""))
                    d_src = str(d_assy.get("source_drawing", ""))
                    if (e_no and e_no == d_no) or (e_no and e_no in d_src):
                        matched_dwg = d_assy
                        break

                if matched_dwg and matched_dwg.get("items") and not e_assy.get("items"):
                    e_assy["items"] = matched_dwg["items"]

                if matched_dwg:
                    # Detect conflicts instead of silently overwriting
                    for field in ("assy_model", "assy_rev", "assy_no"):
                        conflict = self._detect_conflicts(
                            field,
                            e_assy.get(field), "email_summary_table",
                            matched_dwg.get(field), matched_dwg.get("source_drawing", "drawing.pdf")
                        )
                        if conflict:
                            all_conflicts.append(conflict)
                            if conflict.auto_resolution == "USE_B":
                                e_assy[field] = matched_dwg.get(field)
                        elif matched_dwg.get(field) and not e_assy.get(field):
                            e_assy[field] = matched_dwg.get(field)

                    if "evidence" in matched_dwg:
                        e_assy["evidence"].update(matched_dwg["evidence"])

                if not e_assy.get("items"):
                    e_assy["items"] = [{
                        "line_item": 1,
                        "part_number": e_assy["assy_no"],
                        "description": e_assy["assy_model"],
                        "mfr": None,
                        "mpn": None,
                        "qty": 1,
                        "uom": "EA",
                        "eau": e_assy.get("eau", params["eau"]),
                        "evidence": {
                            "part": self._create_evidence(e_assy["assy_no"], ResolutionType.DIRECT, "email_summary_table"),
                            "mpn": self._create_evidence(None, ResolutionType.NOT_AVAILABLE, "email_summary_table", reasoning="Summary table provides customer article number without MPN"),
                            "mfr": self._create_evidence(None, ResolutionType.NOT_AVAILABLE, "email_summary_table", reasoning="No manufacturer specified in table"),
                            "qty": self._create_evidence(1, ResolutionType.DIRECT, "email_summary_table"),
                            "uom": self._create_evidence("EA", ResolutionType.DIRECT, "email_summary_table")
                        }
                    }]
                final_assemblies.append(e_assy)

        elif drawing_assemblies:
            final_assemblies = drawing_assemblies
        elif email_table_assemblies:
            final_assemblies = email_table_assemblies
        else:
            final_assemblies = [{
                "assy_no": params["rfq_number"],
                "assy_rev": "Rev A",
                "assy_model": params["project_title"],
                "eau": params["eau"],
                "target_price": params["target_price"],
                "items": [{
                    "line_item": 1,
                    "part_number": params["rfq_number"],
                    "description": params["project_title"],
                    "mfr": None,
                    "mpn": params["rfq_number"],
                    "qty": 1,
                    "uom": "EA",
                    "evidence": {
                        "part": self._create_evidence(params["rfq_number"], ResolutionType.DIRECT, "email_subject"),
                        "mpn": self._create_evidence(params["rfq_number"], ResolutionType.DIRECT, "email_subject"),
                        "mfr": self._create_evidence(None, ResolutionType.NOT_AVAILABLE, "email_subject"),
                        "qty": self._create_evidence(1, ResolutionType.DIRECT, "email_subject"),
                        "uom": self._create_evidence("EA", ResolutionType.DIRECT, "email_subject")
                    }
                }],
                "evidence": {}
            }]

        # Step 3a: Check Human-Verified BOM Store for known assembly ground truths
        try:
            import copy
            from agents.verified_bom_store import VerifiedBOMStore
            vbs = VerifiedBOMStore()
            for assy in final_assemblies:
                a_no_k = str(assy.get("assy_no", "")).strip()
                v_data = vbs.get_verified_assembly(a_no_k)
                if v_data and v_data.get("items"):
                    print(f"[MultimodalExtractor] [Learned Pattern] Loaded Verified Ground Truth BOM for Assembly '{a_no_k}' ({len(v_data['items'])} components)")
                    assy["items"] = copy.deepcopy(v_data["items"])
                    if v_data.get("assy_model"):
                        assy["assy_model"] = v_data["assy_model"]
                    if v_data.get("assy_rev"):
                        assy["assy_rev"] = v_data["assy_rev"]
        except Exception as v_err:
            print(f"[MultimodalExtractor] Notice loading verified assembly: {v_err}")

        # Propagate per-assembly EAU and Target Price to items
        for assy in final_assemblies:
            a_eau = assy.get("eau", params["eau"])
            a_tp = assy.get("target_price", params["target_price"])
            for it in assy.get("items", []):
                it["eau"] = a_eau
                it["target_price"] = a_tp
        
        # Step 3b: Customer Alternative MPN Auto-Enrichment from BOM Module Library
        self._enrich_components_with_alternative_mpns(final_assemblies, resolved_cust)

        # Step 3c: Apply User Learned Field Corrections from CorrectionStore
        try:
            from agents.correction_store import CorrectionStore
            cs = CorrectionStore()
            applied_corrections_cnt = 0
            for assy in final_assemblies:
                a_no = str(assy.get("assy_no", "")).strip()
                for it in assy.get("items", []):
                    p_no = str(it.get("part_number", "")).strip()
                    desc_str = str(it.get("description", "")).strip()
                    for hint in (f"{a_no}::{p_no}", p_no):
                        if not hint:
                            continue
                        for field in ("mpn", "mfr", "description", "qty", "uom", "part_number"):
                            corr = cs.get_field_correction(hint, field)
                            if corr and corr.get("correct_value"):
                                it[field] = corr["correct_value"]
                                if corr.get("mfr") and field == "mpn":
                                    it["mfr"] = corr["mfr"]
                                applied_corrections_cnt += 1
            if applied_corrections_cnt > 0:
                print(f"[MultimodalExtractor] Applied {applied_corrections_cnt} learned field corrections from CorrectionStore!")
        except Exception as cs_err:
            print(f"[MultimodalExtractor] Notice applying corrections: {cs_err}")

        total_items = sum(len(a.get("items", [])) for a in final_assemblies)

        # Calculate multi-assembly summary ranges for metadata
        eaus = [a.get("eau") for a in final_assemblies if isinstance(a.get("eau"), (int, float))]
        tps = [float(re.sub(r'[^\d.]', '', str(a.get("target_price")))) for a in final_assemblies if re.search(r'\d', str(a.get("target_price", "")))]
        eau_disp = f"{min(eaus):,} - {max(eaus):,} pcs (Total: {sum(eaus):,} pcs)" if len(eaus) > 1 else (f"{eaus[0]:,} pcs" if eaus else (f"{params['eau']:,} pcs" if params.get('eau') else "Not Specified"))
        tp_disp = f"${min(tps):.2f} - ${max(tps):.2f} (per assembly)" if len(tps) > 1 else (f"${tps[0]:.2f}" if tps else params["target_price"])

        is_known = cust_ident.get("is_known_customer", False)
        requires_review = cust_ident.get("requires_human_review", True) or bool(all_conflicts)
        review_reason = cust_ident.get("review_reason", "Human review recommended")
        if all_conflicts:
            review_reason += f" • {len(all_conflicts)} data conflict(s) detected across sources"

        structured_rfq = {
            "rfq_metadata": {
                "source_email_id": email_data.get("id", "EMAIL-001"),
                "sender": sender,
                "received_date": date_str,
                "email_subject": subject,
                "customer_name": resolved_cust,
                "rfq_number": params["rfq_number"],
                "project_title": params["project_title"],
                "commodity": params["commodity"],
                "target_price": tp_disp,
                "eau": eau_disp,
                "eau_ambiguity": params.get("eau_ambiguity", {"is_ambiguous": False, "raw_text": "", "suggested_options": []}),
                "default_moqs": [100, 250, 500, 1000],
                "is_known_customer": is_known,
                "requires_human_review": requires_review,
                "review_reason": review_reason,
                "review_badge": "✅ Pattern Verified" if not requires_review else "⚠️ Requires Human Review"
            },
            "metadata_evidence": meta_evidence,
            "drawings_detected": drawings_metadata,
            "assemblies": final_assemblies,
            "total_items_count": total_items,
            "conflict_candidates": [c.to_dict() for c in all_conflicts]
        }

        return structured_rfq

    def _enrich_components_with_alternative_mpns(self, assemblies, customer_name):
        """
        Enriches extracted BOM component lines with pre-maintained Customer Alternative MPNs & MFRs
        from the BOM module's database (Customer Parts - Alternative MPNs).
        Combines and merges existing extracted MPN/MFR with alternative pairs.
        """
        if not customer_name or customer_name.lower() in ("customer", "unknown", ""):
            return

        alt_path = ""
        try:
            # 1. Dynamic path lookup
            from agents.evidence_schema import BASE_DIR
            alt_dir = os.path.join(BASE_DIR, "ref", "BOM", "AppData", "Customer Parts - Alternative MPNs")
            if not os.path.exists(alt_dir):
                alt_dir = os.path.join("D:\\RadysisAsia MockServer", "BOM", "AppData", "Customer Parts - Alternative MPNs")
            safe_cust = customer_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            cand = os.path.join(alt_dir, safe_cust, "Alternative_MPNs.json")
            if os.path.exists(cand):
                alt_path = cand
            else:
                legacy_cand = os.path.join(alt_dir, f"{safe_cust}.json")
                if os.path.exists(legacy_cand):
                    alt_path = legacy_cand
        except Exception:
            pass

        if not alt_path or not os.path.exists(alt_path):
            return

        try:
            with open(alt_path, "r", encoding="utf-8", errors="replace") as f:
                alt_data = json.load(f)
            parts_db = alt_data.get("Parts", {})
            if not parts_db:
                return

            def _clean_k(p):
                return re.sub(r'[^a-zA-Z0-9]', '', str(p or '')).lower()

            def _merge_mpn_mfr(ext_mpn, ext_mfr, db_mpn, db_mfr):
                e_mpns = [m.strip() for m in str(ext_mpn or '').split(',') if m.strip()]
                e_mfrs = [m.strip() for m in str(ext_mfr or '').split(',') if m.strip()]
                while len(e_mfrs) < len(e_mpns): e_mfrs.append('')
                while len(e_mpns) < len(e_mfrs): e_mpns.append('')

                d_mpns = [m.strip() for m in str(db_mpn or '').split(',') if m.strip()]
                d_mfrs = [m.strip() for m in str(db_mfr or '').split(',') if m.strip()]
                while len(d_mfrs) < len(d_mpns): d_mfrs.append('')
                while len(d_mpns) < len(d_mfrs): d_mpns.append('')

                seen = set()
                merged_pairs = []
                for mpn, mfr in zip(e_mpns, e_mfrs):
                    k = (mpn.upper(), mfr.upper())
                    if k not in seen and mpn:
                        seen.add(k)
                        merged_pairs.append((mpn, mfr))

                for mpn, mfr in zip(d_mpns, d_mfrs):
                    k = (mpn.upper(), mfr.upper())
                    if k not in seen and mpn and mpn.upper() not in [p[0].upper() for p in merged_pairs]:
                        seen.add(k)
                        merged_pairs.append((mpn, mfr))

                return ", ".join([p[0] for p in merged_pairs]), ", ".join([p[1] for p in merged_pairs])

            db_lookup = {}
            for db_p_no, db_val in parts_db.items():
                if isinstance(db_val, dict):
                    db_lookup[_clean_k(db_p_no)] = (db_p_no, db_val)

            for assy in assemblies:
                for it in assy.get("items", []):
                    p_no = str(it.get("part_number") or "").strip()
                    clean_k = _clean_k(p_no)
                    if clean_k in db_lookup:
                        db_p_orig, db_info = db_lookup[clean_k]
                        alt_mpn = db_info.get("Alternative_MPN") or ""
                        alt_mfr = db_info.get("Alternative_MFR") or ""

                        if alt_mpn and str(alt_mpn).strip().lower() not in ("none", "need info", ""):
                            cur_mpn = str(it.get("mpn") or "").strip()
                            cur_mfr = str(it.get("mfr") or "").strip()

                            if cur_mpn:
                                merged_mpn, merged_mfr = _merge_mpn_mfr(cur_mpn, cur_mfr, alt_mpn, alt_mfr)
                            else:
                                merged_mpn, merged_mfr = alt_mpn, alt_mfr

                            it["mpn"] = merged_mpn
                            it["mfr"] = merged_mfr
                            it["has_alternative_mpn"] = True
                            it["alternative_mpn_source"] = f"Customer Alternative Library ({customer_name})"

                            if "evidence" not in it:
                                it["evidence"] = {}
                            it["evidence"]["alternative_mpns"] = {
                                "value": merged_mpn,
                                "mfr": merged_mfr,
                                "source_document": f"Customer Alternative MPN Library ({customer_name})",
                                "zone": "CUSTOMER_AVL_LIBRARY",
                                "confidence": 0.98,
                                "reasoning": f"Auto-merged extracted MPN with approved customer AVL library: {alt_mpn} ({alt_mfr})"
                            }
        except Exception as e:
            print(f"[MultimodalExtractor] Warning enriching alternative MPNs: {e}")


if __name__ == "__main__":
    extractor = MultimodalExtractor()
    staging_tecan = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'ContXs', 'EmailStaging', '47456_Enquiry___Cable___Tecan_-_RS25')
    attachments = []
    if os.path.exists(staging_tecan):
        for f in os.listdir(staging_tecan):
            attachments.append({"filename": f, "path": os.path.join(staging_tecan, f)})

    sample_email = {
        "id": "47456",
        "subject": "Enquiry ~ Cable _ Tecan - RS25-8099",
        "sender": "jessiekong@radysis-asia.com",
        "date": "2026-08-13 09:52:12",
        "body": """
Hi AT,
Here is the enquiry from Tecan.
I did ask for the BOM as some information is not available in the drawings. But they reply just want to look at roughly price from our local sources.
Target price is for our reference.

Tecan Article No	Rev. Level	Description	Estimated annual volume quantity	TP USD
30079632	5	HARNESS 2 VISION	900	73.12
30078993	2	HARNESS 3 VISION (MAX)	1000	53.90
30059436	1	SENSOR LIGHT ASSY SICK WLG4S-3W1132	300	61.10
30059453	2	CABLE RSN 9*0.34MM2 1M DSUB9F-DSUB9M	500	10.30
30077977	0	CABLE RSN 10*0.5MM2 1.75M 2*MFT2*5F	200	7.23
        """,
        "attachments": attachments
    }

    res = extractor.extract_full_rfq(sample_email)
    print("=== MULTIMODAL EXTRACTOR RESULT ===")
    print(f"Customer: {res['rfq_metadata']['customer_name']}")
    print(f"RFQ No: {res['rfq_metadata']['rfq_number']}")
    print(f"Total Assemblies: {len(res['assemblies'])}")
    print(f"Total Items: {res['total_items_count']}")
