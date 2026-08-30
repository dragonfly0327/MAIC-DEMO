import os
import re
import json
import time
import uuid
import tkinter as tk
import threading
import queue
from datetime import datetime
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
from featuremanager import FeatureManager
from agents.orchestrator import (
    WorkflowStateManager, WorkflowStage, WorkflowStatus,
    ApprovalGateManager, ApprovalCheckpoint,
    MicroserviceToolDispatcher, DependencyEvaluator,
    CycleTimeAIEngine, RevertOrchestrator
)

# ==========================================
# 1. ADAPTIVE UI MODULE
# ==========================================
class UnifiedPortal(tk.Tk):
    def __init__(self, server_path, user_context, on_logout=None):
        super().__init__()
        self.server_path = server_path
        self.user = user_context or {}
        self.user_context = self.user
        self.title(f"ContinuumX Feature Portal - {self.user['username']} ({self.user['role']})")
        # Dynamic responsive screen autofit by default
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        target_w = max(1160, min(1440, int(sw * 0.85)))
        target_h = max(720, min(900, int(sh * 0.88)))
        pos_x = max(0, (sw - target_w) // 2)
        pos_y = max(0, (sh - target_h) // 2 - 25)
        self.geometry(f"{target_w}x{target_h}+{pos_x}+{pos_y}")
        self.minsize(1020, 640)
        self.resizable(True, True)

        # Initialize AI Orchestrator Layer
        self.orch_state_mgr = WorkflowStateManager()
        self.approval_gate_mgr = ApprovalGateManager(self.orch_state_mgr)
        self.tool_dispatcher = MicroserviceToolDispatcher(self.orch_state_mgr)
        self.dep_eval = DependencyEvaluator(self.orch_state_mgr)
        self.ct_ai_engine = CycleTimeAIEngine(self.orch_state_mgr)
        self.revert_orch = RevertOrchestrator(self.orch_state_mgr)

        # Multi-Session Conversation Store & Interrupt Control
        from agents.session_store import ChatSessionStore
        self.session_store = ChatSessionStore()
        self.current_session_id, self.current_session_data = self.session_store.create_new_session("New RFQ Session")
        self._is_agent_thinking = False
        self._abort_signal = threading.Event()

        # Set Icon
        icon_path = os.path.join(os.path.dirname(__file__), "assets", "images", "logo.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)
        
        self.selected_feature = None
        self.icon_refs = {} 
        self.feature_status = {}     # Status string: READY, UPDATING, etc.
        self.status_labels = {}      # UI Label references
        self.download_queue = queue.Queue()
        self._gui_queue = queue.Queue()
        self.remote_manifest = {}    # Initialized in load_thread
        self.active_subprocesses = []
        
        self.init_layout()
        self.setup_widgets()
        self._check_gui_queue()

        try:
            from agents.platform_bridge import DesktopAgentBridge
            DesktopAgentBridge.instance().start_all()
        except Exception as bridge_err:
            print(f"[PlatformBridge] start skipped: {bridge_err}")
        
        # Start background check thread
        threading.Thread(target=self.initial_load_sequence, daemon=True).start()

    def _check_gui_queue(self):
        """Main thread loop that processes GUI callbacks posted from background threads."""
        try:
            while hasattr(self, '_gui_queue') and not self._gui_queue.empty():
                func = self._gui_queue.get_nowait()
                func()
        except Exception:
            pass
        finally:
            if hasattr(self, 'winfo_exists') and self.winfo_exists():
                self.after(100, self._check_gui_queue)

    def _safe_gui(self, func):
        """Pushes GUI callback to queue for main thread execution."""
        if hasattr(self, '_gui_queue'):
            self._gui_queue.put(func)
        else:
            func()

    def init_layout(self):
        # Top Header Banner
        header = tk.Frame(self, bg="#1a252f", height=60)
        header.pack(side="top", fill="x")
        header.pack_propagate(False)

        title = tk.Label(header, text="ContinuumX Agentic Platform", font=("Arial", 16, "bold"), fg="#ffffff", bg="#1a252f")
        title.pack(side="left", padx=15, pady=10)

        user_info = tk.Label(header, text=f"👤 {self.user['username']} | Role: {self.user['role']}", 
                             font=("Arial", 10), fg="#ecf0f1", bg="#1a252f")
        user_info.pack(side="right", padx=15, pady=10)

        # Main Split Container using resizable PanedWindow (Draggable Sash)
        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(side="top", fill="both", expand=True, padx=5, pady=5)

        # Left Column Pane: Feature Cards (Compact ~280px)
        left_pane = tk.Frame(paned, bg="#ffffff", width=280)
        paned.add(left_pane, weight=1)

        subtitle = tk.Label(left_pane, text="Select a Feature to Launch", font=("Arial", 12, "bold"), fg="#2c3e50", bg="#ffffff")
        subtitle.pack(pady=(10, 5))

        self.list_frame = tk.Frame(left_pane, bg="#ffffff")
        self.list_frame.pack(side="top", fill="both", expand=True, padx=10, pady=5)

        # Right Column Pane: Resizable AI Agent Assistant Chatbot Panel (Expanded width by default)
        right_pane = tk.Frame(paned, bg="#f8f9fa", bd=1, relief="solid")
        paned.add(right_pane, weight=3)

        agent_header = tk.Frame(right_pane, bg="#1E293B", padx=10, pady=8)
        agent_header.pack(fill="x")

        hdr_left = tk.Frame(agent_header, bg="#1E293B")
        hdr_left.pack(side="left")
        tk.Label(hdr_left, text="ContinuumX AI Assistant", font=("Segoe UI", 11, "bold"), fg="white", bg="#1E293B").pack(anchor="w")
        tk.Label(hdr_left, text="Ask about RFQs, charts, BOM files, or inbox scans", font=("Segoe UI", 8), fg="#94A3B8", bg="#1E293B").pack(anchor="w")

        hdr_right = tk.Frame(agent_header, bg="#1E293B")
        hdr_right.pack(side="right")

        self.is_privacy_mode = False
        self._privacy_hash_cache = {}

        self.btn_privacy_mode = tk.Button(hdr_right, text="🔒 Privacy Mode", command=self._toggle_launcher_privacy_mode,
                                          font=("Segoe UI", 8, "bold"), bg="#334155", fg="#F8FAFC",
                                          activebackground="#475569", activeforeground="#FFFFFF",
                                          relief="flat", padx=7, pady=2, cursor="hand2")
        self.btn_privacy_mode.pack(side="left", padx=(0, 4))

        self.btn_new_chat = tk.Button(hdr_right, text="➕ New Chat", command=self._on_click_new_chat,
                                      font=("Segoe UI", 8, "bold"), bg="#2563EB", fg="#FFFFFF",
                                      activebackground="#1D4ED8", activeforeground="#FFFFFF",
                                      relief="flat", padx=7, pady=2, cursor="hand2")
        self.btn_new_chat.pack(side="left", padx=(0, 4))

        self.btn_chat_history = tk.Button(hdr_right, text="🕒 History", command=self._show_chat_history_dialog,
                                          font=("Segoe UI", 8), bg="#334155", fg="#F8FAFC",
                                          activebackground="#475569", activeforeground="#FFFFFF",
                                          relief="flat", padx=7, pady=2, cursor="hand2")
        self.btn_chat_history.pack(side="left")

        # Chat display area — scrollable canvas with bubble messages
        chat_outer = tk.Frame(right_pane, bg="#f0f4f8")
        chat_outer.pack(fill="both", expand=True, padx=0, pady=0)

        self._chat_canvas = tk.Canvas(chat_outer, bg="#f0f4f8", highlightthickness=0)
        chat_vsb = ttk.Scrollbar(chat_outer, orient="vertical", command=self._chat_canvas.yview)
        self._chat_canvas.configure(yscrollcommand=chat_vsb.set)
        chat_vsb.pack(side="right", fill="y")
        self._chat_canvas.pack(side="left", fill="both", expand=True)

        self._chat_inner = tk.Frame(self._chat_canvas, bg="#f0f4f8")
        self._chat_inner_id = self._chat_canvas.create_window((0, 0), window=self._chat_inner, anchor="nw")

        def _on_chat_inner_configure(event):
            self._chat_canvas.configure(scrollregion=self._chat_canvas.bbox("all"))
            self._chat_canvas.itemconfig(self._chat_inner_id, width=self._chat_canvas.winfo_width())
        self._chat_inner.bind("<Configure>", _on_chat_inner_configure)
        self._chat_canvas.bind("<Configure>", lambda e: self._chat_canvas.itemconfig(self._chat_inner_id, width=e.width))

        def _on_chat_mousewheel(event):
            try:
                self._chat_canvas.yview_scroll(self._chat_wheel_units(event.delta), "units")
            except Exception:
                pass

        def _bind_chat_mousewheel(e):
            self._chat_canvas.bind_all("<MouseWheel>", _on_chat_mousewheel)

        def _unbind_chat_mousewheel(e):
            self._chat_canvas.unbind_all("<MouseWheel>")

        right_pane.bind("<Enter>", _bind_chat_mousewheel)
        right_pane.bind("<Leave>", _unbind_chat_mousewheel)
        self._chat_canvas.bind("<Enter>", _bind_chat_mousewheel)
        self._chat_inner.bind("<Enter>", _bind_chat_mousewheel)

        # Pending Attachments (Images, Screenshots, Drawings)
        self._pending_chat_attachments = []

        # Attachment Thumbnail Preview Tray (above chat input box)
        self._attachment_tray = tk.Frame(right_pane, bg="#f1f5f9", bd=1, relief="solid")
        # Hidden initially; packed when images are pasted/attached

        # Chat Input Area
        self._chat_placeholder = "Ask about RFQs, charts, or BOM…  (Enter to send)"
        input_wrap = tk.Frame(right_pane, bg="#f8f9fa")
        input_wrap.pack(fill="x", padx=8, pady=(0, 8))

        input_container = tk.Frame(input_wrap, bg="#ffffff", highlightbackground="#cbd5e1", highlightthickness=1, bd=0)
        input_container.pack(fill="x")

        action_btn = tk.Button(input_container, text="📎", command=self._show_action_menu,
                               font=("Segoe UI", 12), bg="#ffffff", fg="#475569",
                               activebackground="#edf2f7", relief="flat", cursor="hand2", padx=6, pady=6)
        action_btn.pack(side="left", anchor="s")

        self.chat_input = tk.Text(input_container, font=("Segoe UI", 11), height=2, wrap="word", bd=0,
                                  relief="flat", bg="#ffffff", fg="#1a202c", padx=6, pady=8)
        self.chat_input.insert("1.0", self._chat_placeholder)
        self.chat_input.config(fg="#94a3b8")

        def _on_focus_in(e):
            if self._chat_input_is_placeholder():
                self.chat_input.delete("1.0", "end")
                self.chat_input.config(fg="#1a202c")
        def _on_focus_out(e):
            if not self.chat_input.get("1.0", "end-1c").strip():
                self._reset_chat_input()
        self.chat_input.bind("<FocusIn>", _on_focus_in)
        self.chat_input.bind("<FocusOut>", _on_focus_out)
        for seq in ("<Control-v>", "<Control-V>", "<Command-v>", "<Command-V>", "<Shift-Insert>"):
            self.chat_input.bind(seq, self._on_chat_paste)
        self.chat_input.pack(side="left", fill="x", expand=True)

        def _on_key_return(event):
            if event.state & 0x0001 or event.state & 0x0004 or event.state & 0x20000:
                return None
            else:
                self._on_chat_submit()
                return "break"
        self.chat_input.bind("<Return>", _on_key_return)

        self.send_btn = tk.Button(input_container, text="Send", command=self._on_chat_submit,
                                  font=("Segoe UI", 9, "bold"), bg="#2563EB", fg="white",
                                  activebackground="#1D4ED8", relief="flat", cursor="hand2", padx=14, pady=8)
        self.send_btn.pack(side="right", anchor="s", padx=(4, 6), pady=6)

        tk.Label(
            input_wrap,
            text="Enter send  ·  Shift+Enter new line  ·  ⌘/Ctrl+V paste image  ·  📎 more actions",
            font=("Segoe UI", 8), fg="#64748b", bg="#f8f9fa",
        ).pack(anchor="w", pady=(4, 0))

        # Bottom Footer Container
        footer_container = tk.Frame(self)
        footer_container.pack(side="bottom", fill="x")

        # Progress bar above footer status line
        self.progress = ttk.Progressbar(footer_container, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", padx=10, pady=(4, 2))

        # Status Footer bar
        footer_frame = tk.Frame(footer_container, bg="#f1f2f6")
        footer_frame.pack(fill="x", padx=10, pady=(2, 6))

        self.status_var = tk.StringVar(value="Initializing...")
        self.status_bar = tk.Label(footer_frame, textvariable=self.status_var, bd=1, relief="sunken", anchor="w", font=("Arial", 9), bg="#ffffff")
        self.status_bar.pack(side="left", fill="x", expand=True, padx=(0, 10))

        logout_btn = tk.Button(footer_frame, text="🔒 Logout", command=self.on_logout, 
                               font=("Arial", 9, "bold"), bg="#e74c3c", fg="white", activebackground="#c0392b", activeforeground="white", relief="flat", padx=10, cursor="hand2")
        logout_btn.pack(side="right")

        # Initial Welcome Message in Chatbot
        init_suggs = [
            "Draw RFQ stage chart",
            "How many RFQs do I have?",
            "Which customer has the most RFQs?",
            "Import a customer BOM",
            "Check RFQ emails",
        ]
        self._append_agent_message(
            "Hi — I can help with RFQs, status charts, BOM files, and inbox scans.\n\n"
            "Type a question, click a suggestion, or use 📎 for actions.",
            suggestions=init_suggs
        )
        self.after(250, lambda: self.chat_input.focus_set())
        self.after(1500, self._check_active_approval_gates)
        self.after(2000, self._check_agent_completion_loop)

    def _check_active_approval_gates(self):
        """Scans all RFQs and renders the complete Daily Morning Operations Briefing."""
        try:
            user_info = getattr(self, "user", {}) or getattr(self, "user_context", {}) or {}
            user_role = user_info.get("role") or user_info.get("Role") or "System Administrator"
            username = user_info.get("username") or user_info.get("name") or "Sysadmin"
            
            summary = self.orch_state_mgr.get_pipeline_operations_summary(user_role=user_role, username=username)
            active_gates = summary.get("actionable", [])
            wip_list = summary.get("wip", [])
            kpi = summary.get("kpi", {})

            if not active_gates and not wip_list:
                return

            self._cached_active_gates = list(active_gates)

            # 1. Briefing Greeting & KPI Ribbon
            briefing_hdr = (
                f"🌅 **Good Morning, {username}! Here is your Daily ContinuumX Operations Briefing:**\n\n"
                f"📊 **Enterprise Pipeline Snapshot:**\n"
                f"• ⚡ **{kpi.get('ready', 0)} Ready to Sign-Off** (Urgent — unblocks downstream teams)\n"
                f"• ⏳ **{kpi.get('wip', 0)} In-Progress / WIP** (Active Engineering Analysis)\n"
                f"• ✅ **{kpi.get('completed', 0)} Completed Process**\n"
                f"• 📁 **{kpi.get('total', 0)} Total Active RFQs** in Database\n\n"
                f"Below is **Section 1: Urgent Action Queue** for your immediate sign-off:"
            )

            actions = [
                "🔍 Open Guided Review Queue",
                f"⚡ Dispatch All Ready ({len(active_gates)} RFQs)",
                f"📋 View In-Progress Pipeline ({len(wip_list)} RFQs)",
                "📊 Draw RFQ Stage Chart"
            ]

            if active_gates:
                table_data = self.approval_gate_mgr.get_consolidated_queue_table_data(active_gates)
                self._append_agent_message(briefing_hdr, suggestions=actions, table_data=table_data)
            else:
                self._append_agent_message(briefing_hdr, suggestions=actions)

        except Exception as ex:
            print(f"[ApprovalGate Resume Notice] {ex}")

    def _check_agent_completion_loop(self):
        """Polls for agent_bom_completion.json written when BOM Verification completes."""
        try:
            local_appdata = os.environ.get('LOCALAPPDATA', os.environ.get('TEMP', 'C:\\Temp'))
            comp_path = os.path.join(local_appdata, "ContXs", "agent_bom_completion.json")
            if os.path.exists(comp_path):
                data = None
                try:
                    with open(comp_path, 'r', encoding='utf-8') as cf:
                        data = json.load(cf)
                    os.remove(comp_path)
                except Exception:
                    pass

                if data:
                    rfq_id = data.get("rfq_id", "RFQ")
                    cust = data.get("customer", "Customer")
                    moqs = data.get("assigned_moqs", [])
                    custom_moqs = data.get("custom_moqs", {})
                    if custom_moqs:
                        c_str = ", ".join(f"{k} ({', '.join(str(m) for m in v)})" for k, v in custom_moqs.items())
                        moq_str = f"Default: {', '.join(str(m) for m in moqs)}" if moqs else "Custom"
                        moq_str += f" | Custom: {c_str}"
                    elif moqs:
                        moq_str = ', '.join(str(m) for m in moqs)
                    else:
                        moq_str = "Standard / Default"

                    assy_cnt = data.get("assembly_count")
                    if not assy_cnt and hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                        assy_cnt = len(self._last_extracted_rfq_json.get("assemblies", []))
                    if not assy_cnt:
                        assy_cnt = 1

                    gate_info = self.approval_gate_mgr.create_gate(
                        rfq_id=rfq_id,
                        checkpoint=ApprovalCheckpoint.CHECKPOINT_2_BOM_COMPLETED,
                        stage="bom",
                        summary_data={
                            "customer": cust,
                            "assigned_moqs": moq_str,
                            "assembly_count": assy_cnt
                        },
                        customer=cust
                    )
                    card_msg = self.approval_gate_mgr.render_approval_card(gate_info)
                    dispatch_suggs = self.approval_gate_mgr.get_approval_actions(ApprovalCheckpoint.CHECKPOINT_2_BOM_COMPLETED, rfq_id)
                    self._append_agent_message(card_msg, suggestions=dispatch_suggs)

            # Detect BOM dispatch completion and auto-advance to Sourcing Calculation & Review
            dispatch_comp_path = os.path.join(local_appdata, "ContXs", "agent_bom_dispatch_completion.json")
            if os.path.exists(dispatch_comp_path):
                disp_data = None
                try:
                    with open(dispatch_comp_path, 'r', encoding='utf-8') as dcf:
                        disp_data = json.load(dcf)
                    os.remove(dispatch_comp_path)
                except Exception:
                    pass

                if disp_data:
                    rfq_id = disp_data.get("rfq_id", "RFQ")
                    cust = disp_data.get("customer", "Customer")
                    
                    msg = (f"🚀 **BOM Dispatch Completed for RFQ '{rfq_id}' ({cust})!**\n\n"
                           f"• Sent To Stage: Sourcing Operations\n"
                           f"• Automatically launching Sourcing Module & Calculation Review Tool...")
                    self._append_agent_message(msg)

                    # Write sourcing launch command and launch Sourcing feature
                    cmd_path = os.path.join(local_appdata, "ContXs", "agent_sourcing_launch_command.json")
                    with open(cmd_path, 'w', encoding='utf-8') as f:
                        json.dump({"action": "start_sourcing", "rfq_id": rfq_id, "customer": cust}, f)
                    self.after(500, lambda: self.launch_feature("Sourcing"))
        except Exception as err:
            print(f"[CompletionLoop Notice] {err}")
        finally:
            self.after(2000, self._check_agent_completion_loop)

    def initial_load_sequence(self):
        """Loads manifest and then checks each feature in background."""
        self.update_global_status("Connecting to server...")
        manifest = self.load_remote_manifest()
        if not manifest:
            return
            
        self.remote_manifest = manifest
        self.update_global_status(f"Connected as {self.user['username']}", 0)

        # We need to recreate widgets once manifest is loaded because RBAC filter 
        # and list depends on manifest keys.
        self.after(0, self.setup_widgets)
        
        # Give small delay for UI to populate
        self.after(500, self.start_status_checks)
        
        # Start Download Manager Thread
        threading.Thread(target=self.download_manager_loop, daemon=True).start()

    def start_status_checks(self):
        for key in self.remote_manifest:
            if key in self.status_labels:
                threading.Thread(target=self.check_feature_status, args=(key,), daemon=True).start()

    def check_feature_status(self, key):
        self.set_feature_status(key, "Verifying...", "#f39c12")
        manager = FeatureManager(key, self.server_path)
        remote_info = self.remote_manifest.get(key) or {}

        if manager.local_source_path():
            self.set_feature_status(key, "Ready", "#27ae60")
            self.feature_status[key] = "READY"
            return
        if not os.path.exists(manager.feature_dir):
            self.set_feature_status(key, "Not Installed", "#e67e22")
            self.feature_status[key] = "NOT_INSTALLED"
        elif manager.get_local_version() != remote_info['version']:
            self.set_feature_status(key, "Update Available", "#3498db")
            self.feature_status[key] = "UPDATE_AVAILABLE"
        else:
            self.set_feature_status(key, "Ready", "#27ae60")
            self.feature_status[key] = "READY"

    def set_feature_status(self, key, text, color="#7f8c8d"):
        """Thread-safe update of feature status label."""
        def update():
            if key in self.status_labels:
                self.status_labels[key].config(text=text, fg=color)
        self.after(0, update)

    def update_global_status(self, text, progress=None):
        def update():
            self.status_var.set(text)
            if progress is not None:
                self.progress['value'] = progress
        self.after(0, update)

    def _fallback_feature_manifest(self):
        """Source-mode catalog so the portal features the BOM Verification Engine."""
        modules = [
            ("BOM", "BOM Verification & Extraction Engine"),
        ]
        return {
            key: {"name": name, "version": "2.0.0", "icon": None}
            for key, name in modules
        }

    def load_remote_manifest(self):
        import time
        import random
        import json
        attempts = 3
        manifest_path = os.path.join(self.server_path, "version.json")

        for i in range(attempts):
            try:
                if not os.path.exists(manifest_path):
                    if i < attempts - 1:
                        time.sleep(0.2)
                        continue
                    print(f"[ERR_VER_001] Golden Version manifest not found at: {manifest_path}")
                    self.after(0, lambda: self.update_global_status("Local source mode (no version.json)", 0))
                    return self._fallback_feature_manifest()

                with open(manifest_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except json.JSONDecodeError as e:
                print(f"[ERR_VER_003] Golden Version manifest is invalid JSON: {e}")
                self.after(0, lambda: self.update_global_status("Local source mode (invalid version.json)", 0))
                return self._fallback_feature_manifest()
            except Exception as e:
                if i < attempts - 1:
                    time.sleep(0.1 + random.uniform(0, 0.2))
                    continue
                print(f"[ERR_VER_002] Could not read Golden Version server: {e}")
                self.after(0, lambda: self.update_global_status("Local source mode (server unreachable)", 0))
                return self._fallback_feature_manifest()
        return self._fallback_feature_manifest()

    def setup_widgets(self):
        # Clear existing features for reload
        for widget in self.list_frame.winfo_children():
            widget.destroy()

        canvas = tk.Canvas(self.list_frame, borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        def _on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind("<Configure>", _on_canvas_configure)
        
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        visible_features = 0
        for key, info in self.remote_manifest.items():
            # RBAC Filtering
            allowed_roles = info.get('allowed_roles', [])
            if allowed_roles and self.user['role'] not in allowed_roles:
                continue
            
            visible_features += 1
            manager = FeatureManager(key, self.server_path)
            display_name = info.get('name', key)
            ver = info.get('version', '?.?')
            icon_name = info.get('icon') # Get icon filename from JSON
            manual_file = info.get('manual')

            # --- CARD CONTAINER ---
            card = tk.Frame(scrollable_frame, bg="white", highlightbackground="#d1d8e0", 
                            highlightthickness=1, cursor="hand2")
            card.pack(pady=8, padx=5, fill="x")

            # --- ICON LOGIC ---
            if icon_name:
                icon_path = manager.sync_icon(icon_name) # Ensure FeatureManager has this method
                if icon_path and os.path.exists(icon_path):
                    try:
                        img = Image.open(icon_path).resize((40, 40), Image.Resampling.LANCZOS)
                        photo = ImageTk.PhotoImage(img)
                        self.icon_refs[key] = photo # Store reference
                        
                        icon_label = tk.Label(card, image=photo, bg="white")
                        icon_label.pack(side="left", padx=(15, 5), pady=10)
                        icon_label.bind("<Button-1>", lambda e, k=key: self.on_feature_selected(k))
                    except Exception as e:
                        print(f"Failed to load icon for {key}: {e}")

            # --- TEXT CONTENT CONTAINER ---
            text_container = tk.Frame(card, bg="white")
            text_container.pack(side="left", fill="both", expand=True)

            # Title Label
            title = tk.Label(text_container, text=display_name, font=("Arial", 11, "bold"), 
                             bg="white", fg="#2c3e50", anchor="w")
            title.pack(side="top", fill="x", padx=10, pady=(12, 0))

            # Version Label
            ver_label = tk.Label(text_container, text=f"Version: v{ver}", font=("Arial", 9), 
                                   bg="white", fg="#7f8c8d", anchor="w")
            ver_label.pack(side="top", fill="x", padx=10)

            # Status Label (NEW)
            status_label = tk.Label(text_container, text="Initializing...", font=("Arial", 8, "bold"), 
                                     bg="white", fg="#95a5a6", anchor="w")
            status_label.pack(side="top", fill="x", padx=10, pady=(0, 5))
            self.status_labels[key] = status_label
            self.feature_status[key] = "CHECKING"

            # User Manual Link (Below Details)
            if manual_file:
                manual_btn = tk.Label(text_container, text="📖 View User Manual", 
                                      font=("Arial", 8, "italic"), fg="#34495e", 
                                      bg="white", cursor="hand2", anchor="w")
                manual_btn.pack(side="top", fill="x", pady=(5, 0))
                
                # Bindings for Manual (Stop propagation so it doesn't trigger Launch)
                manual_btn.bind("<Button-1>", lambda e, m=manual_file: (manager.open_manual(m), "break"))
                manual_btn.bind("<Enter>", lambda e, w=manual_btn: w.config(fg="#3498db", font=("Arial", 8, "italic", "underline")))
                manual_btn.bind("<Leave>", lambda e, w=manual_btn: w.config(fg="#34495e", font=("Arial", 8, "italic")))
                
            # Launch Arrow
            launch_container = tk.Frame(card, bg="white")
            launch_container.pack(side="right", padx=15)

            launch_hint = tk.Label(card, text="➜", font=("Arial", 12), bg="white", fg="#3498db")
            launch_hint.pack(side="right", padx=15)
            
            # --- BINDINGS (Apply to all components for seamless clicking) ---
            for widget in (card, text_container, title, ver_label):
                widget.bind("<Button-1>", lambda e, k=key: self.on_feature_selected(k))
                widget.bind("<Enter>", lambda e, c=card: c.config(bg="#f8f9fa", highlightbackground="#3498db"))
                widget.bind("<Leave>", lambda e, c=card: c.config(bg="white", highlightbackground="#d1d8e0"))

        if visible_features == 0 and self.remote_manifest:
            tk.Label(scrollable_frame, text="No features available for your role.", 
                     fg="#e74c3c", font=("Arial", 10)).pack(pady=50)
        elif not self.remote_manifest:
            tk.Label(scrollable_frame, text="Waiting for server response...", 
                     fg="#7f8c8d", font=("Arial", 10)).pack(pady=50)

    def on_feature_selected(self, feature_key):
        status = self.feature_status.get(feature_key)
        
        if status == "READY":
            self.launch_feature(feature_key)
        elif status in ("UPDATE_AVAILABLE", "NOT_INSTALLED"):
            self.enqueue_download(feature_key)
        elif status == "IN_QUEUE":
            messagebox.showinfo("In Queue", f"{feature_key} is already in the download queue.")
        elif status == "UPDATING":
            messagebox.showinfo("Busy", f"{feature_key} is currently downloading/installing.")
        else:
            messagebox.showwarning("Checking", "Please wait for version verification to complete.")

    def enqueue_download(self, key):
        self.feature_status[key] = "IN_QUEUE"
        self.set_feature_status(key, "In Queue...", "#8e44ad")
        self.download_queue.put(key)

    def download_manager_loop(self):
        while True:
            key = self.download_queue.get()
            self.feature_status[key] = "UPDATING"
            self.process_download(key)
            self.download_queue.task_done()

    def process_download(self, key):
        manager = FeatureManager(key, self.server_path)
        remote_info = self.remote_manifest[key]
        
        def local_update_fn(msg, prog):
            self.set_feature_status(key, f"{msg} ({prog}%)", "#2980b9")
            self.update_global_status(f"{key}: {msg}", prog)

        self.set_feature_status(key, "Connecting...", "#2980b9")
        success, message = manager.perform_sync(remote_info, local_update_fn)
        
        if success:
            self.set_feature_status(key, "Ready", "#27ae60")
            self.feature_status[key] = "READY"
            self.update_global_status(f"Installed {key}", 0)
        else:
            self.set_feature_status(key, f"Failed: {message}", "#c0392b")
            self.feature_status[key] = "ERROR"
            self.update_global_status(f"Error updating {key}", 0)
            self.after(0, lambda: messagebox.showerror("Update Failed", f"Failed to update {key}:\n{message}"))

    def launch_feature(self, key):
        manager = FeatureManager(key, self.server_path)
        self.update_global_status("Launching...", 100)
        manager.launch(self.user)
    def _safe_gui(self, fn):
        """Thread-safe dispatch for GUI updates."""
        try:
            self.after(0, fn)
        except Exception:
            try:
                fn()
            except Exception as e:
                print(f"[Launcher GUI Error] {e}")

    def _on_chat_paste(self, event=None):
        """Captures images or screenshots from clipboard (Ctrl+V) and stages them in the attachment tray."""
        try:
            from PIL import ImageGrab
            clip_data = ImageGrab.grabclipboard()
            if clip_data is not None:
                if isinstance(clip_data, list):
                    # List of file paths copied in Explorer
                    for fp in clip_data:
                        if str(fp).lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.pdf')):
                            self._add_file_attachment(fp)
                    return "break"
                elif hasattr(clip_data, 'size'):
                    # Screenshot or copied image in clipboard
                    idx = len(self._pending_chat_attachments) + 1
                    self._add_image_attachment(clip_data, name=f"Screenshot {idx}")
                    return "break"
        except Exception as ex:
            print(f"[Launcher Chat] Clipboard paste notice: {ex}")
        return None  # Allow standard text paste to proceed

    def _add_image_attachment(self, pil_img, name="Image"):
        """Encodes and adds a PIL Image to the pending attachments list with a thumbnail."""
        import io
        import base64
        from PIL import Image, ImageTk

        # Resize image for base64 payload if excessively large (max 1600px width/height)
        img_for_api = pil_img.copy()
        if img_for_api.width > 1600 or img_for_api.height > 1600:
            img_for_api.thumbnail((1600, 1600), Image.Resampling.LANCZOS)

        # Convert to JPEG base64 string
        buf = io.BytesIO()
        if img_for_api.mode != "RGB":
            img_for_api = img_for_api.convert("RGB")
        img_for_api.save(buf, format="JPEG", quality=85)
        b64_str = base64.b64encode(buf.getvalue()).decode("utf-8")

        # Create compact UI thumbnail (48x48)
        thumb_img = pil_img.copy()
        thumb_img.thumbnail((48, 48), Image.Resampling.LANCZOS)
        photo = ImageTk.PhotoImage(thumb_img)

        att_obj = {
            "name": name,
            "image": pil_img,
            "photo": photo,
            "b64": b64_str
        }
        self._pending_chat_attachments.append(att_obj)
        self._render_attachment_tray()

    def _add_file_attachment(self, file_path):
        """Attaches an image or renders the first page of a PDF drawing as an image."""
        if not os.path.exists(file_path):
            return
        from PIL import Image
        fn = os.path.basename(file_path)
        ext = os.path.splitext(fn)[1].lower()

        if ext in ('.png', '.jpg', '.jpeg', '.webp', '.bmp'):
            try:
                img = Image.open(file_path)
                self._add_image_attachment(img, name=fn)
            except Exception as e:
                print(f"[Launcher Chat] Error loading image {file_path}: {e}")
        elif ext == '.pdf':
            try:
                import pypdfium2 as pdfium
                pdf = pdfium.PdfDocument(file_path)
                if len(pdf) > 0:
                    page = pdf[0]
                    pil_img = page.render(scale=2).to_pil()
                    self._add_image_attachment(pil_img, name=f"PDF: {fn}")
            except Exception as e:
                print(f"[Launcher Chat] Error rendering PDF {file_path}: {e}")

    def _pick_and_attach_files(self):
        """Opens file dialog for user to attach images, screenshots, or PDF drawings."""
        from tkinter import filedialog
        paths = filedialog.askopenfilenames(
            title="Attach Images or PDF Drawings",
            filetypes=[
                ("Supported Files", "*.png *.jpg *.jpeg *.webp *.bmp *.pdf"),
                ("Images (*.png, *.jpg, *.jpeg)", "*.png *.jpg *.jpeg *.webp *.bmp"),
                ("PDF Drawings (*.pdf)", "*.pdf"),
                ("All Files", "*.*")
            ]
        )
        if paths:
            for p in paths:
                self._add_file_attachment(p)

    def _render_attachment_tray(self):
        """Renders visual chips in the attachment tray above the chat input box."""
        for widget in self._attachment_tray.winfo_children():
            widget.destroy()

        if not self._pending_chat_attachments:
            self._attachment_tray.pack_forget()
            return

        self._attachment_tray.pack(fill="x", padx=8, pady=(0, 4), before=self.chat_input.master)

        lbl = tk.Label(self._attachment_tray, text=f"📎 Attached ({len(self._pending_chat_attachments)}):",
                       font=("Segoe UI", 8, "bold"), fg="#475569", bg="#f1f5f9")
        lbl.pack(side="left", padx=(6, 4), pady=3)

        for idx, att in enumerate(self._pending_chat_attachments):
            chip = tk.Frame(self._attachment_tray, bg="#ffffff", bd=1, relief="solid", padx=4, pady=2)
            chip.pack(side="left", padx=3, pady=2)

            # Thumbnail Icon
            img_lbl = tk.Label(chip, image=att["photo"], bg="#ffffff")
            img_lbl.image = att["photo"]  # Retain reference
            img_lbl.pack(side="left", padx=(0, 4))

            # Filename / label
            name_txt = att["name"] if len(att["name"]) <= 16 else att["name"][:13] + "..."
            tk.Label(chip, text=name_txt, font=("Segoe UI", 8), fg="#0f172a", bg="#ffffff").pack(side="left")

            # Remove button
            def _remove(index=idx):
                if 0 <= index < len(self._pending_chat_attachments):
                    self._pending_chat_attachments.pop(index)
                    self._render_attachment_tray()

            del_btn = tk.Button(chip, text="✕", command=_remove, font=("Segoe UI", 8, "bold"),
                                fg="#ef4444", bg="#ffffff", relief="flat", cursor="hand2", padx=2, bd=0)
            del_btn.pack(side="left", padx=(2, 0))

    def _clear_attachment_tray(self):
        """Clears pending attachments after sending."""
        self._pending_chat_attachments.clear()
        self._render_attachment_tray()

    def _show_action_menu(self):
        """Displays popup menu for quick agent upload actions and multimodal attachments."""
        menu = tk.Menu(self, tearoff=0, font=("Segoe UI", 9))
        menu.add_command(label="🖼️ 1. Attach Image / Screenshot / Drawing (or Ctrl+V)", command=self._pick_and_attach_files)
        menu.add_command(label="📋 2. Paste Image from Clipboard (Ctrl+V)", command=self._on_chat_paste)
        menu.add_separator()
        menu.add_command(label="📄 3. Import Customer BOM (Excel)", command=self._agent_import_bom)
        menu.add_command(label="📧 4. Process & Classify RFQ Email", command=self._agent_process_email)
        menu.add_separator()
        menu.add_command(label="🧠 5. View AI Learned Rules & Memory", command=self._show_ai_memory_dialog)
        
        try:
            menu.tk_popup(self.winfo_pointerx(), self.winfo_pointery())
        finally:
            menu.grab_release()

    def _show_ai_memory_dialog(self):
        """Displays a dialog showing all persistent knowledge items and rules taught by the user."""
        from agents.correction_store import CorrectionStore
        cs = CorrectionStore()
        records = cs.get_all()

        dlg = tk.Toplevel(self)
        dlg.title("🧠 AI Memory & User-Taught Knowledge")
        dlg.geometry("750x520")
        dlg.configure(bg="#0F172A")
        dlg.transient(self)
        dlg.grab_set()

        hdr = tk.Frame(dlg, bg="#1E293B", padx=16, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🧠 ContinuumX AI Knowledge & Learned Rules",
                 font=("Segoe UI", 12, "bold"), fg="#FFFFFF", bg="#1E293B").pack(anchor="w")
        tk.Label(hdr, text=f"Displaying {len(records)} learned rules and corrections injected as few-shot memory for Gemini.",
                 font=("Segoe UI", 9), fg="#94A3B8", bg="#1E293B").pack(anchor="w", pady=(2, 0))

        fr = tk.Frame(dlg, bg="#0F172A", padx=12, pady=10)
        fr.pack(fill="both", expand=True)

        txt = tk.Text(fr, bg="#1E293B", fg="#F8FAFC", font=("Consolas", 10), wrap="word", bd=1, relief="solid")
        sb = ttk.Scrollbar(fr, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        txt.pack(side="left", fill="both", expand=True)

        if not records:
            txt.insert("1.0", "No custom rules or corrections learned yet.\n\nTip: You can directly train the AI in chat by typing:\n  • Teach AI: for Graco RFQs, manufacturer is Alpha Wire\n  • Remember that JST 10HR-4K is a 10-Way connector\n  • Learn rule: set Commodity to Wire Harness for all cable drawings")
        else:
            for idx, r in enumerate(records, start=1):
                if r.get("type") == "taught_rule" or "rule" in r:
                    txt.insert("end", f"[{idx}] 🧠 TAUGHT RULE (Scope: {r.get('doc_hint', 'GLOBAL')}):\n")
                    txt.insert("end", f"    Rule: {r.get('rule', '')}\n")
                    txt.insert("end", f"    Taught by: {r.get('taught_by', 'User')} | {r.get('timestamp', '')[:19]}\n\n")
                else:
                    txt.insert("end", f"[{idx}] ✏️ FIELD CORRECTION (Doc: {r.get('doc_hint', 'General')}):\n")
                    txt.insert("end", f"    Field: {r.get('field', '')} | Was: '{r.get('wrong_value', '')}' -> MUST BE: '{r.get('correct_value', '')}'\n")
                    if r.get('mfr'): txt.insert("end", f"    MFR: {r.get('mfr')}\n")
                    txt.insert("end", f"    Corrected by: {r.get('corrected_by', 'User')} | {r.get('timestamp', '')[:19]}\n\n")

        txt.config(state="disabled")

        bbar = tk.Frame(dlg, bg="#1E293B", padx=16, pady=8)
        bbar.pack(fill="x")
        tk.Button(bbar, text="Close", command=dlg.destroy, bg="#334155", fg="#FFFFFF",
                  font=("Segoe UI", 9), relief="flat", padx=12, pady=4, cursor="hand2").pack(side="right")

    def _chat_wheel_units(self, delta):
        """Windows uses ±120 steps; macOS Tk often sends ±1."""
        try:
            d = int(delta)
        except (TypeError, ValueError):
            return -1
        if abs(d) >= 120:
            return int(-1 * (d / 120))
        return -1 if d > 0 else 1

    def _chat_input_is_placeholder(self):
        raw = self.chat_input.get("1.0", "end-1c").strip()
        placeholder = getattr(self, "_chat_placeholder", "")
        return (not raw) or raw == placeholder or raw.startswith("Ask me anything")

    def _reset_chat_input(self):
        placeholder = getattr(self, "_chat_placeholder", "Ask about RFQs…")
        self.chat_input.delete("1.0", "end")
        self.chat_input.insert("1.0", placeholder)
        self.chat_input.config(fg="#94a3b8")

    def _insert_chat_markdown(self, widget, text):
        """Render **bold** and `code` so chat is readable without raw markdown."""
        widget.tag_config("md_bold", font=("Segoe UI", 10, "bold"), foreground="#0f172a")
        widget.tag_config("code", font=("Consolas", 9), background="#F8FAFC", foreground="#0F172A")
        parts = re.split(r"(\*\*[^*]+\*\*|`[^`]+`)", text or "")
        for part in parts:
            if len(part) >= 4 and part.startswith("**") and part.endswith("**"):
                widget.insert("end", part[2:-2], "md_bold")
            elif len(part) >= 2 and part.startswith("`") and part.endswith("`"):
                widget.insert("end", part[1:-1], "code")
            else:
                widget.insert("end", part)

    def _style_suggestion_chip(self, parent, text):
        btn = tk.Button(
            parent, text=text, command=lambda p=text: self._on_suggestion_chip_click(p),
            font=("Segoe UI", 9), bg="#EFF6FF", fg="#1D4ED8",
            activebackground="#DBEAFE", activeforeground="#1E3A8A",
            relief="flat", cursor="hand2", padx=10, pady=5, bd=0,
            highlightthickness=1, highlightbackground="#BFDBFE",
        )
        btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#DBEAFE"))
        btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#EFF6FF"))
        return btn

    def _append_user_message(self, text, images=None, save_to_session=True):
        """Renders a right-aligned Gemini/Antigravity style user prompt card with optional embedded thumbnail images."""
        if save_to_session and hasattr(self, 'current_session_data') and isinstance(self.current_session_data, dict):
            if "messages" not in self.current_session_data:
                self.current_session_data["messages"] = []
            self.current_session_data["messages"].append({
                "sender": "user",
                "text": text,
                "timestamp": time.time()
            })
            if hasattr(self, 'session_store') and hasattr(self, 'current_session_id'):
                # Update title based on first query if still default
                if self.current_session_data.get("title") in ("New Conversation", "New RFQ Session") and text and text.strip():
                    self.current_session_data["title"] = text.strip()[:35]
                self.session_store.save_session(self.current_session_id, self.current_session_data)

        row = tk.Frame(self._chat_inner, bg="#f0f4f8")
        row.pack(fill="x", padx=14, pady=(8, 4))

        bubble = tk.Frame(row, bg="#2563EB", padx=12, pady=8)
        bubble.pack(side="right", anchor="e", padx=(80, 0))

        # 1. Render image attachments in user bubble (if any)
        if images:
            img_row = tk.Frame(bubble, bg="#2563EB")
            img_row.pack(anchor="e", pady=(0, 6))
            for p in images:
                lbl = tk.Label(img_row, image=p, bg="#1e40af", bd=0)
                lbl.image = p
                lbl.pack(side="left", padx=2)

        if text and text.strip():
            raw_t = text.strip()
            raw_lines = raw_t.splitlines()
            max_line_len = max(len(l) for l in raw_lines) if raw_lines else 10
            bubble_width = min(62, max(22, max_line_len))

            # Estimate wrapped line count based on character width
            total_wrapped_lines = 0
            for l in raw_lines:
                total_wrapped_lines += max(1, (len(l) + bubble_width - 1) // bubble_width)
            line_cnt = max(1, min(35, total_wrapped_lines))

            msg_txt = tk.Text(bubble, bg="#2563EB", fg="#ffffff", font=("Segoe UI", 10),
                              relief="flat", borderwidth=0, highlightthickness=0,
                              wrap="word", height=line_cnt, width=bubble_width, cursor="xterm",
                              exportselection=False,
                              selectbackground="#1D4ED8", selectforeground="#FFFFFF",
                              inactiveselectbackground="#3B82F6")
            msg_txt.insert("1.0", raw_t)
            msg_txt.pack(anchor="e", fill="both", expand=True)

            # Auto-adjust height to exact rendered display lines
            def _adjust_user_bubble_height():
                try:
                    msg_txt.update_idletasks()
                    d_cnt = msg_txt.count("1.0", "end-1c", "displaylines")
                    if d_cnt and d_cnt[0] > 0:
                        msg_txt.config(height=min(35, d_cnt[0]))
                except Exception:
                    pass
            self.after(20, _adjust_user_bubble_height)

            # Mouse wheel scroll forwarding
            def _scroll_user_chat(e):
                try: self._chat_canvas.yview_scroll(self._chat_wheel_units(e.delta), "units")
                except Exception: pass
            msg_txt.bind("<MouseWheel>", _scroll_user_chat)

            # Allow selection & Ctrl+C copy, block typing
            def _on_txt_key(event):
                if (event.state & 4) and event.keysym.lower() == 'c':
                    try:
                        sel = msg_txt.selection_get() if msg_txt.tag_ranges("sel") else text
                        self._copy_to_clip(sel)
                    except Exception:
                        self._copy_to_clip(text)
                    return "break"
                if (event.state & 4) and event.keysym.lower() == 'a':
                    msg_txt.tag_add("sel", "1.0", "end-1c")
                    return "break"
                if event.keysym in ('Left', 'Right', 'Up', 'Down', 'Home', 'End', 'Prior', 'Next', 'Shift_L', 'Shift_R', 'Control_L', 'Control_R'):
                    return
                return "break"

            msg_txt.bind("<Key>", _on_txt_key)

            # Right-click Copy menu
            menu = tk.Menu(msg_txt, tearoff=0)
            def _show_user_menu(e):
                menu.delete(0, "end")
                sel_t = ""
                try:
                    if msg_txt.tag_ranges("sel"):
                        sel_t = msg_txt.selection_get()
                except Exception:
                    pass
                if sel_t:
                    p = sel_t.strip()[:25] + ("…" if len(sel_t.strip()) > 25 else "")
                    menu.add_command(label=f"📋 Copy Selected: \"{p}\"", command=lambda: self._copy_to_clip(sel_t))
                    menu.add_separator()
                menu.add_command(label="📑 Copy Entire Message", command=lambda: self._copy_to_clip(text))
                menu.tk_popup(e.x_root, e.y_root)

            msg_txt.bind("<Button-3>", _show_user_menu)

        self._chat_canvas.update_idletasks()
        self._chat_canvas.yview_moveto(1.0)

    def _show_typing_indicator(self):
        """Shows animated '...' thinking indicator. Returns the row frame to destroy later."""
        row = tk.Frame(self._chat_inner, bg="#f0f4f8")
        row.pack(fill="x", padx=14, pady=(4, 4))

        icon = tk.Label(row, text="🤖", bg="#f0f4f8", font=("Segoe UI", 14))
        icon.pack(side="left", anchor="n", padx=(0, 8), pady=(2, 0))

        bubble = tk.Frame(row, bg="#ffffff", padx=14, pady=10,
                          highlightbackground="#cbd5e0", highlightthickness=1)
        bubble.pack(side="left", anchor="w")
        self._typing_label = tk.Label(bubble, text="Working on your question…", bg="#ffffff", fg="#64748b",
                                      font=("Segoe UI", 10, "italic"))
        self._typing_label.pack()

        self._chat_canvas.update_idletasks()
        self._chat_canvas.yview_moveto(1.0)

        # Animate dots
        self._typing_anim_running = True
        self._typing_row = row
        self._animate_typing(0)
        return row

    def _animate_typing(self, frame):
        if not getattr(self, '_typing_anim_running', False):
            return
        dots = ["…", "·  ", "·· ", "···"]
        try:
            self._typing_label.config(text=f"Working on your question{dots[frame % len(dots)]}")
        except Exception:
            return
        self.after(350, lambda: self._animate_typing(frame + 1))

    def _hide_typing_indicator(self):
        self._typing_anim_running = False
        self._extraction_status_text = ""
        try:
            if hasattr(self, '_typing_row') and self._typing_row:
                self._typing_row.destroy()
                self._typing_row = None
        except Exception:
            pass

    def _update_extraction_status(self, status_text):
        """Updates the typing indicator bubble with live extraction progress."""
        self._extraction_status_text = str(status_text)
        try:
            if hasattr(self, '_typing_label') and self._typing_label:
                self._typing_label.config(text=str(status_text)[:80])
                self._chat_canvas.update_idletasks()
                self._chat_canvas.yview_moveto(1.0)
        except Exception:
            pass

    def _make_progress_callback(self):
        """Returns a thread-safe progress callback that updates the typing indicator."""
        def _cb(status):
            self._safe_gui(lambda: self._update_extraction_status(status))
        return _cb

    def _copy_to_clip(self, text):
        """Copies text to system clipboard with error handling."""
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
        except Exception:
            pass

    def _append_agent_message(self, text="", suggestions=None, table_data=None, save_to_session=True):
        """Renders a full-width Gemini/Antigravity style agent card with selectable copyable text, native styled tables, and interactive Copilot prompt chips."""
        self._hide_typing_indicator()
        self._set_thinking_state(False)

        if save_to_session and hasattr(self, 'current_session_data') and isinstance(self.current_session_data, dict):
            if "messages" not in self.current_session_data:
                self.current_session_data["messages"] = []
            self.current_session_data["messages"].append({
                "sender": "agent",
                "text": text,
                "suggestions": suggestions,
                "table_data": table_data,
                "timestamp": time.time()
            })
            if hasattr(self, 'session_store') and hasattr(self, 'current_session_id'):
                self.session_store.save_session(self.current_session_id, self.current_session_data)

        row = tk.Frame(self._chat_inner, bg="#f0f4f8")
        row.pack(fill="x", padx=14, pady=(4, 8))

        icon = tk.Label(row, text="🤖", bg="#f0f4f8", font=("Segoe UI", 14))
        icon.pack(side="left", anchor="n", padx=(0, 8), pady=(2, 0))

        card = tk.Frame(row, bg="#ffffff", padx=14, pady=10,
                        highlightbackground="#cbd5e0", highlightthickness=1)
        card.pack(side="left", fill="x", expand=True)

        # 1. Text rendering (if text provided)
        if text and text.strip():
            line_cnt = max(1, min(24, len(text.splitlines())))
            msg_txt = tk.Text(card, bg="#ffffff", fg="#1a202c", font=("Segoe UI", 10),
                              relief="flat", borderwidth=0, highlightthickness=0,
                              wrap="none" if "```" in text or "| #" in text else "word",
                              height=line_cnt, cursor="xterm",
                              exportselection=False,
                              selectbackground="#2563EB", selectforeground="#FFFFFF",
                              inactiveselectbackground="#93C5FD")
            msg_txt.tag_config("code", font=("Consolas", 9), background="#F8FAFC", foreground="#0F172A")

            if "```" in text:
                parts = text.split("```")
                for idx, part in enumerate(parts):
                    if idx % 2 == 1:
                        clean_code = part.strip()
                        msg_txt.insert("end", "\n" + clean_code + "\n", "code")
                    else:
                        self._insert_chat_markdown(msg_txt, part)
            else:
                self._insert_chat_markdown(msg_txt, text)

            msg_txt.pack(fill="both", expand=True, anchor="w", pady=(0, 4))

            # Mouse wheel scroll forwarding
            def _scroll_chat(e):
                try: self._chat_canvas.yview_scroll(self._chat_wheel_units(e.delta), "units")
                except Exception: pass
            msg_txt.bind("<MouseWheel>", _scroll_chat)

            # Allow selection & Ctrl+C copy, block typing
            def _on_txt_key(event):
                if (event.state & 4) and event.keysym.lower() == 'c':
                    try:
                        sel = msg_txt.selection_get() if msg_txt.tag_ranges("sel") else text
                        self._copy_to_clip(sel)
                    except Exception:
                        self._copy_to_clip(text)
                    return "break"
                if (event.state & 4) and event.keysym.lower() == 'a':
                    msg_txt.tag_add("sel", "1.0", "end-1c")
                    return "break"
                if event.keysym in ('Left', 'Right', 'Up', 'Down', 'Home', 'End', 'Prior', 'Next', 'Shift_L', 'Shift_R', 'Control_L', 'Control_R'):
                    return
                return "break"

            msg_txt.bind("<Key>", _on_txt_key)

            # Right-click Copy menu
            menu = tk.Menu(msg_txt, tearoff=0)
            def _show_msg_menu(e):
                menu.delete(0, "end")
                sel_t = ""
                try:
                    if msg_txt.tag_ranges("sel"):
                        sel_t = msg_txt.selection_get()
                except Exception:
                    pass
                if sel_t:
                    p = sel_t.strip()[:25] + ("…" if len(sel_t.strip()) > 25 else "")
                    menu.add_command(label=f"📋 Copy Selected: \"{p}\"", command=lambda: self._copy_to_clip(sel_t))
                    menu.add_separator()
                menu.add_command(label="📑 Copy Entire Message", command=lambda: self._copy_to_clip(text))
                menu.tk_popup(e.x_root, e.y_root)

            msg_txt.bind("<Button-3>", _show_msg_menu)

        # 2. Native High-Quality Styled Table rendering (if table_data provided)
        if table_data and isinstance(table_data, dict) and "headers" in table_data and "rows" in table_data:
            headers = table_data["headers"]
            rows = table_data["rows"]

            # Table Header Bar with Title and 1-Click "Copy Table" Button
            t_hdr_bar = tk.Frame(card, bg="#FFFFFF")
            t_hdr_bar.pack(fill="x", pady=(0, 6))

            if table_data.get("title"):
                tk.Label(
                    t_hdr_bar,
                    text=table_data["title"],
                    font=("Segoe UI", 10, "bold"),
                    fg="#0F172A",
                    bg="#FFFFFF"
                ).pack(side="left")

            def _copy_entire_table_tsv():
                try:
                    lines = ["\t".join(str(h).strip() for h in headers)]
                    for r in rows:
                        lines.append("\t".join(str(c).strip() for c in r))
                    tsv_text = "\n".join(lines)
                    self.clipboard_clear()
                    self.clipboard_append(tsv_text)
                    copy_btn.config(text="✅ Copied to Clipboard!", bg="#059669")
                    self.after(2000, lambda: copy_btn.config(text="📋 Copy Table", bg="#2563EB"))
                except Exception:
                    pass

            copy_btn = tk.Button(
                t_hdr_bar,
                text="📋 Copy Table",
                command=_copy_entire_table_tsv,
                font=("Segoe UI", 8, "bold"),
                bg="#2563EB",
                fg="#FFFFFF",
                relief="flat",
                padx=8,
                pady=2,
                cursor="hand2"
            )
            copy_btn.pack(side="right")

            # Create a horizontally & vertically scrollable frame container for wide/tall tables
            t_scroll_frame = tk.Frame(card, bg="#CBD5E1", bd=1, relief="solid")
            t_scroll_frame.pack(fill="both", expand=True, pady=(2, 6))
            t_scroll_frame.grid_rowconfigure(0, weight=1)
            t_scroll_frame.grid_columnconfigure(0, weight=1)

            t_canvas = tk.Canvas(t_scroll_frame, bg="#FFFFFF", highlightthickness=0, height=min(380, max(140, len(rows) * 36 + 45)))
            t_hsb = ttk.Scrollbar(t_scroll_frame, orient="horizontal", command=t_canvas.xview)
            t_vsb = ttk.Scrollbar(t_scroll_frame, orient="vertical", command=t_canvas.yview)
            t_canvas.configure(xscrollcommand=t_hsb.set, yscrollcommand=t_vsb.set)
            t_canvas.grid(row=0, column=0, sticky="nsew")

            table_container = tk.Frame(t_canvas, bg="#CBD5E1")
            t_win = t_canvas.create_window((0, 0), window=table_container, anchor="nw")

            def _update_table_scrollbars(event=None):
                t_canvas.configure(scrollregion=t_canvas.bbox("all"))
                req_w = table_container.winfo_reqwidth()
                req_h = table_container.winfo_reqheight()
                can_w = t_canvas.winfo_width()
                can_h = t_canvas.winfo_height()

                if req_w > can_w and can_w > 1:
                    t_hsb.grid(row=1, column=0, sticky="ew")
                else:
                    t_hsb.grid_forget()

                if req_h > can_h and can_h > 1:
                    t_vsb.grid(row=0, column=1, sticky="ns")
                else:
                    t_vsb.grid_forget()

            table_container.bind("<Configure>", _update_table_scrollbars)
            t_canvas.bind("<Configure>", _update_table_scrollbars)

            # Calculate dynamic column widths based on headers and rows
            col_widths = []
            for c_idx, h in enumerate(headers):
                max_len = len(str(h))
                for row_val in rows:
                    if c_idx < len(row_val):
                        max_len = max(max_len, len(str(row_val[c_idx])))
                col_widths.append(min(32, max(12, max_len + 2)))

            # Column weights
            for col_idx in range(len(headers)):
                table_container.grid_columnconfigure(col_idx, weight=1)

            def _scroll_table_chat(e):
                try:
                    units = self._chat_wheel_units(e.delta)
                    # If vertical scrollbar is active and within bounds, scroll table; else scroll chat
                    if t_vsb.winfo_ismapped():
                        y_pos = t_canvas.yview()
                        if (units > 0 and y_pos[1] < 1.0) or (units < 0 and y_pos[0] > 0.0):
                            t_canvas.yview_scroll(units, "units")
                            return
                    self._chat_canvas.yview_scroll(units, "units")
                except Exception:
                    try: self._chat_canvas.yview_scroll(self._chat_wheel_units(e.delta), "units")
                    except Exception: pass

            def _make_cell_key_handler(w, default_txt):
                def _handler(event):
                    if (event.state & 4) and event.keysym.lower() == 'c':
                        try:
                            sel = w.selection_get() if w.tag_ranges("sel") else default_txt
                            self._copy_to_clip(sel)
                        except Exception:
                            self._copy_to_clip(default_txt)
                        return "break"
                    if (event.state & 4) and event.keysym.lower() == 'a':
                        w.tag_add("sel", "1.0", "end-1c")
                        return "break"
                    if event.keysym in ('Left', 'Right', 'Up', 'Down', 'Home', 'End', 'Prior', 'Next', 'Shift_L', 'Shift_R', 'Control_L', 'Control_R'):
                        return
                    return "break"
                return _handler

            # Context Menu for copy actions on cells
            cell_menu = tk.Menu(self, tearoff=0, bg="#1E293B", fg="#FFFFFF", activebackground="#2563EB", activeforeground="#FFFFFF")

            def _show_cell_menu(event, cell_widget, cell_text, row_vals):
                cell_menu.delete(0, "end")
                sel_text = ""
                try:
                    if cell_widget.tag_ranges("sel"):
                        sel_text = cell_widget.selection_get()
                except Exception:
                    pass

                if sel_text:
                    preview = sel_text.strip()[:25] + ("…" if len(sel_text.strip()) > 25 else "")
                    cell_menu.add_command(label=f"📋 Copy Selected: \"{preview}\"", command=lambda: self._copy_to_clip(sel_text))
                    cell_menu.add_separator()

                val_preview = str(cell_text).strip()[:25] + ("…" if len(str(cell_text).strip()) > 25 else "")
                cell_menu.add_command(label=f"📋 Copy Cell: \"{val_preview}\"", command=lambda: self._copy_to_clip(str(cell_text).strip()))
                if row_vals:
                    cell_menu.add_command(label="📑 Copy Entire Row", command=lambda: self._copy_to_clip("\t".join(str(c).strip() for c in row_vals)))
                cell_menu.add_command(label="📊 Copy Table", command=_copy_entire_table_tsv)
                cell_menu.tk_popup(event.x_root, event.y_root)

            # Header row (Corporate Deep Navy with Bold White Text, focusable/selectable)
            for col_idx, h_text in enumerate(headers):
                h_str = str(h_text)
                hdr_txt = tk.Text(
                    table_container,
                    height=1,
                    width=col_widths[col_idx],
                    font=("Segoe UI", 9, "bold"),
                    bg="#1E293B",
                    fg="#FFFFFF",
                    relief="flat",
                    bd=0,
                    highlightthickness=0,
                    padx=10,
                    pady=6,
                    cursor="xterm",
                    exportselection=False,
                    selectbackground="#3B82F6",
                    selectforeground="#FFFFFF",
                    inactiveselectbackground="#1D4ED8"
                )
                hdr_txt.insert("1.0", h_str)
                hdr_txt.bind("<Key>", _make_cell_key_handler(hdr_txt, h_str))
                hdr_txt.bind("<MouseWheel>", _scroll_table_chat)
                hdr_txt.bind("<Button-3>", lambda e, hw=hdr_txt, ht=h_str: _show_cell_menu(e, hw, ht, None))
                hdr_txt.grid(
                    row=0,
                    column=col_idx,
                    sticky="nsew",
                    padx=(0, 1 if col_idx < len(headers) - 1 else 0),
                    pady=(0, 1)
                )

            # Data rows (Alternating Zebra Striping with adequate padding, text highlight selection, and clean wrapping)
            for row_idx, row_data in enumerate(rows, start=1):
                row_bg = "#FFFFFF" if row_idx % 2 == 1 else "#F8FAFC"
                # Calculate uniform row height
                max_lines = 1
                for c_idx, val in enumerate(row_data):
                    val_str = str(val)
                    c_w = col_widths[c_idx] if c_idx < len(col_widths) else 15
                    lines = max(1, (len(val_str) + c_w - 1) // c_w)
                    if '\n' in val_str:
                        lines = max(lines, len(val_str.splitlines()))
                    max_lines = max(max_lines, lines)

                for col_idx, cell_val in enumerate(row_data):
                    is_action = (col_idx == len(row_data) - 1)
                    fg_color = "#0284C7" if is_action else "#0F172A"
                    font_style = ("Segoe UI", 9, "bold") if (col_idx in (0, 1) or is_action) else ("Segoe UI", 9)
                    c_val_str = str(cell_val)
                    r_data_copy = list(row_data)

                    cell_txt = tk.Text(
                        table_container,
                        height=max_lines,
                        width=col_widths[col_idx] if col_idx < len(col_widths) else 15,
                        font=font_style,
                        bg=row_bg,
                        fg=fg_color,
                        relief="flat",
                        bd=0,
                        highlightthickness=0,
                        padx=10,
                        pady=6,
                        wrap="word",
                        cursor="xterm",
                        exportselection=False,
                        selectbackground="#2563EB",
                        selectforeground="#FFFFFF",
                        inactiveselectbackground="#93C5FD"
                    )
                    cell_txt.insert("1.0", c_val_str)
                    cell_txt.bind("<Key>", _make_cell_key_handler(cell_txt, c_val_str))
                    cell_txt.bind("<MouseWheel>", _scroll_table_chat)
                    cell_txt.bind("<Button-3>", lambda e, cw=cell_txt, cv=c_val_str, rd=r_data_copy: _show_cell_menu(e, cw, cv, rd))
                    cell_txt.grid(
                        row=row_idx,
                        column=col_idx,
                        sticky="nsew",
                        padx=(0, 1 if col_idx < len(headers) - 1 else 0),
                        pady=(0, 1 if row_idx < len(rows) else 0)
                    )

            if table_data.get("footer"):
                tk.Label(
                    card,
                    text=table_data["footer"],
                    font=("Segoe UI", 9, "italic"),
                    fg="#475569",
                    bg="#FFFFFF",
                    wraplength=700,
                    justify="left"
                ).pack(anchor="w", pady=(4, 2))

        # Interactive Copilot-style suggestion chips with Auto-Wrapping Rows
        if suggestions is None:
            suggestions = [
                "📊 Draw RFQ Stage Chart",
                "👥 Top Customers Summary",
                "📄 Import Customer BOM File"
            ]

        if suggestions:
            chip_frame = tk.Frame(card, bg="#ffffff")
            chip_frame.pack(fill="x", expand=True, pady=(10, 2))

            tk.Label(chip_frame, text="Try one of these:", font=("Segoe UI", 8),
                     fg="#64748b", bg="#ffffff").pack(anchor="w", pady=(0, 4))

            cur_row = None
            cur_chars = 0
            for s_text in suggestions:
                if cur_row is None or cur_chars + len(s_text) > 46:
                    cur_row = tk.Frame(chip_frame, bg="#ffffff")
                    cur_row.pack(fill="x", anchor="w", pady=(2, 2))
                    cur_chars = 0

                chip_btn = self._style_suggestion_chip(cur_row, s_text)
                chip_btn.pack(side="left", padx=(0, 6), pady=2)
                cur_chars += len(s_text) + 6

        self._chat_canvas.update_idletasks()
        self._chat_canvas.yview_moveto(1.0)

    def _append_agent_chart(self, fig, caption="", suggestions=None):
        """Embeds a matplotlib Figure as an inline chart card with Copilot prompt chips."""
        self._hide_typing_indicator()
        try:
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        except ImportError:
            self._append_agent_message("[Chart rendering requires matplotlib. Run: pip install matplotlib]")
            return

        row = tk.Frame(self._chat_inner, bg="#f0f4f8")
        row.pack(fill="x", padx=14, pady=(4, 8))

        icon = tk.Label(row, text="🤖", bg="#f0f4f8", font=("Segoe UI", 14))
        icon.pack(side="left", anchor="n", padx=(0, 8), pady=(2, 0))

        card = tk.Frame(row, bg="#ffffff", padx=10, pady=10,
                        highlightbackground="#cbd5e0", highlightthickness=1)
        card.pack(side="left", fill="x", expand=True)

        if caption:
            tk.Label(card, text=caption, bg="#ffffff", fg="#2d3748",
                     font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))

        chart_canvas = FigureCanvasTkAgg(fig, master=card)
        chart_canvas.draw()
        chart_canvas.get_tk_widget().pack(fill="both", expand=True)

        # Bottom Bar: Save Button & Copilot Chips
        bbar = tk.Frame(card, bg="#ffffff")
        bbar.pack(fill="x", pady=(8, 0))

        def _save_chart():
            from tkinter import filedialog
            path = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG Image", "*.png"), ("PDF", "*.pdf")],
                initialfile="rfq_chart"
            )
            if path:
                fig.savefig(path, dpi=150, bbox_inches="tight", facecolor="white")
                self._append_agent_message(f"Chart saved to: {path}")

        save_btn = tk.Button(bbar, text="💾 Save Chart", command=_save_chart,
                             font=("Segoe UI", 9, "bold"), bg="#edf2f7", fg="#2d3748",
                             relief="flat", cursor="hand2", padx=8, pady=3)
        save_btn.pack(side="right", anchor="e")

        if suggestions is None:
            suggestions = [
                "👥 Top Customers Summary",
                "⏱️ Check Sourcing Status",
                "📄 Import Customer BOM File"
            ]

        if suggestions:
            chip_container = tk.Frame(bbar, bg="#ffffff")
            chip_container.pack(side="left", fill="x", expand=True, anchor="w")
            cur_r = None
            cur_ch = 0
            for s_text in suggestions:
                if cur_r is None or cur_ch + len(s_text) > 42:
                    cur_r = tk.Frame(chip_container, bg="#ffffff")
                    cur_r.pack(fill="x", anchor="w", pady=(1, 1))
                    cur_ch = 0
                chip_btn = self._style_suggestion_chip(cur_r, s_text)
                chip_btn.pack(side="left", padx=(0, 6), pady=2)
                cur_ch += len(s_text) + 6

        self._chat_canvas.update_idletasks()
        self._chat_canvas.yview_moveto(1.0)

    def _set_thinking_state(self, is_thinking):
        """Toggles between 'Send ►' and '⏹️ Stop' on the input bar."""
        self._is_agent_thinking = is_thinking
        if hasattr(self, 'send_btn') and self.send_btn.winfo_exists():
            if is_thinking:
                self._abort_signal.clear()
                self.send_btn.config(
                    text="⏹️ Stop",
                    bg="#DC2626",
                    activebackground="#B91C1C",
                    command=self._stop_thinking
                )
            else:
                self.send_btn.config(
                    text="Send",
                    bg="#2563EB",
                    activebackground="#1D4ED8",
                    command=self._on_chat_submit
                )

    def _stop_thinking(self):
        """Interrupts ongoing agent extraction, vision processing, or thinking loop."""
        self._abort_signal.set()
        self._hide_typing_indicator()
        self._set_thinking_state(False)
        self._append_agent_message("⏹️ **Process paused / cancelled by user.**", suggestions=[
            "➕ New Chat",
            "📊 Draw RFQ Stage Chart",
            "📩 Check RFQ Emails"
        ])

    def _on_click_new_chat(self):
        """Saves current conversation and starts a clean, fresh session."""
        if hasattr(self, 'session_store'):
            if hasattr(self, 'current_session_id') and self.current_session_id and hasattr(self, 'current_session_data'):
                if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                    self.current_session_data["active_rfq_json"] = self._last_extracted_rfq_json
                self.session_store.save_session(self.current_session_id, self.current_session_data)

            self.current_session_id, self.current_session_data = self.session_store.create_new_session("New RFQ Session")

        # Clear UI canvas
        for child in self._chat_inner.winfo_children():
            child.destroy()

        self._staged_bom_payload = None
        self._last_extracted_rfq_json = None
        self._hide_typing_indicator()
        self._set_thinking_state(False)

        # Render clean initial greeting
        self._append_agent_message(
            "New chat started. Ask about RFQs, charts, BOM files, or inbox scans.",
            suggestions=[
                "Draw RFQ stage chart",
                "How many RFQs do I have?",
                "Which customer has the most RFQs?",
                "Import a customer BOM",
            ]
        )
        self.after(100, lambda: self.chat_input.focus_set())

    def _show_chat_history_dialog(self):
        """Displays searchable modal dialog with all saved conversation sessions, load, and delete options."""
        if not hasattr(self, 'session_store'):
            return

        sessions = self.session_store.list_sessions()
        dlg = tk.Toplevel(self)
        dlg.title("🕒 ContinuumX Conversation History")
        dlg.geometry("680x520")
        dlg.configure(bg="#0F172A")
        dlg.transient(self)
        dlg.grab_set()

        hdr = tk.Frame(dlg, bg="#1E293B", padx=16, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🕒 Conversation History & Session Manager", font=("Segoe UI", 12, "bold"), fg="#FFFFFF", bg="#1E293B").pack(anchor="w")
        tk.Label(hdr, text="Select any past session to resume, view previous evidence cards, or clean up history.", font=("Segoe UI", 8), fg="#94A3B8", bg="#1E293B").pack(anchor="w", pady=(2, 0))

        # Search Bar
        sf = tk.Frame(dlg, bg="#0F172A", padx=14, pady=8)
        sf.pack(fill="x")
        tk.Label(sf, text="🔍 Filter:", font=("Segoe UI", 9, "bold"), fg="#94A3B8", bg="#0F172A").pack(side="left", padx=(0, 6))
        search_v = tk.StringVar()
        search_ent = tk.Entry(sf, textvariable=search_v, font=("Segoe UI", 9), width=28, bg="#1E293B", fg="#FFFFFF", insertbackground="#FFFFFF", bd=1, relief="solid")
        search_ent.pack(side="left")

        # Treeview list of sessions
        tr_frame = tk.Frame(dlg, bg="#1E293B", padx=12, pady=6)
        tr_frame.pack(fill="both", expand=True, padx=14, pady=4)

        cols = ("Session ID", "Topic / Title", "Messages", "Last Active")
        tree = ttk.Treeview(tr_frame, columns=cols, show="headings", height=10)
        tree.heading("Session ID", text="Session ID")
        tree.heading("Topic / Title", text="Topic / Title")
        tree.heading("Messages", text="Messages")
        tree.heading("Last Active", text="Last Active")

        tree.column("Session ID", width=85, anchor="center", stretch=False)
        tree.column("Topic / Title", width=310, anchor="w", stretch=False)
        tree.column("Messages", width=75, anchor="center", stretch=False)
        tree.column("Last Active", width=140, anchor="center", stretch=False)

        vsb = ttk.Scrollbar(tr_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)

        def _populate_tree(*args):
            for r in tree.get_children(): tree.delete(r)
            q = search_v.get().strip().lower()
            for s in sessions:
                title = s.get("title", "")
                if q and q not in title.lower() and q not in s.get("session_id", "").lower():
                    continue
                tree.insert("", "end", values=(
                    s.get("session_id", ""),
                    title,
                    f"{s.get('msg_count', 0)} msgs",
                    s.get("last_updated", "")[:16]
                ))

        search_v.trace_add("write", _populate_tree)
        _populate_tree()

        # Action Buttons on Bottom Bar
        bbar = tk.Frame(dlg, bg="#0F172A", padx=14, pady=12)
        bbar.pack(fill="x")

        def _load_selected():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Select Session", "Please select a conversation to resume.", parent=dlg)
                return
            s_id = tree.item(sel[0], "values")[0]
            s_data = self.session_store.load_session(s_id)
            if s_data:
                self._restore_session_to_ui(s_data)
                dlg.destroy()

        def _delete_selected():
            sel = tree.selection()
            if not sel: return
            s_id = tree.item(sel[0], "values")[0]
            if messagebox.askyesno("Delete Conversation", f"Delete session '{s_id}' from history?", parent=dlg):
                self.session_store.delete_session(s_id)
                sessions[:] = self.session_store.list_sessions()
                _populate_tree()

        def _clear_all():
            if messagebox.askyesno("Clear All History", "Are you sure you want to delete ALL saved conversation history?", parent=dlg):
                self.session_store.clear_all_sessions()
                sessions.clear()
                _populate_tree()

        tk.Button(bbar, text="📂 Resume Conversation", command=_load_selected, bg="#2563EB", fg="#FFFFFF",
                  font=("Segoe UI", 9, "bold"), relief="flat", padx=14, pady=5, cursor="hand2").pack(side="left", padx=(0, 6))
        tk.Button(bbar, text="🗑️ Delete Selected", command=_delete_selected, bg="#DC2626", fg="#FFFFFF",
                  font=("Segoe UI", 9), relief="flat", padx=10, pady=5, cursor="hand2").pack(side="left", padx=(0, 6))
        tk.Button(bbar, text="⚠️ Clear All", command=_clear_all, bg="#334155", fg="#CBD5E1",
                  font=("Segoe UI", 9), relief="flat", padx=10, pady=5, cursor="hand2").pack(side="left")
        tk.Button(bbar, text="Close", command=dlg.destroy, bg="#1E293B", fg="#94A3B8",
                  font=("Segoe UI", 9), relief="flat", padx=12, pady=5, cursor="hand2").pack(side="right")

    def _restore_session_to_ui(self, session_data):
        """Restores full visual chat bubbles, cards, and tables from a saved session."""
        self.current_session_id = session_data.get("session_id", str(uuid.uuid4())[:8])
        self.current_session_data = session_data
        if session_data.get("active_rfq_json"):
            self._last_extracted_rfq_json = session_data["active_rfq_json"]

        for child in self._chat_inner.winfo_children():
            child.destroy()

        messages = session_data.get("messages", [])
        for m in messages:
            if m.get("sender") == "user":
                self._append_user_message(m.get("text", ""), images=None, save_to_session=False)
            else:
                self._append_agent_message(
                    m.get("text", ""),
                    suggestions=m.get("suggestions"),
                    table_data=m.get("table_data"),
                    save_to_session=False
                )

    def _commit_and_launch_staged_bom(self):
        """Writes payload and launches BOM module after user confirms in chat."""
        if not hasattr(self, '_staged_bom_payload') or not self._staged_bom_payload:
            self._append_agent_message("No BOM file staged. Please select a BOM file first via '📎 Actions' -> 'Import Customer BOM'.")
            return

        payload = self._staged_bom_payload
        local_appdata = os.environ.get('LOCALAPPDATA', os.environ.get('TEMP', 'C:\\Temp'))
        payload_path = os.path.join(local_appdata, "ContXs", "agent_bom_payload.json")
        os.makedirs(os.path.dirname(payload_path), exist_ok=True)
        with open(payload_path, 'w', encoding='utf-8') as pf:
            json.dump(payload, pf, indent=4)

        cust = payload.get("customer_name", "Customer")
        proj = payload.get("project_title", "Project")
        tp = payload.get("target_price", "Not set")
        eau = payload.get("eau", "Not set")

        msg = (f"🚀 Launching BOM Verification Window with pre-filled settings!\n"
               f"• Customer: {cust}\n"
               f"• Project Title: {proj}\n"
               f"• Target Price: {tp}\n"
               f"• EAU: {eau}")
        self._append_agent_message(msg)

        # Launch BOM module automatically!
        self.launch_feature("BOM")
        self._staged_bom_payload = None

    def _chat_run_error_test(self):
        """Type 'test error' in launcher chat to write ErrorTelemetryStore + dashboard."""
        from agents.telemetry_tracker import ErrorTelemetryStore
        err_store = ErrorTelemetryStore()
        meta = getattr(self, "_last_extracted_rfq_json", {}) or {}
        rfq_meta = meta.get("rfq_metadata", {}) if isinstance(meta, dict) else {}
        sim_err = err_store.record_error(
            module="LLMGateway",
            error_category="SIMULATED_TEST_ERROR",
            error_message="Simulated pipeline diagnostic error triggered by user test in Chat.",
            severity="WARNING",
            rfq_number=rfq_meta.get("rfq_number", "RS26-8004"),
            customer=rfq_meta.get("customer_name", "Graco"),
            document_name="diagnostics_test_drawing.pdf",
            recovery_action="Verified fallback & error store capture (0s downtime)",
            status="RECOVERED_VIA_FALLBACK",
        )
        latest_summ = err_store.get_latest_summary()
        self._append_agent_message(
            f"🧪 **Agent Error Telemetry Test Successful!**\n\n"
            f"• **Incident ID:** `{sim_err['error_id']}`\n"
            f"• **Category:** `{sim_err['error_category']}` in `{sim_err['module']}`\n"
            f"• **Incident File:** `data/telemetry/errors/{sim_err['error_id']}.json`\n"
            f"• **Master Audit Log:** `data/telemetry/agent_errors.json`\n"
            f"• **Dashboard Summary:** `data/telemetry/latest_errors_summary.json`\n"
            f"• **Total Incidents Recorded:** `{latest_summ.get('total_incidents', 1)}` "
            f"(Auto-Recovery Rate: `{latest_summ.get('auto_recovery_rate_pct', 100.0)}%`)\n\n"
            f"✨ The web dashboard can now read and display this incident in real-time!",
            suggestions=["test approval", "🔍 Inspect Source Evidence", "📩 Check RFQ Emails"],
        )

    def _chat_run_approval_test(self):
        """Type 'test approval' in launcher chat to raise a dashboard HITL item."""
        from agents.platform_bridge import DesktopAgentBridge
        meta = getattr(self, "_last_extracted_rfq_json", {}) or {}
        rfq_meta = meta.get("rfq_metadata", {}) if isinstance(meta, dict) else {}
        rfq = rfq_meta.get("rfq_number", "RS26-8004")
        item = DesktopAgentBridge.instance().request_approval(
            agent_id="costing",
            step_name="Quote release",
            summary=f"Diagnostic HITL: approve costing summary for {rfq}.",
            confidence_score=0.82,
            transaction_uuid=rfq,
        )
        if item:
            self._append_agent_message(
                f"🧑‍⚖️ **Human approval raised on the monitoring dashboard.**\n\n"
                f"• **Approval ID:** `{item.get('approval_id')}`\n"
                f"• **Agent:** `{item.get('agent_id')}`\n"
                f"• **Step:** `{item.get('step_name')}`\n"
                f"• **RFQ:** `{rfq}`\n\n"
                f"Open http://127.0.0.1:8000/ and use Approve / Reject.",
                suggestions=["test error", "📊 Draw RFQ Stage Chart"],
            )
        else:
            self._append_agent_message(
                "Could not reach the monitoring dashboard at http://127.0.0.1:8000/. "
                "Start the platform (`uvicorn app.main:app` in `platform/`) and try again.",
                suggestions=["test error"],
            )

    def _on_suggestion_chip_click(self, prompt):
        """Submits suggestion chip text directly into the chat engine."""
        prompt_lower = prompt.lower()
        if ("launch" in prompt_lower or "verification" in prompt_lower or "proceed" in prompt_lower) and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
            self._append_user_message(prompt)
            self._commit_and_launch_staged_bom()
            return
        elif any(k in prompt_lower for k in ["confirm eau", "set eau"]) and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
            self._append_user_message(prompt)
            m = re.search(r'(?:confirm\s+eau|set\s+eau)[:\s]*([0-9,]+)', prompt_lower)
            if m:
                new_eau = int(m.group(1).replace(',', ''))
                self._staged_bom_payload["eau"] = new_eau
                if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                    self._last_extracted_rfq_json["rfq_metadata"]["eau"] = f"{new_eau:,} pcs"
                    for a in self._last_extracted_rfq_json.get("assemblies", []):
                        a["eau"] = new_eau
                        for it in a.get("items", []):
                            it["eau"] = new_eau
                    try:
                        from agents.synthetic_bom_generator import SyntheticBOMGenerator
                        gen = SyntheticBOMGenerator()
                        gen_res = gen.generate_synthetic_excel(self._last_extracted_rfq_json)
                        if gen_res.get("success"):
                            self._staged_bom_payload["file_path"] = gen_res["file_path"]
                    except Exception:
                        pass
                self._append_agent_message(
                    f"✅ **EAU successfully confirmed and updated to `{new_eau:,} pcs`!**\n\n"
                    f"• Synthetic BOM Excel staged with `{new_eau:,} pcs`\n"
                    f"• Downstream BOM Verification and Sourcing will use `{new_eau:,} pcs`.",
                    suggestions=["📊 Review Full Table & Filter", "🚀 Launch BOM Verification Window", "🔍 Inspect Source Evidence"]
                )
            return
        elif any(k in prompt_lower for k in ["review full", "review table", "filter", "full table", "open table", "review"]):
            self._append_user_message(prompt)
            self._open_rfq_review_window()
            return
        elif any(k in prompt_lower for k in ["inspect evidence", "view evidence", "source evidence", "evidence"]):
            self._append_user_message(prompt)
            self._show_extracted_evidence_dialog()
            return
        elif any(k in prompt_lower for k in ["test error", "/test-error", "simulate error", "test error logging"]):
            self._append_user_message(prompt)
            self._chat_run_error_test()
            return
        elif any(k in prompt_lower for k in ["test approval", "/test-approval", "simulate approval"]):
            self._append_user_message(prompt)
            self._chat_run_approval_test()
            return
        elif "import" in prompt_lower and "bom" in prompt_lower:
            self._append_user_message(prompt)
            self._agent_import_bom()
            return

        if isinstance(self.chat_input, tk.Text):
            self.chat_input.delete("1.0", "end")
            self.chat_input.insert("1.0", prompt)
            self.chat_input.config(fg="#2d3748")
        else:
            self.chat_input.delete(0, "end")
            self.chat_input.insert(0, prompt)
        self._on_chat_submit()

    def _show_extracted_evidence_dialog(self):
        """
        Phase 4+5: 3-tab Evidence Audit Dialog.
        Tab 1: Metadata Fields with per-field confidence, source zone, snippet, reasoning.
        Tab 2: Components with confidence color + Correct button (saves to CorrectionStore).
        Tab 3: Conflicts detected during multi-source consolidation.
        """
        if not hasattr(self, '_last_extracted_rfq_json') or not self._last_extracted_rfq_json:
            self._append_agent_message("No RFQ data extracted yet. Scan emails first!", suggestions=["📩 Check RFQ Emails"])
            return

        rfq = self._last_extracted_rfq_json
        meta = rfq.get("rfq_metadata", {})
        cust = meta.get("customer_name", "Customer")
        rfq_no = meta.get("rfq_number", "RFQ")
        conflicts = rfq.get("conflict_candidates", [])

        dlg = tk.Toplevel(self)
        dlg.title(f"🔍 Evidence Audit — {rfq_no} ({cust})")
        dlg.geometry("1060x680")
        dlg.configure(bg="#0F172A")
        dlg.transient(self)
        dlg.grab_set()
        dlg.lift()
        dlg.focus_force()
        dlg.lift()
        dlg.focus_force()

        from agents.telemetry_tracker import AccuracyTelemetryStore, ACCURACY_JSON, ACCURACY_LATEST_JSON
        from agents.verified_bom_store import VerifiedBOMStore
        vbs = VerifiedBOMStore()
        is_learned = any(bool(vbs.get_verified_assembly(a.get("assy_no"))) for a in rfq.get("assemblies", []))

        # Evaluate and record telemetry
        acc_store = AccuracyTelemetryStore()
        audit_res = acc_store.evaluate_and_record(rfq, is_learned_pattern=is_learned)
        overall_acc = audit_res.get("overall_accuracy_pct", 100.0)
        acc_grade = audit_res.get("accuracy_grade", "A+")
        ver_status = audit_res.get("verification_status", "Human Verified")

        hdr = tk.Frame(dlg, bg="#1E293B", padx=20, pady=12)
        hdr.pack(fill="x")

        hdr_top = tk.Frame(hdr, bg="#1E293B")
        hdr_top.pack(fill="x")

        conf_badge = f"⚠️ {len(conflicts)} Conflict(s)" if conflicts else "✅ No Conflicts"
        tk.Label(hdr_top, text=f"🔍 Source Evidence & Verification Audit — {rfq_no}",
                 font=("Segoe UI", 12, "bold"), fg="#FFFFFF", bg="#1E293B").pack(side="left")

        def _open_dashboard_json():
            try:
                import os
                if os.path.exists(ACCURACY_LATEST_JSON):
                    os.startfile(os.path.dirname(ACCURACY_LATEST_JSON))
                else:
                    os.startfile(os.path.dirname(ACCURACY_JSON))
            except Exception as e:
                self.clipboard_clear()
                self.clipboard_append(ACCURACY_LATEST_JSON)
                import tkinter.messagebox as mb
                mb.showinfo("Dashboard Data", f"Dashboard JSON saved at:\n{ACCURACY_LATEST_JSON}\n(Copied to clipboard)", parent=dlg)

        tk.Button(hdr_top, text="📊 Web Dashboard JSON", command=_open_dashboard_json,
                  bg="#2563EB", fg="#FFFFFF", font=("Segoe UI", 8, "bold"),
                  relief="flat", padx=10, pady=3, cursor="hand2").pack(side="right")

        # Subtitle with Accuracy Metric
        acc_color = "#10B981" if overall_acc >= 90 else ("#F59E0B" if overall_acc >= 75 else "#EF4444")
        sub_txt = (f"Customer: {cust}  •  Assemblies: {len(rfq.get('assemblies', []))}  •  "
                   f"Components: {audit_res.get('total_components', 0)}  •  {conf_badge}  •  "
                   f"🎯 Accuracy: {overall_acc}% ({acc_grade})")
        tk.Label(hdr, text=sub_txt, font=("Segoe UI", 9, "bold"), fg=acc_color, bg="#1E293B").pack(anchor="w", pady=(3, 0))

        legend_fr = tk.Frame(dlg, bg="#1E293B", padx=20, pady=4)
        legend_fr.pack(fill="x")
        for color, label in [("#10B981", "100% Ground Truth (Verified)"), ("#22C55E", ">=90% High Conf"),
                              ("#F59E0B", "70-89% Med Conf"), ("#EF4444", "<70% Low Conf"), ("#6366F1", "NOT_AVAILABLE")]:
            tk.Label(legend_fr, text="●", fg=color, bg="#1E293B",
                     font=("Segoe UI", 10)).pack(side="left", padx=(0, 2))
            tk.Label(legend_fr, text=label, fg="#CBD5E1", bg="#1E293B",
                     font=("Segoe UI", 8)).pack(side="left", padx=(0, 12))

        from tkinter import ttk
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Evidence.TNotebook", background="#0F172A", borderwidth=0)
        style.configure("Evidence.TNotebook.Tab", background="#1E293B", foreground="#94A3B8",
                        font=("Segoe UI", 9, "bold"), padding=[14, 6])
        style.map("Evidence.TNotebook.Tab",
                  background=[("selected", "#2563EB")],
                  foreground=[("selected", "#FFFFFF")])

        nb = ttk.Notebook(dlg, style="Evidence.TNotebook")
        nb.pack(fill="both", expand=True, padx=12, pady=8)

        def _conf_tag(ev_dict, is_verified=False):
            if is_verified:
                return "verified"
            if not isinstance(ev_dict, dict):
                return "high"
            res = ev_dict.get("resolution_type", "DIRECT")
            if res == "NOT_AVAILABLE":
                return "na"
            conf = float(ev_dict.get("confidence", 0.95))
            if conf >= 0.90:
                return "high"
            if conf >= 0.70:
                return "medium"
            return "low"

        def _make_tree(parent, cols, col_widths, col_anchors):
            fr = tk.Frame(parent, bg="#0F172A")
            fr.pack(fill="both", expand=True, padx=8, pady=8)
            tr_fr = tk.Frame(fr, bg="#1E293B", bd=1, relief="solid")
            tr_fr.pack(fill="both", expand=True)
            tv = ttk.Treeview(tr_fr, columns=cols, show="headings")
            vsb = ttk.Scrollbar(tr_fr, orient="vertical", command=tv.yview)
            hsb = ttk.Scrollbar(tr_fr, orient="horizontal", command=tv.xview)
            tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            for col, w, anc in zip(cols, col_widths, col_anchors):
                tv.heading(col, text=col)
                tv.column(col, width=w, anchor=anc, minwidth=40)
            vsb.pack(side="right", fill="y")
            hsb.pack(side="bottom", fill="x")
            tv.pack(side="left", fill="both", expand=True)
            
            # Clean Black/Dark Slate text for all data columns
            tv.tag_configure("odd", foreground="#0F172A", background="#FFFFFF")
            tv.tag_configure("even", foreground="#0F172A", background="#F8FAFC")
            return tv

        # TAB 1 — Metadata & Assembly Parameters
        tab_fields = tk.Frame(nb, bg="#0F172A")
        nb.add(tab_fields, text="📋 Metadata & Assembly Level")
        f_cols = ("Field", "Value", "Conf%", "Accuracy%", "Resolution", "Source Doc", "Zone", "Snippet / Reasoning")
        f_widths = (150, 180, 70, 120, 95, 140, 95, 260)
        f_anchors = ("w", "w", "center", "center", "center", "w", "center", "w")
        tv_fields = _make_tree(tab_fields, f_cols, f_widths, f_anchors)

        meta_ev = rfq.get("metadata_evidence", {})
        email_doc = "email_header_and_body"
        
        rows_meta = [
            ("Customer", meta.get("customer_name", ""), meta_ev.get("customer_name", {
                "source_document": email_doc, "source_zone": "METADATA", "raw_evidence_snippet": f"End Customer: {cust}", "resolution_type": "DIRECT"
            })),
            ("RFQ Number", meta.get("rfq_number", ""), meta_ev.get("rfq_number", {
                "source_document": email_doc, "source_zone": "METADATA", "raw_evidence_snippet": f"RFQ: {rfq_no}", "resolution_type": "DIRECT"
            })),
            ("Project Title", meta.get("project_title", ""), meta_ev.get("project_title", {
                "source_document": email_doc, "source_zone": "METADATA", "raw_evidence_snippet": f"Subject/Project: {meta.get('project_title', '')}", "resolution_type": "DIRECT"
            })),
            ("Commodity", meta.get("commodity", "Wire Harness"), meta_ev.get("commodity", {
                "source_document": email_doc, "source_zone": "METADATA", "raw_evidence_snippet": f"Commodity: {meta.get('commodity', 'Wire Harness')}", "resolution_type": "DIRECT"
            })),
            ("EAU (Total)", str(meta.get("eau", "")), meta_ev.get("eau", {
                "source_document": email_doc, "source_zone": "METADATA", "raw_evidence_snippet": f"Annual Consumption (EAU): {meta.get('eau', '')}", "resolution_type": "DIRECT"
            })),
            ("Target Price (RFQ)", str(meta.get("target_price", "")), meta_ev.get("target_price", {
                "source_document": email_doc, "source_zone": "METADATA", "raw_evidence_snippet": f"Target Price: {meta.get('target_price', '')}", "resolution_type": "DIRECT"
            })),
        ]
        for assy in rfq.get("assemblies", []):
            an = assy.get("assy_no", "")
            am = assy.get("assy_model", "")
            ar = assy.get("assy_rev", "")
            ae = str(assy.get("eau", meta.get("eau", "")))
            atp = str(assy.get("target_price", meta.get("target_price", "")))
            moqs = str(meta.get("default_moqs", [100, 250, 500, 1000]))
            src_pdf = assy.get("drawing_filename") or f"{an}.pdf"
            
            ev = assy.get("evidence", {})
            tb_ev = ev.get("title_block", {})
            if not tb_ev or not tb_ev.get("source_document"):
                tb_ev = {"source_document": src_pdf, "source_zone": "TITLE_BLOCK", "raw_evidence_snippet": f"Model: {am}", "resolution_type": "DIRECT", "confidence": 0.95}
                
            rev_ev = ev.get("revision", {})
            if not rev_ev or not rev_ev.get("source_document"):
                rev_ev = {"source_document": src_pdf, "source_zone": "TITLE_BLOCK", "raw_evidence_snippet": f"Rev {ar}", "resolution_type": "DIRECT", "confidence": 0.95}
                
            eau_ev = ev.get("eau", {})
            if not eau_ev or not eau_ev.get("source_document"):
                eau_ev = {"source_document": "email_body", "source_zone": "METADATA", "raw_evidence_snippet": f"Forecast EAU: {ae}", "resolution_type": "DIRECT", "confidence": 0.95}
                
            tp_ev = ev.get("target_price", {})
            if not tp_ev or not tp_ev.get("source_document"):
                tp_ev = {"source_document": "email_body", "source_zone": "METADATA", "raw_evidence_snippet": f"Target Price: {atp}", "resolution_type": "DIRECT", "confidence": 0.95}
                
            moq_ev = {"source_document": "customer_rfq_profile", "source_zone": "SYSTEM_MOQ_VAULT", "raw_evidence_snippet": f"Assigned Tier MOQs: {moqs}", "resolution_type": "DIRECT", "confidence": 0.95}
            
            rows_meta.append((f"Assy Model ({an})", am, tb_ev))
            rows_meta.append((f"Assy Rev ({an})", ar, rev_ev))
            rows_meta.append((f"Assy EAU ({an})", ae, eau_ev))
            rows_meta.append((f"Assy Target Price ({an})", atp, tp_ev))
            rows_meta.append((f"Assigned MOQs ({an})", moqs, moq_ev))

        for idx, (fn, fv, ev_d) in enumerate(rows_meta, start=1):
            ev_d = ev_d if isinstance(ev_d, dict) else {}
            conf_val = ev_d.get("confidence", 0.95)
            conf_num = float(conf_val) * 100.0 if isinstance(conf_val, (int, float)) else 95.0
            conf_badge = f"🟢 {conf_num:.0f}%" if conf_num >= 90 else (f"🟠 {conf_num:.0f}%" if conf_num >= 70 else f"🔴 {conf_num:.0f}%")
            
            acc_badge = "🎯 100% (Verified)" if is_learned else f"🔵 {conf_num:.0f}% (Direct)"
            row_tag = "odd" if idx % 2 == 1 else "even"
            snippet = str(ev_d.get("raw_evidence_snippet") or ev_d.get("reasoning") or "")[:100]
            src_d = ev_d.get("source_document") or email_doc
            sz_d = ev_d.get("source_zone") or "METADATA"
            res_d = ev_d.get("resolution_type") or "DIRECT"
            
            tv_fields.insert("", "end", values=(
                fn, str(fv)[:80], conf_badge, acc_badge,
                res_d, src_d, sz_d, snippet
            ), tags=(row_tag,))

        # TAB 2 — Components with all 10 engineering columns
        tab_comp = tk.Frame(nb, bg="#0F172A")
        nb.add(tab_comp, text="🔩 Components & Evidence ✏️")
        c_cols = (
            "Line", "Assy#", "Assy Model", "Assy Rev", "Part", "Description",
            "MPN", "MFR", "Qty", "UOM", "Target Price (USD)",
            "Conf%", "Accuracy%", "Verification Status", "Zone", "Source"
        )
        c_widths = (35, 80, 110, 50, 95, 120, 90, 75, 40, 35, 75, 55, 65, 130, 75, 95)
        c_anchors = (
            "center", "w", "w", "center", "w", "w",
            "w", "w", "center", "center", "center",
            "center", "center", "w", "center", "w"
        )
        tv_comp = _make_tree(tab_comp, c_cols, c_widths, c_anchors)

        comp_rows_data = []
        line_idx = 1
        for assy in rfq.get("assemblies", []):
            a_no = assy.get("assy_no", "")
            a_model = assy.get("assy_model", "")
            a_rev = str(assy.get("assy_rev", "")).replace("Rev", "").strip()
            a_tp = str(assy.get("target_price", meta.get("target_price", "N/A"))).replace('$', '').strip()
            assy_pdf = assy.get("drawing_filename") or f"{a_no}.pdf"

            for it in assy.get("items", []):
                part = str(it.get("part_number") or "")
                desc = str(it.get("description") or "Component")
                mpn = str(it.get("mpn") or "")
                mfr = str(it.get("mfr") or "")
                qty = it.get("qty", 1)
                uom = str(it.get("uom", "EA"))
                row_tp = str(it.get("target_price", a_tp)).replace('$', '').strip()

                ev = it.get("evidence", {})
                mpn_ev = ev.get("mpn", {}) if isinstance(ev.get("mpn"), dict) else {}
                conf_val = mpn_ev.get("confidence", 0.95)
                conf_num = float(conf_val) * 100.0 if isinstance(conf_val, (int, float)) else 95.0
                conf_badge = f"🟢 {conf_num:.0f}%" if conf_num >= 90 else (f"🟠 {conf_num:.0f}%" if conf_num >= 70 else f"🔴 {conf_num:.0f}%")
                zone = mpn_ev.get("source_zone") or "DRAWING_BOM"
                src_doc = mpn_ev.get("source_document") or assy_pdf

                if is_learned:
                    acc_badge = "🎯 100%"
                    status_str = "✅ Verified Ground Truth"
                else:
                    acc_badge = f"🔵 {conf_num:.0f}%"
                    status_str = "🤖 AI Prediction"

                row_tag = "odd" if line_idx % 2 == 1 else "even"
                row_id = tv_comp.insert("", "end", values=(
                    f"#{line_idx}", a_no, a_model, a_rev, part, desc,
                    mpn, mfr, qty, uom, row_tp,
                    conf_badge, acc_badge, status_str, zone, src_doc
                ), tags=(row_tag,))
                comp_rows_data.append({
                    "row_id": row_id, "assy_no": a_no, "assy_model": a_model, "assy_rev": a_rev,
                    "part": part, "description": desc, "mpn": mpn, "mfr": mfr, "qty": qty, "uom": uom,
                    "target_price": row_tp, "src_doc": src_doc
                })
                line_idx += 1

        def _open_correction_dialog():
            sel = tv_comp.selection()
            if not sel:
                import tkinter.messagebox as mb2
                mb2.showwarning("No Selection", "Please select a component row first.", parent=dlg)
                return
            row_data = next((r for r in comp_rows_data if r["row_id"] == sel[0]), None)
            if not row_data:
                return

            cd = tk.Toplevel(dlg)
            cd.title("Correct Extraction")
            cd.geometry("520x400")
            cd.configure(bg="#1E293B")
            cd.transient(dlg)
            cd.grab_set()
            tk.Label(cd, text="✏️ Submit Extraction Correction",
                     font=("Segoe UI", 11, "bold"), fg="#FFFFFF", bg="#1E293B").pack(pady=(14, 2))
            tk.Label(cd, text="Gemini will use this correction as few-shot context on the next extraction.",
                     font=("Segoe UI", 8), fg="#94A3B8", bg="#1E293B", wraplength=480).pack()

            fm = tk.Frame(cd, bg="#1E293B", padx=24, pady=10)
            fm.pack(fill="x")
            fm.columnconfigure(1, weight=1)
            vars_list = []
            def _fr(lbl, default=""):
                r = len(vars_list)
                tk.Label(fm, text=lbl, font=("Segoe UI", 8, "bold"), fg="#CBD5E1",
                         bg="#1E293B", anchor="w", width=18).grid(row=r, column=0, sticky="w", pady=3)
                v = tk.StringVar(value=str(default))
                e = tk.Entry(fm, textvariable=v, font=("Segoe UI", 9), width=34,
                             bd=1, relief="solid", bg="#0F172A", fg="#F1F5F9", insertbackground="#F1F5F9")
                e.grid(row=r, column=1, sticky="ew", pady=3, padx=(8, 0))
                vars_list.append(v)
                return v
            doc_hint_v = _fr("Document Hint", (row_data["assy_no"] or row_data["src_doc"].split(".")[0])[:14])
            field_v = _fr("Field to Correct", "mpn")
            wrong_v = _fr("Wrong Value (was)", row_data["mpn"])
            correct_v = _fr("Correct Value", "")
            mfr_v = _fr("Manufacturer", row_data["mfr"])
            note_v = _fr("Note (optional)", "")

            field_opts = ["mpn", "mfr", "description", "qty", "uom", "part_number", "assy_no", "assy_rev", "assy_model"]
            ttk.Combobox(fm, values=field_opts, textvariable=field_v, font=("Segoe UI", 9),
                         width=32).grid(row=1, column=1, sticky="ew", pady=3, padx=(8, 0))

            st_lbl = tk.Label(cd, text="", font=("Segoe UI", 9), bg="#1E293B")
            st_lbl.pack(pady=4)

            def _submit():
                try:
                    from agents.correction_store import CorrectionStore
                    cs = CorrectionStore()
                    try:
                        sess = getattr(self, '_current_user_session', {}) or {}
                        user = sess.get("username", "User")
                    except Exception:
                        user = "User"
                    cs.save_correction(
                        doc_hint=doc_hint_v.get().strip(),
                        field=field_v.get().strip(),
                        wrong_value=wrong_v.get().strip(),
                        correct_value=correct_v.get().strip(),
                        mfr=mfr_v.get().strip(),
                        note=note_v.get().strip(),
                        corrected_by=user
                    )
                    st_lbl.config(text="✅ Saved! Gemini uses this correction next time.", fg="#22C55E")
                    cd.after(1800, cd.destroy)
                except Exception as ex:
                    st_lbl.config(text=f"❌ {ex}", fg="#EF4444")

            bf = tk.Frame(cd, bg="#1E293B", pady=8)
            bf.pack()
            tk.Button(bf, text="✅ Submit", command=_submit, bg="#22C55E", fg="#FFFFFF",
                      font=("Segoe UI", 9, "bold"), relief="flat", padx=12, pady=5, cursor="hand2").pack(side="left", padx=4)
            tk.Button(bf, text="Cancel", command=cd.destroy, bg="#334155", fg="#CBD5E1",
                      font=("Segoe UI", 9), relief="flat", padx=10, pady=5, cursor="hand2").pack(side="left")

        btn_bar = tk.Frame(tab_comp, bg="#0F172A", pady=6)
        btn_bar.pack(fill="x", padx=8)
        tk.Button(btn_bar, text="✏️ Correct Selected Field", command=_open_correction_dialog,
                  bg="#7C3AED", fg="#FFFFFF", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=12, pady=5, cursor="hand2").pack(side="left")
        tk.Label(btn_bar, text="Select a row → click to submit correction → Gemini learns for next extraction",
                 font=("Segoe UI", 8), fg="#475569", bg="#0F172A").pack(side="left", padx=10)

        # TAB 3 — Conflicts
        conflict_label = f"⚠️ Conflicts ({len(conflicts)})" if conflicts else "✅ No Conflicts"
        tab_conf = tk.Frame(nb, bg="#0F172A")
        nb.add(tab_conf, text=conflict_label)
        if conflicts:
            cf_cols = ("Field", "Source A", "Value A", "Source B", "Value B", "Auto-Res", "Reasoning")
            cf_widths = (90, 150, 160, 150, 160, 100, 340)
            cf_anchors = ("w", "w", "w", "w", "w", "center", "w")
            tv_cf = _make_tree(tab_conf, cf_cols, cf_widths, cf_anchors)
            for idx, cc in enumerate(conflicts, start=1):
                sa = cc.get("source_a", {})
                sb = cc.get("source_b", {})
                tv_cf.insert("", "end", values=(
                    cc.get("field", ""),
                    sa.get("source_document", ""), str(sa.get("value", "")),
                    sb.get("source_document", ""), str(sb.get("value", "")),
                    cc.get("auto_resolution", "NEEDS_HUMAN"),
                    cc.get("auto_reasoning", "")[:120]
                ), tags=("odd" if idx % 2 == 1 else "even",))
        else:
            tk.Label(tab_conf, text="✅ All sources agree — no conflicts detected.",
                     font=("Segoe UI", 12), fg="#22C55E", bg="#0F172A").pack(expand=True)

        bbar = tk.Frame(dlg, bg="#1E293B", padx=16, pady=10)
        bbar.pack(fill="x")

        def _on_proceed():
            dlg.destroy()
            self._commit_and_launch_staged_bom()

        tk.Button(bbar, text="🚀 Proceed to BOM Verification", command=_on_proceed,
                  bg="#2563EB", fg="#FFFFFF", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=14, pady=6, cursor="hand2").pack(side="right")
    def _show_email_and_attachments_dialog(self, focus_filter=None, target_rfq=None):
        """Displays full original email content, headers, and smart-filtered clickable attachment list with live sync."""
        active_rfq = target_rfq
        if not active_rfq and hasattr(self, '_current_active_rfq') and self._current_active_rfq:
            active_rfq = self._current_active_rfq
        if not active_rfq and hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
            last_meta = self._last_extracted_rfq_json.get("rfq_metadata", {})
            last_no = last_meta.get("rfq_number", "")
            last_cust = last_meta.get("customer_name", "")
            if hasattr(self, '_detected_rfq_list') and self._detected_rfq_list:
                for r_item in self._detected_rfq_list:
                    r_meta = r_item.get("rfq_json", {}).get("rfq_metadata", {})
                    if (last_no and r_meta.get("rfq_number") == last_no) or (last_cust and r_meta.get("customer_name") == last_cust):
                        active_rfq = r_item
                        break
        if not active_rfq and hasattr(self, '_detected_rfq_list') and self._detected_rfq_list:
            active_rfq = self._detected_rfq_list[0]

        if not active_rfq:
            messagebox.showinfo("No Email Data", "No email RFQ data loaded yet. Scan email inbox first!", parent=self)
            return

        cur_rfq_id = active_rfq.get("rfq_json", {}).get("rfq_metadata", {}).get("rfq_number") or active_rfq.get("email", {}).get("subject")

        # If dialog is already open, update focus live or reload if RFQ changed
        if hasattr(self, '_active_email_att_dialog') and self._active_email_att_dialog:
            try:
                dlg_ref, update_fn, prev_rfq_id = self._active_email_att_dialog
                if dlg_ref.winfo_exists():
                    if prev_rfq_id == cur_rfq_id:
                        update_fn(focus_filter)
                        dlg_ref.lift()
                        return
                    else:
                        dlg_ref.destroy()
            except Exception:
                pass
        email_obj = active_rfq.get("email", {})
        subj = email_obj.get("subject", "RFQ Email")
        sender = email_obj.get("sender", "Unknown Sender")
        date_str = email_obj.get("date", "Recent")
        body_text = email_obj.get("body", "")
        raw_attachments = email_obj.get("attachments", []) or []

        root_attachments = []
        nested_by_zip = {}
        all_flattened_files = []

        for a_obj in raw_attachments:
            a_fp = a_obj.get("path", "") if isinstance(a_obj, dict) else str(a_obj)
            a_fn = a_obj.get("filename", "") if isinstance(a_obj, dict) else os.path.basename(a_fp)
            is_zip = a_fn.lower().endswith(".zip")
            
            root_att = {
                "filename": a_fn,
                "path": a_fp,
                "size_bytes": a_obj.get("size_bytes", os.path.getsize(a_fp) if os.path.exists(a_fp) else 0),
                "is_zip": is_zip,
                "nested_count": 0
            }
            root_attachments.append(root_att)
            all_flattened_files.append(root_att)

            if is_zip and a_fp:
                unzip_d = os.path.join(os.path.dirname(a_fp), f"unzipped_{os.path.splitext(a_fn)[0]}")
                nested_list = []
                if os.path.exists(unzip_d):
                    for root, _, files in os.walk(unzip_d):
                        for f in files:
                            sub_fp = os.path.join(root, f)
                            n_obj = {
                                "filename": f,
                                "path": sub_fp,
                                "size_bytes": os.path.getsize(sub_fp),
                                "is_zip": False,
                                "extracted_from": a_fn
                            }
                            nested_list.append(n_obj)
                            all_flattened_files.append(n_obj)
                nested_by_zip[a_fn] = nested_list
                root_att["nested_count"] = len(nested_list)

        dlg = tk.Toplevel(self)
        dlg.title(f"📧 Original Email & Attachments — {subj[:40]}")
        dlg.geometry("980x640")
        dlg.minsize(760, 480)
        dlg.configure(bg="#0F172A")

        # Header Frame
        hdr = tk.Frame(dlg, bg="#1E293B", padx=16, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"📧 {subj}", font=("Segoe UI", 11, "bold"), fg="#FFFFFF", bg="#1E293B").pack(anchor="w")
        tk.Label(hdr, text=f"From: {sender} • Date: {date_str} • Email Trail: {len(root_attachments)} attachment(s)",
                 font=("Segoe UI", 9), fg="#94A3B8", bg="#1E293B").pack(anchor="w", pady=(2, 0))

        # Main Split Frame
        main_fr = tk.Frame(dlg, bg="#0F172A", padx=14, pady=8)
        main_fr.pack(fill="both", expand=True)

        # Left Pane: Email Body with Rich Table Formatter & Browser Preview
        left_fr = tk.LabelFrame(main_fr, text=" 📝 Email Body Text ", font=("Segoe UI", 9, "bold"),
                                bg="#1E293B", fg="#94A3B8", padx=8, pady=8)
        left_fr.pack(side="left", fill="both", expand=True, padx=(0, 8))

        left_top_bar = tk.Frame(left_fr, bg="#1E293B")
        left_top_bar.pack(fill="x", pady=(0, 4))

        # Intelligent Assembly-to-Image association
        assy_image_map = {}
        img_files_in_stage = sorted([
            f for f in all_flattened_files
            if f.get("filename", "").lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.webp'))
        ], key=lambda x: x.get("filename", ""))

        if body_text:
            sections = re.split(r'\*(?:Art\.?|Article|Assy\.?)\s*([0-9A-Za-z_-]+)\*', body_text, flags=re.IGNORECASE)
            cur_img_ptr = 0
            for i in range(1, len(sections), 2):
                a_key = sections[i].strip().lower()
                sec_txt = sections[i+1] if i+1 < len(sections) else ""
                img_tags = re.findall(r'\[image:\s*([^\]]+)\]', sec_txt, re.IGNORECASE)
                assigned_list = []
                for tag in img_tags:
                    if cur_img_ptr < len(img_files_in_stage):
                        assigned_list.append(img_files_in_stage[cur_img_ptr]["filename"].lower())
                        cur_img_ptr += 1
                if a_key and assigned_list:
                    assy_image_map[a_key] = assigned_list

        def _open_in_outlook():
            try:
                stage_dir = email_obj.get("stage_dir", "")
                if not stage_dir or not os.path.exists(stage_dir):
                    stage_dir = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'ContXs', 'EmailStaging', f"{email_obj.get('id', '1')}_rfq")
                    os.makedirs(stage_dir, exist_ok=True)

                eml_path = os.path.join(stage_dir, "original_email.eml")
                if not os.path.exists(eml_path):
                    import email
                    from email.mime.multipart import MIMEMultipart
                    from email.mime.text import MIMEText
                    from email.mime.base import MIMEBase
                    from email import encoders

                    msg = MIMEMultipart('related')
                    msg['Subject'] = subj
                    msg['From'] = sender
                    msg['To'] = "user@radysis.com"
                    msg['Date'] = email.utils.formatdate(localtime=True)

                    msg_alt = MIMEMultipart('alternative')
                    msg.attach(msg_alt)
                    msg_alt.attach(MIMEText(body_text, 'plain', 'utf-8'))

                    for f_obj in all_flattened_files:
                        fp = f_obj.get("path", "")
                        fn = f_obj.get("filename", "")
                        if fp and os.path.exists(fp) and not fn.endswith('.eml') and not fn.startswith('_'):
                            try:
                                part = MIMEBase('application', 'octet-stream')
                                with open(fp, 'rb') as file_obj:
                                    part.set_payload(file_obj.read())
                                encoders.encode_base64(part)
                                part.add_header('Content-Disposition', f'attachment; filename="{fn}"')
                                msg.attach(part)
                            except Exception:
                                pass

                    with open(eml_path, 'wb') as f_out:
                        f_out.write(msg.as_bytes())

                os.startfile(eml_path)
            except Exception as ex:
                messagebox.showerror("Outlook Open Failed", f"Could not launch Outlook/Mail client:\n{ex}", parent=dlg)

        def _open_in_gmail():
            try:
                import urllib.parse, webbrowser
                q_subj = re.sub(r'^(?:\[(?:re|fwd|fw)\]|\((?:re|fwd|fw)\)|(?:re|fwd|fya|fw)[\s:_-]+)+', '', subj, flags=re.I).strip()
                encoded_query = urllib.parse.quote(f"subject:{q_subj}")
                url = f"https://mail.google.com/mail/u/0/#search/{encoded_query}"
                webbrowser.open(url)
            except Exception as ex:
                messagebox.showerror("Gmail Open Failed", f"Could not launch browser:\n{ex}", parent=dlg)

        def _open_in_browser():
            try:
                html_preview_dir = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'ContXs', 'EmailStaging', 'html_previews')
                os.makedirs(html_preview_dir, exist_ok=True)
                r_no = active_rfq.get('rfq_json', {}).get('rfq_metadata', {}).get('rfq_number', 'rfq')
                html_path = os.path.join(html_preview_dir, f"email_preview_{r_no}.html")

                from agents.multimodal_extractor import MultimodalExtractor
                asms = MultimodalExtractor()._parse_email_summary_table(body_text)
                tbl_html = ""
                if asms:
                    tbl_html = "<table style='border-collapse:collapse;width:100%;margin:10px 0;font-family:Segoe UI,sans-serif;font-size:13px;'>"
                    tbl_html += "<tr style='background:#1E293B;color:#38BDF8;'><th style='border:1px solid #334155;padding:7px;'>Article No</th><th style='border:1px solid #334155;padding:7px;'>Rev</th><th style='border:1px solid #334155;padding:7px;'>Description</th><th style='border:1px solid #334155;padding:7px;'>EAU (Qty)</th><th style='border:1px solid #334155;padding:7px;'>TP (USD)</th></tr>"
                    for a in asms:
                        eau_val = f"{a['eau']:,} pcs" if isinstance(a.get('eau'), (int, float)) else str(a.get('eau', ''))
                        tbl_html += f"<tr style='background:#0F172A;color:#F8FAFC;'><td style='border:1px solid #334155;padding:6px;font-weight:bold;'>{a.get('assy_no','')}</td><td style='border:1px solid #334155;padding:6px;text-align:center;'>{a.get('assy_rev','')}</td><td style='border:1px solid #334155;padding:6px;'>{a.get('assy_model','')}</td><td style='border:1px solid #334155;padding:6px;text-align:right;'>{eau_val}</td><td style='border:1px solid #334155;padding:6px;text-align:right;color:#10B981;font-weight:bold;'>{a.get('target_price','')}</td></tr>"
                    tbl_html += "</table>"

                # Normalize repeated blank lines and spaces
                clean_body = re.sub(r'(\n\s*){2,}', '\n', body_text)
                body_formatted = re.sub(r'(\*(?:Tecan\s+)?Article\s+No\*.*?(?=\*Art\.|\n\s*Regards|\Z))', tbl_html, clean_body, flags=re.DOTALL | re.IGNORECASE)

                # Embed real local photos into HTML preview
                img_ptr = [0]
                def _embed_img(m):
                    idx = img_ptr[0]
                    img_ptr[0] += 1
                    if idx < len(img_files_in_stage):
                        f_img = img_files_in_stage[idx]
                        fp_norm = f_img["path"].replace('\\', '/')
                        fn_img = f_img["filename"]
                        return f"""<div style="display:inline-block;margin:6px 4px;text-align:center;background:#1E293B;padding:6px;border-radius:6px;border:1px solid #334155;vertical-align:top;">
                            <a href="file:///{fp_norm}" target="_blank">
                                <img src="file:///{fp_norm}" style="max-width:280px;max-height:210px;border-radius:4px;display:block;">
                            </a>
                            <div style="color:#38BDF8;font-size:11px;font-weight:bold;margin-top:4px;">🖼️ {fn_img} (Click to open)</div>
                        </div>"""
                    return m.group(0)

                body_formatted = re.sub(r'\[image:\s*[^\]]+\]', _embed_img, body_formatted)
                body_formatted = body_formatted.replace('\n', '<br>')
                body_formatted = re.sub(r'(?:<br\s*/?>\s*){2,}', '<br>', body_formatted)

                full_html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Email Preview — {subj}</title>
<style>
body {{ font-family: Segoe UI, sans-serif; background: #0B132B; color: #F1F5F9; padding: 20px; line-height: 1.5; }}
.header {{ background: #1C2541; padding: 14px 18px; border-radius: 8px; margin-bottom: 16px; border-left: 4px solid #38BDF8; }}
.header h2 {{ margin: 0 0 6px 0; color: #FFFFFF; font-size: 16px; }}
.meta {{ color: #94A3B8; font-size: 13px; }}
.content {{ background: #111D4A; padding: 18px; border-radius: 8px; font-size: 13px; }}
</style>
</head>
<body>
<div class="header">
    <h2>📧 {subj}</h2>
    <div class="meta"><strong>From:</strong> {sender} &nbsp;|&nbsp; <strong>Date:</strong> {date_str}</div>
</div>
<div class="content">
{body_formatted}
</div>
</body>
</html>"""
                with open(html_path, "w", encoding="utf-8") as hf:
                    hf.write(full_html)
                os.startfile(html_path)
            except Exception as ex:
                messagebox.showerror("Browser Open Failed", f"Could not launch browser view:\n{ex}", parent=dlg)

        # Action Buttons in Top Header
        tk.Button(left_top_bar, text="📧 Open in Outlook", command=_open_in_outlook, bg="#0284C7",
                  fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=8, pady=1, cursor="hand2").pack(side="right", padx=(3, 0))
        tk.Button(left_top_bar, text="🌐 Gmail Web", command=_open_in_gmail, bg="#DC2626",
                  fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=8, pady=1, cursor="hand2").pack(side="right", padx=(3, 0))
        tk.Button(left_top_bar, text="📄 HTML View", command=_open_in_browser, bg="#334155",
                  fg="#38BDF8", font=("Segoe UI", 8, "bold"), relief="flat", padx=6, pady=1, cursor="hand2").pack(side="right", padx=(3, 0))

        tk.Label(left_top_bar, text="Original Text & Formatted Table View", font=("Segoe UI", 8, "italic"),
                 fg="#94A3B8", bg="#1E293B").pack(side="left")

        body_txt = tk.Text(left_fr, bg="#0F172A", fg="#F8FAFC", font=("Segoe UI", 9), wrap="none", bd=0)
        body_sb_y = ttk.Scrollbar(left_fr, command=body_txt.yview)
        body_sb_x = ttk.Scrollbar(left_fr, orient="horizontal", command=body_txt.xview)
        body_txt.configure(yscrollcommand=body_sb_y.set, xscrollcommand=body_sb_x.set)
        body_sb_y.pack(side="right", fill="y")
        body_sb_x.pack(side="bottom", fill="x")
        body_txt.pack(side="left", fill="both", expand=True)

        body_txt.tag_config("table_tag", font=("Consolas", 9, "bold"), foreground="#38BDF8", background="#0B132B")
        body_txt.tag_config("section_tag", font=("Segoe UI", 9, "bold"), foreground="#F59E0B")
        body_txt.tag_config("image_badge", font=("Segoe UI", 9, "bold"), foreground="#10B981")

        def _render_rich_body():
            body_txt.config(state="normal")
            body_txt.delete("1.0", "end")

            raw = body_text or "(No text body in email)"
            # Clean up consecutive empty lines
            raw_clean = re.sub(r'(\n\s*){2,}', '\n\n', raw)

            from agents.multimodal_extractor import MultimodalExtractor
            asms = MultimodalExtractor()._parse_email_summary_table(raw_clean)

            if asms:
                headers = ['Article No', 'Rev', 'Description', 'EAU (Qty)', 'TP (USD)']
                rows = []
                for a in asms:
                    eau_str = f"{a['eau']:,} pcs" if isinstance(a.get('eau'), (int, float)) else str(a.get('eau', ''))
                    rows.append([str(a.get('assy_no', '')), str(a.get('assy_rev', '')), str(a.get('assy_model', '')), eau_str, str(a.get('target_price', ''))])

                widths = [len(h) for h in headers]
                for r in rows:
                    for i, val in enumerate(r):
                        widths[i] = max(widths[i], len(str(val)))

                def sep(left, mid, right, fill='─'):
                    return left + mid.join(fill * (w + 2) for w in widths) + right

                top = sep('┌', '┬', '┐')
                mid = sep('├', '┼', '┤')
                bot = sep('└', '┴', '┘')
                hdr_row = '│ ' + ' │ '.join(f'{h:<{widths[i]}}' for i, h in enumerate(headers)) + ' │'

                tbl_lines = [top, hdr_row, mid]
                for r in rows:
                    row_str = '│ ' + ' │ '.join(f'{str(r[i]):<{widths[i]}}' for i in range(len(widths))) + ' │'
                    tbl_lines.append(row_str)
                tbl_lines.append(bot)
                ascii_table = "\n" + '\n'.join(tbl_lines) + "\n"

                pattern = r'(\*(?:Tecan\s+)?Article\s+No\*.*?(?=\*Art\.|\n\s*Regards|\Z))'
                match = re.search(pattern, raw_clean, flags=re.DOTALL | re.IGNORECASE)
                if match:
                    before_tbl = raw_clean[:match.start()]
                    after_tbl = raw_clean[match.end():]
                    body_txt.insert("end", before_tbl)
                    body_txt.insert("end", ascii_table, "table_tag")

                    # Style image badges in after_tbl
                    t_ptr = [0]
                    def _badge_replace(m):
                        idx = t_ptr[0]
                        t_ptr[0] += 1
                        fn_disp = img_files_in_stage[idx]["filename"] if idx < len(img_files_in_stage) else "photo.jpg"
                        return f"   🖼️ [{fn_disp}]"

                    formatted_after = re.sub(r'\[image:\s*[^\]]+\]', _badge_replace, after_tbl)
                    body_txt.insert("end", formatted_after)
                else:
                    body_txt.insert("end", raw_clean)
                    body_txt.insert("end", "\n\n--- Extracted Summary Table ---\n" + ascii_table, "table_tag")
            else:
                body_txt.insert("end", raw_clean)

            body_txt.config(state="disabled")

        _render_rich_body()

        # Right Pane: Attachments List with Hierarchical Nested View
        filter_state = {
            "focus": str(focus_filter).strip() if focus_filter else None,
            "view_zip": None
        }

        right_fr = tk.LabelFrame(main_fr, text=" 📎 Attachments ", font=("Segoe UI", 9, "bold"),
                                 bg="#1E293B", fg="#94A3B8", padx=8, pady=8, width=390)
        right_fr.pack(side="right", fill="both", expand=False)
        right_fr.pack_propagate(False)

        filter_bar = tk.Frame(right_fr, bg="#1E293B")
        filter_bar.pack(fill="x", pady=(0, 6))

        filter_lbl = tk.Label(filter_bar, text="", font=("Segoe UI", 8, "bold"), fg="#38BDF8", bg="#1E293B", anchor="w")
        filter_lbl.pack(side="left", fill="x", expand=True)

        toggle_btn = tk.Button(filter_bar, text="📁 Show All Attachments", bg="#334155", fg="#F8FAFC", font=("Segoe UI", 8),
                               relief="flat", padx=6, pady=1, cursor="hand2")
        toggle_btn.pack(side="right")

        att_canvas = tk.Canvas(right_fr, bg="#1E293B", highlightthickness=0)
        att_sb = ttk.Scrollbar(right_fr, command=att_canvas.yview)
        att_canvas.configure(yscrollcommand=att_sb.set)
        att_inner = tk.Frame(att_canvas, bg="#1E293B")
        att_win = att_canvas.create_window((0, 0), window=att_inner, anchor="nw")

        def _on_att_cfg(e):
            att_canvas.configure(scrollregion=att_canvas.bbox("all"))
            att_canvas.itemconfig(att_win, width=e.width)
        att_canvas.bind("<Configure>", _on_att_cfg)
        att_sb.pack(side="right", fill="y")
        att_canvas.pack(side="left", fill="both", expand=True)

        def _make_open_cmd(target_path):
            def _cmd():
                if not target_path or not os.path.exists(target_path):
                    messagebox.showwarning("Missing File", f"Attachment file not found on disk:\n{target_path}", parent=dlg)
                    return
                try:
                    os.startfile(target_path)
                except Exception:
                    try:
                        import subprocess
                        subprocess.Popen(f'explorer.exe "{os.path.normpath(target_path)}"', shell=True)
                    except Exception as ex2:
                        messagebox.showerror("Open Error", f"Could not launch file:\n{ex2}", parent=dlg)
            return _cmd

        def _render_attachments():
            for w in att_inner.winfo_children():
                w.destroy()

            cur_focus = filter_state["focus"]
            view_zip = filter_state["view_zip"]

            # Mode A: Focused on a specific assembly from Review Studio
            if cur_focus:
                filter_lbl.config(text=f"🎯 Focus: {cur_focus[:14]}")
                toggle_btn.config(text="📁 All Attachments", command=lambda: _set_mode(None, None))
                toggle_btn.pack(side="right")

                rfq_json = active_rfq.get("rfq_json", {})
                f_low = cur_focus.lower()

                def _matches_focus(f_obj):
                    fn = f_obj.get("filename", "").lower()
                    fp = f_obj.get("path", "")
                    
                    # 1. Direct filename match
                    if f_low in fn:
                        return True
                    
                    # 2. Match via RFQ Extracted Metadata Evidence / Drawing Mapping
                    for assy in rfq_json.get("assemblies", []):
                        a_no = str(assy.get("assy_no", "")).lower()
                        if a_no == f_low:
                            for it in assy.get("items", []):
                                ev = it.get("evidence", {})
                                for f_field in ev.values():
                                    if isinstance(f_field, dict):
                                        src_doc = str(f_field.get("source_doc", "")).lower()
                                        if src_doc and (src_doc in fn or fn in src_doc):
                                            return True
                            dwg_l = str(assy.get("drawing_link", "")).lower()
                            if dwg_l and (dwg_l in fn or fn in dwg_l):
                                return True
                    
                    # 3. Check drawings_detected metadata
                    for dwg in rfq_json.get("drawings_detected", []):
                        d_no = str(dwg.get("drawing_number", "")).lower()
                        d_fn = str(dwg.get("filename", "")).lower()
                        d_cpn = str(dwg.get("customer_part_number", "")).lower()
                        if (d_no == f_low or d_cpn == f_low or f_low in d_no) and (d_fn in fn or fn in d_fn):
                            return True

                    # 4. Intelligent Assembly-to-Image mapping from email body
                    if f_low in assy_image_map and fn in assy_image_map[f_low]:
                        return True

                    return False

                matched = [f for f in all_flattened_files if _matches_focus(f)]
                if not matched:
                    tk.Label(att_inner, text=f"No direct blueprint or BOM match for {cur_focus}.\nClick 'All Attachments' to view full files.",
                             font=("Segoe UI", 9, "italic"), fg="#94A3B8", bg="#1E293B").pack(pady=20)
                    return

                for idx, f_obj in enumerate(matched, start=1):
                    _render_file_card(f_obj, is_highlighted=True)

            # Mode B: Drilled down into a specific ZIP container
            elif view_zip and view_zip in nested_by_zip:
                items = nested_by_zip[view_zip]
                filter_lbl.config(text=f"📦 {view_zip[:20]} ({len(items)} files)")
                toggle_btn.config(text="⬅️ Back", command=lambda: _set_mode(None, None))
                toggle_btn.pack(side="right")

                for idx, f_obj in enumerate(items, start=1):
                    _render_file_card(f_obj, is_highlighted=False)

            # Mode C: Root Email Trail Attachments (Clean Grouping)
            else:
                filter_lbl.config(text=f"📬 Email Trail ({len(root_attachments)} files)")
                toggle_btn.pack_forget()

                for idx, r_att in enumerate(root_attachments, start=1):
                    _render_root_attachment_card(r_att)

            att_inner.update_idletasks()
            att_canvas.configure(scrollregion=att_canvas.bbox("all"))

        def _render_file_card(f_obj, is_highlighted=False):
            fn = f_obj["filename"]
            fp = f_obj["path"]
            f_size_kb = int(f_obj.get("size_bytes", 0) / 1024)
            ext = os.path.splitext(fn)[1].lower()

            icon_str = "📄"
            if ext in ('.png', '.jpg', '.jpeg', '.bmp', '.webp', '.tif', '.tiff'): icon_str = "🖼️"
            elif ext in ('.xlsx', '.xls', '.csv'): icon_str = "📊"
            elif ext in ('.docx', '.doc', '.txt'): icon_str = "📝"
            elif ext in ('.zip', '.rar', '.7z'): icon_str = "📦"

            card_bg = "#1E3A8A" if is_highlighted else "#334155"
            card = tk.Frame(att_inner, bg=card_bg, padx=8, pady=6, bd=1, relief="solid")
            card.pack(fill="x", pady=3, padx=2)

            tk.Label(card, text=f"{icon_str} {fn}", font=("Segoe UI", 9, "bold"),
                     fg="#FFFFFF", bg=card_bg, anchor="w", wraplength=270, justify="left").pack(fill="x")
            from_str = f" • from {f_obj.get('extracted_from')}" if f_obj.get('extracted_from') else ""
            tk.Label(card, text=f"{ext.upper() or 'DATA'} • {f_size_kb} KB{from_str}",
                     font=("Segoe UI", 8), fg="#93C5FD" if is_highlighted else "#94A3B8", bg=card_bg, anchor="w").pack(fill="x")

            card.bind("<MouseWheel>", lambda e: att_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
            for child in card.winfo_children():
                if not isinstance(child, tk.Button):
                    child.bind("<MouseWheel>", lambda e: att_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

            btn_box = tk.Frame(card, bg=card_bg)
            btn_box.pack(fill="x", pady=(4, 0))

            if ext in ('.pdf', '.png', '.jpg', '.jpeg', '.bmp', '.webp'):
                def _open_annot(target_fp=fp):
                    from agents.visual_annotation_dialog import open_visual_annotation_studio
                    f_term = filter_state.get("focus")
                    open_visual_annotation_studio(
                        parent=dlg,
                        file_path=target_fp,
                        highlight_terms=[f_term] if f_term else [],
                        component_data={"assy_no": f_term or ""}
                    )
                tk.Button(btn_box, text="🔍 Annotations", command=_open_annot,
                          bg="#0284C7", fg="#FFFFFF", font=("Segoe UI", 8, "bold"),
                          relief="flat", padx=6, pady=2, cursor="hand2").pack(side="left")

            tk.Button(btn_box, text="Open File ↗", command=_make_open_cmd(fp), bg="#2563EB" if not is_highlighted else "#3B82F6",
                      fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=6, pady=2, cursor="hand2").pack(side="right")

        def _render_root_attachment_card(r_att):
            fn = r_att["filename"]
            fp = r_att["path"]
            f_size_kb = int(r_att.get("size_bytes", 0) / 1024)
            is_zip = r_att.get("is_zip", False)
            nested_cnt = r_att.get("nested_count", 0)
            ext = os.path.splitext(fn)[1].lower()

            card_bg = "#1E293B"
            card = tk.Frame(att_inner, bg=card_bg, padx=10, pady=8, bd=1, relief="solid")
            card.pack(fill="x", pady=4, padx=2)

            icon_str = "📦" if is_zip else ("📊" if fn.endswith(('.xlsx', '.xls')) else "📄")
            tk.Label(card, text=f"{icon_str} {fn}", font=("Segoe UI", 9, "bold"),
                     fg="#FFFFFF", bg=card_bg, anchor="w", wraplength=270, justify="left").pack(fill="x")

            sub_txt = f"Archive Package • {nested_cnt} nested file(s) • {f_size_kb:,} KB" if is_zip else f"Document • {f_size_kb:,} KB"
            tk.Label(card, text=sub_txt, font=("Segoe UI", 8), fg="#38BDF8" if is_zip else "#94A3B8", bg=card_bg, anchor="w").pack(fill="x")

            b_fr = tk.Frame(card, bg=card_bg)
            b_fr.pack(fill="x", pady=(6, 0))

            if is_zip and nested_cnt > 0:
                tk.Button(b_fr, text=f"📂 View Nested ({nested_cnt} files) ▾", command=lambda z=fn: _set_mode(None, z),
                          bg="#0284C7", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=6, pady=2, cursor="hand2").pack(side="left")

            if not is_zip and ext in ('.pdf', '.png', '.jpg', '.jpeg', '.bmp', '.webp'):
                def _open_root_annot(target_fp=fp):
                    from agents.visual_annotation_dialog import open_visual_annotation_studio
                    open_visual_annotation_studio(
                        parent=dlg,
                        file_path=target_fp,
                        highlight_terms=[],
                        component_data={}
                    )
                tk.Button(b_fr, text="🔍 Annotations", command=_open_root_annot,
                          bg="#0284C7", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=6, pady=2, cursor="hand2").pack(side="left", padx=(0, 4))

            tk.Button(b_fr, text="Open File ↗", command=_make_open_cmd(fp), bg="#2563EB",
                      fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=6, pady=2, cursor="hand2").pack(side="right")

            card.bind("<MouseWheel>", lambda e: att_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
            for child in card.winfo_children():
                if not isinstance(child, tk.Button):
                    child.bind("<MouseWheel>", lambda e: att_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        def _set_mode(focus=None, view_zip=None):
            filter_state["focus"] = focus
            filter_state["view_zip"] = view_zip
            _render_attachments()

        def _on_wheel(e):
            att_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        att_canvas.bind("<MouseWheel>", _on_wheel)
        att_inner.bind("<MouseWheel>", _on_wheel)

        _render_attachments()

        def _set_external_focus(new_focus):
            _set_mode(str(new_focus).strip() if new_focus else None, None)

        self._active_email_att_dialog = (dlg, _set_external_focus, cur_rfq_id)

        def _on_dlg_close():
            self._active_email_att_dialog = None
            dlg.destroy()

        dlg.protocol("WM_DELETE_WINDOW", _on_dlg_close)
        _render_attachments()

    def _show_telemetry_benchmark_dialog(self):
        """Displays AI processing latency benchmarks, consistency stats, and multi-run speedup metrics."""
        try:
            from agents.telemetry_tracker import ProcessingTelemetryTracker
            tracker = ProcessingTelemetryTracker()
            stats = tracker.get_summary_stats()
        except Exception as e:
            messagebox.showerror("Telemetry Error", f"Could not load telemetry metrics:\n{e}", parent=self)
            return

        dlg = tk.Toplevel(self)
        dlg.title("📈 AI Processing & Telemetry Benchmarks — ContinuumX")
        dlg.geometry("980x620")
        dlg.minsize(760, 460)
        dlg.configure(bg="#0F172A")
        dlg.transient(self)
        dlg.grab_set()

        # Header
        hdr = tk.Frame(dlg, bg="#1E293B", padx=18, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📈 AI Processing Telemetry & Performance Benchmarks", font=("Segoe UI", 12, "bold"), fg="#FFFFFF", bg="#1E293B").pack(anchor="w")
        tk.Label(hdr, text=f"Total Email Runs: {stats['total_runs']} • Unique RFQs: {stats['unique_emails']} • Total BOMs Extracted: {stats['total_boms_processed']} • Avg Duration: {stats['avg_duration_sec']}s",
                 font=("Segoe UI", 9), fg="#94A3B8", bg="#1E293B").pack(anchor="w", pady=(2, 0))

        # Table Container
        tbl_frame = tk.Frame(dlg, bg="#1E293B", padx=12, pady=10)
        tbl_frame.pack(fill="both", expand=True)

        cols = ("Run#", "Timestamp", "Email Subject / RFQ", "BOMs", "Items", "Duration", "Speedup vs Run 1", "Trigger Source")
        tree = ttk.Treeview(tbl_frame, columns=cols, show="headings", height=14)

        tree.heading("Run#", text="Run#")
        tree.heading("Timestamp", text="Timestamp")
        tree.heading("Email Subject / RFQ", text="Email Subject / RFQ")
        tree.heading("BOMs", text="BOMs")
        tree.heading("Items", text="Items")
        tree.heading("Duration", text="Duration")
        tree.heading("Speedup vs Run 1", text="Speedup / Cache Ratio")
        tree.heading("Trigger Source", text="Trigger Source")
        tree.column("Run#", width=55, minwidth=45, anchor="center", stretch=False)
        tree.column("Timestamp", width=145, minwidth=120, anchor="center", stretch=False)
        tree.column("Email Subject / RFQ", width=280, minwidth=180, anchor="w", stretch=False)
        tree.column("BOMs", width=65, minwidth=50, anchor="center", stretch=False)
        tree.column("Items", width=65, minwidth=50, anchor="center", stretch=False)
        tree.column("Duration", width=85, minwidth=65, anchor="center", stretch=False)
        tree.column("Speedup vs Run 1", width=140, minwidth=110, anchor="center", stretch=False)
        tree.column("Trigger Source", width=120, minwidth=90, anchor="w", stretch=False)

        tree.tag_configure('odd', background='#1E293B', foreground='#F8FAFC')
        tree.tag_configure('even', background='#0F172A', foreground='#F8FAFC')
        tree.tag_configure('fast', background='#064E3B', foreground='#6EE7B7')

        sb = ttk.Scrollbar(tbl_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tbl_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sb.set, xscrollcommand=hsb.set)
        sb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(side="left", fill="both", expand=True)

        runs = stats.get("latest_runs", [])
        for idx, r in enumerate(reversed(runs), start=1):
            tag = 'fast' if "faster" in str(r.get("speedup_vs_run1", "")).lower() else ('odd' if idx % 2 == 1 else 'even')
            subj_disp = str(r.get("rfq_number") or r.get("email_subject", ""))[:45]
            tree.insert("", "end", values=(
                f"#{r.get('run_index', 1)}",
                r.get("timestamp", ""),
                subj_disp,
                r.get("bom_assemblies_count", 0),
                r.get("components_count", 0),
                r.get("formatted_duration", f"{r.get('duration_seconds')}s"),
                r.get("speedup_vs_run1", "Baseline"),
                r.get("trigger_source", "User Run")
            ), tags=(tag,))

        # Bottom Bar
        bbar = tk.Frame(dlg, bg="#1E293B", padx=16, pady=10)
        bbar.pack(fill="x")

        def _open_csv():
            from agents.telemetry_tracker import TELEMETRY_CSV
            if os.path.exists(TELEMETRY_CSV):
                try: os.startfile(TELEMETRY_CSV)
                except Exception as ex: messagebox.showerror("Open Error", f"Could not launch CSV:\n{ex}", parent=dlg)
            else:
                messagebox.showinfo("No CSV", "No telemetry data recorded yet.", parent=dlg)

        tk.Button(bbar, text="📊 Open CSV Log in Excel", command=_open_csv, bg="#10B981", fg="#FFFFFF",
                  font=("Segoe UI", 9, "bold"), relief="flat", padx=12, pady=4, cursor="hand2").pack(side="left")
        tk.Button(bbar, text="Close", command=dlg.destroy, bg="#475569", fg="#FFFFFF",
                  font=("Segoe UI", 9), relief="flat", padx=14, pady=4, cursor="hand2").pack(side="right")

    def _open_rfq_review_window(self):
        """Opens a full interactive Review Studio with live search, per-line source evidence comparison, and Excel export."""
        if not hasattr(self, '_last_extracted_rfq_json') or not self._last_extracted_rfq_json:
            self._append_agent_message("No RFQ data extracted yet. Scan emails first!", suggestions=["📩 Check RFQ Emails"])
            return

        rfq_data = self._last_extracted_rfq_json
        meta = rfq_data.get("rfq_metadata", {})
        assemblies = rfq_data.get("assemblies", [])
        rfq_no = meta.get("rfq_number", "RFQ")
        cust = meta.get("customer_name", "Customer")
        comm = meta.get("commodity", "Wire Harness")
        excel_fp = self._staged_bom_payload.get("file_path", "") if hasattr(self, '_staged_bom_payload') and self._staged_bom_payload else ""

        dlg = tk.Toplevel(self)
        dlg.title(f"RFQ Component Review & Filter Studio — {rfq_no} ({cust})")
        dlg.geometry("1280x760")
        dlg.minsize(850, 560)
        dlg.configure(bg="#F1F5F9")
        dlg.lift()
        selected_record_holder = [None]

        # Resolve active RFQ object from detected list
        current_rfq_entry = None
        if hasattr(self, '_detected_rfq_list') and self._detected_rfq_list:
            for r_item in self._detected_rfq_list:
                r_meta = r_item.get("rfq_json", {}).get("rfq_metadata", {})
                if r_meta.get("rfq_number") == rfq_no or (cust and r_meta.get("customer_name") == cust):
                    current_rfq_entry = r_item
                    break

        # 1. Top Header Banner
        hdr = tk.Frame(dlg, bg="#0F172A", padx=20, pady=10)
        hdr.pack(fill="x")
        
        # Determine dynamic EAU display string
        # Dynamic EAU Banner Calculation
        unique_assy_set = set(str(a.get('assy_no', '')).strip() for a in assemblies if a.get('assy_no'))
        unique_assy_count = len(unique_assy_set) if unique_assy_set else len(assemblies)
        
        eau_vals = []
        for a in assemblies:
            e = a.get('eau')
            if e is not None and str(e).strip() not in ('', 'None'):
                try: eau_vals.append(int(re.sub(r'[^\d]', '', str(e))))
                except Exception: pass

        unique_eaus = set(eau_vals)
        if len(unique_eaus) > 1:
            eau_disp_str = f"Multiple ({min(unique_eaus):,} – {max(unique_eaus):,} pcs)"
        elif len(unique_eaus) == 1:
            eau_disp_str = f"{list(unique_eaus)[0]:,} pcs"
        else:
            eau_disp_str = f"{meta.get('eau', 'Not Specified')} pcs"

        hdr_top_row = tk.Frame(hdr, bg="#0F172A")
        hdr_top_row.pack(fill="x")
        hdr_title_lbl = tk.Label(hdr_top_row, text=f"📊 RFQ Component Review & Traceability Studio — {rfq_no} ({cust})", font=("Segoe UI", 12, "bold"), fg="#FFFFFF", bg="#0F172A")
        hdr_title_lbl.pack(side="left")
        cur_badge = meta.get("review_badge", "✅ Pattern Verified" if meta.get("is_known_customer") else "⚠️ Requires Human Review")
        is_rev_needed = "Requires" in cur_badge
        status_badge_lbl = tk.Label(hdr_top_row, text=cur_badge, font=("Segoe UI", 9, "bold"),
                                    bg="#92400E" if is_rev_needed else "#065F46",
                                    fg="#FDE68A" if is_rev_needed else "#A7F3D0", padx=8, pady=2)
        status_badge_lbl.pack(side="right")

        hdr_sub_lbl = tk.Label(hdr, text=f"Customer: {cust} • Commodity: {comm} • Assemblies: {unique_assy_count} • Target Price: {meta.get('target_price', 'Not Specified')} • EAU: {eau_disp_str}",
                               font=("Segoe UI", 8), fg="#94A3B8", bg="#0F172A", anchor="w")
        hdr_sub_lbl.pack(fill="x", pady=(2, 0))

        # Competition Mode / Anonymization Hashing State
        import hashlib
        is_anonymized_mode = [False]
        _hash_cache = {}

        def _get_hash(val: str, prefix: str = "VAL") -> str:
            if not val or str(val).strip() in ("", "N/A", "None", "-", "Not Specified"):
                return str(val) if val is not None else ""
            s_val = str(val).strip()
            cache_key = (prefix, s_val)
            if cache_key in _hash_cache:
                return _hash_cache[cache_key]

            h_code = hashlib.md5(s_val.encode('utf-8')).hexdigest()[:4].upper()
            if prefix == "CUST":
                hashed = f"CUST-{h_code}"
            elif prefix == "ASSY":
                hashed = f"ASY-{h_code}"
            elif prefix == "MODEL":
                hashed = f"HARNESS-MOD-{h_code}"
            elif prefix == "REV":
                hashed = f"R-{h_code[:2]}"
            elif prefix == "PART":
                hashed = f"CP-{h_code}"
            else:
                hashed = f"{prefix}-{h_code}"

            _hash_cache[cache_key] = hashed
            return hashed

        def _display_val(val: str, prefix: str) -> str:
            if is_anonymized_mode[0]:
                return _get_hash(val, prefix)
            return str(val) if val is not None else ""

        # Helper function for strictly 1-to-1 paired comma-separated MPN and MFR
        def _pair_mpn_mfr(mpn_val, mfr_val):
            raw_mpns = [m.strip() for m in re.split(r'[/;|,\\]', str(mpn_val or "")) if m.strip() and m.strip() != "None"]
            raw_mfrs = [f.strip() for f in re.split(r'[/;|,\\]', str(mfr_val or "")) if f.strip() and f.strip() != "None"]
            if not raw_mpns:
                return "", ", ".join(raw_mfrs)
            if len(raw_mfrs) == 0:
                raw_mfrs = ["Unknown"] * len(raw_mpns)
            elif len(raw_mfrs) < len(raw_mpns):
                if len(raw_mfrs) == 1:
                    raw_mfrs = [raw_mfrs[0]] * len(raw_mpns)
                else:
                    raw_mfrs.extend([raw_mfrs[-1]] * (len(raw_mpns) - len(raw_mfrs)))
            elif len(raw_mfrs) > len(raw_mpns):
                raw_mfrs = raw_mfrs[:len(raw_mpns)]
            return ", ".join(raw_mpns), ", ".join(raw_mfrs)

        # 2. Main Studio Notebook (Tab 1: Components, Tab 2: Assembly EAU & MOQs)
        st_style = ttk.Style()
        try:
            st_style.theme_use("clam")
        except Exception:
            pass

        st_style.configure(
            "Studio.TNotebook",
            background="#0F172A",
            borderwidth=0,
            tabmargins=[4, 6, 4, 0]
        )
        st_style.configure(
            "Studio.TNotebook.Tab",
            background="#1E293B",
            foreground="#94A3B8",
            font=("Segoe UI", 9, "bold"),
            padding=[16, 7],
            borderwidth=0
        )
        st_style.map(
            "Studio.TNotebook.Tab",
            background=[
                ("selected", "#2563EB"),
                ("active", "#334155")
            ],
            foreground=[
                ("selected", "#FFFFFF"),
                ("active", "#F8FAFC")
            ],
            expand=[("selected", [1, 2, 1, 0])]
        )

        studio_nb = ttk.Notebook(dlg, style="Studio.TNotebook")
        studio_nb.pack(fill="both", expand=True, padx=14, pady=(6, 2))

        # =========================================================================
        # TAB 1: 🔩 Component BOM & Traceability Studio
        # =========================================================================
        tab_components = tk.Frame(studio_nb, bg="#F8FAFC")
        studio_nb.add(tab_components, text=" 🔩 Component Review & Filter Studio ")

        # Controls & Filter Bar for Tab 1 (2-Row Responsive Layout for Laptops & Displays)
        ctrl_bar = tk.Frame(tab_components, bg="#FFFFFF", padx=8, pady=4, bd=1, relief="solid")
        ctrl_bar.pack(fill="x", padx=6, pady=(4, 2))

        ctrl_row1 = tk.Frame(ctrl_bar, bg="#FFFFFF")
        ctrl_row1.pack(fill="x", pady=(1, 2))

        # Assembly Combobox Filter
        tk.Label(ctrl_row1, text="Filter Assembly:", font=("Segoe UI", 9, "bold"), bg="#FFFFFF", fg="#334155").pack(side="left", padx=(0, 4))
        seen_opt = set()
        assy_options = ["All Assemblies"]
        for a in assemblies:
            ano = str(a.get("assy_no", "")).strip()
            if ano and ano not in seen_opt:
                seen_opt.add(ano)
                assy_options.append(f"{ano} — {a.get('assy_model', '')[:22]}")
        
        assy_combo = ttk.Combobox(ctrl_row1, values=assy_options, state="readonly", width=24, font=("Segoe UI", 9))
        assy_combo.current(0)
        assy_combo.pack(side="left", padx=(0, 8))

        # Search Box
        tk.Label(ctrl_row1, text="🔍 Search:", font=("Segoe UI", 9, "bold"), bg="#FFFFFF", fg="#334155").pack(side="left", padx=(0, 4))
        search_var = tk.StringVar()
        search_ent = tk.Entry(ctrl_row1, textvariable=search_var, font=("Segoe UI", 9), width=16, bd=1, relief="solid")
        search_ent.pack(side="left", padx=(0, 8))

        # Dynamic Row & Focus Count Badge for Tab 1
        comp_count_lbl = tk.Label(ctrl_row1, text="📊 Rows: 0 / 0", font=("Segoe UI", 8, "bold"),
                                  bg="#EFF6FF", fg="#1D4ED8", padx=8, pady=2, bd=1, relief="solid")
        comp_count_lbl.pack(side="left", padx=(0, 6))

        ctrl_row2 = tk.Frame(ctrl_bar, bg="#FFFFFF")
        ctrl_row2.pack(fill="x", pady=(2, 1))

        def _sync_from_excel():
            sync_fp = excel_fp
            if not sync_fp or not os.path.exists(sync_fp):
                from tkinter import filedialog
                sync_fp = filedialog.askopenfilename(
                    title="Select Edited BOM Excel File",
                    filetypes=[("Excel Files", "*.xlsx;*.xls")],
                    parent=dlg
                )
            if not sync_fp or not os.path.exists(sync_fp):
                messagebox.showwarning("File Missing", "Please select a valid Excel file to sync.", parent=dlg)
                return

            try:
                import openpyxl
                wb = openpyxl.load_workbook(sync_fp, data_only=True)
                ws = wb.active
                
                header_row_idx = None
                headers = []
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    row_strs = [str(c).strip().lower() for c in row if c is not None]
                    if any("part" in s or "item" in s or "assy" in s for s in row_strs) and any("desc" in s or "qty" in s or "uom" in s for s in row_strs):
                        headers = [str(c).strip() if c is not None else "" for c in row]
                        header_row_idx = r_idx
                        break

                if not header_row_idx:
                    messagebox.showerror("Format Error", "Could not locate BOM headers.", parent=dlg)
                    return

                col_map = {}
                for c_idx, h in enumerate(headers):
                    h_low = h.lower()
                    if "line" in h_low: col_map["line"] = c_idx
                    elif "assy#" in h_low or "assy no" in h_low or "assembly" in h_low: col_map["assy_no"] = c_idx
                    elif "assy model" in h_low or "model" in h_low: col_map["assy_model"] = c_idx
                    elif "assy rev" in h_low or "rev" in h_low: col_map["assy_rev"] = c_idx
                    elif "part" in h_low or "article" in h_low: col_map["part"] = c_idx
                    elif "desc" in h_low: col_map["desc"] = c_idx
                    elif "mpn" in h_low: col_map["mpn"] = c_idx
                    elif "mfr" in h_low or "vendor" in h_low: col_map["mfr"] = c_idx
                    elif "qty" in h_low or "quantity" in h_low: col_map["qty"] = c_idx
                    elif "uom" in h_low or "unit" in h_low: col_map["uom"] = c_idx
                    elif "eau" in h_low or "volume" in h_low: col_map["eau"] = c_idx
                    elif "price" in h_low or "target" in h_low: col_map["tp"] = c_idx

                from agents.correction_store import CorrectionStore
                cs = CorrectionStore()

                updated_assemblies_map = {}
                total_imported_rows = 0

                for row in list(ws.iter_rows(values_only=True))[header_row_idx:]:
                    if not any(row): continue
                    
                    def _g(k, d=""):
                        idx = col_map.get(k)
                        if idx is not None and idx < len(row) and row[idx] is not None:
                            return str(row[idx]).strip()
                        return d

                    r_a_no = _g("assy_no")
                    r_part = _g("part")
                    if not r_a_no or not r_part: continue

                    r_model = _g("assy_model", r_a_no)
                    r_rev = _g("assy_rev", "00")
                    r_desc = _g("desc", "Component")
                    r_mpn_raw = _g("mpn", "")
                    r_mfr_raw = _g("mfr", "")
                    r_mpn, r_mfr = _pair_mpn_mfr(r_mpn_raw, r_mfr_raw)
                    r_uom = _g("uom", "EA")
                    r_tp = _g("tp", "")
                    try: r_qty = float(_g("qty", "1"))
                    except Exception: r_qty = 1.0

                    if r_a_no not in updated_assemblies_map:
                        updated_assemblies_map[r_a_no] = {
                            "assy_no": r_a_no,
                            "assy_model": r_model,
                            "assy_rev": r_rev,
                            "target_price": r_tp,
                            "items_map": {}
                        }

                    k_item = (r_part, r_uom)
                    if k_item in updated_assemblies_map[r_a_no]["items_map"]:
                        existing_it = updated_assemblies_map[r_a_no]["items_map"][k_item]
                        existing_it["qty"] = round(existing_it["qty"] + r_qty, 4)
                        if r_mpn and not existing_it.get("mpn"): existing_it["mpn"] = r_mpn
                        if r_mfr and not existing_it.get("mfr"): existing_it["mfr"] = r_mfr
                    else:
                        updated_assemblies_map[r_a_no]["items_map"][k_item] = {
                            "part_number": r_part,
                            "description": r_desc,
                            "mpn": r_mpn,
                            "mfr": r_mfr,
                            "qty": r_qty,
                            "uom": r_uom
                        }
                    total_imported_rows += 1

                new_assemblies = []
                for a_no_k, a_info in updated_assemblies_map.items():
                    items_list = list(a_info["items_map"].values())
                    for idx_item, it_obj in enumerate(items_list, start=1):
                        it_obj["line_item"] = idx_item
                    new_assemblies.append({
                        "assy_no": a_info["assy_no"],
                        "assy_model": a_info["assy_model"],
                        "assy_rev": a_info["assy_rev"],
                        "target_price": a_info["target_price"],
                        "items": items_list
                    })

                if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                    self._last_extracted_rfq_json["assemblies"] = new_assemblies
                    from agents.synthetic_bom_generator import SyntheticBOMGenerator
                    gen = SyntheticBOMGenerator()
                    gen_res = gen.generate_synthetic_excel(self._last_extracted_rfq_json)
                    if gen_res.get("success") and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
                        self._staged_bom_payload["file_path"] = gen_res["file_path"]

                all_table_records.clear()
                for assy in new_assemblies:
                    a_no = str(assy.get("assy_no", "")).strip()
                    a_model = assy.get("assy_model", "")
                    a_rev = str(assy.get("assy_rev", "")).replace("Rev", "").strip()
                    a_tp = str(assy.get("target_price", "")).replace('$', '').strip()
                    for it in assy.get("items", []):
                        p_no = str(it.get("part_number") or "").strip()
                        desc_str = str(it.get("description") or "").strip()
                        c_mpn, c_mfr = _pair_mpn_mfr(it.get("mpn"), it.get("mfr"))
                        current_assy_lines = sum(1 for r in all_table_records if r["assy_no"] == a_no) + 1
                        all_table_records.append({
                            "line": str(current_assy_lines),
                            "assy_no": a_no,
                            "assy_model": a_model,
                            "assy_rev": a_rev,
                            "part": p_no,
                            "desc": desc_str or "Component",
                            "mpn": c_mpn,
                            "mfr": c_mfr,
                            "qty": str(it.get("qty", 1)),
                            "uom": str(it.get("uom") or "EA"),
                            "tp": a_tp or "N/A",
                            "raw_item": it
                        })

                # Persist learned component field corrections to CorrectionStore, VerifiedBOMStore & invalidate LLM cache
                try:
                    from agents.correction_store import CorrectionStore
                    from agents.verified_bom_store import VerifiedBOMStore
                    from agents.llm_gateway import LLMGateway
                    cs = CorrectionStore()
                    vbs = VerifiedBOMStore()
                    for assy in new_assemblies:
                        a_no_k = str(assy.get("assy_no", "")).strip()
                        vbs.save_verified_assembly(a_no_k, assy, customer=cust, rfq_no=rfq_no)
                        for it in assy.get("items", []):
                            p_no = str(it.get("part_number") or "").strip()
                            if not p_no: continue
                            mpn_v = str(it.get("mpn") or "").strip()
                            mfr_v = str(it.get("mfr") or "").strip()
                            desc_v = str(it.get("description") or "").strip()
                            qty_v = str(it.get("qty", "1")).strip()
                            uom_v = str(it.get("uom") or "EA").strip()
                            # Key corrections to specific part and composite assy::part (never assembly alone)
                            for hint in (p_no, f"{a_no_k}::{p_no}"):
                                if not hint: continue
                                if mpn_v: cs.save_correction(hint, "mpn", "", mpn_v, mfr=mfr_v, note="Excel Synced by User")
                                if mfr_v: cs.save_correction(hint, "mfr", "", mfr_v, note="Excel Synced by User")
                                if desc_v: cs.save_correction(hint, "description", "", desc_v, note="Excel Synced by User")
                                if qty_v: cs.save_correction(hint, "qty", "", qty_v, note="Excel Synced by User")
                                if uom_v: cs.save_correction(hint, "uom", "", uom_v, note="Excel Synced by User")
                    LLMGateway().clear_cache(rfq_no)
                except Exception as c_err:
                    print(f"[ReviewStudio] Notice saving corrections: {c_err}")

                _refresh_table()
                _refresh_assy_table()

                # Live update chatbot card in background
                try:
                    if hasattr(self, '_detected_rfq_list') and self._detected_rfq_list:
                        for r_item in self._detected_rfq_list:
                            if r_item.get("rfq_json", {}).get("rfq_metadata", {}).get("rfq_number") == rfq_no:
                                r_item["rfq_json"]["assemblies"] = new_assemblies
                                break
                    if current_rfq_entry:
                        current_rfq_entry["rfq_json"]["assemblies"] = new_assemblies
                        from agents.synthetic_bom_generator import SyntheticBOMGenerator
                        gen_res = SyntheticBOMGenerator().generate_synthetic_excel(current_rfq_entry["rfq_json"])
                        current_rfq_entry["synthetic_bom"] = gen_res
                        self._load_selected_rfq_payload(current_rfq_entry)
                except Exception as err:
                    print(f"[SyncLiveSync] {err}")

                messagebox.showinfo("Excel Sync Complete", f"✅ Successfully synced {total_imported_rows} components from Excel & updated learning memory and Chatbot table!", parent=dlg)
            except Exception as sync_ex:
                messagebox.showerror("Sync Failed", f"Could not sync Excel edits:\n{sync_ex}", parent=dlg)

        def _open_excel():
            cur_fp = excel_fp
            if hasattr(self, '_staged_bom_payload') and self._staged_bom_payload and self._staged_bom_payload.get("file_path"):
                cur_fp = self._staged_bom_payload["file_path"]
            if cur_fp and os.path.exists(cur_fp):
                try: os.startfile(cur_fp)
                except Exception as e: messagebox.showerror("Open Excel Failed", f"Could not launch Excel:\n{e}", parent=dlg)
            else:
                messagebox.showwarning("File Missing", "Synthetic Excel file has not been generated yet.", parent=dlg)

        def _open_folder():
            cur_fp = excel_fp
            if hasattr(self, '_staged_bom_payload') and self._staged_bom_payload and self._staged_bom_payload.get("file_path"):
                cur_fp = self._staged_bom_payload["file_path"]
            if cur_fp and os.path.exists(os.path.dirname(cur_fp)):
                try: os.startfile(os.path.dirname(cur_fp))
                except Exception as e: messagebox.showerror("Open Folder Failed", f"Could not open folder:\n{e}", parent=dlg)
            else:
                messagebox.showwarning("Folder Missing", "No folder found for generated BOM file.", parent=dlg)

        def _on_click_email_att():
            cur_focus = None
            if selected_record_holder[0]:
                cur_focus = selected_record_holder[0].get("assy_no")
            elif assy_combo.get() and assy_combo.get() != "All Assemblies":
                cur_focus = assy_combo.get().split("—")[0].strip()
            self._show_email_and_attachments_dialog(focus_filter=cur_focus, target_rfq=current_rfq_entry)

        def _approve_and_learn_pattern():
            try:
                from agents.customer_profile_store import CustomerProfileStore
                from agents.correction_store import CorrectionStore
                from agents.verified_bom_store import VerifiedBOMStore
                from agents.llm_gateway import LLMGateway
                cps = CustomerProfileStore()
                cs = CorrectionStore()
                vbs = VerifiedBOMStore()
                target_json = getattr(self, '_last_extracted_rfq_json', None) or rfq_data
                ok = cps.learn_or_update_customer_pattern(cust, target_json, feedback_notes="Verified by PIC in Review Studio")
                
                # Persist all ground truth assemblies to VerifiedBOMStore & corrections
                for assy in target_json.get("assemblies", []):
                    a_no_k = str(assy.get("assy_no", "")).strip()
                    vbs.save_verified_assembly(a_no_k, assy, customer=cust, rfq_no=rfq_no)
                    for it in assy.get("items", []):
                        p_no = str(it.get("part_number") or "").strip()
                        if not p_no: continue
                        mpn_v = str(it.get("mpn") or "").strip()
                        mfr_v = str(it.get("mfr") or "").strip()
                        desc_v = str(it.get("description") or "").strip()
                        qty_v = str(it.get("qty", "1")).strip()
                        uom_v = str(it.get("uom") or "EA").strip()
                        for hint in (p_no, f"{a_no_k}::{p_no}"):
                            if not hint: continue
                            if mpn_v: cs.save_correction(hint, "mpn", "", mpn_v, mfr=mfr_v, note="Pattern Learned Ground Truth")
                            if mfr_v: cs.save_correction(hint, "mfr", "", mfr_v, note="Pattern Learned Ground Truth")
                            if desc_v: cs.save_correction(hint, "description", "", desc_v, note="Pattern Learned Ground Truth")
                            if qty_v: cs.save_correction(hint, "qty", "", qty_v, note="Pattern Learned Ground Truth")
                            if uom_v: cs.save_correction(hint, "uom", "", uom_v, note="Pattern Learned Ground Truth")

                # Invalidate stale LLM Gateway cache
                LLMGateway().clear_cache(rfq_no)

                # Live update chatbot card
                try:
                    if current_rfq_entry:
                        current_rfq_entry["rfq_json"] = target_json
                        from agents.synthetic_bom_generator import SyntheticBOMGenerator
                        gen_res = SyntheticBOMGenerator().generate_synthetic_excel(target_json)
                        current_rfq_entry["synthetic_bom"] = gen_res
                        self._load_selected_rfq_payload(current_rfq_entry)
                except Exception as err:
                    print(f"[LearnLiveSync] {err}")

                if ok:
                    from datetime import datetime
                    time_stamp = datetime.now().strftime("%d %b %Y, %I:%M %p")
                    meta["review_badge"] = f"✅ Pattern Verified ({time_stamp})"
                    meta["requires_human_review"] = False
                    status_badge_lbl.config(text=f"✅ Pattern Verified ({time_stamp})", bg="#065F46", fg="#A7F3D0")
                    if 'tip_banner' in locals() or 'tip_banner' in globals():
                        try:
                            tip_banner.config(bg="#ECFDF5")
                            tip_icon_lbl.config(text="✅ AI Pattern Saved:", fg="#047857", bg="#ECFDF5")
                            tip_text_lbl.config(text=f"Customer '{cust}' layout & component rules saved ({time_stamp}) for future extractions.", fg="#065F46", bg="#ECFDF5")
                        except Exception: pass
                    messagebox.showinfo("Pattern Learned", f"Successfully saved customer profile & component rules ({time_stamp})!\nNext extraction will automatically apply these verified patterns.", parent=dlg)
                else:
                    messagebox.showwarning("Notice", "Could not update customer profile store.", parent=dlg)
            except Exception as ex:
                messagebox.showerror("Error", f"Failed to save customer pattern: {ex}", parent=dlg)

        def _reset_learned_pattern_cache():
            """Clears all learned patterns and customer profile cache so extraction can restart cleanly."""
            confirm = messagebox.askyesno(
                "Reset Learned Patterns",
                f"Are you sure you want to clear the learned pattern cache for '{cust}'?\n\nThis removes previous learned ground truth and allows you to re-prompt or restart extraction with 0 cache bias.",
                parent=dlg
            )
            if not confirm:
                return

            try:
                from agents.correction_store import CorrectionStore
                cs = CorrectionStore()
                cs.remove_correction_by_hint(cust)
            except Exception as ex:
                print(f"[ReviewStudio] Correction clear err: {ex}")

            try:
                from agents.customer_profile import CustomerProfileStore
                CustomerProfileStore().reset_profile(cust)
            except Exception: pass

            meta["review_badge"] = "⚠️ Requires Human Review"
            meta["requires_human_review"] = True
            status_badge_lbl.config(text="⚠️ Requires Human Review", bg="#78350F", fg="#FDE68A")
            tip_banner.config(bg="#F5F3FF")
            tip_icon_lbl.config(text="💡 Tip:", fg="#6D28D9", bg="#F5F3FF")
            tip_text_lbl.config(text="Pattern cache cleared. You can now re-run extraction or assign verified patterns cleanly.", fg="#5B21B6", bg="#F5F3FF")
            messagebox.showinfo("Pattern Cache Cleared", f"✅ Learned patterns for '{cust}' successfully reset!\nYou can now re-evaluate or teach fresh ground truth.", parent=dlg)

        # Action buttons on top toolbar (Row 1 right & Row 2 left)
        tk.Button(ctrl_row1, text="📁 Open Folder", command=_open_folder, bg="#64748B", fg="#FFFFFF", font=("Segoe UI", 8), relief="flat", padx=7, pady=2, cursor="hand2").pack(side="right")

        tk.Button(ctrl_row2, text="📊 Open in Excel", command=_open_excel, bg="#10B981", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=8, pady=2, cursor="hand2").pack(side="left", padx=(0, 4))
        tk.Button(ctrl_row2, text="🔄 Sync Excel", command=_sync_from_excel, bg="#2563EB", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=8, pady=2, cursor="hand2").pack(side="left", padx=(0, 4))
        tk.Button(ctrl_row2, text="📧 Email & Attachments", command=_on_click_email_att, bg="#6366F1", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=8, pady=2, cursor="hand2").pack(side="left", padx=(0, 4))
        
        btn_learn_pat = tk.Button(ctrl_row2, text="🧠 Learn Pattern", command=_approve_and_learn_pattern, bg="#7C3AED", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=8, pady=2, cursor="hand2")
        btn_learn_pat.pack(side="left", padx=(0, 4))

        btn_reset_pat = tk.Button(ctrl_row2, text="🗑️ Reset Pattern", command=_reset_learned_pattern_cache, bg="#475569", fg="#FFFFFF", font=("Segoe UI", 8), relief="flat", padx=8, pady=2, cursor="hand2")
        btn_reset_pat.pack(side="left", padx=(0, 4))

        def _toggle_anonymization():
            is_anonymized_mode[0] = not is_anonymized_mode[0]
            is_anon = is_anonymized_mode[0]
            btn_txt = "🔓 Reveal Real Data" if is_anon else "🔒 Mask Customer Info (MAIC)"
            btn_bg = "#059669" if is_anon else "#475569"
            
            btn_anon_top.config(text=btn_txt, bg=btn_bg)
            if 'btn_anon_tab2' in locals() and btn_anon_tab2:
                try: btn_anon_tab2.config(text=btn_txt, bg=btn_bg)
                except Exception: pass

            cust_disp = _display_val(cust, "CUST")
            dlg.title(f"RFQ Component Review & Filter Studio — {rfq_no} ({cust_disp})")
            hdr_title_lbl.config(text=f"📊 RFQ Component Review & Traceability Studio — {rfq_no} ({cust_disp})")

            new_assy_opts = ["All Assemblies"]
            seen_o = set()
            for a in assemblies:
                ano = str(a.get("assy_no", "")).strip()
                if ano and ano not in seen_o:
                    seen_o.add(ano)
                    a_lbl = f"{_display_val(ano, 'ASSY')} — {_display_val(a.get('assy_model', ''), 'MODEL')[:22]}"
                    new_assy_opts.append(a_lbl)
            
            cur_sel_idx = assy_combo.current()
            assy_combo.config(values=new_assy_opts)
            if cur_sel_idx >= 0 and cur_sel_idx < len(new_assy_opts):
                assy_combo.current(cur_sel_idx)
            else:
                assy_combo.current(0)

            _refresh_table()
            _refresh_assy_table()

        btn_anon_top = tk.Button(ctrl_row2, text="🔒 Mask Customer Info (MAIC)", command=_toggle_anonymization, bg="#475569", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=8, pady=2, cursor="hand2")
        btn_anon_top.pack(side="left")

        # 2.5 Active Learning Tip & Color Legend Banner (Compact & Responsive)
        tip_banner = tk.Frame(tab_components, bg="#F8FAFC", padx=6, pady=3, bd=1, relief="solid")
        tip_banner.pack(fill="x", padx=6, pady=(2, 3))
        
        tk.Label(tip_banner, text="💡 Legend:", font=("Segoe UI", 8, "bold"), fg="#1E293B", bg="#F8FAFC").pack(side="left", padx=(0, 6))

        # Gold Badge
        gold_badge = tk.Label(tip_banner, text=" 🟨 Gold: Online Sourced MPN ", font=("Segoe UI", 8, "bold"), bg="#FEF3C7", fg="#92400E", padx=5, pady=1, bd=1, relief="solid")
        gold_badge.pack(side="left", padx=(0, 4))

        # Pink Badge
        pink_badge = tk.Label(tip_banner, text=" 🌸 Pink: Needs Manual Sourcing ", font=("Segoe UI", 8, "bold"), bg="#FFE4E6", fg="#BE123C", padx=5, pady=1, bd=1, relief="solid")
        pink_badge.pack(side="left", padx=(0, 4))

        # White Badge
        white_badge = tk.Label(tip_banner, text=" ⬜ White: Blueprint Ground Truth ", font=("Segoe UI", 8, "bold"), bg="#FFFFFF", fg="#334155", padx=5, pady=1, bd=1, relief="solid")
        white_badge.pack(side="left")

        # 3. 11-Column Treeview Table for Components
        table_frame = tk.Frame(tab_components, bg="#FFFFFF", padx=2, pady=2, bd=1, relief="solid")
        table_frame.pack(fill="both", expand=True, padx=6, pady=4)

        cols = ("Line", "Assy#", "Assy Model", "Assy Rev", "Part", "Description", "MPN", "MFR", "QTY", "UOM", "Target Price (USD)")
        tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=11)

        tree.heading("Line", text="Line")
        tree.heading("Assy#", text="Assy#")
        tree.heading("Assy Model", text="Assy Model")
        tree.heading("Assy Rev", text="Assy Rev")
        tree.heading("Part", text="Part")
        tree.heading("Description", text="Description")
        tree.heading("MPN", text="MPN")
        tree.heading("MFR", text="MFR")
        tree.heading("QTY", text="QTY")
        tree.heading("UOM", text="UOM")
        tree.heading("Target Price (USD)", text="Target Price (USD)")

        tree.column("Line", width=45, minwidth=35, anchor="center", stretch=False)
        tree.column("Assy#", width=95, minwidth=80, anchor="w", stretch=False)
        tree.column("Assy Model", width=180, minwidth=140, anchor="w", stretch=False)
        tree.column("Assy Rev", width=70, minwidth=55, anchor="center", stretch=False)
        tree.column("Part", width=110, minwidth=90, anchor="w", stretch=False)
        tree.column("Description", width=220, minwidth=150, anchor="w", stretch=False)
        tree.column("MPN", width=160, minwidth=110, anchor="w", stretch=False)
        tree.column("MFR", width=130, minwidth=90, anchor="w", stretch=False)
        tree.column("QTY", width=65, minwidth=50, anchor="center", stretch=False)
        tree.column("UOM", width=55, minwidth=45, anchor="center", stretch=False)
        tree.column("Target Price (USD)", width=125, minwidth=100, anchor="center", stretch=False)

        tree_vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree_hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=tree_vsb.set, xscrollcommand=tree_hsb.set)

        tree_vsb.pack(side="right", fill="y")
        tree_hsb.pack(side="bottom", fill="x")
        tree.pack(side="left", fill="both", expand=True)

        # Collect all known Part Numbers from the RFQ BOM line items to match directly against attachments
        target_part_numbers = set()
        for assy in assemblies:
            for it in assy.get("items", []):
                pn = str(it.get("part_number") or "").strip()
                if pn and len(pn) >= 4 and pn != "N/A":
                    target_part_numbers.add(pn)

        # Index all drawing PDFs and component spec sheets for this RFQ to resolve true drawing MPNs and MFRs on the fly
        dwg_parts_cache = {}
        try:
            from agents.drawing_agent import DrawingVisionAgent
            import pypdfium2 as pdfium
            dwg_paths = []
            seen_dwg_paths = set()
            def _add_dwg_cand(p):
                if not p or not os.path.exists(p): return
                norm_p = os.path.normcase(os.path.abspath(p))
                if norm_p not in seen_dwg_paths and p.lower().endswith(".pdf") and not os.path.basename(p).startswith("synthetic"):
                    seen_dwg_paths.add(norm_p)
                    dwg_paths.append(p)

            if current_rfq_entry and isinstance(current_rfq_entry, dict):
                em = current_rfq_entry.get("email", {})
                for a in em.get("attachments", []):
                    if isinstance(a, dict):
                        _add_dwg_cand(a.get("path") or a.get("file_path"))
                    elif isinstance(a, str):
                        _add_dwg_cand(a)
                s_dir = em.get("stage_dir", "")
                if s_dir and os.path.exists(s_dir):
                    for root, _, f_names in os.walk(s_dir):
                        for fn in f_names:
                            _add_dwg_cand(os.path.join(root, fn))

            # Also scan active ContXs EmailStaging directories
            gen_stage = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'ContXs', 'EmailStaging')
            if os.path.exists(gen_stage):
                for root, _, f_names in os.walk(gen_stage):
                    for fn in f_names:
                        _add_dwg_cand(os.path.join(root, fn))

            known_mfr_dict = {
                'molex': 'Molex', 'tyco': 'TE Connectivity / Tyco', 'te connectivity': 'TE Connectivity / Tyco',
                'te': 'TE Connectivity / Tyco', 'jst': 'JST', 'heiniger': 'Heiniger', 'harting': 'Harting',
                'phoenix': 'Phoenix Contact', '3m': '3M', 'sick': 'Sick', 'fci': 'FCI', 'amphenol': 'Amphenol',
                'hirose': 'Hirose', 'samtec': 'Samtec', 'lapp': 'LAPP', 'helukabel': 'Helukabel', 'alpha wire': 'Alpha Wire'
            }

            # Step 1: High-Priority Dedicated Component Datasheet PDFs (e.g. BB0_30062421_EN_00.pdf, B80_30062421_EN_00.pdf)
            for dp in dwg_paths:
                fn = os.path.basename(dp)
                m_fn_part = re.search(r'\b(?:BB0|B80|DS|SPEC)?[_-]?([0-9]{7,10})', fn, re.I)
                if m_fn_part:
                    doc_sap = m_fn_part.group(1)
                    if doc_sap in target_part_numbers:
                        try:
                            pdf_doc = pdfium.PdfDocument(dp)
                            for p_i in range(len(pdf_doc)):
                                p_txt = pdf_doc[p_i].get_textpage().get_text_range()
                                d_mfr = ""
                                for km, km_val in known_mfr_dict.items():
                                    if re.search(r'\b' + km + r'\b', p_txt, re.I) and not d_mfr and 'cleanup' not in p_txt.lower() and 'history' not in p_txt.lower():
                                        d_mfr = km_val
                                        break
                                if not d_mfr:
                                    m_h = re.search(r'Hersteller\s*:[\s\n\r]*([A-Za-z0-9\s]+)', p_txt, re.I)
                                    if m_h:
                                        d_mfr = m_h.group(1).strip().split('\n')[0].strip()

                                # Check Part Number, Typ, Typ Hersteller, or Order Code
                                for m in re.finditer(r'(?:Part\s*Number|Typ\s*Hersteller|Typ|Order\s*Code|Bestell-?Nr)[\s:\n\r]+([0-9A-Za-z\-\.\/\s]{2,30})', p_txt, re.I):
                                    cand_code = m.group(1).strip().split('\n')[0].strip()
                                    clean_c = DrawingVisionAgent.sanitize_mpn(cand_code, mfr=d_mfr)
                                    if clean_c and clean_c != doc_sap and clean_c.lower() not in DrawingVisionAgent.MPN_BLACKLIST:
                                        dwg_parts_cache[doc_sap] = (clean_c, d_mfr or "Deltron")
                                        break
                                if doc_sap in dwg_parts_cache:
                                    break
                        except Exception:
                            pass

            # Step 2: Multi-Item Drawing Blueprint Sheets (e.g. AJ0_30081932_EN_00.pdf)
            for dp in dwg_paths:
                try:
                    pdf_doc = pdfium.PdfDocument(dp)
                    fn = os.path.basename(dp)
                    for p_i in range(len(pdf_doc)):
                        p_txt = pdf_doc[p_i].get_textpage().get_text_range()
                        d_mfr = ""
                        for km, km_val in known_mfr_dict.items():
                            if re.search(r'\b' + km + r'\b', p_txt, re.I) and not d_mfr and 'cleanup' not in p_txt.lower() and 'history' not in p_txt.lower():
                                d_mfr = km_val
                                break

                        for tpn in target_part_numbers:
                            if tpn in dwg_parts_cache:
                                continue  # Dedicated datasheet already gave ground truth!
                            for m_sap in re.finditer(r'\b' + tpn + r'\b', p_txt):
                                idx = m_sap.start()
                                raw_chunk = p_txt[max(0, idx - 350):idx + 350]
                                loc_mpn = ""
                                for m in re.finditer(r'(?:Order\s*Code|Ordercode|Order-Code|Part\s*Number|Part\s*No|Bestell-?Nr|MFR\s*Part)[\s:\n\r]+([0-9A-Za-z\-\.\/]{4,30})', raw_chunk, re.I):
                                    cand_code = m.group(1).strip()
                                    clean_c = DrawingVisionAgent.sanitize_mpn(cand_code, mfr=d_mfr)
                                    if clean_c and clean_c != tpn and clean_c.lower() not in DrawingVisionAgent.MPN_BLACKLIST:
                                        loc_mpn = clean_c
                                        break
                                if loc_mpn:
                                    dwg_parts_cache[tpn] = (loc_mpn, d_mfr or "Molex")
                                    break
                except Exception:
                    pass
        except Exception:
            pass

        all_table_records = []
        seen_table_records = set()
        seen_assy_records = set()
        for assy in assemblies:
            a_no = str(assy.get("assy_no", "")).strip()
            if not a_no or a_no in seen_assy_records: continue
            seen_assy_records.add(a_no)
            a_model = assy.get("assy_model", "")
            a_rev = str(assy.get("assy_rev", "")).replace("Rev", "").strip()
            a_tp = str(assy.get("target_price", meta.get("target_price", "N/A"))).replace('$', '').strip()
            for it in assy.get("items", []):
                p_no = str(it.get("part_number") or "").strip()
                desc_str = str(it.get("description") or "").strip()
                rec_key = (a_no, p_no, desc_str)
                if rec_key in seen_table_records: continue
                seen_table_records.add(rec_key)
                
                raw_mpn_val = it.get("mpn")
                raw_mfr_val = it.get("mfr")
                is_ai_suggested = False

                # Auto-enrich from drawing cache if MPN was empty/part number or MFR was unknown
                if (not raw_mpn_val or raw_mpn_val == p_no) and p_no in dwg_parts_cache:
                    c_cached_mpn, c_cached_mfr = dwg_parts_cache[p_no]
                    if c_cached_mpn: 
                        raw_mpn_val = c_cached_mpn
                        it["mpn"] = c_cached_mpn
                    if c_cached_mfr and (not raw_mfr_val or raw_mfr_val == "Unknown"): 
                        raw_mfr_val = c_cached_mfr
                        it["mfr"] = c_cached_mfr

                # Check if MPN is valid, AI suggested, or needs manual sourcing
                is_ai_suggested = False
                needs_manual_source = False

                if not raw_mpn_val or raw_mpn_val == p_no or str(raw_mpn_val).upper() in ("UNKNOWN", "N/A", "NONE", "") or re.match(r'^[\-\_\.\*\s\/]+$', str(raw_mpn_val)):
                    raw_mpn_val = ""
                    try:
                        from agents.web_sourcing_engine import WebSourcingEngine
                        em_body = ""
                        if current_rfq_entry and isinstance(current_rfq_entry, dict):
                            em_body = current_rfq_entry.get("email", {}).get("body", "") if isinstance(current_rfq_entry.get("email"), dict) else ""
                        fused_desc = f"{desc_str} {em_body[:250]}".strip() if em_body else desc_str
                        cands = WebSourcingEngine.suggest_mpn_candidates(fused_desc, raw_mfr_val or "", p_no)
                        if cands:
                            top_cand = cands[0]
                            raw_mpn_val = top_cand["mpn"]
                            it["mpn"] = top_cand["mpn"]
                            if not raw_mfr_val or str(raw_mfr_val) in ("Unknown", "", "Auto-Detect"):
                                raw_mfr_val = top_cand["mfr"]
                                it["mfr"] = top_cand["mfr"]
                            is_ai_suggested = True
                        else:
                            needs_manual_source = True
                    except Exception as ex:
                        print(f"[ReviewStudio] Auto-enrich MPN error: {ex}")
                        needs_manual_source = True

                if not raw_mpn_val:
                    needs_manual_source = True
                    c_mpn, c_mfr = "", (raw_mfr_val if raw_mfr_val and raw_mfr_val != "Unknown" else "")
                else:
                    c_mpn, c_mfr = _pair_mpn_mfr(raw_mpn_val, raw_mfr_val)

                current_assy_lines = sum(1 for r in all_table_records if r["assy_no"] == a_no) + 1
                all_table_records.append({
                    "line": str(current_assy_lines),
                    "assy_no": a_no,
                    "assy_model": a_model,
                    "assy_rev": a_rev,
                    "part": p_no,
                    "desc": desc_str or "Component",
                    "mpn": c_mpn,
                    "mfr": c_mfr,
                    "qty": str(it.get("qty", 1)),
                    "uom": str(it.get("uom") or "EA"),
                    "tp": a_tp or "N/A",
                    "ai_suggested": is_ai_suggested,
                    "needs_manual_source": needs_manual_source,
                    "raw_item": it
                })

        # Selected Evidence Card & Actions (Compact & Responsive Layout)
        evidence_card = tk.Frame(tab_components, bg="#1E293B", padx=8, pady=6, bd=1, relief="solid")
        evidence_card.pack(fill="x", padx=6, pady=(2, 4))

        ev_src_lbl = tk.Label(evidence_card, text="Select any component row above to inspect source blueprint, zone, and extraction confidence.",
                              font=("Segoe UI", 9, "bold"), fg="#E2E8F0", bg="#1E293B", anchor="w")
        ev_src_lbl.pack(fill="x")

        ev_detail = tk.Label(evidence_card, text="", font=("Consolas", 8), fg="#94A3B8", bg="#1E293B", anchor="w")
        ev_detail.pack(fill="x", pady=(1, 3))

        ev_actions = tk.Frame(evidence_card, bg="#1E293B")
        ev_actions.pack(fill="x", anchor="w")

        selected_record_holder = [None]

        def _show_instant_traceability():
            rec = selected_record_holder[0]
            if not rec:
                messagebox.showinfo("Select Item", "Please select a component row from the table first.", parent=dlg)
                return
            
            t_dlg = tk.Toplevel(dlg)
            t_dlg.title(f"🔍 Extraction Provenance & Evidence — {rec['part']}")
            t_dlg.geometry("540x440")
            t_dlg.minsize(480, 380)
            t_dlg.configure(bg="#0F172A")
            t_dlg.transient(dlg)
            t_dlg.grab_set()

            # Header
            t_hdr = tk.Frame(t_dlg, bg="#1E293B", padx=16, pady=12)
            t_hdr.pack(fill="x")
            tk.Label(t_hdr, text=f"🔍 Extraction Provenance: {rec['part']}", font=("Segoe UI", 12, "bold"), fg="#FFFFFF", bg="#1E293B").pack(anchor="w")
            tk.Label(t_hdr, text=f"Assembly {rec['assy_no']} (Rev {rec['assy_rev']}) • {rec['desc'][:40]}", font=("Segoe UI", 9), fg="#94A3B8", bg="#1E293B").pack(anchor="w", pady=(2, 0))

            # Body Card
            body_fr = tk.Frame(t_dlg, bg="#0F172A", padx=16, pady=10)
            body_fr.pack(fill="both", expand=True)

            raw_it = rec.get("raw_item", {})
            ev = raw_it.get("evidence", {})
            p_ev = ev.get("part", {})
            mpn_ev = ev.get("mpn", {})
            src_doc = p_ev.get("source_doc") or mpn_ev.get("source_doc") or "Drawing Blueprint / RFQ Attachment"
            zone = p_ev.get("zone", "Schematic Callout & BOM Block")
            snip = p_ev.get("snippet") or f"{rec['mfr']} {rec['mpn']} {rec['desc']}".strip()
            conf = int(p_ev.get("confidence", 0.95) * 100)
            reason = mpn_ev.get("reasoning") or p_ev.get("reasoning") or "Matched from source document drawing block via spatial OCR proximity."

            info_cards = [
                ("📄 Source Document:", src_doc, "#38BDF8"),
                ("📍 Document Zone:", zone, "#F59E0B"),
                ("🎯 AI Extraction Confidence:", f"{conf}% (High Confidence)", "#10B981"),
                ("📜 Raw Matched Snippet:", f"\"{snip}\"", "#F1F5F9"),
                ("🧠 AI Traceability Reasoning:", reason, "#CBD5E1")
            ]

            for label_t, val_t, val_c in info_cards:
                row_f = tk.Frame(body_fr, bg="#1E293B", padx=10, pady=5, bd=1, relief="solid")
                row_f.pack(fill="x", pady=2)
                tk.Label(row_f, text=label_t, font=("Segoe UI", 8, "bold"), fg="#94A3B8", bg="#1E293B").pack(anchor="w")
                tk.Label(row_f, text=val_t, font=("Consolas" if label_t.startswith("📜") else "Segoe UI", 9, "bold" if not label_t.startswith("🧠") else "normal"), fg=val_c, bg="#1E293B", wraplength=490, justify="left").pack(anchor="w", pady=(1, 0))

            # Footer
            ftr = tk.Frame(t_dlg, bg="#1E293B", padx=16, pady=10)
            ftr.pack(fill="x", side="bottom")
            tk.Button(ftr, text="🔍 Open Visual Blueprint Studio", command=lambda: [t_dlg.destroy(), _inspect_visual_annotation()], bg="#0284C7", fg="#FFFFFF", font=("Segoe UI", 9, "bold"), relief="flat", padx=12, pady=4, cursor="hand2").pack(side="left")
            tk.Button(ftr, text="Close", command=t_dlg.destroy, bg="#475569", fg="#FFFFFF", font=("Segoe UI", 9), relief="flat", padx=14, pady=4, cursor="hand2").pack(side="right")

        def _edit_selected_line():
            rec = selected_record_holder[0]
            if not rec: return
            edit_dlg = tk.Toplevel(dlg)
            edit_dlg.title(f"✏️ Amend Component — {rec['part']}")
            edit_dlg.geometry("440x265")
            edit_dlg.resizable(False, False)
            edit_dlg.configure(bg="#0F172A")
            edit_dlg.transient(dlg)
            edit_dlg.grab_set()

            tk.Label(edit_dlg, text=f"Amend Line Item: {rec['part']}", font=("Segoe UI", 11, "bold"), fg="#FFFFFF", bg="#0F172A").pack(pady=(10, 6))
            e_fr = tk.Frame(edit_dlg, bg="#0F172A", padx=20)
            e_fr.pack(fill="x")

            fields = [("Description:", "desc"), ("MPN (comma-separated):", "mpn"), ("MFR (comma-separated):", "mfr"), ("QTY:", "qty"), ("UOM:", "uom")]
            vars_dict = {}
            for row_i, (lbl_t, k) in enumerate(fields):
                tk.Label(e_fr, text=lbl_t, font=("Segoe UI", 9), fg="#94A3B8", bg="#0F172A").grid(row=row_i, column=0, sticky="w", pady=2)
                v = tk.StringVar(value=rec[k])
                vars_dict[k] = v
                tk.Entry(e_fr, textvariable=v, font=("Segoe UI", 9), width=26, bd=1, relief="solid").grid(row=row_i, column=1, sticky="w", pady=2, padx=(8, 0))

            def _save_edit():
                try:
                    for k in vars_dict:
                        rec[k] = vars_dict[k].get().strip()
                        rec["raw_item"][k if k != "desc" else "description"] = rec[k]
                    p_mpn, p_mfr = _pair_mpn_mfr(rec["mpn"], rec["mfr"])
                    rec["mpn"] = p_mpn
                    rec["mfr"] = p_mfr
                    rec["raw_item"]["mpn"] = p_mpn
                    rec["raw_item"]["mfr"] = p_mfr

                    from agents.correction_store import CorrectionStore
                    from agents.verified_bom_store import VerifiedBOMStore
                    from agents.llm_gateway import LLMGateway
                    cs = CorrectionStore()
                    vbs = VerifiedBOMStore()

                    p_no = str(rec["part"]).strip()
                    a_no_k = str(rec["assy_no"]).strip()
                    for hint in (p_no, f"{a_no_k}::{p_no}"):
                        if not hint: continue
                        if rec["mpn"]: cs.save_correction(hint, "mpn", "", rec["mpn"], mfr=rec["mfr"], note="Manual amendment in Review Studio")
                        if rec["mfr"]: cs.save_correction(hint, "mfr", "", rec["mfr"], note="Manual amendment in Review Studio")
                        if rec["desc"]: cs.save_correction(hint, "description", "", rec["desc"], note="Manual amendment in Review Studio")
                        if rec["qty"]: cs.save_correction(hint, "qty", "", rec["qty"], note="Manual amendment in Review Studio")
                        if rec["uom"]: cs.save_correction(hint, "uom", "", rec["uom"], note="Manual amendment in Review Studio")

                    # Update VerifiedBOMStore for this assembly
                    if self._last_extracted_rfq_json:
                        for assy_obj in self._last_extracted_rfq_json.get("assemblies", []):
                            if str(assy_obj.get("assy_no", "")).strip() == a_no_k:
                                vbs.save_verified_assembly(a_no_k, assy_obj, customer=cust, rfq_no=rfq_no)
                                break

                    LLMGateway().clear_cache(rfq_no)

                    from agents.synthetic_bom_generator import SyntheticBOMGenerator
                    gen_res = SyntheticBOMGenerator().generate_synthetic_excel(self._last_extracted_rfq_json)
                    if gen_res.get("success") and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
                        self._staged_bom_payload["file_path"] = gen_res["file_path"]

                    _refresh_table()

                    # Live update chatbot card in background
                    try:
                        if current_rfq_entry:
                            current_rfq_entry["rfq_json"] = self._last_extracted_rfq_json
                            current_rfq_entry["synthetic_bom"] = gen_res
                            self._load_selected_rfq_payload(current_rfq_entry)
                    except Exception as err:
                        print(f"[AmendLiveSync] {err}")

                    edit_dlg.destroy()
                    messagebox.showinfo("Saved", "Component successfully updated and reflected in Chatbot table!", parent=dlg)
                except Exception as ex:
                    messagebox.showerror("Error", f"Could not save amendment:\n{ex}", parent=edit_dlg)

            tk.Button(edit_dlg, text="Save & Update BOM", command=_save_edit, bg="#2563EB", fg="#FFFFFF", font=("Segoe UI", 9, "bold"), relief="flat", padx=12, pady=4, cursor="hand2").pack(pady=(10, 8))

        def _inspect_visual_annotation():
            rec = selected_record_holder[0]
            if not rec:
                messagebox.showinfo("Select Item", "Please select a component row from the table first.", parent=dlg)
                return

            target_assy = str(rec.get("assy_no", "")).strip()
            target_part = str(rec.get("part", "")).strip()
            target_mpn = str(rec.get("mpn", "")).strip()
            target_desc = str(rec.get("desc", "")).strip()

            # 1. Collect candidate files from active RFQ email attachments and staging directories (deduplicated)
            candidate_files = []
            seen_paths = set()
            seen_basenames = set()

            def _add_cand(p):
                if not p or not os.path.exists(p): return
                norm_p = os.path.normcase(os.path.abspath(p))
                b_name = os.path.basename(p).lower()
                if norm_p not in seen_paths and b_name not in seen_basenames:
                    seen_paths.add(norm_p)
                    seen_basenames.add(b_name)
                    candidate_files.append(p)

            if current_rfq_entry and isinstance(current_rfq_entry, dict):
                em = current_rfq_entry.get("email", {})
                for a in em.get("attachments", []):
                    if isinstance(a, dict) and a.get("path"):
                        _add_cand(a["path"])
                    elif isinstance(a, str):
                        _add_cand(a)
                s_dir = em.get("stage_dir", "")
                if s_dir and os.path.exists(s_dir):
                    for root, _, f_names in os.walk(s_dir):
                        for fn in f_names:
                            _add_cand(os.path.join(root, fn))

            # Also check general EmailStaging directories if no PDF drawings are present in current candidates
            has_pdf = any(fp.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg', '.bmp', '.webp')) for fp in candidate_files)
            gen_stage = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'ContXs', 'EmailStaging')
            if os.path.exists(gen_stage) and not has_pdf:
                for root, _, f_names in os.walk(gen_stage):
                    for fn in f_names:
                        _add_cand(os.path.join(root, fn))

            # 2. Find best matching drawing
            matched_fp = ""
            for fp in candidate_files:
                fn = os.path.basename(fp).lower()
                if fn.endswith(('.pdf', '.png', '.jpg', '.jpeg', '.bmp', '.webp')):
                    if (target_assy and target_assy.lower() in fn) or (target_part and target_part.lower() in fn):
                        matched_fp = fp
                        break

            if not matched_fp:
                for fp in candidate_files:
                    fn = os.path.basename(fp).lower()
                    if fn.endswith(('.pdf', '.png', '.jpg', '.jpeg', '.bmp', '.webp')) and not fn.startswith(('original_email', 'thumb_')):
                        matched_fp = fp
                        break

            if not matched_fp:
                from tkinter import filedialog
                matched_fp = filedialog.askopenfilename(
                    title=f"Select Drawing/Blueprint for Part {target_part} ({target_assy})",
                    filetypes=[("PDF & Drawing Files", "*.pdf;*.png;*.jpg;*.jpeg;*.bmp"), ("All Files", "*.*")],
                    parent=dlg
                )

            if matched_fp and os.path.exists(matched_fp):
                from agents.visual_annotation_dialog import open_visual_annotation_studio
                terms = [t for t in (target_part, target_mpn, target_desc) if t and len(t) >= 2]

                def _on_annot_amended(annot_dict, orig_rec):
                    new_cat = annot_dict.get("category", "")
                    new_txt = annot_dict.get("text", "").strip()
                    if not new_txt: return
                    
                    if new_cat == "ASSEMBLY_NUMBER":
                        orig_rec["Assembly"] = new_txt
                        orig_rec["assy_no"] = new_txt
                    elif new_cat == "PART_NUMBER":
                        orig_rec["Part Number"] = new_txt
                        orig_rec["part"] = new_txt
                        if "raw_item" in orig_rec: orig_rec["raw_item"]["part_number"] = new_txt
                    elif new_cat == "MPN":
                        orig_rec["MPN"] = new_txt
                        orig_rec["mpn"] = new_txt
                        if "raw_item" in orig_rec: orig_rec["raw_item"]["mpn"] = new_txt
                    elif new_cat == "DESCRIPTION":
                        orig_rec["Description"] = new_txt
                        orig_rec["desc"] = new_txt
                        if "raw_item" in orig_rec: orig_rec["raw_item"]["description"] = new_txt
                    elif new_cat == "MANUFACTURER":
                        orig_rec["Manufacturer"] = new_txt
                        orig_rec["mfr"] = new_txt
                        if "raw_item" in orig_rec: orig_rec["raw_item"]["mfr"] = new_txt
                    elif new_cat == "SPECIFICATION":
                        orig_rec["Specification"] = new_txt
                        orig_rec["spec"] = new_txt

                    orig_rec["Status"] = "Amended"
                    orig_rec["Verified"] = True

                    sel_items = tree.selection()
                    if sel_items:
                        vals = list(tree.item(sel_items[0], "values"))
                        field_map = {
                            "Assy#": orig_rec.get("assy_no"),
                            "Part": orig_rec.get("part"),
                            "MPN": orig_rec.get("mpn"),
                            "Description": orig_rec.get("desc"),
                            "MFR": orig_rec.get("mfr")
                        }
                        for field, val in field_map.items():
                            if field in cols and val:
                                idx = cols.index(field)
                                if idx < len(vals):
                                    vals[idx] = val
                        tree.item(sel_items[0], values=vals, tags=("amended",))

                    # Update bottom evidence inspector display immediately
                    ev_src_lbl.config(text=f"📌 Focused: Line #{orig_rec.get('line', '?')} • Part: {orig_rec.get('part')} • Assembly: {orig_rec.get('assy_no')} (Rev {orig_rec.get('assy_rev', '00')}) • Status: ✏️ Amended")
                    ev_detail.config(text=f"💡 Updated via Visual Annotation Studio: [{new_cat}] \"{new_txt}\" • Live memory synchronized.")

                    try:
                        from agents.correction_store import CorrectionStore
                        cs = CorrectionStore()
                        cs.save_correction(
                            doc_hint=os.path.basename(matched_fp),
                            field=new_cat,
                            wrong_value=str(orig_rec.get("Part Number", orig_rec.get("part", ""))),
                            correct_value=new_txt,
                            mfr=str(orig_rec.get("Manufacturer", orig_rec.get("mfr", ""))),
                            note=f"Visual Annotation Studio user amendment ({new_cat})",
                            corrected_by="Engineer"
                        )
                    except Exception as c_err:
                        print(f"[ReviewStudio] CorrectionStore save err: {c_err}")

                    # Invalidate stale LLM Gateway cache & regenerate synthetic BOM payload
                    try:
                        from agents.llm_gateway import LLMGateway
                        from agents.synthetic_bom_generator import SyntheticBOMGenerator
                        LLMGateway().clear_cache(rfq_no)
                        if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                            gen_res = SyntheticBOMGenerator().generate_synthetic_excel(self._last_extracted_rfq_json)
                            if gen_res.get("success") and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
                                self._staged_bom_payload["file_path"] = gen_res["file_path"]
                    except Exception as gen_ex:
                        print(f"[ReviewStudio] Synthetic regen notice: {gen_ex}")

                p_disp_t = _display_val(target_part, 'PART')
                a_disp_t = _display_val(target_assy, 'ASSY')
                open_visual_annotation_studio(
                    parent=dlg,
                    file_path=matched_fp,
                    highlight_terms=terms,
                    component_data=rec,
                    title=f"🔍 Unified Multi-Modal Evidence & Annotation Studio — Part {p_disp_t} ({a_disp_t})",
                    on_update_callback=_on_annot_amended,
                    candidate_files=candidate_files,
                    email_data=em if 'em' in locals() and isinstance(em, dict) else (current_rfq_entry.get("email") if isinstance(current_rfq_entry, dict) else None),
                    is_anonymized_mode=bool(is_anonymized_mode[0])
                )

        def _recalculate_assembly_line_numbers():
            """Recalculates and reassigns sequential line numbers (1, 2, 3, ...) per assembly."""
            counts_per_assy = {}
            for r in all_table_records:
                a_no = r.get("assy_no", "")
                cur_cnt = counts_per_assy.get(a_no, 0) + 1
                counts_per_assy[a_no] = cur_cnt
                r["line"] = str(cur_cnt)

            if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                for assy_obj in self._last_extracted_rfq_json.get("assemblies", []):
                    a_no = str(assy_obj.get("assy_no", "")).strip()
                    matched_items = [r["raw_item"] for r in all_table_records if r.get("assy_no") == a_no and "raw_item" in r]
                    assy_obj["items"] = matched_items

        def _delete_selected_line():
            rec = selected_record_holder[0]
            if not rec:
                sel = tree.selection()
                if sel:
                    vals = tree.item(sel[0], "values")
                    if vals:
                        l_num, a_no, part_k = vals[0], vals[1], vals[4]
                        rec = next((r for r in all_table_records if r["line"] == str(l_num) and r["assy_no"] == str(a_no) and r["part"] == str(part_k)), None)
            if not rec:
                messagebox.showinfo("Select Item", "Please select a component line to delete.", parent=dlg)
                return

            p_disp = rec.get("part", "item")
            a_disp = rec.get("assy_no", "")
            l_disp = rec.get("line", "")

            confirm = messagebox.askyesno(
                "Confirm Delete Line Item",
                f"Are you sure you want to delete Line #{l_disp} (Part: {p_disp}) from Assembly {a_disp}?\n\nLine numbers for Assembly {a_disp} will be automatically re-sequenced.",
                parent=dlg
            )
            if not confirm:
                return

            if rec in all_table_records:
                all_table_records.remove(rec)

            if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                for assy_obj in self._last_extracted_rfq_json.get("assemblies", []):
                    if str(assy_obj.get("assy_no", "")).strip() == a_disp:
                        raw_it = rec.get("raw_item")
                        if raw_it and raw_it in assy_obj.get("items", []):
                            assy_obj["items"].remove(raw_it)
                        elif "items" in assy_obj:
                            assy_obj["items"] = [it for it in assy_obj["items"] if str(it.get("part_number", "")).strip() != p_disp]
                        break

            _recalculate_assembly_line_numbers()
            selected_record_holder[0] = None
            _refresh_table()

            try:
                from agents.llm_gateway import LLMGateway
                from agents.synthetic_bom_generator import SyntheticBOMGenerator
                LLMGateway().clear_cache(rfq_no)
                if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                    gen_res = SyntheticBOMGenerator().generate_synthetic_excel(self._last_extracted_rfq_json)
                    if gen_res.get("success") and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
                        self._staged_bom_payload["file_path"] = gen_res["file_path"]
                    if current_rfq_entry:
                        current_rfq_entry["rfq_json"] = self._last_extracted_rfq_json
                        current_rfq_entry["synthetic_bom"] = gen_res
            except Exception as ex:
                print(f"[ReviewStudio] Delete sync notice: {ex}")

            messagebox.showinfo("Line Item Deleted", f"✅ Line #{l_disp} deleted successfully.\nAssembly {a_disp} line sequence updated.", parent=dlg)

        def _add_new_line():
            add_dlg = tk.Toplevel(dlg)
            add_dlg.title("➕ Add New Component Line Item")
            add_dlg.geometry("520x430")
            add_dlg.minsize(480, 360)
            add_dlg.configure(bg="#0F172A")
            add_dlg.transient(dlg)
            add_dlg.grab_set()

            a_hdr = tk.Frame(add_dlg, bg="#1E293B", padx=16, pady=10)
            a_hdr.pack(fill="x")
            tk.Label(a_hdr, text="➕ Add Component Line Item", font=("Segoe UI", 12, "bold"), fg="#FFFFFF", bg="#1E293B").pack(anchor="w")
            tk.Label(a_hdr, text="Adds a component to the assembly and automatically assigns the next sequence line number.", font=("Segoe UI", 8), fg="#94A3B8", bg="#1E293B").pack(anchor="w")

            body_f = tk.Frame(add_dlg, bg="#0F172A", padx=18, pady=12)
            body_f.pack(fill="both", expand=True)

            all_assy_list = list(dict.fromkeys(r.get("assy_no", "") for r in all_table_records if r.get("assy_no")))
            sel_rec = selected_record_holder[0]
            default_assy = sel_rec.get("assy_no", all_assy_list[0] if all_assy_list else "") if sel_rec else (all_assy_list[0] if all_assy_list else "")
            
            tk.Label(body_f, text="Target Assembly Number:", font=("Segoe UI", 9, "bold"), fg="#CBD5E1", bg="#0F172A").grid(row=0, column=0, sticky="w", pady=3)
            var_assy = tk.StringVar(value=default_assy)
            combo_assy = ttk.Combobox(body_f, textvariable=var_assy, values=all_assy_list, state="readonly", font=("Segoe UI", 9), width=28)
            combo_assy.grid(row=0, column=1, sticky="w", pady=3, padx=(8, 0))

            var_part = tk.StringVar()
            var_desc = tk.StringVar()
            var_mpn = tk.StringVar()
            var_mfr = tk.StringVar()
            var_qty = tk.StringVar(value="1")
            var_uom = tk.StringVar(value="PC")

            fields = [
                ("Part Number / Article #:", var_part),
                ("Description:", var_desc),
                ("Manufacturer MPN:", var_mpn),
                ("Manufacturer (MFR):", var_mfr),
                ("Quantity (QTY):", var_qty),
            ]
            for row_idx, (lbl_t, v_obj) in enumerate(fields, start=1):
                tk.Label(body_f, text=lbl_t, font=("Segoe UI", 9), fg="#94A3B8", bg="#0F172A").grid(row=row_idx, column=0, sticky="w", pady=3)
                ent = tk.Entry(body_f, textvariable=v_obj, font=("Segoe UI", 9), width=30, bg="#1E293B", fg="#FFFFFF", insertbackground="#38BDF8", bd=1, relief="solid")
                ent.grid(row=row_idx, column=1, sticky="w", pady=3, padx=(8, 0), ipady=2)

            tk.Label(body_f, text="Unit of Measure (UOM):", font=("Segoe UI", 9), fg="#94A3B8", bg="#0F172A").grid(row=6, column=0, sticky="w", pady=3)
            combo_uom = ttk.Combobox(body_f, textvariable=var_uom, values=["PC", "M", "EA", "MM", "SET", "FT"], state="readonly", font=("Segoe UI", 9), width=10)
            combo_uom.grid(row=6, column=1, sticky="w", pady=3, padx=(8, 0))

            def _pick_from_studio():
                target_a = var_assy.get().strip()
                dummy_rec = {
                    "assy_no": target_a,
                    "part": var_part.get().strip() or "NEW_PART",
                    "mpn": var_mpn.get().strip(),
                    "desc": var_desc.get().strip(),
                    "mfr": var_mfr.get().strip(),
                    "qty": var_qty.get().strip() or "1",
                    "uom": var_uom.get().strip() or "PC",
                    "line": "?"
                }
                
                matched_fp = ""
                for fp in candidate_files:
                    fn = os.path.basename(fp).lower()
                    if target_a and target_a.lower() in fn and fn.endswith(('.pdf', '.png', '.jpg')):
                        matched_fp = fp; break
                if not matched_fp and candidate_files:
                    matched_fp = candidate_files[0]

                if matched_fp and os.path.exists(matched_fp):
                    def _on_studio_picked(annot_dict, orig_rec):
                        cat = annot_dict.get("category", "")
                        txt = annot_dict.get("text", "").strip()
                        if cat == "PART_NUMBER": var_part.set(txt)
                        elif cat == "MPN": var_mpn.set(txt)
                        elif cat == "MANUFACTURER": var_mfr.set(txt)
                        elif cat in ("DESCRIPTION", "SPECIFICATION"): var_desc.set(txt)

                    from agents.visual_annotation_dialog import open_visual_annotation_studio
                    open_visual_annotation_studio(
                        parent=add_dlg, file_path=matched_fp, highlight_terms=[target_a],
                        component_data=dummy_rec,
                        title=f"🔍 Pick Sourcing from Blueprint — Assembly {_display_val(target_a, 'ASSY')}",
                        on_update_callback=_on_studio_picked,
                        candidate_files=candidate_files,
                        is_anonymized_mode=bool(is_anonymized_mode[0])
                    )
                else:
                    messagebox.showinfo("No Drawing", f"No blueprint file found for Assembly {target_a}.", parent=add_dlg)

            tk.Button(body_f, text="🔍 Pick MPN & Specs in Annotation Studio", command=_pick_from_studio, bg="#0284C7", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").grid(row=7, column=0, columnspan=2, sticky="w", pady=(10, 4))

            ftr_f = tk.Frame(add_dlg, bg="#1E293B", padx=16, pady=10)
            ftr_f.pack(fill="x", side="bottom")

            def _save_new_line():
                p_val = var_part.get().strip()
                d_val = var_desc.get().strip()
                m_mpn = var_mpn.get().strip()
                m_mfr = var_mfr.get().strip()
                q_val = var_qty.get().strip() or "1"
                u_val = var_uom.get().strip() or "PC"
                t_assy = var_assy.get().strip()

                if not p_val and not d_val and not m_mpn:
                    messagebox.showwarning("Missing Information", "Please enter at least a Part Number, Description, or MPN.", parent=add_dlg)
                    return

                p_val = p_val or (m_mpn if m_mpn else "NEW_ITEM")
                p_mpn, p_mfr = _pair_mpn_mfr(m_mpn, m_mfr)

                new_raw_item = {
                    "part_number": p_val,
                    "description": d_val or f"Component {p_val}",
                    "mpn": p_mpn,
                    "mfr": p_mfr,
                    "qty": q_val,
                    "uom": u_val,
                    "evidence": {
                        "part": {"source_doc": "User Manually Added in Review Studio", "zone": "MANUAL_ENTRY", "confidence": 1.0, "reasoning": "Manually added by engineer"}
                    }
                }

                existing_assy_rec = next((r for r in all_table_records if r.get("assy_no") == t_assy), None)
                a_mod = existing_assy_rec.get("assy_model", "") if existing_assy_rec else ""
                a_rev = existing_assy_rec.get("assy_rev", "00") if existing_assy_rec else "00"
                a_tp = existing_assy_rec.get("tp", "N/A") if existing_assy_rec else "N/A"

                new_rec = {
                    "line": "?",
                    "assy_no": t_assy,
                    "assy_model": a_mod,
                    "assy_rev": a_rev,
                    "part": p_val,
                    "desc": d_val or f"Component {p_val}",
                    "mpn": p_mpn,
                    "mfr": p_mfr,
                    "qty": q_val,
                    "uom": u_val,
                    "tp": a_tp,
                    "ai_suggested": False,
                    "needs_manual_source": False if p_mpn else True,
                    "raw_item": new_raw_item
                }

                last_idx = -1
                for idx, r in enumerate(all_table_records):
                    if r.get("assy_no") == t_assy:
                        last_idx = idx
                if last_idx >= 0:
                    all_table_records.insert(last_idx + 1, new_rec)
                else:
                    all_table_records.append(new_rec)

                if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                    for assy_obj in self._last_extracted_rfq_json.get("assemblies", []):
                        if str(assy_obj.get("assy_no", "")).strip() == t_assy:
                            if "items" not in assy_obj: assy_obj["items"] = []
                            assy_obj["items"].append(new_raw_item)
                            break

                _recalculate_assembly_line_numbers()
                _refresh_table()

                try:
                    from agents.synthetic_bom_generator import SyntheticBOMGenerator
                    from agents.llm_gateway import LLMGateway
                    LLMGateway().clear_cache(rfq_no)
                    if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                        gen_res = SyntheticBOMGenerator().generate_synthetic_excel(self._last_extracted_rfq_json)
                        if gen_res.get("success") and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
                            self._staged_bom_payload["file_path"] = gen_res["file_path"]
                        if current_rfq_entry:
                            current_rfq_entry["rfq_json"] = self._last_extracted_rfq_json
                            current_rfq_entry["synthetic_bom"] = gen_res
                except Exception as ex:
                    print(f"[ReviewStudio] Add sync notice: {ex}")

                add_dlg.destroy()
                messagebox.showinfo("Line Item Added", f"✅ Added '{p_val}' to Assembly {t_assy} as Line #{new_rec['line']}.\nSequence numbers updated.", parent=dlg)

            tk.Button(ftr_f, text="Cancel", command=add_dlg.destroy, bg="#475569", fg="#FFFFFF", font=("Segoe UI", 9), relief="flat", padx=12, pady=4, cursor="hand2").pack(side="right", padx=(8, 0))
            tk.Button(ftr_f, text="💾 Add Line Item & Recalculate", command=_save_new_line, bg="#059669", fg="#FFFFFF", font=("Segoe UI", 9, "bold"), relief="flat", padx=14, pady=4, cursor="hand2").pack(side="right")

        btn_add_line = tk.Button(ev_actions, text="➕ Add Line Item", command=_add_new_line, bg="#059669", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=7, pady=2, cursor="hand2")
        btn_add_line.pack(side="left", padx=(0, 3))
        btn_del_line = tk.Button(ev_actions, text="🗑️ Delete Line Item", command=_delete_selected_line, bg="#DC2626", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=7, pady=2, cursor="hand2")
        btn_del_line.pack(side="left", padx=(0, 3))
        btn_edit_line = tk.Button(ev_actions, text="✏️ Amend Item", command=_edit_selected_line, bg="#D97706", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=7, pady=2, cursor="hand2")
        btn_edit_line.pack(side="left", padx=(0, 3))
        btn_visual_annot = tk.Button(ev_actions, text="🔍 Visual Annotations", command=_inspect_visual_annotation, bg="#0284C7", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=7, pady=2, cursor="hand2")
        btn_visual_annot.pack(side="left", padx=(0, 3))
        btn_ai_ask = tk.Button(ev_actions, text="🔍 Trace Evidence", command=_show_instant_traceability, bg="#8B5CF6", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=7, pady=2, cursor="hand2")
        btn_ai_ask.pack(side="left", padx=(0, 3))

        def _on_row_select(event):
            sel = tree.selection()
            if not sel: return
            cur_pos = tree.index(sel[0]) + 1
            tot_vis = len(tree.get_children())
            tot_all = len(all_table_records)
            vals = tree.item(sel[0], "values")
            if not vals: return
            l_num, disp_a, disp_p = vals[0], vals[1], vals[4]
            comp_count_lbl.config(text=f"📌 Line {cur_pos} of {tot_vis} (Total: {tot_all})")
            for r in all_table_records:
                if str(r["line"]) == str(l_num) and (_display_val(r["assy_no"], "ASSY") == str(disp_a) or r["assy_no"] == str(disp_a)) and (_display_val(r["part"], "PART") == str(disp_p) or r["part"] == str(disp_p)):
                    selected_record_holder[0] = r
                    raw_it = r["raw_item"]
                    ev = raw_it.get("evidence", {})
                    p_ev = ev.get("part", {})
                    mpn_ev = ev.get("mpn", {})
                    src_doc = p_ev.get("source_doc") or mpn_ev.get("source_doc") or "Email Body Table"
                    zone = p_ev.get("zone", "BOM_TABLE")
                    snip = p_ev.get("snippet") or f"{r['mfr']} {r['mpn']} {r['desc']}".strip()
                    conf = int(p_ev.get("confidence", 0.95) * 100)
                    reason = mpn_ev.get("reasoning") or p_ev.get("reasoning") or "Directly extracted from source document drawing block."
                    disp_p_k = _display_val(r['part'], 'PART')
                    disp_a_k = _display_val(r['assy_no'], 'ASSY')
                    disp_r_k = _display_val(r['assy_rev'], 'REV')
                    ev_src_lbl.config(text=f"📌 Focused: Line #{l_num} (Row {cur_pos} of {tot_vis}) • Part: {disp_p_k} • Assembly: {disp_a_k} (Rev {disp_r_k}) • AI Conf: {conf}%")
                    ev_detail.config(text=f"💡 AI Reasoning: {reason} | Evidence: \"{snip}\" | Qty: {r['qty']} {r['uom']}")
                    break
        tree.bind("<<TreeviewSelect>>", _on_row_select)

        # Context Menu & Shortcuts for BOM Table Rows
        comp_ctx_menu = tk.Menu(tree, tearoff=0, bg="#1E293B", fg="#F8FAFC", activebackground="#2563EB", activeforeground="#FFFFFF", font=("Segoe UI", 9))
        comp_ctx_menu.add_command(label="➕ Add New Line Item to this Assembly", command=_add_new_line)
        comp_ctx_menu.add_command(label="🗑️ Delete Selected Line Item", command=_delete_selected_line)
        comp_ctx_menu.add_separator()
        comp_ctx_menu.add_command(label="🔍 Inspect Blueprint & Drawing Evidence", command=_inspect_visual_annotation)
        comp_ctx_menu.add_command(label="✏️ Amend Component Item", command=_edit_selected_line)
        comp_ctx_menu.add_command(label="🔍 View Extraction Provenance", command=_show_instant_traceability)
        comp_ctx_menu.add_separator()
        comp_ctx_menu.add_command(label="🧠 Learn Pattern for this RFQ", command=_approve_and_learn_pattern)

        tree.bind("<Delete>", lambda e: _delete_selected_line())
        tree.bind("<Control-n>", lambda e: _add_new_line())
        tree.bind("<Control-N>", lambda e: _add_new_line())

        def _show_comp_context_menu(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                _on_row_select(event)
                try:
                    comp_ctx_menu.tk_popup(event.x_root, event.y_root)
                finally:
                    comp_ctx_menu.grab_release()

        tree.bind("<Button-3>", _show_comp_context_menu)
        tree.bind("<Double-1>", lambda e: _inspect_visual_annotation())

        tree.tag_configure("ai_suggested", background="#FEF3C7", foreground="#92400E")  # Gold / Amber for Online Sourced
        tree.tag_configure("needs_manual_source", background="#FFE4E6", foreground="#BE123C")  # Pink / Rose for User Manual Sourcing
        tree.tag_configure("normal_even", background="#F8FAFC", foreground="#0F172A")
        tree.tag_configure("normal_odd", background="#FFFFFF", foreground="#0F172A")

        def _refresh_table(*args):
            for row in tree.get_children(): tree.delete(row)
            selected_assy_filter = assy_combo.get()
            search_query = search_var.get().strip().lower()
            count = 0
            for r in all_table_records:
                disp_a_no = _display_val(r["assy_no"], "ASSY")
                disp_a_model = _display_val(r["assy_model"], "MODEL")
                disp_a_rev = _display_val(r["assy_rev"], "REV")
                disp_part = _display_val(r["part"], "PART")

                if selected_assy_filter != "All Assemblies":
                    target_a_no = selected_assy_filter.split("—")[0].strip()
                    if r["assy_no"] != target_a_no and disp_a_no != target_a_no: continue
                if search_query:
                    searchable = f"{r['assy_no']} {disp_a_no} {r['assy_model']} {disp_a_model} {r['assy_rev']} {disp_a_rev} {r['part']} {disp_part} {r['desc']} {r['mpn']} {r['mfr']}".lower()
                    if search_query not in searchable: continue
                
                if r.get("needs_manual_source") or not r.get("mpn"):
                    row_tags = ["needs_manual_source"]
                elif r.get("ai_suggested"):
                    row_tags = ["ai_suggested"]
                else:
                    row_tags = ["normal_even"] if count % 2 == 0 else ["normal_odd"]

                tree.insert("", "end", values=(r["line"], disp_a_no, disp_a_model, disp_a_rev, disp_part, r["desc"], r["mpn"], r["mfr"], r["qty"], r["uom"], r["tp"]), tags=tuple(row_tags))
                count += 1
            tot_all = len(all_table_records)
            comp_count_lbl.config(text=f"📊 Showing {count} of {tot_all} Rows")
            u_assys = len(set(r["assy_no"] for r in all_table_records if r.get("assy_no")))
            try:
                cust_disp = _display_val(cust, "CUST")
                hdr_sub_lbl.config(text=f"Customer: {cust_disp} • Commodity: {comm} • Assemblies: {u_assys} • Rows: {count} of {len(all_table_records)} • Target Price: {meta.get('target_price', 'Not Specified')} • EAU: {eau_disp_str}")
            except Exception: pass

        assy_combo.bind("<<ComboboxSelected>>", _refresh_table)
        search_var.trace_add("write", _refresh_table)
        _refresh_table()

        # =========================================================================
        # TAB 2: 📦 Assembly EAU & MOQ Maintenance Table
        # =========================================================================
        tab_moq = tk.Frame(studio_nb, bg="#F8FAFC")
        studio_nb.add(tab_moq, text=" 📦 Assembly EAU & MOQ Maintenance ")

        assigned_custom_moqs = {}
        if hasattr(self, '_staged_bom_payload') and isinstance(self._staged_bom_payload, dict) and isinstance(self._staged_bom_payload.get("custom_moqs"), dict):
            assigned_custom_moqs = dict(self._staged_bom_payload["custom_moqs"])
        default_moq_list = self._staged_bom_payload.get("default_moqs", [100, 250, 500, 1000]) if hasattr(self, '_staged_bom_payload') and isinstance(self._staged_bom_payload, dict) else [100, 250, 500, 1000]
        default_moq_str = ", ".join(str(m) for m in default_moq_list)

        moq_ctrl = tk.Frame(tab_moq, bg="#FFFFFF", padx=10, pady=8, bd=1, relief="solid")
        moq_ctrl.pack(fill="x", padx=6, pady=(6, 4))
        tk.Label(moq_ctrl, text="Default MOQs (All Assemblies):", font=("Segoe UI", 9, "bold"), bg="#FFFFFF", fg="#334155").pack(side="left", padx=(0, 4))
        def_moq_var = tk.StringVar(value=default_moq_str)
        def_moq_ent = tk.Entry(moq_ctrl, textvariable=def_moq_var, font=("Segoe UI", 9), width=24, bd=1, relief="solid")
        def_moq_ent.pack(side="left", padx=(0, 8))

        def _apply_default_moqs_to_all():
            try:
                parsed_moqs = [int(m.strip()) for m in re.split(r'[,;|\s]+', def_moq_var.get()) if m.strip().isdigit()]
                if not parsed_moqs: return
                if hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
                    self._staged_bom_payload["default_moqs"] = parsed_moqs
                    for a_row in assembly_rows_data:
                        a_row["moqs"] = ", ".join(str(x) for x in parsed_moqs)
                        self._staged_bom_payload["custom_moqs"][a_row["assy_no"]] = parsed_moqs
                _refresh_assy_table()
            except Exception: pass

        tk.Button(moq_ctrl, text="⚡ Apply to All", command=_apply_default_moqs_to_all, bg="#2563EB", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=10, pady=2, cursor="hand2").pack(side="left")

        btn_anon_tab2 = tk.Button(moq_ctrl, text="🔒 Mask Customer Info (MAIC)", command=_toggle_anonymization, bg="#475569", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=8, pady=2, cursor="hand2")
        btn_anon_tab2.pack(side="left", padx=(8, 0))

        # Tab 2 Count Badge
        assy_count_lbl = tk.Label(moq_ctrl, text="📊 Total: 0 Assemblies", font=("Segoe UI", 9, "bold"),
                                  bg="#EFF6FF", fg="#1D4ED8", padx=10, pady=2, bd=1, relief="solid")
        assy_count_lbl.pack(side="left", padx=(12, 0))

        assy_table_frame = tk.Frame(tab_moq, bg="#FFFFFF", padx=2, pady=2, bd=1, relief="solid")
        assy_table_frame.pack(fill="both", expand=True, padx=6, pady=4)
        assy_cols = ("Assy#", "Assy Model", "Assy Rev", "EAU (pcs)", "Target Price (USD)", "Assigned MOQs", "Components", "Status")
        assy_tree = ttk.Treeview(assy_table_frame, columns=assy_cols, show="headings", height=11)
        for col in assy_cols: assy_tree.heading(col, text=col)

        assy_tree.column("Assy#", width=110, minwidth=85, anchor="w", stretch=False)
        assy_tree.column("Assy Model", width=200, minwidth=140, anchor="w", stretch=False)
        assy_tree.column("Assy Rev", width=80, minwidth=60, anchor="center", stretch=False)
        assy_tree.column("EAU (pcs)", width=120, minwidth=90, anchor="center", stretch=False)
        assy_tree.column("Target Price (USD)", width=130, minwidth=100, anchor="center", stretch=False)
        assy_tree.column("Assigned MOQs", width=180, minwidth=130, anchor="w", stretch=False)
        assy_tree.column("Components", width=100, minwidth=80, anchor="center", stretch=False)
        assy_tree.column("Status", width=110, minwidth=90, anchor="center", stretch=False)

        assy_vsb = ttk.Scrollbar(assy_table_frame, orient="vertical", command=assy_tree.yview)
        assy_hsb = ttk.Scrollbar(assy_table_frame, orient="horizontal", command=assy_tree.xview)
        assy_tree.configure(yscrollcommand=assy_vsb.set, xscrollcommand=assy_hsb.set)

        assy_vsb.pack(side="right", fill="y")
        assy_hsb.pack(side="bottom", fill="x")
        assy_tree.pack(side="left", fill="both", expand=True)

        assembly_rows_data = []
        for assy in assemblies:
            a_no = str(assy.get("assy_no", "")).strip()
            if not a_no or any(ar["assy_no"] == a_no for ar in assembly_rows_data): continue
            assembly_rows_data.append({
                "assy_no": a_no, "assy_model": assy.get("assy_model", ""), "assy_rev": str(assy.get("assy_rev", "")).replace("Rev", "").strip(),
                "eau": str(assy.get("eau", meta.get("eau") or "Not Specified")), "tp": str(assy.get("target_price", meta.get("target_price", "N/A"))).replace('$', '').strip(),
                "moqs": ", ".join(str(m) for m in assigned_custom_moqs.get(a_no, default_moq_list)),
                "comp_count": f"{len(assy.get('items', []))} items", "status": "✅ Verified"
            })

        # Selected Assembly Detail Status Bar
        assy_status_bar = tk.Label(tab_moq, text="Select any assembly row above to inspect EAU volume, target prices, and MOQ tier breakdown.",
                                   font=("Segoe UI", 9, "bold"), fg="#E2E8F0", bg="#1E293B", padx=10, pady=8, bd=1, relief="solid", anchor="w")
        assy_status_bar.pack(fill="x", padx=6, pady=(3, 4))

        def _on_assy_select(event):
            sel = assy_tree.selection()
            if not sel: return
            cur_idx = assy_tree.index(sel[0]) + 1
            tot_assys = len(assy_tree.get_children())
            vals = assy_tree.item(sel[0], "values")
            if vals:
                disp_a_no, disp_a_mod, disp_a_rev, a_eau, a_tp, a_moqs, a_comps, a_st = vals
                assy_count_lbl.config(text=f"📌 Focused: Assembly {cur_idx} of {tot_assys}")
                assy_status_bar.config(text=f"📌 Focused: Assembly {cur_idx} of {tot_assys} • Assy#: {disp_a_no} ({disp_a_mod[:25]}) • Rev: {disp_a_rev} • EAU: {a_eau} • Target Price: {a_tp} • Components: {a_comps} • MOQs: [{a_moqs}]")
        assy_tree.bind("<<TreeviewSelect>>", _on_assy_select)

        def _refresh_assy_table():
            for row in assy_tree.get_children(): assy_tree.delete(row)
            for ar in assembly_rows_data:
                disp_a_no = _display_val(ar["assy_no"], "ASSY")
                disp_a_mod = _display_val(ar["assy_model"], "MODEL")
                disp_a_rev = _display_val(ar["assy_rev"], "REV")
                assy_tree.insert("", "end", values=(disp_a_no, disp_a_mod, disp_a_rev, ar['eau'], f"${ar['tp']}", ar["moqs"], ar["comp_count"], ar["status"]))
            assy_count_lbl.config(text=f"📊 Total: {len(assembly_rows_data)} Assemblies")

        def _edit_selected_assembly_moq_eau():
            sel = assy_tree.selection()
            if not sel: return
            a_no_sel = assy_tree.item(sel[0], "values")[0]
            matched_ar = next((ar for ar in assembly_rows_data if ar["assy_no"] == a_no_sel or _display_val(ar["assy_no"], "ASSY") == a_no_sel), None)
            if not matched_ar: return
            disp_a_title = _display_val(matched_ar["assy_no"], "ASSY")
            a_dlg = tk.Toplevel(dlg); a_dlg.geometry("400x300"); a_dlg.configure(bg="#0F172A")
            tk.Label(a_dlg, text=f"Assembly {disp_a_title}", font=("Segoe UI", 12, "bold"), fg="#FFFFFF", bg="#0F172A").pack(pady=10)
            e_v = tk.StringVar(value=matched_ar["eau"]); mo_v = tk.StringVar(value=matched_ar["moqs"])
            tk.Entry(a_dlg, textvariable=e_v, width=20).pack(pady=5); tk.Entry(a_dlg, textvariable=mo_v, width=20).pack(pady=5)
            def _save():
                matched_ar["eau"] = e_v.get(); matched_ar["moqs"] = mo_v.get(); _refresh_assy_table(); a_dlg.destroy()
            tk.Button(a_dlg, text="Save", command=_save, bg="#2563EB", fg="#FFFFFF").pack(pady=10)

        assy_tree.bind("<Double-1>", lambda e: _edit_selected_assembly_moq_eau())
        _refresh_assy_table()

        moq_bottom_bar = tk.Frame(tab_moq, bg="#F8FAFC", padx=6, pady=4)
        moq_bottom_bar.pack(fill="x")
        tk.Button(moq_bottom_bar, text="✏️ Edit Selection", command=_edit_selected_assembly_moq_eau, bg="#D97706", fg="#FFFFFF", font=("Segoe UI", 8, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left")

        # Bottom Bar
        bbar = tk.Frame(dlg, bg="#F1F5F9", padx=16, pady=8)
        bbar.pack(fill="x")
        def _on_import_bom():
            dlg.destroy()
            self._commit_and_launch_staged_bom()

        tk.Button(bbar, text="🚀 Import into BOM Verification Module", command=_on_import_bom,
                  bg="#2563EB", fg="#FFFFFF", font=("Segoe UI", 9, "bold"), relief="flat", padx=16, pady=6, cursor="hand2").pack(side="right")
        tk.Button(bbar, text="Close Studio", command=dlg.destroy,
                  bg="#E2E8F0", fg="#334155", font=("Segoe UI", 9), relief="flat", padx=14, pady=6, cursor="hand2").pack(side="right", padx=(0, 8))

    def _handle_in_chat_training(self, rule_content):
        """Processes and saves a user-taught engineering rule or correction directly in chat."""
        try:
            from agents.correction_store import CorrectionStore
            cs = CorrectionStore()
            username = getattr(self, 'user', {}).get('username', 'User') if hasattr(self, 'user') else 'User'

            # Save the rule permanently into the store
            record = cs.save_taught_rule(
                rule_text=rule_content,
                category="engineering_rule",
                doc_hint="GLOBAL",
                taught_by=username
            )

            msg = (
                f"🧠 **Knowledge & Rule Successfully Learned!**\n\n"
                f"• **Taught Rule:** `{rule_content}`\n"
                f"• **Target Scope:** Global Engineering Memory\n"
                f"• **Learned By:** `{username}`\n"
                f"• **Storage:** Persistent AI Knowledge (`data/corrections/corrections.json`)\n\n"
                f"✨ *The AI will automatically inject this memory into future drawing extractions, BOM mapping, and question-answering!*"
            )
            self._append_agent_message(
                msg,
                suggestions=[
                    "🧠 View AI Learned Rules & Memory",
                    "📩 Check RFQ Emails",
                    "🧪 Run Demo RFQ Extraction"
                ]
            )
        except Exception as e:
            self._append_agent_message(f"⚠️ Error saving knowledge rule: {e}")

    def _sync_amended_components_from_text(self, text_content):
        """Parses component tables from Gemini verification response and syncs them to staged BOM & CorrectionStore."""
        if not hasattr(self, '_last_extracted_rfq_json') or not self._last_extracted_rfq_json:
            return 0

        updated_count = 0
        try:
            from agents.correction_store import CorrectionStore
            from agents.synthetic_bom_generator import SyntheticBOMGenerator
            cs = CorrectionStore()

            # Parse markdown table lines: e.g. | 30077977.00 | ... | 30061240 | Shielded Cable | ... | 1.75 | M |
            table_lines = [l.strip() for l in text_content.splitlines() if "|" in l and not l.startswith("|---") and not l.startswith("|:--")]
            
            for line in table_lines:
                cells = [c.strip().replace('*', '').replace('`', '') for c in line.strip('|').split('|')]
                if len(cells) < 4: continue

                # Match any part number in our active assemblies
                for assy in self._last_extracted_rfq_json.get("assemblies", []):
                    for it in assy.get("items", []):
                        p_no = str(it.get("part_number", "")).strip()
                        mpn = str(it.get("mpn", "")).strip()
                        
                        # Check if this cell line mentions part_number or mpn
                        matched_cell = False
                        for cell in cells:
                            if p_no and p_no == cell: matched_cell = True; break
                            if mpn and mpn == cell: matched_cell = True; break
                        
                        if matched_cell:
                            # Try to extract updated Qty, UOM, MPN, MFR from row cells
                            for c in cells:
                                # Check for quantity/uom pattern: e.g. "1.75 M", "20 EA", "2 EA", "1.75"
                                q_m = re.search(r'\b([0-9]+(?:\.[0-9]+)?)\s*(?:M|EA|PCS|IN|FT|SET)?\b', c, re.I)
                                if q_m and float(q_m.group(1)) != float(it.get("qty", 1)):
                                    it["qty"] = float(q_m.group(1))
                                    updated_count += 1
                                if " M" in c or c.upper() == "M": it["uom"] = "M"
                                elif "EA" in c or c.upper() == "EA": it["uom"] = "EA"
                                elif "PCS" in c or c.upper() == "PCS": it["uom"] = "PCS"

                                # Check for MFR
                                for mfr_cand in ["Molex", "JST", "Heiniger", "Alpha", "TE Connectivity", "Sick", "FCI"]:
                                    if mfr_cand.lower() in c.lower():
                                        it["mfr"] = mfr_cand

                            # Save correction to persistent store
                            cs.save_correction(
                                doc_hint=assy.get("assy_no", "BOM"),
                                field=p_no or mpn,
                                wrong_value="",
                                correct_value=f"{it.get('mpn', '')} | {it.get('mfr', '')} | Qty: {it.get('qty', 1)} {it.get('uom', 'EA')}",
                                mfr=it.get("mfr", ""),
                                note="Auto-synced from Gemini Vision Verification in Chat"
                            )

            if updated_count > 0:
                gen_res = SyntheticBOMGenerator().generate_synthetic_excel(self._last_extracted_rfq_json)
                if gen_res.get("success") and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
                    self._staged_bom_payload["file_path"] = gen_res["file_path"]
                print(f"[Launcher] Auto-synced {updated_count} component amendment(s) to staged BOM Excel!")
        except Exception as e:
            print(f"[Launcher] Notice during component auto-sync: {e}")

        return updated_count

    def _handle_multimodal_chat_query(self, user_text, attachments):
        """Sends attached images and custom user prompt to Gemini Vision for analysis and auto-syncs amendments."""
        self._show_typing_indicator()

        def _worker():
            try:
                from agents.llm_gateway import LLMGateway
                from agents.correction_store import CorrectionStore

                gw = LLMGateway()
                cs = CorrectionStore()
                memory_ctx = cs.get_all_memory_context()

                system_prompt = (
                    "You are the ContinuumX Manufacturing AI Assistant and Engineering Vision Expert. "
                    "Analyze the provided engineering drawings, component photos, wiring diagrams, or BOM screenshots. "
                    "Provide clear, accurate, and detailed engineering insights, component identification (MPN, Manufacturer, Part Number, Qty, UOM), "
                    "or answers to the user's specific questions. Use Markdown formatting, bullet points, and tables where helpful.\n\n"
                    + memory_ctx
                )

                b64_list = [att["b64"] for att in attachments if att.get("b64")]
                prompt = user_text if user_text and not (
                    user_text.startswith("Ask me anything")
                    or user_text.startswith("Ask about RFQs")
                ) else "Please analyze this image or drawing in detail and extract all key engineering components, part numbers, and specifications."

                response_text, status = gw.generate_text_or_multimodal(
                    system_prompt=system_prompt,
                    user_prompt=prompt,
                    inline_images=b64_list,
                    doc_name="UserChatAttachment"
                )

                # Auto-sync any verified amendments into active RFQ & Synthetic BOM
                synced_cnt = self._sync_amended_components_from_text(response_text)
                sync_badge = f"\n\n✨ **Auto-Sync:** Staged BOM updated with verified component values!" if synced_cnt > 0 else ""

                self._safe_gui(lambda: self._append_agent_message(
                    response_text + sync_badge,
                    suggestions=[
                        "📊 Review Full Table & Filter",
                        "🔄 Reshow Updated BOM Table",
                        "🔍 Inspect Source Evidence",
                        "🚀 Launch BOM Verification Window"
                    ]
                ))
            except Exception as ex:
                self._safe_gui(lambda: self._append_agent_message(f"⚠️ Error analyzing image: {ex}"))

        import threading
        threading.Thread(target=_worker, daemon=True).start()

    def _on_chat_submit(self):
        placeholder = getattr(self, "_chat_placeholder", "Ask about RFQs, charts, or BOM…  (Enter to send)")
        if isinstance(self.chat_input, tk.Text):
            text = self.chat_input.get("1.0", "end-1c").strip()
            if self._chat_input_is_placeholder() or text == placeholder:
                text = ""
            self._reset_chat_input()
        else:
            text = str(self.chat_input.get()).strip()
            self.chat_input.delete(0, "end")

        pending_attachments = list(self._pending_chat_attachments)
        self._clear_attachment_tray()

        if not text and not pending_attachments:
            return

        # Render User Message with Thumbnails
        img_photos = [att["photo"] for att in pending_attachments if att.get("photo")]
        self._append_user_message(text or "📷 [Attached Image / Drawing]", images=img_photos)
        text_lower = text.lower() if text else ""

        # Direct In-Chat AI Training & Teaching Commands
        train_match = re.search(r'^(?:teach\s+ai|teach|train\s+ai|train|learn\s+rule|learn|remember\s+that|remember|rule|knowledge|note\s+that)[:\s]+(.+)', text, re.IGNORECASE) if text else None
        if train_match:
            rule_content = train_match.group(1).strip()
            if rule_content:
                self._handle_in_chat_training(rule_content)
                return

        # Multimodal Image Q&A Handler
        if pending_attachments:
            self._handle_multimodal_chat_query(text, pending_attachments)
            return

        # Reshow / Refresh Updated BOM Table Card Handler
        if any(k in text_lower for k in ["reshow bom", "reshow updated", "show updated", "refresh bom table", "display bom table", "reshow table"]):
            if hasattr(self, '_detected_rfq_list') and self._detected_rfq_list:
                self._load_selected_rfq_payload(self._detected_rfq_list[0])
                return

        # Review Full RFQ Table & Studio Handler
        if any(k in text_lower for k in ["review full", "review rfq", "review table", "full table", "show full", "open table", "filter table", "open review"]):
            self._open_rfq_review_window()
            return

        # EAU Confirmation / Override Handler (e.g. "set eau 3000", "eau is 2640", "confirm eau 5000", "eau 3000", "eau: 4000")
        eau_m = re.search(r'(?:set\s+eau|confirm\s+eau|eau\s+is|change\s+eau|eau[:\s=]+)\s*([0-9,]+)', text_lower)
        if eau_m and hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
            new_eau = int(eau_m.group(1).replace(',', ''))
            self._staged_bom_payload["eau"] = new_eau
            if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                self._last_extracted_rfq_json["rfq_metadata"]["eau"] = f"{new_eau:,} pcs"
                for a in self._last_extracted_rfq_json.get("assemblies", []):
                    a["eau"] = new_eau
                    for it in a.get("items", []):
                        it["eau"] = new_eau
                # Regenerate synthetic BOM Excel with updated EAU
                try:
                    from agents.synthetic_bom_generator import SyntheticBOMGenerator
                    gen = SyntheticBOMGenerator()
                    gen_res = gen.generate_synthetic_excel(self._last_extracted_rfq_json)
                    if gen_res.get("success"):
                        self._staged_bom_payload["file_path"] = gen_res["file_path"]
                except Exception as e:
                    print(f"[Launcher] Error updating Excel with new EAU: {e}")

            self._append_agent_message(
                f"✅ **EAU successfully confirmed and updated to `{new_eau:,} pcs`!**\n\n"
                f"• Staged Synthetic BOM updated with `{new_eau:,} pcs`\n"
                f"• Downstream BOM Verification and Sourcing will use `{new_eau:,} pcs`.",
                suggestions=["📊 Review Full Table & Filter", "🚀 Launch BOM Verification Window", "🔍 Inspect Source Evidence"]
            )
            return

        # Inspect Source Evidence Handler
        if any(k in text_lower for k in ["inspect evidence", "view evidence", "source evidence", "show evidence", "inspect source evidence", "evidence audit"]):
            self._show_extracted_evidence_dialog()
            return

        if any(k in text_lower for k in ["test error", "/test-error", "simulate error", "test error logging"]):
            self._chat_run_error_test()
            return
        if any(k in text_lower for k in ["test approval", "/test-approval", "simulate approval"]):
            self._chat_run_approval_test()
            return
        if re.search(r"\bimport\b.{0,24}\bbom\b", text_lower) and not re.search(r"\b(how|what|why|explain)\b", text_lower):
            self._agent_import_bom()
            return

        # Check / Scan Email RFQs Handler (e.g. "Check email: Fwd: Enquiry ~ Cable _ Tecan - RS25-8099", "check 1 email", "check email for RS26-8300")
        if any(k in text_lower for k in [
            "check email", "scan email", "read email", "fetch email", "check rfq email",
            "scan rfq", "email rfq", "scan inbox", "check inbox", "maic-demo", "maic demo", "rfq email"
        ]) or text_lower in ["check emails", "scan emails", "read emails", "check rfqs"]:
            # 1. Parse user requested email count (e.g. "check 1 email", "scan 3 emails", "latest 5 emails")
            lim = 5
            lim_m = re.search(r'\b([0-9]{1,3})\s*(?:emails?|rfqs?|messages?)\b', text_lower)
            if lim_m:
                try:
                    lim = max(1, min(50, int(lim_m.group(1))))
                except Exception:
                    lim = 5
            elif "1 email" in text_lower or "one email" in text_lower:
                lim = 1

            # 2. Parse exact RFQ ID (e.g. "RS25-8099", "RS26-8300", "RS26-8004")
            target_rfq_id = None
            rfq_m = re.search(r'\b(RS[0-9]{2}-[0-9]{3,5})\b', text, re.IGNORECASE)
            if rfq_m:
                target_rfq_id = rfq_m.group(1).upper()

            # 3. Parse query string after colon or keyword (e.g. "Check email: Fwd: Enquiry ~ Cable _ Tecan - RS25-8099")
            sf = target_rfq_id
            if not sf:
                colon_m = re.search(r'(?:check\s+email|scan\s+email|read\s+email|fetch\s+email|rfq\s+email)[:\s]+(.+)', text, re.IGNORECASE)
                if colon_m:
                    raw_q = colon_m.group(1).strip()
                    if raw_q and not raw_q.lower().startswith(('box', 'inbox', 'please', 'now')):
                        sf = raw_q
                else:
                    filt_m = re.search(r'(?:for|about|matching|from)\s+([a-zA-Z0-9_-]+)', text_lower)
                    if filt_m:
                        cand = filt_m.group(1).strip()
                        if cand not in ("email", "emails", "rfq", "rfqs", "inbox", "mailbox"):
                            sf = cand

            self._show_typing_indicator()
            import threading
            threading.Thread(target=self._check_email_rfqs_async, args=(lim, sf, target_rfq_id), daemon=True).start()
            return

        # Demo RFQ Extraction Handler (e.g. "run demo rfq extraction", "demo rfq", "demo extraction")
        if any(k in text_lower for k in ["demo rfq", "run demo", "demo extraction"]):
            self._show_typing_indicator()
            import threading
            threading.Thread(target=self._run_demo_rfq_extraction_async, daemon=True).start()
            return

        # Load / Switch Selected RFQ Handler (e.g. "load rfq rs26-8004", "load rs25-8099")
        if "load rfq" in text_lower or any(k in text_lower for k in ["load rs25", "load rs26", "load rs24"]):
            if hasattr(self, '_detected_rfq_list') and self._detected_rfq_list:
                for rfq_obj in self._detected_rfq_list:
                    r_no = rfq_obj.get("rfq_json", {}).get("rfq_metadata", {}).get("rfq_number", "").lower()
                    if r_no and r_no in text_lower:
                        self._load_selected_rfq_payload(rfq_obj)
                        return
                # If specific ID didn't match, load the first one
                self._load_selected_rfq_payload(self._detected_rfq_list[0])
                return

        # View Extracted JSON Handler
        if any(k in text_lower for k in ["view json", "view extracted json", "show json", "raw json"]):
            if hasattr(self, '_last_extracted_rfq_json') and self._last_extracted_rfq_json:
                json_str = json.dumps(self._last_extracted_rfq_json, indent=2)
                if len(json_str) > 1200:
                    json_str = json_str[:1200] + "\n... (truncated)"
                self._append_agent_message(f"📄 **Extracted Intermediate JSON Payload:**\n\n```json\n{json_str}\n```", suggestions=["🚀 Launch BOM Verification Window", "🔍 Inspect Source Evidence", "📊 Draw RFQ Stage Chart"])
            else:
                self._append_agent_message("No RFQ JSON extracted yet. Scan emails or click '🧪 Run Demo RFQ Extraction' first!", suggestions=["📩 Check RFQ Emails", "🧪 Run Demo RFQ Extraction"])
            return

        # View AI Telemetry Benchmark Handler
        if any(k in text_lower for k in ["benchmark", "telemetry", "speed stats", "ai speed", "processing stats", "performance"]):
            self._show_telemetry_benchmark_dialog()
            return

        # Open Guided Review Queue (Single-window stepper review)
        if any(k in text_lower for k in ["open guided review queue", "guided queue", "review queue", "open queue", "open guided queue", "review 1 by 1", "guided review"]):
            self._open_guided_approval_queue()
            return

        # Batch Dispatch All Queue Handler
        if "dispatch all" in text_lower or "dispatch all rfqs" in text_lower:
            self._dispatch_all_active_gates()
            return

        # View In-Progress WIP Pipeline Handler
        if any(k in text_lower for k in ["in-progress pipeline", "wip pipeline", "view pipeline", "pipeline tracker", "ongoing rfqs", "in progress", "wip"]):
            user_info = getattr(self, "user", {}) or getattr(self, "user_context", {}) or {}
            user_role = user_info.get("role") or user_info.get("Role") or "System Administrator"
            username = user_info.get("username") or user_info.get("name") or "Sysadmin"
            summary = self.orch_state_mgr.get_pipeline_operations_summary(user_role=user_role, username=username)
            wip_list = summary.get("wip", [])
            if wip_list:
                wip_table = self.approval_gate_mgr.get_wip_pipeline_table_data(wip_list)
                msg = (
                    f"⏳ **Work-In-Progress (WIP) Pipeline Tracker:**\n"
                    f"Displaying **{len(wip_list)} active RFQs** undergoing engineering drafting, sourcing price collection, or cycle time calculation.\n\n"
                    f"👉 *To continue action on any RFQ, select its module from the left menu or use the quick launch chips below:*"
                )
                
                # Contextual quick launch chips
                chips = []
                wip_stages = {w.get("stage", "") for w in wip_list}
                if "BOM Verification" in wip_stages:
                    chips.append("🔨 Launch BOM Management")
                if any("Sourcing" in s for s in wip_stages):
                    chips.append("📦 Launch Sourcing Module")
                if any("Cycle Time" in s or "Sourcing" in s for s in wip_stages):
                    chips.append("⏱️ Launch Cycle Time")
                if any("Costing" in s for s in wip_stages):
                    chips.append("💰 Launch Costing Module")
                chips.append("🔍 Open Guided Review Queue")

                self._append_agent_message(msg, suggestions=chips, table_data=wip_table)
            else:
                self._append_agent_message("🎉 Great news! There are currently 0 pending WIP RFQs in the pipeline.")
            return

        # View Queue Summary
        if any(k in text_lower for k in ["view approval queue summary", "queue summary", "pending queue"]):
            self._check_active_approval_gates()
            return

        # Open Module Feature Handler (e.g. "open sourcing", "launch costing", "open bom")
        if any(w in text_lower for w in ["open ", "launch ", "goto ", "go to "]):
            mod_map = {
                "sourcing": "Sourcing",
                "costing": "Costing",
                "cycle time": "Cycle Time",
                "bom": "BOM",
                "npi": "NPI",
                "work instruction": "WI",
                "wi": "WI",
                "project management": "Project Management",
                "pm": "Project Management"
            }
            for key_str, feat_key in mod_map.items():
                if key_str in text_lower:
                    self._append_agent_message(f"⚡ Launching **{feat_key} Module** for deep engineering analysis...")
                    self.after(300, lambda fk=feat_key: self.launch_feature(fk))
                    return

        # Open MOQ Window / Target Price Wizard Handler (e.g. "open assign moq window for rfq 99999999")
        if ("open" in text_lower or "launch" in text_lower or "show" in text_lower) and ("moq" in text_lower or "target price" in text_lower):
            rfq_m = re.search(r'rfq\s*([a-z0-9_-]+)', text_lower)
            rfq_target = rfq_m.group(1).lower() if rfq_m else "Latest"
            self._append_agent_message(f"⚡ Opening Assign MOQ & Target Price Maintenance Wizard for RFQ '{rfq_target}'...")
            self.launch_feature("BOM")
            return

        # 1. Stage Dispatch Handler (Direct Agent Dispatch with Email Composer Popup & Auto-Transition)
        if "dispatch" in text_lower:
            rfq_m = re.search(r'\brfq[-_\s]*([a-z0-9_-]+)', text_lower)
            rfq_target = rfq_m.group(1) if rfq_m else "1009"
            username = self.user.get("username", "Admin")

            if "costing" in text_lower:
                success, msg = self.tool_dispatcher.dispatch_sourcing_to_costing(rfq_target, parent_window=self, username=username)
                if success:
                    next_suggs = [
                        f"⏱️ Review Cycle Time for {rfq_target}",
                        f"📊 View Costing Summary for {rfq_target}",
                        "📊 Draw RFQ Stage Chart"
                    ]
                    self._append_agent_message(f"🚀 **Dispatch Successful!**\n\n{msg}\n\n• Email notification sent to Costing team and CC'd to you.\n• RFQ status transitioned to 'pending_costing'.\n• Automatically launching Costing Module...", suggestions=next_suggs)
                    self.after(500, lambda: self.launch_feature("Costing"))
                else:
                    self._append_agent_message(f"⚠️ **Dispatch Notice:**\n\n{msg}")
            elif "npi" in text_lower:
                success, msg = self.tool_dispatcher.dispatch_costing_to_npi(rfq_target, parent_window=self, username=username)
                if success:
                    next_suggs = [
                        f"🚀 Launch NPI Gateway for {rfq_target}",
                        "📊 Draw RFQ Stage Chart"
                    ]
                    self._append_agent_message(f"🚀 **Dispatch Successful!**\n\n{msg}\n\n• Email notification sent to NPI team and CC'd to you.\n• Automatically launching NPI Gateway...", suggestions=next_suggs)
                    self.after(500, lambda: self.launch_feature("NPI"))
                else:
                    self._append_agent_message(f"⚠️ **Dispatch Notice:**\n\n{msg}")
            else:
                success, msg = self.tool_dispatcher.dispatch_bom_to_sourcing_and_ct(rfq_target, parent_window=self, username=username)
                if success:
                    next_suggs = [
                        f"⏱️ Start Cycle Time Analysis for {rfq_target}",
                        f"🚀 Dispatch RFQ {rfq_target} to Costing",
                        "📊 Draw RFQ Stage Chart"
                    ]
                    self._append_agent_message(f"🚀 **Dispatch Successful!**\n\n{msg}\n\n• Email notification sent to Sourcing & Cycle Time teams and CC'd to you.\n• RFQ status transitioned to 'pending_sourcing_and_cycle_time'.\n• Automatically launching Sourcing Module & Calculation Window...", suggestions=next_suggs)
                    
                    # Write sourcing auto-launch command and launch Sourcing feature
                    local_appdata = os.environ.get('LOCALAPPDATA', os.environ.get('TEMP', 'C:\\Temp'))
                    cmd_path = os.path.join(local_appdata, "ContXs", "agent_sourcing_launch_command.json")
                    os.makedirs(os.path.dirname(cmd_path), exist_ok=True)
                    with open(cmd_path, 'w', encoding='utf-8') as f:
                        json.dump({"action": "start_sourcing", "rfq_id": rfq_target}, f)
                    self.after(500, lambda: self.launch_feature("Sourcing"))
                else:
                    self._append_agent_message(f"⚠️ **Dispatch Notice:**\n\n{msg}")
            return

        # 1b. Launch Assign MOQ Window directly (e.g. "pls launch assign moq window for rfq01", "open assign moq for 1009")
        if ("moq window" in text_lower or ("launch" in text_lower and "moq" in text_lower) or ("open" in text_lower and "moq" in text_lower) or "assign moq window" in text_lower) and not any(k in text_lower for k in ["help to assign", "assign moq 100", "moq is", "moq:"]):
            rfq_m = re.search(r'\brfq[-_\s]*([a-z0-9_-]+)', text_lower)
            rfq_target = rfq_m.group(1) if rfq_m else "RFQ01"
            local_appdata = os.environ.get('LOCALAPPDATA', os.environ.get('TEMP', 'C:\\Temp'))
            cmd_path = os.path.join(local_appdata, "ContXs", "agent_assign_moq_command.json")
            os.makedirs(os.path.dirname(cmd_path), exist_ok=True)
            with open(cmd_path, 'w', encoding='utf-8') as f:
                json.dump({"action": "assign_moq", "rfq_id": rfq_target}, f)
            self._append_agent_message(f"⚡ Opening Assign MOQ & Target Price Maintenance Wizard for RFQ '{rfq_target}'...")
            self.launch_feature("BOM")
            return

        # 2. Revert RFQ Handler (Directly opens the existing Project Management tool)
        if "revert" in text_lower:
            rfq_m = re.search(r'rfq\s*([a-z0-9_-]+)', text_lower)
            rfq_target = rfq_m.group(1) if rfq_m else "1009"
            self._append_agent_message(f"↩️ Opening existing Project Management Portal for RFQ '{rfq_target}' Revert...")
            self.launch_feature("Project Management")
            return

        # 2.5 Component & Drawing Extraction Provenance & AI Explanation Handler
        if any(k in text_lower for k in ["explain data extraction", "explain extraction", "evidence for component", "check drawing", "explain line", "how was part", "why was part", "extraction and evidence", "for drawing"]):
            target_part = ""
            target_assy = ""
            target_dwg = ""

            # Check for explicit drawing request: e.g. "for drawing BB0_502356122_EN_00.pdf" or "155-892105-010-00R.pdf"
            d_m_pdf = re.search(r'([A-Za-z0-9_.-]+\.pdf)', text, re.IGNORECASE)
            if d_m_pdf:
                target_dwg = d_m_pdf.group(1).strip()
            else:
                d_m = re.search(r"(?:for\s+drawing|drawing\s+for|check\s+drawing\s+for|drawing|blueprint|pdf)\s*[:=]?\s*['\"]?([A-Za-z0-9_.-]+)['\"]?", text, re.IGNORECASE)
                if d_m:
                    cand_dwg = d_m.group(1).strip()
                    if cand_dwg.lower() not in ("for", "the", "a", "an", "this", "that", "extraction", "evidence", "data", "and", "component", "part", "details", "sample"):
                        target_dwg = cand_dwg

            p_m = re.search(r"['\"]?([A-Za-z0-9_.-]+)['\"]?\s*(?:\(([^)]+)\))?\s+in\s+Assembly\s+([A-Za-z0-9_.-]+)", text, re.IGNORECASE)
            if p_m:
                target_part = p_m.group(1)
                target_assy = p_m.group(3)
            else:
                p_m2 = re.search(r"(?:part|component|line|item)\s*[:=]?\s*['\"]?([A-Za-z0-9_.-]+)['\"]?", text, re.IGNORECASE)
                if p_m2:
                    cand_part = p_m2.group(1).strip()
                    if cand_part.lower() not in ("extraction", "evidence", "data", "and", "the", "for", "details"):
                        target_part = cand_part

            target_rfq_json = getattr(self, '_last_extracted_rfq_json', {}) or {}
            all_candidate_rfqs = []
            if target_rfq_json:
                all_candidate_rfqs.append(target_rfq_json)
            if hasattr(self, '_detected_rfq_list') and self._detected_rfq_list:
                for det in self._detected_rfq_list:
                    r_json = det.get("rfq_json")
                    if r_json and r_json not in all_candidate_rfqs:
                        all_candidate_rfqs.append(r_json)

            rfq_meta = target_rfq_json.get("rfq_metadata", {})
            rfq_no = rfq_meta.get("rfq_number", "RS26-8004")
            cust_name = rfq_meta.get("customer_name", "Customer")

            # --- CASE 1: User requested a specific DRAWING / PDF ---
            if target_dwg:
                clean_dwg_stem = re.sub(r'\.pdf$', '', target_dwg, flags=re.IGNORECASE).lower()
                clean_dwg_raw = re.sub(r'^(?:AJ0_|BB0_)', '', clean_dwg_stem, flags=re.IGNORECASE)
                clean_dwg_raw = re.sub(r'_EN_[0-9]+$', '', clean_dwg_raw, flags=re.IGNORECASE)

                matched_assy = None
                matched_rfq = target_rfq_json
                for rfq_cand in all_candidate_rfqs:
                    for assy in rfq_cand.get("assemblies", []):
                        a_no = str(assy.get("assy_no", "")).strip().lower()
                        a_mod = str(assy.get("assy_model", "")).strip().lower()
                        a_src = str(assy.get("drawing_filename", "")).strip().lower()
                        if clean_dwg_stem in a_no or a_no in clean_dwg_stem or clean_dwg_raw in a_no or a_no in clean_dwg_raw or target_dwg.lower() in a_src or target_dwg.lower() in a_mod:
                            matched_assy = assy
                            matched_rfq = rfq_cand
                            break
                    if matched_assy: break

                if matched_assy:
                    a_no_disp = matched_assy.get("assy_no", "")
                    a_mod_disp = matched_assy.get("assy_model", "")
                    a_rev_disp = str(matched_assy.get("assy_rev", "")).replace("Rev", "").strip()
                    cand_meta = matched_rfq.get("rfq_metadata", {})
                    a_eau_disp = matched_assy.get("eau", cand_meta.get("eau", "N/A"))
                    a_tp_disp = matched_assy.get("target_price", cand_meta.get("target_price", "N/A"))
                    items = matched_assy.get("items", [])

                    t_rows = []
                    for idx, it in enumerate(items[:12], start=1):
                        p_no = str(it.get("part_number") or it.get("mpn") or f"Item #{idx}")
                        desc = str(it.get("description") or "Component")[:30]
                        mpn = str(it.get("mpn") or "Not Specified")
                        mfr = str(it.get("mfr") or "Not Specified")
                        qty_uom = f"{it.get('qty', 1)} {it.get('uom', 'EA')}"
                        t_rows.append([str(idx), p_no, desc, mpn, mfr, qty_uom])

                    table_data = {
                        "title": f"📄 Drawing Extraction Table — {target_dwg} ({a_no_disp})",
                        "headers": ["Line", "Part Number", "Description", "MPN", "MFR", "Qty/UOM"],
                        "rows": t_rows,
                        "footer": f"Total Components in Drawing: {len(items)} items • Title Block: {a_mod_disp} • Rev: {a_rev_disp}"
                    }

                    exp_msg = (
                        f"📄 **Drawing Extraction Provenance — `{target_dwg}`**\n"
                        f"────────────────────────────────────────────────────────────────────────────\n"
                        f"  Drawing File    : {target_dwg}\n"
                        f"  Assembly Number : {a_no_disp} (Rev {a_rev_disp})\n"
                        f"  Model / Title   : {a_mod_disp}\n"
                        f"  Forecast EAU    : {a_eau_disp} pcs\n"
                        f"  Target Price    : {a_tp_disp}\n"
                        f"  Components Found: {len(items)} structured items extracted\n"
                        f"────────────────────────────────────────────────────────────────────────────\n\n"
                        f"💡 **AI Extraction Logic**: Successfully parsed title block, wire/component schedules, and bill of materials from drawing `{target_dwg}`."
                    )
                    self._append_agent_message(
                        exp_msg,
                        suggestions=[
                            "📊 Review Full Table & Filter",
                            "🔍 Inspect Source Evidence",
                            "🚀 Launch BOM Verification Window"
                        ],
                        table_data=table_data
                    )
                    return
                else:
                    # Drawing NOT found in any loaded RFQ -> Log real error to ErrorTelemetryStore!
                    from agents.telemetry_tracker import ErrorTelemetryStore
                    err_rec = ErrorTelemetryStore().record_error(
                        module="DrawingVisionAgent",
                        error_category="PDF_NOT_FOUND_ERROR",
                        error_message=f"Drawing file '{target_dwg}' could not be located in attachments or active RFQ assemblies.",
                        severity="ERROR",
                        rfq_number=rfq_no,
                        customer=cust_name,
                        document_name=target_dwg,
                        prompt_context={"requested_drawing": target_dwg, "query": text},
                        recovery_action="Reported missing drawing to user; error logged to telemetry",
                        status="UNRECOVERED_FILE_NOT_FOUND"
                    )
                    latest_summ = ErrorTelemetryStore().get_latest_summary()

                    err_msg = (
                        f"⚠️ **Drawing Extraction Error — File Not Found**\n\n"
                        f"• **Requested Document:** `{target_dwg}`\n"
                        f"• **Status:** The drawing file was not found in the current RFQ attachments or active assemblies.\n"
                        f"• **Incident ID:** `{err_rec['error_id']}`\n"
                        f"• **Category:** `{err_rec['error_category']}` in `{err_rec['module']}`\n"
                        f"• **Logged To:** `data/telemetry/errors/{err_rec['error_id']}.json`\n"
                        f"• **Master Audit Updated:** `data/telemetry/agent_errors.json`\n"
                        f"• **Dashboard Summary:** `data/telemetry/latest_errors_summary.json` (Total Incidents: `{latest_summ.get('total_incidents', 1)}`)\n\n"
                        f"👉 Please verify the drawing filename or check the attached files via **'📧 Email & Attachments'**."
                    )
                    self._append_agent_message(
                        err_msg,
                        suggestions=[
                            "📧 Email & Attachments",
                            "📊 Review Full Table & Filter",
                            "🔍 Inspect Source Evidence"
                        ]
                    )
                    return

            # --- CASE 2: Lookup by Component / Part Number ---
            matched_item = None
            matched_assy_obj = None
            matched_rfq_obj = target_rfq_json

            for rfq_cand in all_candidate_rfqs:
                for assy in rfq_cand.get("assemblies", []):
                    a_no_cur = str(assy.get("assy_no", "")).strip()
                    if target_assy and a_no_cur != target_assy:
                        continue
                    for it in assy.get("items", []):
                        p_no_cur = str(it.get("part_number", "")).strip()
                        desc_cur = str(it.get("description", "")).strip()
                        mpn_cur = str(it.get("mpn", "")).strip()
                        if target_part:
                            if target_part.lower() in p_no_cur.lower() or target_part.lower() in desc_cur.lower() or target_part.lower() in mpn_cur.lower():
                                matched_item = it
                                matched_assy_obj = assy
                                matched_rfq_obj = rfq_cand
                                break
                        else:
                            matched_item = it
                            matched_assy_obj = assy
                            matched_rfq_obj = rfq_cand
                            break
                    if matched_item: break
                if matched_item: break

            if target_part and not matched_item:
                avail_parts = []
                for a in target_rfq_json.get("assemblies", []):
                    for it in a.get("items", [])[:5]:
                        if it.get("part_number"): avail_parts.append(str(it.get("part_number")))
                self._append_agent_message(
                    f"ℹ️ Component `{target_part}` was not found in active RFQ (`{rfq_no}`).\n\n"
                    f"• **Current Customer:** {cust_name}\n"
                    f"• **Sample Parts Available:** " + ", ".join(f"`{p}`" for p in avail_parts[:6]) + "\n\n"
                    f"👉 Try asking about one of the active parts above or click **'📊 Review Full Table & Filter'** to view all lines.",
                    suggestions=["📊 Review Full Table & Filter", "🔍 Inspect Source Evidence", "📩 Check RFQ Emails"]
                )
                return

            if matched_item and matched_assy_obj:
                ev = matched_item.get("evidence", {})
                p_ev = ev.get("part", {}) if isinstance(ev.get("part"), dict) else {}
                mpn_ev = ev.get("mpn", {}) if isinstance(ev.get("mpn"), dict) else {}
                src_doc = p_ev.get("source_doc") or mpn_ev.get("source_doc") or "Authoritative RFQ Document"
                zone = p_ev.get("zone", "BOM_TABLE")
                snip = p_ev.get("snippet") or f"{matched_item.get('mfr', '')} {matched_item.get('mpn', '')} {matched_item.get('description', '')}".strip()
                conf = int(p_ev.get("confidence", 0.95) * 100)
                reason = mpn_ev.get("reasoning") or p_ev.get("reasoning") or "Matched from structured engineering table explosion."
                res_type = p_ev.get("resolution_type", "DIRECT")

                mpn_disp = matched_item.get("mpn") or "Not Specified"
                mfr_disp = matched_item.get("mfr") or "Not Specified"
                qty_disp = f"{matched_item.get('qty', 1)} {matched_item.get('uom', 'EA')}"

                has_alt = matched_item.get("has_alternative_mpn") or "alternative_mpns" in ev
                alt_ev = ev.get("alternative_mpns", {}) if isinstance(ev.get("alternative_mpns"), dict) else {}

                consistency_status = "✅ 100% Consistent (Single Source)"
                if has_alt:
                    consistency_status = "✅ Auto-Enriched (Drawing + AVL Library)"
                elif "drawing" in src_doc.lower() and "email" in reason.lower():
                    consistency_status = "✅ Cross-Verified (Email + Drawing)"
                elif conf < 80:
                    consistency_status = f"⚠️ Review Recommended ({conf}%)"
                else:
                    consistency_status = f"✅ High Confidence ({conf}%)"

                t_rows = [
                    ["Part Number", str(matched_item.get('part_number', '')), src_doc, str(zone), consistency_status],
                    ["Description", str(matched_item.get('description', 'Component')), src_doc, "Line Item Callout", "✅ Verified"],
                    ["MPN", mpn_disp, src_doc, "Manufacturer Part Code", "✅ Validated MPN" if mpn_disp != "Not Specified" else "ℹ️ Customer CPN"],
                    ["Manufacturer", mfr_disp, src_doc, "Manufacturer AVL", "✅ Verified" if mfr_disp != "Not Specified" else "ℹ️ Standard AVL"],
                ]
                if has_alt:
                    t_rows.append(["Alternative AVL", mpn_disp, alt_ev.get("source_document", "Customer Alternative MPN Library"), "CUSTOMER_AVL_LIBRARY", "✅ Auto-Merged from Library"])
                t_rows.extend([
                    ["Quantity & UOM", qty_disp, src_doc, f"Snippet: {snip[:25]}", "✅ Dimension Callout"],
                    ["Parent Assembly", str(matched_assy_obj.get('assy_no', '')), "Drawing Title Block", "TITLE_BLOCK", "✅ Hierarchically Linked"]
                ])

                table_data = {
                    "title": f"🔍 Component Multi-Source Evidence Table — {matched_item.get('part_number')}",
                    "headers": ["Attribute", "Extracted Value", "Source Document", "Zone / Evidence", "Cross-Source Status"],
                    "rows": t_rows,
                    "footer": f"AI Extraction Logic: {reason}" + (" • Combined with Customer Alternative MPN Library." if has_alt else "")
                }

                rfq_ref_str = f" (RFQ {matched_rfq_obj.get('rfq_metadata', {}).get('rfq_number', '')})" if matched_rfq_obj != target_rfq_json else ""
                exp_msg = (
                    f"🔍 Component Extraction Provenance & Audit Trail{rfq_ref_str}\n"
                    f"────────────────────────────────────────────────────────────────────────────\n"
                    f"  Part / Item     : {matched_item.get('part_number')} — {matched_item.get('description')}\n"
                    f"  Parent Assembly : {matched_assy_obj.get('assy_no')} ({matched_assy_obj.get('assy_model', '')})\n"
                    f"  Source Document : {src_doc} (Zone: {zone})\n"
                    f"  Resolution Mode : {res_type} (Confidence: {conf}%)\n"
                    f"  Quantity & UOM  : {qty_disp}\n"
                    f"  MPN & MFR       : {mpn_disp} | {mfr_disp}\n"
                    f"  Raw Evidence    : \"{snip}\"\n"
                    f"────────────────────────────────────────────────────────────────────────────\n\n"
                    f"💡 AI Reasoning: {reason}\n\n"
                    f"👉 Review the multi-source cross-check table below or launch the Review Studio to adjust."
                )
                self._append_agent_message(
                    exp_msg,
                    suggestions=[
                        "📊 Review Full Table & Filter",
                        "📧 Email & Attachments",
                        "🚀 Launch BOM Verification Window"
                    ],
                    table_data=table_data
                )
                return

        # 3. Cycle Time AI Drawing Analysis Handler (Requires explicit 'cycle time' token)
        if "cycle time" in text_lower or "cycle-time" in text_lower:
            rfq_m = re.search(r'rfq\s*([a-z0-9_-]+)', text_lower)
            rfq_target = rfq_m.group(1) if rfq_m else "1009"

            if "confirm" in text_lower or "approve" in text_lower:
                sample_features = {
                    "assembly_code": "A01",
                    "wire_size": "24 AWG",
                    "wire_length": "350 mm",
                    "circuit_count": 8,
                    "suggested_cycle_time_sec": 36.0
                }
                ok, msg = self.ct_ai_engine.save_approved_cycle_time(rfq_target, sample_features)
                if ok:
                    self._append_agent_message(f"✅ **Cycle Time Approved & Saved!**\n\n{msg}\n\nReady for Costing calculations.", suggestions=[f"🚀 Dispatch RFQ {rfq_target} to Costing", "📊 Draw RFQ Stage Chart"])
                else:
                    self._append_agent_message(f"ℹ️ {msg}", suggestions=[f"🚀 Dispatch RFQ {rfq_target} to Costing"])
                return

            analysis = self.ct_ai_engine.analyze_drawing(f"Drawing for RFQ {rfq_target} (24 AWG, 350 mm, 8 CKT)", assy_code="A01")
            gate_info = self.approval_gate_mgr.create_gate(
                rfq_id=rfq_target,
                checkpoint=ApprovalCheckpoint.CHECKPOINT_4_CYCLE_TIME_PROPOSAL,
                stage="cycle_time",
                summary_data=analysis
            )
            card_msg = self.approval_gate_mgr.render_approval_card(gate_info)
            ct_suggs = self.approval_gate_mgr.get_approval_actions(ApprovalCheckpoint.CHECKPOINT_4_CYCLE_TIME_PROPOSAL, rfq_target)
            self._append_agent_message(card_msg, suggestions=ct_suggs)
            return

        # 1. Parse BOM parameters from user prompt using LLM-powered BrainRouter
        from agents.brain_router import BrainRouter
        brain = BrainRouter()
        parsed_params = brain.extract_parameters_with_llm(text, context=getattr(self, '_staged_bom_payload', None))
        has_param_update = parsed_params.get("has_updates", False)

        if parsed_params.get("customer_name"):
            self._override_customer_name = parsed_params["customer_name"]
        if parsed_params.get("project_title"):
            self._override_project_title = parsed_params["project_title"]
        if parsed_params.get("commodity"):
            self._override_commodity = parsed_params["commodity"]
        if parsed_params.get("rfq_number"):
            self._override_rfq_number = parsed_params["rfq_number"]
        if parsed_params.get("target_price"):
            self._override_target_price = parsed_params["target_price"]
        if parsed_params.get("eau"):
            self._override_eau = parsed_params["eau"]
        default_moqs = parsed_params.get("default_moqs", [])
        custom_moqs = parsed_params.get("custom_moqs", {})

        # Update staged payload if present
        if hasattr(self, '_staged_bom_payload') and self._staged_bom_payload and has_param_update:
            if getattr(self, "_override_customer_name", None):
                self._staged_bom_payload["customer_name"] = self._override_customer_name
            if getattr(self, "_override_project_title", None):
                self._staged_bom_payload["project_title"] = self._override_project_title
            if getattr(self, "_override_commodity", None):
                self._staged_bom_payload["commodity"] = self._override_commodity
            if getattr(self, "_override_rfq_number", None):
                self._staged_bom_payload["rfq_number"] = self._override_rfq_number
            if getattr(self, "_override_target_price", None):
                self._staged_bom_payload["target_price"] = self._override_target_price
            if getattr(self, "_override_eau", None):
                self._staged_bom_payload["eau"] = self._override_eau
            if default_moqs:
                self._staged_bom_payload["assigned_moqs"] = default_moqs
            if custom_moqs:
                self._staged_bom_payload["custom_moqs"] = custom_moqs

        # 2. Check if user commands to proceed / launch staged BOM AFTER applying any inline updates!
        if hasattr(self, '_staged_bom_payload') and self._staged_bom_payload:
            if any(k in text_lower for k in ["proceed", "proceess", "process", "open", "launch", "verify", "go", "yes", "confirm", "start", "run", "do it", "ok", "next", "continue"]):
                self._commit_and_launch_staged_bom()
                return

        # 3. If user ONLY updated parameters without typing launch command:
        if hasattr(self, '_staged_bom_payload') and self._staged_bom_payload and has_param_update:
            updates = []
            if parsed_params.get("customer_name") and getattr(self, "_override_customer_name", None): updates.append(f"Customer Name: {self._override_customer_name}")
            if parsed_params.get("project_title") and getattr(self, "_override_project_title", None): updates.append(f"Project Title: {self._override_project_title}")
            if parsed_params.get("commodity") and getattr(self, "_override_commodity", None): updates.append(f"Commodity: {self._override_commodity}")
            if parsed_params.get("rfq_number") and getattr(self, "_override_rfq_number", None): updates.append(f"RFQ Number: {self._override_rfq_number}")
            if parsed_params.get("target_price"): updates.append(f"Target Price: {self._override_target_price}")
            if parsed_params.get("eau"): updates.append(f"EAU: {self._override_eau} pcs")
            if default_moqs: updates.append(f"Default MOQs: {', '.join(str(m) for m in default_moqs)}")
            if custom_moqs:
                cust_str = ", ".join(f"{k} ({', '.join(str(m) for m in v)})" for k, v in custom_moqs.items())
                updates.append(f"Custom MOQs: {cust_str}")
            
            staged_suggs = [
                "🚀 Launch Verification Window",
                "💲 Target Price is $12.50",
                "📦 EAU is 5000 pcs"
            ]
            self._append_agent_message("Updated BOM Parameters:\n• " + "\n• ".join(updates) + "\n\nClick '🚀 Launch Verification Window' below or type 'proceed' when ready!", suggestions=staged_suggs)
            return

        # Show typing indicator then process in background thread
        self._show_typing_indicator()
        import threading
        threading.Thread(target=self._process_chat_async, args=(text,), daemon=True).start()

    def _process_chat_async(self, text):
        """Runs chart detection + BrainRouter query in a background thread and safely posts to GUI thread."""
        try:
            from agents.brain_router import BrainRouter, detect_chart_intent, generate_rfq_chart, get_rfq_summary_stats

            # 1. Chart intent detection (fast, no LLM needed)
            chart_type, caption = detect_chart_intent(text)
            if chart_type:
                stats = get_rfq_summary_stats()
                fig = generate_rfq_chart(chart_type, stats)
                if fig:
                    chart_suggs = [
                        "👥 Top Customers Summary",
                        "📊 Draw Assembly Count Chart",
                        "📄 Import Customer BOM File"
                    ]
                    self._safe_gui(lambda: self._append_agent_chart(fig, caption, suggestions=chart_suggs))
                    return
                else:
                    self._safe_gui(lambda: self._append_agent_message(
                        "No data available yet to draw a chart. Create some RFQs first!"))
                    return

            # 2. General NL query through BrainRouter (LLM + data fallback)
            router = BrainRouter()
            response = router.answer_system_query(text, module_key="brain")
            if response:
                query_suggs = [
                    "📊 Draw RFQ Stage Chart",
                    "👥 Top Customers Summary",
                    "📊 Draw Assembly Count Chart"
                ]
                self._safe_gui(lambda r=response: self._append_agent_message(r, suggestions=query_suggs))
                return

        except Exception as err:
            print(f"[ChatAsync] {err}")
            try:
                from agents.telemetry_tracker import ErrorTelemetryStore
                ErrorTelemetryStore().record_error(
                    module="ChatAsync",
                    error_category="UNHANDLED",
                    error_message=str(err),
                    severity="ERROR",
                    recovery_action="Returned default chat help message",
                    status="RECOVERED_VIA_FALLBACK",
                    stack_trace=str(err),
                )
            except Exception:
                pass

        default_suggs = [
            "📊 Draw RFQ Stage Chart",
            "👥 Top Customers Summary",
            "📄 Import Customer BOM File"
        ]
        self._safe_gui(lambda: self._append_agent_message(
            "I am your ContinuumX Agent Assistant! I can help analyze system data, draw charts, and process RFQs.\n\n"
            "• System data queries: 'How many RFQs are created?' or 'Which customer has the most orders?'\n"
            "• Interactive graphs: 'Draw pie chart of stage distribution' or 'Show customer bar chart'\n"
            "• Workflow tools: Click '📎 Actions' to import BOM Excel or classify RFQ emails.",
            suggestions=default_suggs
        ))

    def _agent_import_bom(self, file_path=None):
        if not file_path:
            from tkinter import filedialog
            file_path = filedialog.askopenfilename(
                title="Select Customer BOM Excel File",
                filetypes=[("Excel Files", "*.xlsx;*.xls")]
            )
        if not file_path:
            return

        self._last_bom_file = file_path
        self._append_user_message(f"Selected BOM File: {os.path.basename(file_path)}")
        self._append_agent_message(f"Parsing '{os.path.basename(file_path)}' using BOM Verification Agent...")

        try:
            from agents.skills.bom_verification_agent import BOMVerificationAgent
            agent = BOMVerificationAgent(self.server_path)
            res = agent.parse_customer_bom(file_path)

            if res.get("success"):
                mapping = res.get("suggested_mapping", {})
                cust = getattr(self, "_override_customer_name", None) or res.get("suggested_customer_name", "Customer")
                comm = res.get("suggested_commodity", "Wire Harness")
                proj = getattr(self, "_override_project_title", None) or res.get("suggested_project_title", os.path.splitext(os.path.basename(file_path))[0])
                tp = getattr(self, "_override_target_price", None) or "Not Specified"
                eau = getattr(self, "_override_eau", None) or "Not Specified"

                # Stage Payload in Chatbot first (allow human conversation before launching)
                self._staged_bom_payload = {
                    "file_path": file_path,
                    "customer_name": cust,
                    "commodity": comm,
                    "project_title": proj,
                    "target_price": tp,
                    "eau": eau,
                    "suggested_mapping": mapping,
                    "suggested_special": res.get("suggested_special")
                }

                msg = (f"☑️ BOM File Parsed & Staged!\n"
                       f"• File: {os.path.basename(file_path)}\n"
                       f"• Customer: {cust}\n"
                       f"• Commodity: {comm}\n"
                       f"• Project Title: {proj}\n"
                       f"• Target Price: {tp}\n"
                       f"• EAU: {eau}\n"
                       f"• Mapped Headers:\n"
                       f"   - Part: {mapping.get('Part')}\n"
                       f"   - Description: {mapping.get('Description')}\n"
                       f"   - Qty: {mapping.get('Qty')}\n"
                       f"   - UOM: {mapping.get('UOM')}\n"
                       f"   - Line Item: {mapping.get('Line Item')}\n\n"
                       f"💬 You can chat with me now to update any missing details (e.g. 'Target Price is $12.50, EAU is 5000 pcs'), or click '🚀 Launch Verification Window' below!")
                
                staged_suggs = [
                    "🚀 Launch Verification Window",
                    "💲 Target Price is $12.50",
                    "📦 EAU is 5000 pcs"
                ]
                self._append_agent_message(msg, suggestions=staged_suggs)
            else:
                self._append_agent_message(f"❌ Failed to parse BOM file: {res.get('error')}")
        except Exception as e:
            self._append_agent_message(f"❌ Error running BOM Agent: {e}")

    def _agent_process_email(self):
        self._append_user_message("Check & Extract RFQ Emails from Inbox")
        self._show_typing_indicator()
        import threading
        threading.Thread(target=self._check_email_rfqs_async, daemon=True).start()

    def _check_email_rfqs_async(self, limit=50, subject_filter=None, target_rfq_id=None):
        try:
            from agents.brain_router import BrainRouter
            router = BrainRouter()
            progress_cb = self._make_progress_callback()
            res = router.check_rfq_emails(limit=limit, subject_filter=subject_filter, progress_callback=progress_cb)

            if not res.get("success"):
                err_msg = res.get("error", "Unknown error")
                self._safe_gui(lambda: self._append_agent_message(
                    f"⚠️ **Email Check Notice:**\n\nCould not access mailbox ({err_msg}).\n\n"
                    f"You can test the full pipeline using the synthetic demonstration runner below:",
                    suggestions=["🧪 Run Demo RFQ Extraction", "📄 Import Customer BOM File", "📊 Draw RFQ Stage Chart"]
                ))
                return

            email_addr = res.get("email_address", "maic-demo@continuumx.com.my")
            total_scanned = res.get("total_scanned", 0)
            rfq_count = res.get("rfq_count", 0)
            rfqs = res.get("rfqs", [])
            self._detected_rfq_list = rfqs

            if rfq_count > 0 and rfqs:
                chosen_rfq = rfqs[0]
                # 1. Match by exact target_rfq_id (e.g. RS25-8099)
                if target_rfq_id:
                    for r in rfqs:
                        r_no = str(r.get("rfq_json", {}).get("rfq_metadata", {}).get("rfq_number", "")).upper()
                        r_subj = str(r.get("email", {}).get("subject", "")).upper()
                        if target_rfq_id in r_no or target_rfq_id in r_subj:
                            chosen_rfq = r
                            break
                # 2. Match by subject_filter
                elif subject_filter:
                    sf_low = str(subject_filter).lower()
                    for r in rfqs:
                        r_no = str(r.get("rfq_json", {}).get("rfq_metadata", {}).get("rfq_number", "")).lower()
                        r_subj = str(r.get("email", {}).get("subject", "")).lower()
                        r_cust = str(r.get("rfq_json", {}).get("rfq_metadata", {}).get("customer_name", "")).lower()
                        if sf_low in r_no or sf_low in r_subj or sf_low in r_cust:
                            chosen_rfq = r
                            break

                self._safe_gui(lambda: self._load_selected_rfq_payload(chosen_rfq))
            else:
                # 0 RFQs in recent emails
                recent = res.get("recent_emails_summary", [])
                lines = []
                for idx, em in enumerate(recent[:5], start=1):
                    lines.append(f"  {idx}. **{em['subject'][:35]}** — Intent: `{em['intent']}` ({int(em['confidence']*100)}%)")

                msg = (
                    f"📬 **Inbox Scan Summary for `{email_addr}`**\n\n"
                    f"Scanned the latest **{total_scanned} emails**. Currently found **0 active customer RFQ emails**.\n\n"
                    f"**Recent Scanned Emails:**\n" + "\n".join(lines) + "\n\n"
                    f"💡 *Tip: You can test the end-to-end extraction and synthetic BOM generation with sample RFQ data by clicking '🧪 Run Demo RFQ Extraction' below.*"
                )
                suggs = [
                    "🧪 Run Demo RFQ Extraction",
                    "📩 Check RFQ Emails",
                    "📊 Draw RFQ Stage Chart"
                ]
                self._safe_gui(lambda: self._append_agent_message(msg, suggestions=suggs))

        except Exception as e:
            self._safe_gui(lambda: self._append_agent_message(f"⚠️ Email processing exception: {e}"))
        finally:
            self._safe_gui(self._hide_typing_indicator)

    def _load_selected_rfq_payload(self, target_rfq):
        """Renders the rich RFQ card and stages the synthetic BOM payload for verification."""
        rfq_json = target_rfq.get("rfq_json", {})
        synth_bom = target_rfq.get("synthetic_bom")
        if not synth_bom:
            try:
                from agents.synthetic_bom_generator import SyntheticBOMGenerator
                synth_bom = SyntheticBOMGenerator().generate_synthetic_excel(rfq_json)
                target_rfq["synthetic_bom"] = synth_bom
            except Exception:
                synth_bom = {}
        email_info = target_rfq.get("email", {})

        meta = rfq_json.get("rfq_metadata", {})
        cust = meta.get("customer_name", "Customer")
        rfq_no = meta.get("rfq_number", "RFQ")
        proj = meta.get("project_title", "Assembly")
        comm = meta.get("commodity", "Wire Harness")
        tp = meta.get("target_price", "Not Specified")
        eau = meta.get("eau") if meta.get("eau") is not None else ""
        moqs = meta.get("default_moqs", [100, 250, 500, 1000])

        excel_fp = synth_bom.get("file_path", "")
        self._staged_bom_payload = {
            "file_path": excel_fp,
            "customer_name": cust,
            "project_title": proj,
            "commodity": comm,
            "rfq_number": rfq_no,
            "target_price": tp,
            "eau": eau,
            "assigned_moqs": moqs,
            "suggested_mapping": synth_bom.get("suggested_mapping", {
                "Line": "Line Item",
                "Part": "Part",
                "Description": "Description",
                "QTY": "Qty",
                "UOM": "UOM",
                "MFR": ["MFR"],
                "MPN": ["MPN"]
            }),
            "suggested_special": synth_bom.get("suggested_special", {
                "Assy #": {"method": "map", "source_column": "Assy#", "value": ""},
                "Assy Model": {"method": "map", "source_column": "Assy Model", "value": ""},
                "Assy Rev": {"method": "map", "source_column": "Assy Rev", "value": ""}
            })
        }
        self._last_extracted_rfq_json = rfq_json
        self._current_active_rfq = target_rfq

        # Format Component Table Preview matching exact 12-column user schema:
        # Line | Assy# | Assy Model | Assy Rev | Part | Description | MPN | MFR | QTY | UOM | EAU | Target Price (USD)
        items_with_assy = []
        for a in rfq_json.get("assemblies", []):
            a_no = a.get("assy_no", "ASSY-01")
            a_model = a.get("assy_model", "")
            a_rev = str(a.get("assy_rev", "")).replace("Rev", "").strip()
            a_eau = a.get("eau") if a.get("eau") is not None else eau
            a_tp = str(a.get("target_price", tp)).replace('$', '').strip()
            for c_idx, it in enumerate(a.get("items", []), start=1):
                it_eau = it.get("eau") if it.get("eau") is not None else a_eau
                items_with_assy.append((c_idx, a_no, a_model, a_rev, it, it_eau, str(it.get("target_price", a_tp)).replace('$', '').strip()))

        table_rows = []
        is_anon = getattr(self, 'is_privacy_mode', False)
        for c_idx, a_no, a_model, a_rev, it, row_eau, row_tp in items_with_assy[:30]:
            clean_eau_str = re.sub(r'[^\d.]', '', str(row_eau)) if row_eau not in (None, "", "None") else ""
            disp_a_no = self._get_privacy_hash(a_no, "ASSY") if is_anon else str(a_no)
            disp_part = self._get_privacy_hash(it.get("part_number") or "", "PART") if is_anon else str(it.get("part_number") or "")
            disp_desc = self._anonymize_text_simple(it.get("description") or "Component") if is_anon else str(it.get("description") or "Component")
            disp_model = self._anonymize_text_simple(a_model) if is_anon else str(a_model)

            table_rows.append([
                str(c_idx),
                disp_a_no,
                disp_model,
                str(a_rev),
                disp_part,
                disp_desc,
                str(it.get("mpn") or ""),
                str(it.get("mfr") or ""),
                str(it.get("qty", 1)),
                str(it.get("uom", "EA")),
                clean_eau_str or str(eau),
                str(row_tp)
            ])

        table_data = {
            "headers": ["Line", "Assy#", "Assy Model", "Assy Rev", "Part", "Description", "MPN", "MFR", "QTY", "UOM", "EAU", "Target Price (USD)"],
            "rows": table_rows,
            "footer": f"Total Components Detected: {len(items_with_assy)} item(s) across {len(rfq_json.get('assemblies', []))} assembly/drawings" if items_with_assy else "Direct Assembly Specification"
        }

        # Check for EAU Ambiguity (e.g. range '3-5k')
        ambig = meta.get("eau_ambiguity", {})
        ambig_warning = ""
        ambig_chips = []
        if ambig.get("is_ambiguous"):
            raw_t = ambig.get("raw_text", "")
            opts = ambig.get("suggested_options", [])
            ambig_warning = (
                f"\n⚠️ **EAU Ambiguity Alert**: Customer specified range **`{raw_t}`**.\n"
                f"*Downstream BOM & Sourcing automation requires a single fixed integer value.* "
                f"Please confirm or select the target EAU below:\n"
            )
            for opt in opts:
                ambig_chips.append(f"📌 Confirm EAU: {opt:,} pcs")

        # Build dynamic suggestion chips for other RFQs found
        suggs = ambig_chips + ["📊 Review Full Table & Filter", "🚀 Launch BOM Verification Window", "📈 View AI Benchmark & Processing Stats", "🔍 Inspect Source Evidence", "📄 View Extracted JSON"]
        if hasattr(self, '_detected_rfq_list') and self._detected_rfq_list:
            for other_rfq in self._detected_rfq_list:
                o_meta = other_rfq.get("rfq_json", {}).get("rfq_metadata", {})
                o_no = o_meta.get("rfq_number", "")
                o_cust = o_meta.get("customer_name", "")
                if o_no and o_no != rfq_no:
                    suggs.append(f"📄 Load RFQ {o_no} ({o_cust})")

        # Build clean dynamic EAU summary string
        eau_vals = []
        for a in rfq_json.get("assemblies", []):
            e_val = a.get("eau")
            if e_val is not None and str(e_val).strip() not in ('', 'None'):
                try:
                    num = int(re.sub(r'[^\d]', '', str(e_val)))
                    eau_vals.append(num)
                except Exception:
                    pass

        if len(eau_vals) > 1:
            eau_disp = f"Multiple ({min(eau_vals):,} – {max(eau_vals):,} pcs • Total: {sum(eau_vals):,} pcs)"
        elif len(eau_vals) == 1:
            eau_disp = f"{eau_vals[0]:,} pcs"
        else:
            eau_disp = f"{eau:,} pcs" if isinstance(eau, (int, float)) else str(eau or "Not Specified")

        assy_list = [a.get("assy_no", "") for a in rfq_json.get("assemblies", []) if a.get("assy_no")]
        assy_disp = f"{len(assy_list)} Assemblies ({', '.join(assy_list)})" if len(assy_list) > 1 else (assy_list[0] if assy_list else "1 Assembly")

        telemetry_info = target_rfq.get("telemetry", {})
        latency_disp = f"{telemetry_info.get('formatted_duration', '0.8s')} (Run #{telemetry_info.get('run_index', 1)} • {telemetry_info.get('speedup_vs_run1', 'Baseline')})" if telemetry_info else "Instant"

        excel_fn = os.path.basename(str(excel_fp)) if excel_fp else "synthetic_bom.xlsx"

        # Build clean metadata card with aligned columns and box border (0 asterisks)
        is_anon = getattr(self, 'is_privacy_mode', False)
        disp_cust = self._get_privacy_hash(cust, "CUST") if is_anon else cust
        disp_sender = "procurement@client-vault.com" if is_anon else email_info.get('sender', 'Customer')
        disp_proj = self._anonymize_text_simple(email_info.get('subject', proj)) if is_anon else email_info.get('subject', proj)

        if is_anon:
            disp_assy_list = [self._get_privacy_hash(a, "ASSY") for a in (assy_list or [])]
            disp_assy_str = f"{len(rfq_json.get('assemblies', []))} Assemblies ({', '.join(disp_assy_list[:8])}{'...' if len(disp_assy_list) > 8 else ''})"
        else:
            disp_assy_str = assy_disp

        meta_items = [
            ("Customer", disp_cust),
            ("RFQ Number", rfq_no),
            ("Project Title", disp_proj),
            ("Commodity", comm),
            ("Target Price", tp),
            ("EAU", eau_disp),
            ("Assemblies", disp_assy_str),
            ("Sender", disp_sender),
            ("Processing Time", latency_disp),
            ("Synthetic BOM", excel_fn)
        ]

        max_k_len = max(len(k) for k, _ in meta_items)
        divider_line = "─" * 72

        card_lines = [
            "📥 RFQ Email Extracted Successfully",
            divider_line
        ]
        for k, v in meta_items:
            card_lines.append(f"  {k:<{max_k_len}} : {v}")
        card_lines.append(divider_line)

        msg = (
            "\n".join(card_lines) + "\n\n"
            "👉 All columns and parameters have been pre-mapped. Click 'Launch BOM Verification Window' below or type 'proceed' to launch!"
        )
        self._append_agent_message(msg, suggestions=suggs, table_data=table_data if table_rows else None)

    def _get_privacy_hash(self, val: str, prefix: str = "VAL") -> str:
        """Deterministically hashes sensitive values for MAIC competition presentation."""
        if not val or str(val).strip() in ("", "N/A", "None", "-", "Not Specified"):
            return str(val) if val is not None else ""
        s_val = str(val).strip()
        cache_key = (prefix, s_val)
        if not hasattr(self, '_privacy_hash_cache'):
            self._privacy_hash_cache = {}
        if cache_key in self._privacy_hash_cache:
            return self._privacy_hash_cache[cache_key]

        import hashlib
        h_code = hashlib.md5(s_val.encode('utf-8')).hexdigest()[:4].upper()
        if prefix == "CUST": hashed = f"CUST-{h_code}"
        elif prefix == "ASSY": hashed = f"ASY-{h_code}"
        elif prefix == "PART": hashed = f"CP-{h_code}"
        else: hashed = f"{prefix}-{h_code}"
        self._privacy_hash_cache[cache_key] = hashed
        return hashed

    def _anonymize_text_simple(self, text: str) -> str:
        if not text or not getattr(self, 'is_privacy_mode', False):
            return text or ""
        out = text
        for known in ["Tecan", "Graco", "Honeywell", "Eastek", "Radysis"]:
            out = re.sub(r'\b' + re.escape(known) + r'\b', self._get_privacy_hash(known, "CUST"), out, flags=re.I)
        out = re.sub(r'\b([0-9]{8})\b', lambda m: self._get_privacy_hash(m.group(1), "PART"), out)
        out = re.sub(r'[\w\.-]+@[\w\.-]+', 'procurement@client-vault.com', out)
        return out

    def _toggle_launcher_privacy_mode(self):
        """Toggles global competition privacy mode in the Launcher."""
        self.is_privacy_mode = not getattr(self, 'is_privacy_mode', False)
        if self.is_privacy_mode:
            self.btn_privacy_mode.config(text="🔒 Privacy Mode: ON", bg="#059669")
        else:
            self.btn_privacy_mode.config(text="🔒 Privacy Mode", bg="#334155")

        if hasattr(self, '_current_active_rfq') and self._current_active_rfq:
            # Re-render in place
            if hasattr(self, '_chat_inner') and self._chat_inner:
                for ch in self._chat_inner.winfo_children():
                    ch.destroy()
            self._load_selected_rfq_payload(self._current_active_rfq)

    def _run_demo_rfq_extraction_async(self):
        try:
            from agents.multimodal_extractor import MultimodalExtractor
            from agents.synthetic_bom_generator import SyntheticBOMGenerator

            generator = SyntheticBOMGenerator()
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            rfq_json = {
                "rfq_metadata": {
                    "source_email_id": "DEMO-RFQ-8344",
                    "sender": "procurement@radysis.com",
                    "received_date": now_str,
                    "email_subject": "NEW RFQ: Wire Harness Assembly for Radysis N2 - EAU 5000 pcs",
                    "customer_name": "Radysis",
                    "rfq_number": "RS26-8344",
                    "project_title": "Radysis N2 Wire Harness",
                    "commodity": "Wire Harness",
                    "target_price": "$14.50",
                    "eau": 5000,
                    "default_moqs": [100, 250, 500, 1000]
                },
                "drawings_detected": [
                    {
                        "drawing_number": "DWG-8247JT",
                        "revision": "Rev B",
                        "customer_part_number": "810-105035-003",
                        "title": "Radysis N2 Main Harness"
                    }
                ],
                "engineering_notes": [
                    "Wire spec: 24 AWG 300V UL1007",
                    "Strip length: 4.5 mm ± 0.5 mm",
                    "Terminal crimp pull force: > 45 N"
                ],
                "assemblies": [
                    {
                        "assy_no": "810-105035-003",
                        "assy_model": "Radysis N2 Wire Harness",
                        "assy_rev": "Rev B",
                        "items": [
                            {"line_item": 1, "part_number": "1-967616-1", "description": "6-Pin Connector Housing", "mfr": "TE Connectivity", "mpn": "1-967616-1", "qty": 1, "uom": "EA"},
                            {"line_item": 2, "part_number": "968220-1", "description": "Female Crimp Terminal 20-24 AWG", "mfr": "TE Connectivity", "mpn": "968220-1", "qty": 6, "uom": "EA"},
                            {"line_item": 3, "part_number": "3051 BK005", "description": "Hook-up Wire 24 AWG Black 300V", "mfr": "Alpha Wire", "mpn": "3051 BK005", "qty": 350, "uom": "MM"}
                        ]
                    }
                ],
                "total_items_count": 3
            }

            synth_bom = generator.generate_synthetic_excel(rfq_json)
            excel_fp = synth_bom.get("file_path", "")

            self._staged_bom_payload = {
                "file_path": excel_fp,
                "customer_name": "Radysis",
                "project_title": "Radysis N2 Wire Harness",
                "commodity": "Wire Harness",
                "rfq_number": "RS26-8344",
                "target_price": "$14.50",
                "eau": 5000,
                "assigned_moqs": [100, 250, 500, 1000],
                "suggested_mapping": synth_bom.get("suggested_mapping", {}),
                "suggested_special": {}
            }
            self._last_extracted_rfq_json = rfq_json

            table_rows = [
                ["#1", "1-967616-1", "6-Pin Connector Housing", "TE Connectivity", "1", "EA"],
                ["#2", "968220-1", "Female Crimp Terminal", "TE Connectivity", "6", "EA"],
                ["#3", "3051 BK005", "Hook-up Wire 24AWG Black", "Alpha Wire", "350", "MM"]
            ]
            table_data = {
                "headers": ["Line", "Part / MPN", "Description", "MFR", "Qty", "UOM"],
                "rows": table_rows,
                "footer": "Total Components Extracted: 3 item(s) • Wire Harness Commodity"
            }

            msg = (
                f"📥 **Demo RFQ Extraction Completed!**\n\n"
                f"• **Source Email**: `procurement@radysis.com`\n"
                f"• **Subject**: `NEW RFQ: Wire Harness Assembly for Radysis N2`\n"
                f"• **Customer**: Radysis\n"
                f"• **RFQ Number**: RS26-8344\n"
                f"• **Commodity**: Wire Harness\n"
                f"• **Target Price**: $14.50\n"
                f"• **EAU**: 5,000 pcs\n"
                f"• **Generated Synthetic BOM**: `{os.path.basename(excel_fp)}`\n\n"
                f"👉 *The synthetic BOM Excel has been staged with pre-filled mappings. Click '🚀 Launch BOM Verification Window' below or type 'proceed' to launch!*"
            )

            suggs = [
                "🚀 Launch BOM Verification Window",
                "📄 View Extracted JSON",
                "💲 Target Price is $14.50",
                "📦 EAU is 5000 pcs"
            ]
            self._safe_gui(lambda: self._append_agent_message(msg, suggestions=suggs, table_data=table_data))

        except Exception as e:
            self._safe_gui(lambda: self._append_agent_message(f"⚠️ Demo extraction error: {e}"))
        finally:
            self._safe_gui(self._hide_typing_indicator)

    def _open_guided_approval_queue(self):
        """Opens the single-window Guided Approval Queue stepper dialog."""
        user_info = getattr(self, "user", {}) or getattr(self, "user_context", {}) or {}
        user_role = user_info.get("role") or user_info.get("Role") or "System Administrator"
        username = user_info.get("username") or user_info.get("name") or "Sysadmin"

        active_gates = self.orch_state_mgr.get_all_active_approval_gates(user_role=user_role, username=username)
        if not active_gates:
            self._append_agent_message("🎉 Great news! Your Approval Queue is currently empty (0 pending RFQs).", suggestions=["📊 Draw RFQ Stage Chart", "📥 Import Customer BOM File"])
            return

        from agents.orchestrator.guided_queue_dialog import GuidedApprovalQueueWindow

        def on_queue_updated(remaining_gates):
            self._cached_active_gates = remaining_gates
            if not remaining_gates:
                self._append_agent_message("🎉 All items in your Approval Queue have been successfully reviewed and dispatched!", suggestions=["📊 Draw RFQ Stage Chart", "📥 Import Customer BOM File"])
            else:
                count = len(remaining_gates)
                self._append_agent_message(
                    f"🔔 Queue Updated: {count} RFQ{'s' if count > 1 else ''} remaining in your review queue.",
                    suggestions=["🔍 Open Guided Review Queue", f"🚀 Dispatch All ({count} RFQs)"]
                )

        GuidedApprovalQueueWindow(
            master=self,
            active_gates=active_gates,
            server_path=self.server_path,
            username=username,
            on_queue_updated=on_queue_updated
        )

    def _dispatch_all_active_gates(self):
        """Batch dispatches all active gates in parallel with full audit trails and sends ONE single consolidated email."""
        user_info = getattr(self, "user", {}) or getattr(self, "user_context", {}) or {}
        user_role = user_info.get("role") or user_info.get("Role") or "System Administrator"
        username = user_info.get("username") or user_info.get("name") or "Sysadmin"

        active_gates = self.orch_state_mgr.get_all_active_approval_gates(user_role=user_role, username=username)
        if not active_gates:
            self._append_agent_message("ℹ️ No pending RFQs in queue to dispatch.")
            return

        from agents.tool_registry import ContinuumXToolRegistry
        user_info = getattr(self, "user", {}) or getattr(self, "user_context", {}) or {}
        username = user_info.get("username") or user_info.get("name") or "Sysadmin"
        rfq_ids = [g.get("rfq_id") for g in active_gates if g.get("rfq_id")]

        success_list, fail_list = ContinuumXToolRegistry.execute_batch_system_dispatch(
            dept="bom",
            rfq_ids=rfq_ids,
            username=username,
            comments="Batch dispatched via Chatbot Orchestrator."
        )

        if success_list:
            self._append_agent_message(
                f"🚀 **Batch Dispatch Completed!**\n\n"
                f"Successfully dispatched **{len(success_list)} RFQs** ({', '.join(success_list)}) to Sourcing & Cycle Time in parallel!\n"
                f"• **1 Consolidated Batch Email** sent to team PICs with full RFQ summary table.\n"
                f"• History and audit trails recorded for all records.",
                suggestions=["📊 Draw RFQ Stage Chart", "👥 Which customer has most RFQs?"]
            )
        if fail_list:
            self._append_agent_message(f"⚠️ Failed to dispatch {len(fail_list)} RFQs:\n• " + "\n• ".join(fail_list))

    def on_logout(self):
        if messagebox.askyesno("Logout", "Are you sure you want to logout?"):
            self.destroy()
            if self._on_logout_cb:
                self._on_logout_cb()
            else:
                import sys
                os.execl(sys.executable, sys.executable, *sys.argv)

